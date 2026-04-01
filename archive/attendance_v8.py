
import winreg
import win32com.client
import pickle
import logging
import psutil
import os
import sys
import time
import socket
import json
import threading
from datetime import datetime, timedelta
from pathlib import Path            
import faiss




def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    # Create logs directory in user's home folder
    logs_dir = os.path.join(os.path.expanduser('~'), '.attendance_logs')
    os.makedirs(logs_dir, exist_ok=True)
    
    if relative_path.endswith('.log'):
        # Return log file path in logs directory
        return os.path.join(logs_dir, os.path.basename(relative_path))
    
    # For other resources
    return os.path.join(base_path, relative_path)



class SingleInstanceManager:
    def __init__(self, port=12345):
        self.port = port
        self.lock_file = Path.home() / ".attendance_lock"
        self.attendance_file = Path.home() / ".last_attendance"
        self.sock = None
        
        # Add debug logging file
        self.log_file = Path.home() / "attendance_debug.log"
        self.setup_logging()

    def setup_logging(self):
        """Setup detailed logging"""
        logging.basicConfig(
            filename=str(self.log_file),
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )

    def check_instance(self):
        """Check if another instance is running and manage attendance cooldown"""
        try:
            # Log startup type
            is_manual = '--priority' not in sys.argv
            logging.debug(f"Application starting - Manual launch: {is_manual}")
            
            # Try to create a socket server
            self.sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
            
            # Log current PID and port
            logging.debug(f"Attempting to bind port {self.port} for PID {os.getpid()}")
            
            self.sock.bind(('localhost', self.port))
            logging.debug("Successfully bound to port - No other instance running")
            
            # Check attendance cooldown
            if self.check_attendance_cooldown():
                logging.debug("Within attendance cooldown period - preventing new instance")
                self.show_debug_message("Application blocked: Recent attendance detected")
                return False
                
            # Create lock file
            self.create_lock_file()
            return True
            
        except socket.error as e:
            logging.debug(f"Port binding failed - Another instance likely running: {e}")
            self.show_debug_message("Application blocked: Another instance is running")
            return False

    def check_attendance_cooldown(self):
        """Check if attendance was recorded within last 1 minute"""
        try:
            if not self.attendance_file.exists():
                return False
                
            with open(self.attendance_file, 'r') as f:
                data = json.load(f)
                last_attendance = datetime.fromisoformat(data['last_attendance'])
                
            # Check if within 5 minutes
            return datetime.now() - last_attendance < timedelta(minutes=1)
            
        except Exception as e:
            logging.error(f"Error checking attendance cooldown: {e}")
            return False

    def show_debug_message(self, message):
        """Show debug message to user"""
        try:
            import tkinter as tk
            from tkinter import messagebox
            
            root = tk.Tk()
            root.withdraw()  # Hide the main window
            
            # Show debug info
            debug_info = (
                f"{message}\n\n"
                f"Current PID: {os.getpid()}\n"
                f"Lock file exists: {self.lock_file.exists()}\n"
                f"Last attendance file exists: {self.attendance_file.exists()}"
            )
            
            messagebox.showinfo("Debug Info", debug_info)
            root.destroy()
            
        except Exception as e:
            logging.error(f"Error showing debug message: {e}")

    def create_lock_file(self):
        """Create a lock file with current process info"""
        try:
            data = {
                'pid': os.getpid(),
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'manual_launch': '--priority' not in sys.argv
            }
            with open(self.lock_file, 'w') as f:
                json.dump(data, f)
        except Exception as e:
            logging.error(f"Error creating lock file: {e}")

    def record_attendance(self):
        """Record timestamp of successful attendance"""
        try:
            data = {
                'last_attendance': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'pid': os.getpid()
            }
            with open(self.attendance_file, 'w') as f:
                json.dump(data, f)
            
            logging.debug(f"Recorded attendance at {data['last_attendance']}")
            
            # Show confirmation
            self.show_debug_message(
                "Attendance recorded successfully!\n"
                "System will block new instances for 5 minutes."
            )
            
        except Exception as e:
            logging.error(f"Error recording attendance time: {e}")

    def dump_system_state(self):
        """Dump complete system state for debugging"""
        try:
            state = {
                'pid': os.getpid(),
                'executable': sys.executable,
                'port': self.port,
                'lock_exists': self.lock_file.exists(),
                'attendance_exists': self.attendance_file.exists(),
                'running_instances': self.check_running_processes(),
                'command_line': sys.argv,
                'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            }
            
            debug_path = Path.home() / 'attendance_state.json'
            with open(debug_path, 'w') as f:
                json.dump(state, f, indent=2)
            
            logging.debug(f"System state dumped to {debug_path}")
            
        except Exception as e:
            logging.error(f"Error dumping system state: {e}")

    def check_running_processes(self):
        """Debug function to list all running instances"""
        try:
            import psutil
            current_name = Path(sys.executable).name
            instances = []
            
            for proc in psutil.process_iter(['pid', 'name', 'create_time']):
                try:
                    if proc.info['name'] == current_name:
                        instances.append({
                            'pid': proc.info['pid'],
                            'created': datetime.fromtimestamp(proc.info['create_time']).strftime('%Y-%m-%d %H:%M:%S')
                        })
                except (psutil.NoSuchProcess, psutil.AccessDenied):
                    pass
                    
            logging.debug(f"Found instances: {instances}")
            return instances
            
        except Exception as e:
            logging.error(f"Error checking processes: {e}")
            return []

    def cleanup(self):
        """Enhanced cleanup with error handling"""
        try:
            logging.debug("Starting cleanup process")
            
            if self.sock:
                try:
                    self.sock.close()
                    logging.debug("Socket closed successfully")
                except Exception as e:
                    logging.error(f"Error closing socket: {e}")
            
            if self.lock_file.exists():
                try:
                    self.lock_file.unlink()
                    logging.debug("Lock file removed successfully")
                except Exception as e:
                    logging.error(f"Error removing lock file: {e}")
                    
        except Exception as e:
            logging.error(f"Error during cleanup: {e}")
       
class StartupAndAdminManager:
    def __init__(self):
        self.app_name = "FaceRecognitionAttendance"
        self.exe_path = sys.executable
        self.settings_file = Path.home() / ".startup_settings"
        self.registry_key_path = "Software\\Microsoft\\Windows\\CurrentVersion\\Run"

    def initialize(self):
        """Initialize if settings aren't configured"""
        if not self.check_settings_configured():
            success, message = self.add_to_registry()
            if success:
                self.setup_startup()
                self.save_settings_status(True)
            else:
                logging.error(f"Registry setup failed: {message}")

    def check_settings_configured(self):
        """Check if startup settings are already configured"""
        try:
            if self.settings_file.exists():
                with open(self.settings_file, 'rb') as f:
                    if pickle.load(f):
                        registry_status = self.check_registry_entry()
                        shortcut_status = self.verify_startup_shortcut()
                        return "Missing" not in registry_status and shortcut_status
            return False
        except Exception as e:
            logging.error(f"Settings check error: {e}")
            return False

    def add_to_registry(self):
        """Add to user's registry RUN key (no admin needed)"""
        try:
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                self.registry_key_path,
                0,
                winreg.KEY_SET_VALUE
            )
            
            winreg.SetValueEx(
                key,
                self.app_name,
                0,
                winreg.REG_SZ,
                f'"{self.exe_path}" --priority high'
            )
            
            winreg.CloseKey(key)
            return True, "Added to user registry"
        except Exception as e:
            return False, f"Registry error: {str(e)}"

    def check_registry_entry(self):
        """Verify current registry entry"""
        try:
            key = winreg.OpenKey(
                winreg.HKEY_CURRENT_USER,
                self.registry_key_path,
                0,
                winreg.KEY_READ
            )
            
            try:
                value, _ = winreg.QueryValueEx(key, self.app_name)
                return f"Registry entry: Exists ({os.path.basename(value)})"
            except FileNotFoundError:
                return "Registry entry: Missing"
            finally:
                winreg.CloseKey(key)
        except Exception as e:
            return f"Registry check failed: {str(e)}"

    def verify_startup_shortcut(self):
        """Verify startup shortcut exists and is correct"""
        try:
            shell = win32com.client.Dispatch("WScript.Shell")
            startup_folder = shell.SpecialFolders("Startup")
            shortcut_path = os.path.join(startup_folder, f"{self.app_name}.lnk")
            
            if os.path.exists(shortcut_path):
                shortcut = shell.CreateShortCut(shortcut_path)
                return (shortcut.Targetpath == self.exe_path and 
                        shortcut.Arguments == "--priority high")
            return False
        except Exception as e:
            logging.error(f"Shortcut verification error: {e}")
            return False

    def save_settings_status(self, status):
        """Save startup settings status"""
        try:
            with open(self.settings_file, 'wb') as f:
                pickle.dump(status, f)
        except Exception as e:
            logging.error(f"Error saving settings status: {e}")

    def setup_startup(self):
        """Configure startup methods"""
        try:
            # Create startup shortcut if it doesn't exist
            if not self.verify_startup_shortcut():
                shell = win32com.client.Dispatch("WScript.Shell")
                startup_folder = shell.SpecialFolders("Startup")
                shortcut_path = os.path.join(startup_folder, f"{self.app_name}.lnk")
                shortcut = shell.CreateShortCut(shortcut_path)
                shortcut.Targetpath = self.exe_path
                shortcut.Arguments = "--priority high"
                shortcut.WindowStyle = 1
                shortcut.save()

            # Set process priority
            import psutil
            process = psutil.Process()
            process.nice(psutil.HIGH_PRIORITY_CLASS)

        except Exception as e:
            logging.error(f"Startup configuration error: {e}")

            

from tkinter import ttk
import shutil
from tkinter import filedialog

import tkinter as tk
from tkinter import font as tkfont

import tkinter.messagebox as messagebox
import tkinter.simpledialog as simpledialog
import tkinter.ttk as ttk

import cv2
import numpy as np
from sklearn.preprocessing import normalize
import os
import time
import faiss
import tkinter as tk
from tkinter import messagebox, simpledialog
import pandas as pd
from collections import defaultdict
import csv
from datetime import datetime, timedelta, time as datetime_time
import threading
import sqlite3
import pytz
import boto3
import logging
import json
from queue import Queue
from PIL import Image, ImageTk
from typing import Optional, Tuple, Dict
from config import aws_access_key, aws_secret_key, rds_host, rds_password,  rds_database, rds_user, rds_port
import ntplib
import psycopg2
from psycopg2 import pool
import socket
import uuid
from pathlib import Path

import filecmp
from threading import Lock 
import traceback
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from dateutil import parser
import sys
import asyncio

import random
import math

import threading
import subprocess
from http.server import HTTPServer, BaseHTTPRequestHandler
import qrcode
import requests


logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)



def validate_filepath(filepath):
    """Validate the file path for Windows."""
    try:
        # Normalize the path to handle Windows backslashes
        filepath = os.path.normpath(filepath)
        
        # Check path length (Windows limit)
        if len(filepath) > 260:
            raise ValueError("File path exceeds maximum length of 260 characters.")
            
        # Get the filename only
        filename = os.path.basename(filepath)
        
        # Check for invalid characters in filename only
        invalid_chars = '<>:"|?*'
        if any(char in filename for char in invalid_chars):
            raise ValueError("Filename contains invalid characters.")
            
        # Create directory if it doesn't exist
        directory = os.path.dirname(filepath)
        if not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
            
        return True
        
    except Exception as e:
        logging.error(f"Path validation error: {e}")
        raise ValueError(f"Invalid file path: {str(e)}")
    
    



class FaceRecognitionSystem:
    def __init__(self, main_window=None, db_pool=None, app=None):

        self.username = None  # Initialize username
        self.main_window = main_window  # Store main_window
        
        if main_window:
            self.root = main_window
        else:
            self.root = tk.Tk()
         
        self.root.title("Face Recognition System")
        self.root.geometry("600x430")
        self.index = None
        self.user_embeddings = {}
        self.db_pool = db_pool
        self.db_pool = DatabasePool.get_instance()        
  
        self.app = app
        self.current_frame = None
        
        # Ensure db_dir exists
        self.db_dir = os.path.join(os.path.expanduser('~'), 'db')
        if not os.path.exists(self.db_dir):
            os.makedirs(self.db_dir)
            
        self.load_faiss_index()
        self.setup_ui()



    def _is_username_exists(self, username):
        """Check if username exists in either local or AWS database"""
        try:
            # Convert input username to uppercase for consistency
            username = username.upper()
            
            # Check local database
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            cur.execute("SELECT 1 FROM users WHERE username = ?", (username,))
            if cur.fetchone():
                return True

            # Check AWS if online
            if self.db_pool.is_online and self.db_pool.aws_pool:
                aws_conn = self.db_pool.get_aws_connection()
                if aws_conn:
                    try:
                        aws_cur = aws_conn.cursor()
                        aws_cur.execute("SELECT 1 FROM users WHERE username = %s", (username,))
                        if aws_cur.fetchone():
                            return True
                    finally:
                        self.db_pool.return_aws_connection(aws_conn)
            
            return False
            
        except Exception as e:
            logging.error(f"Error checking username existence: {e}")
            return False


    
    def register_new_user(self, capture_with_camera=True):
        """Handle face capture and user registration"""
        try:
            # First check if company is selected in local database
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            
            # Create company_info table if it doesn't exist
            cur.execute("""
                CREATE TABLE IF NOT EXISTS company_info (
                    company_id TEXT PRIMARY KEY,
                    company_name TEXT NOT NULL,
                    password TEXT NOT NULL,
                    address_road TEXT,
                    address_city TEXT,
                    contact_person_name TEXT,
                    contact_person_designation TEXT,
                    contact_person_number TEXT,
                    created_at TEXT,
                    last_updated TEXT
                )
            """)
            
            # Check if any company exists
            cur.execute("SELECT company_name, company_id FROM company_info LIMIT 1")
            company_data = cur.fetchone()
            
            if not company_data:
                messagebox.showinfo("Company Required", "You did not have selected any company. Please select a company first then continue...")
                # Call company management window
                self.manage_company_info()
                return
            
            company_name, company_id = company_data


                
            # Create registration window
            register_window = tk.Toplevel(self.main_window)
            register_window.title("Register New User")
            register_window.geometry("500x600")
            register_window.transient(self.main_window)
            register_window.grab_set()

            # FIX: Use app's centering method or fallback
            try:
                if self.app and hasattr(self.app, 'center_window_robust'):
                    self.app.center_window_robust(register_window, 500, 600)
                else:
                    # Fallback centering
                    register_window.update_idletasks()
                    screen_width = register_window.winfo_screenwidth()
                    screen_height = register_window.winfo_screenheight()
                    x = (screen_width - 500) // 2
                    y = (screen_height - 600) // 2
                    register_window.geometry(f"500x600+{x}+{y}")
                    register_window.lift()
                    register_window.focus_force()
            except Exception as center_error:
                logging.warning(f"Could not center registration window: {center_error}")
            # Set icon if available
                    

            # Create main frame with padding
            main_frame = tk.Frame(register_window, padx=30, pady=30)  # Increased padding
            main_frame.pack(fill='both', expand=True)

            # Title
            title_label = tk.Label(main_frame, text="Register New User", font=("Helvetica", 18, "bold"))
            title_label.pack(pady=(0, 30))  # Increased spacing

            # Username field
            username_frame = tk.Frame(main_frame)
            username_frame.pack(fill='x', pady=10)  # Increased spacing
            tk.Label(username_frame, text="Username:", font=("Helvetica", 10)).pack(side='left')
            username_var = tk.StringVar()
            username_entry = tk.Entry(username_frame, textvariable=username_var, font=("Helvetica", 10))
            username_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            # Company name (read-only, auto-selected)
            company_frame = tk.Frame(main_frame)
            company_frame.pack(fill='x', pady=10)  # Increased spacing
            tk.Label(company_frame, text="Company:", font=("Helvetica", 10)).pack(side='left')
            company_var = tk.StringVar(value=company_name)
            company_entry = tk.Entry(company_frame, textvariable=company_var, state='readonly', font=("Helvetica", 10))
            company_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            # Designation field
            designation_frame = tk.Frame(main_frame)
            designation_frame.pack(fill='x', pady=10)  # Increased spacing
            tk.Label(designation_frame, text="Designation:", font=("Helvetica", 10)).pack(side='left')
            designation_var = tk.StringVar()
            designation_entry = tk.Entry(designation_frame, textvariable=designation_var, font=("Helvetica", 10))
            designation_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            # Password field
            password_frame = tk.Frame(main_frame)
            password_frame.pack(fill='x', pady=10)  # Increased spacing
            tk.Label(password_frame, text="Password:", font=("Helvetica", 10)).pack(side='left')
            password_var = tk.StringVar()
            password_entry = tk.Entry(password_frame, textvariable=password_var, show="*", font=("Helvetica", 10))
            password_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            
            def proceed_with_registration():
                username = username_var.get().strip()
                password = password_var.get().strip()
                designation = designation_var.get().strip()
                
                if not all([username, password, designation]):
                    messagebox.showerror("Error", "Please fill in all required fields!")
                    return
                
                # Store values for later use
                self.username = username
                self.temp_password = password
                self.temp_designation = designation
                self.temp_company_name = company_name
                self.temp_company_uuid = company_id
                
                register_window.destroy()
                
                # Check if user exists before starting capture
                if self._is_username_exists(self.username):  
                    # Create popup window asking if existing employee
                    confirm_window = tk.Toplevel()
                    confirm_window.title("Existing Username")
                    confirm_window.geometry("300x150")
                    confirm_window.transient(self.main_window)
                    confirm_window.grab_set()
                    
                    # Center the confirmation window
                    self.center_window(confirm_window, 300, 150)
                    
                    tk.Label(
                        confirm_window, 
                        text="This username already exists.\nAre you an existing/old employee?",
                        font=('Arial', 10),
                        pady=10
                    ).pack()
                    
                    def handle_yes():
                        confirm_window.destroy()
                        # Get existing user data for this username
                        conn = self.db_pool.get_local_connection()
                        cur = conn.cursor()
                        cur.execute("SELECT company_user_uuid FROM users WHERE username = ?", (self.username,))
                        result = cur.fetchone()
                        
                        if not result:
                            messagebox.showerror("Error", "User data not found!")
                            return
                        
                        existing_company_user_uuid = result[0]

                        # Initialize FAISS index if it doesn't exist
                        if self.index is None:
                            self.index = faiss.IndexFlatL2(512)
                                    
                        # Remove old face embeddings from FAISS if they exist
                        if existing_company_user_uuid in self.user_embeddings:
                            # Get indices to remove
                            indices_to_remove = self.user_embeddings[existing_company_user_uuid]
                            
                            # Create new FAISS index without these embeddings
                            new_index = faiss.IndexFlatL2(512)
                            all_vectors = []
                            
                            # Reconstruct index without the removed user's embeddings
                            for uid, indices in self.user_embeddings.items():
                                if uid != existing_company_user_uuid:
                                    for idx in indices:
                                        vector = self.index.reconstruct(idx)
                                        all_vectors.append(vector)
                            
                            if all_vectors:
                                new_index.add(np.array(all_vectors))
                            
                            # Update index and mappings
                            self.index = new_index
                            del self.user_embeddings[existing_company_user_uuid]
                        
                        # Start capture process
                        self.update_status("Starting capture process...")
                        frames, embeddings = self.capture_faces_continuous(self.username)

                        if len(frames) < 20:
                            messagebox.showinfo("Incomplete", f"Only captured {len(frames)}/20 valid images. Please try again.")
                            return

                        try:
                            # Process embeddings and save user
                            embeddings_array = np.array(embeddings, dtype=np.float32)
                            norms = np.linalg.norm(embeddings_array, axis=1)
                            norms[norms == 0] = 1
                            normalized_embeddings = embeddings_array / norms[:, np.newaxis]

                            # Add normalized embeddings to index
                            current_total = self.index.ntotal
                            self.index.add(normalized_embeddings)

                            # Store user mapping with existing company_user_uuid
                            self.user_embeddings[existing_company_user_uuid] = list(range(current_total, 
                                                                    current_total + len(normalized_embeddings)))

                            # Save index and mappings to db_dir
                            faiss_index_path = os.path.join(self.db_pool.db_dir, 'face_index.faiss')
                            faiss.write_index(self.index, faiss_index_path)
                            self.save_user_mappings()

                            # Generate new user_uuid for existing user
                            user_uuid = str(uuid.uuid4())

                            # Update user in local database
                            conn = self.db_pool.get_local_connection()
                            cur = conn.cursor()
                            try:
                                cur.execute("""
                                    UPDATE users 
                                    SET face_encoding = ?,
                                        password = ?,
                                        user_uuid = ?,
                                        company_name = ?,
                                        company_uuid = ?,
                                        user_designation = ?,
                                        last_updated = strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc')
                                    WHERE company_user_uuid = ?
                                """, (serialize_face_embedding(normalized_embeddings[0]), 
                                    self.temp_password,
                                    user_uuid,
                                    self.temp_company_name,
                                    self.temp_company_uuid,
                                    self.temp_designation,
                                    existing_company_user_uuid))
                                conn.commit()
                            except Exception as e:
                                logging.error(f"Error updating local database: {e}")
                                conn.rollback()
                                raise

                            # Update AWS if online
                            if self.db_pool.is_online:
                                aws_conn = self.db_pool.get_aws_connection()
                                if aws_conn:
                                    try:
                                        aws_cur = aws_conn.cursor()
                                        aws_cur.execute("""
                                            UPDATE users 
                                            SET face_encoding = %s,
                                                password = %s,
                                                user_uuid = %s,
                                                company_name = %s,
                                                company_uuid = %s,
                                                user_designation = %s,
                                                last_updated = TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS')
                                            WHERE company_user_uuid = %s
                                        """, (serialize_face_embedding(normalized_embeddings[0]), 
                                            self.temp_password,
                                            user_uuid,
                                            self.temp_company_name,
                                            self.temp_company_uuid,
                                            self.temp_designation,
                                            existing_company_user_uuid))
                                        aws_conn.commit()
                                    except Exception as e:
                                        logging.error(f"Error updating AWS: {e}")
                                        aws_conn.rollback()
                                        raise
                                    finally:
                                        self.db_pool.return_aws_connection(aws_conn)

                            messagebox.showinfo("Success", "User updated successfully!")
                            self.update_status("User registration completed successfully.")

                        except Exception as e:
                            logging.error(f"Error during user update: {e}")
                            messagebox.showerror("Error", "Update failed. Please try again.")
                            self.update_status("User registration failed.")
                    
                    def handle_no():
                        confirm_window.destroy()
                        # Continue with normal registration flow as a new user
                        self._complete_new_user_registration(capture_with_camera)
                    
                    button_frame = tk.Frame(confirm_window)
                    button_frame.pack(pady=10)
                    
                    tk.Button(
                        button_frame,
                        text="Yes",
                        command=handle_yes,
                        width=10,
                        height=2
                    ).pack(side='left', padx=5)
                    
                    tk.Button(
                        button_frame,
                        text="No",
                        command=handle_no,
                        width=10,
                        height=2
                    ).pack(side='right', padx=5)
                    
                    return

                # If username doesn't exist, proceed with normal registration
                self._complete_new_user_registration(capture_with_camera)
            
            # Register button
         
            register_button = tk.Button(
                main_frame,
                text="Proceed with Camera Registration",
                command=proceed_with_registration,
                width=30,  # Increased button width
                height=2,
                font=("Helvetica", 10, "bold")
            )
            register_button.pack(pady=30)  # Increased spacing
            
            # Focus on username entry
            username_entry.focus()
            
        except Exception as e:
            logging.error(f"Error in register_new_user: {e}")
            messagebox.showerror("Error", "Failed to start registration process")


            
            
    def _complete_new_user_registration(self, capture_with_camera):
        """Complete the registration process for a new user"""
        try:
            if capture_with_camera:
                self.update_status("Starting capture process...")
                frames, embeddings = self.capture_faces_continuous(self.username)
                
                if len(frames) < 20:
                    messagebox.showinfo("Incomplete", f"Only captured {len(frames)}/20 valid images. Please try again.")
                    return

                try:
                    # Process embeddings and save user
                    embeddings_array = np.array(embeddings, dtype=np.float32)
                    norms = np.linalg.norm(embeddings_array, axis=1)
                    norms[norms == 0] = 1
                    normalized_embeddings = embeddings_array / norms[:, np.newaxis]

                    # Initialize FAISS index if needed
                    if self.index is None:
                        self.index = faiss.IndexFlatL2(512)

                    # Generate new UUIDs
                    company_user_uuid = str(uuid.uuid4())
                    user_uuid = str(uuid.uuid4())

                    # Add normalized embeddings to index
                    current_total = self.index.ntotal
                    self.index.add(normalized_embeddings)

                    # Store user mapping with new company_user_uuid
                    self.user_embeddings[company_user_uuid] = list(range(current_total, 
                                                            current_total + len(normalized_embeddings)))

                    # Save index and mappings to db_dir
                    faiss_index_path = os.path.join(self.db_pool.db_dir, 'face_index.faiss')
                    faiss.write_index(self.index, faiss_index_path)
                    self.save_user_mappings()

                    # Save to local database
                    conn = self.db_pool.get_local_connection()
                    cur = conn.cursor()
                    try:
                        cur.execute("""
                            INSERT INTO users (company_user_uuid, username, password, face_encoding, 
                                            user_uuid, company_name, company_uuid, 
                                            user_designation, created_at, last_updated)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc'), strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc'))
                        """, (company_user_uuid, self.username, self.temp_password, serialize_face_embedding(normalized_embeddings[0]),
                            user_uuid, self.temp_company_name, self.temp_company_uuid,
                            self.temp_designation))
                        conn.commit()
                    except Exception as e:
                        logging.error(f"Error saving to local database: {e}")
                        conn.rollback()
                        raise

                    # Save to AWS if online
                    if self.db_pool.is_online:
                        aws_conn = self.db_pool.get_aws_connection()
                        if aws_conn:
                            try:
                                aws_cur = aws_conn.cursor()
                                aws_cur.execute("""
                                    INSERT INTO users (company_user_uuid, username, password, face_encoding,
                                                    user_uuid, company_name, company_uuid,
                                                    user_designation, created_at, last_updated)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 
                                            TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'), 
                                            TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'))
                                    ON CONFLICT (company_user_uuid) DO UPDATE SET
                                        password = EXCLUDED.password,
                                        face_encoding = EXCLUDED.face_encoding,
                                        user_uuid = EXCLUDED.user_uuid,
                                        company_name = EXCLUDED.company_name,
                                        company_uuid = EXCLUDED.company_uuid,
                                        user_designation = EXCLUDED.user_designation,
                                        last_updated = TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS')
                                """, (company_user_uuid, self.username, self.temp_password, serialize_face_embedding(normalized_embeddings[0]),
                                    user_uuid, self.temp_company_name, self.temp_company_uuid,
                                    self.temp_designation))
                                aws_conn.commit()
                            except Exception as e:
                                logging.error(f"Error saving to AWS: {e}")
                                aws_conn.rollback()
                                raise
                            finally:
                                self.db_pool.return_aws_connection(aws_conn)

                    messagebox.showinfo("Success", "User registered successfully!")
                    self.update_status("User registration completed successfully.")

                except Exception as e:
                    logging.error(f"Error during user registration: {e}")
                    messagebox.showerror("Error", "Registration failed. Please try again.")
                    self.update_status("User registration failed.")
            
            else:
                # Registration without camera
                try:
                    # Generate new UUIDs
                    company_user_uuid = str(uuid.uuid4())
                    user_uuid = str(uuid.uuid4())
                    
                    # Create blank face encoding
                    blank_encoding = np.zeros(512)
                    
                    # Save to local database
                    conn = self.db_pool.get_local_connection()
                    cur = conn.cursor()
                    try:
                        cur.execute("""
                            INSERT INTO users (company_user_uuid, username, password, face_encoding, user_uuid, 
                                            company_name, company_uuid, user_designation, created_at, last_updated)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc'), strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc'))
                        """, (company_user_uuid, self.username, self.temp_password, serialize_face_embedding(blank_encoding),
                            user_uuid, self.temp_company_name, self.temp_company_uuid,
                            self.temp_designation))
                        conn.commit()
                    except Exception as e:
                        logging.error(f"Error saving to local database: {e}")
                        conn.rollback()
                        raise
                    
                    # Save to AWS if online
                    if self.db_pool.is_online:
                        aws_conn = self.db_pool.get_aws_connection()
                        if aws_conn:
                            try:
                                aws_cur = aws_conn.cursor()
                                aws_cur.execute("""
                                    INSERT INTO users (company_user_uuid, username, password, face_encoding,
                                                    user_uuid, company_name, company_uuid,
                                                    user_designation, created_at, last_updated)
                                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 
                                            TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'), 
                                            TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'))
                                """, (company_user_uuid, self.username, self.temp_password, serialize_face_embedding(blank_encoding),
                                    user_uuid, self.temp_company_name, self.temp_company_uuid,
                                    self.temp_designation))
                                aws_conn.commit()
                            except Exception as e:
                                logging.error(f"Error saving to AWS: {e}")
                                aws_conn.rollback()
                                raise
                            finally:
                                self.db_pool.return_aws_connection(aws_conn)
                    
                    messagebox.showinfo("Success", "User registered successfully!")
                    self.update_status("User registration completed successfully.")
                    
                except Exception as e:
                    logging.error(f"Error during user registration: {e}")
                    messagebox.showerror("Error", "Registration failed. Please try again.")
                    self.update_status("User registration failed.")
        
        except Exception as e:
            logging.error(f"Error in _complete_new_user_registration: {e}")
            messagebox.showerror("Error", "Registration process failed")


    
               
    def setup_ui(self):
        # Create status label
        self.status_label = tk.Label(self.root, text="", wraplength=400)
        self.status_label.pack(pady=10)


    def center_window(self, window, width=None, height=None):
        """Center a window on the screen - delegates to robust version"""
        try:
            if width and height:
                self.center_window_robust(window, width, height)
            else:
                # Get dimensions from window
                window.update_idletasks()
                if width is None:
                    width = window.winfo_reqwidth() or 400
                if height is None:
                    height = window.winfo_reqheight() or 300
                self.center_window_robust(window, width, height)
        except Exception as e:
            logging.error(f"Error in center_window: {e}")

    def _center_window_fallback(self, window, width=None, height=None):
        """Fallback window centering method"""
        try:
            if not window or not window.winfo_exists():
                return
                
            window.update_idletasks()
            
            if width is None:
                width = window.winfo_reqwidth()
                if width <= 1:
                    width = 400
                    
            if height is None:
                height = window.winfo_reqheight()
                if height <= 1:
                    height = 300
            
            # Get screen dimensions
            screen_width = window.winfo_screenwidth()
            screen_height = window.winfo_screenheight()
            
            # Calculate position
            x = max(0, (screen_width - width) // 2)
            y = max(0, (screen_height - height) // 2)
            
            # Set window position
            window.geometry(f"{width}x{height}+{x}+{y}")
            
            # Bring to front
            window.lift()
            window.focus_force()
            
        except Exception as e:
            logging.error(f"Error in fallback window centering: {e}")

    
    def capture_faces_continuous(self, username, password=None):
        """Capture multiple face embeddings with troubleshooting for camera issues."""
        cap = None
        instruction_window = None
        
        try:
            # Store both username and password as class attributes for later use in registration
            self.current_username = username
            self.temp_password = password  # Store password for database operations
                            

                        
            # Show an initial message before attempting camera access
            status_window = np.zeros((600, 1200, 3), dtype=np.uint8)
            cv2.namedWindow('Face Capture Status', cv2.WINDOW_NORMAL)
            cv2.resizeWindow('Face Capture Status', 1200, 600)
            
            # CRITICAL FIX: Make OpenCV window always on top
            cv2.setWindowProperty('Face Capture Status', cv2.WND_PROP_TOPMOST, 1)
            
            cv2.putText(status_window, "Initializing camera... Please wait", (300, 300), 
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
            cv2.imshow('Face Capture Status', status_window)
            cv2.waitKey(1)  # Process the window event
            
            # Log the attempt to open camera
            logging.info("Attempting to open camera...")
            
            # Try opening the camera with explicit device ID
            for camera_id in range(3):  # Try first 3 camera IDs (0, 1, 2)
                logging.info(f"Trying camera ID: {camera_id}")
                cv2.putText(status_window, f"Trying camera ID: {camera_id}", (300, 350), 
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 255, 255), 2)
                cv2.imshow('Face Capture Status', status_window)
                cv2.waitKey(1)
                
                cap = cv2.VideoCapture(camera_id)
                if cap.isOpened():
                    logging.info(f"Successfully opened camera ID: {camera_id}")
                    break
                else:
                    cap.release()
                    cap = None
            
            if cap is None or not cap.isOpened():
                error_msg = "Failed to open any camera. Please check your camera connection."
                logging.error(error_msg)
                cv2.putText(status_window, error_msg, (200, 400), 
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                cv2.imshow('Face Capture Status', status_window)
                cv2.waitKey(5000)  # Show error for 5 seconds
                return [], []
                
            # Camera is now open, try to get a test frame
            logging.info("Camera opened, attempting to read frame...")
            cv2.putText(status_window, "Camera connected! Reading test frame...", (200, 400), 
                    cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            cv2.imshow('Face Capture Status', status_window)
            cv2.waitKey(1)
            
            # Configure camera with essential settings
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)
            
            # Test reading from camera
            ret, test_frame = cap.read()
            if not ret or test_frame is None:
                error_msg = "Camera opened but could not read frames. Please check camera permissions."
                logging.error(error_msg)
                cv2.putText(status_window, error_msg, (200, 400), 
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                cv2.imshow('Face Capture Status', status_window)
                cv2.waitKey(5000)  # Show error for 5 seconds
                return [], []
                
            # We have a valid frame, show it
            test_frame_resized = cv2.resize(test_frame, (600, 600))
            status_window[0:600, 300:900] = test_frame_resized
            cv2.putText(status_window, "Camera working! Initializing face detection...", 
                    (200, 550), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
            cv2.imshow('Face Capture Status', status_window)
            cv2.waitKey(1)
            
            # Create instruction window now that we know camera works
            try:
                logging.info("Creating instruction window...")
                instruction_window, instruction_text = self.create_instruction_window()
                instruction_text.insert(tk.END, "Camera connected! Initializing face detection...")
                instruction_window.update()
                
                # CRITICAL FIX: Make instruction window always on top
                instruction_window.attributes('-topmost', True)
                instruction_window.lift()
                instruction_window.focus_force()
                
            except Exception as e:
                logging.error(f"Failed to create instruction window: {e}")
                # Continue without instruction window if it fails
            
            # Test face detection with the first frame
            logging.info("Testing face detection...")
            try:
                face_detected, _ = self.verify_face_detection(test_frame)
                if face_detected:
                    status_msg = "Face detection working! Starting capture sequence..."
                    cv2.putText(status_window, status_msg, (200, 580), 
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
                    if instruction_window:
                        instruction_text.delete(1.0, tk.END)
                        instruction_text.insert(tk.END, status_msg)
                        instruction_window.update()
                else:
                    status_msg = "No face detected. Please position yourself in front of the camera."
                    cv2.putText(status_window, status_msg, (150, 580), 
                            cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 165, 255), 2)
                    if instruction_window:
                        instruction_text.delete(1.0, tk.END)
                        instruction_text.insert(tk.END, status_msg)
                        instruction_window.update()
            except Exception as e:
                logging.error(f"Face detection test failed: {e}")
                status_msg = "Face detection initialization failed. Continuing without verification."
                cv2.putText(status_window, status_msg, (150, 580), 
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)
                if instruction_window:
                    instruction_text.delete(1.0, tk.END)
                    instruction_text.insert(tk.END, status_msg)
                    instruction_window.update()
            
            cv2.imshow('Face Capture Status', status_window)
            cv2.waitKey(1000)  # Show status for 1 second
            
            # Initialize capture variables
            captured_frames = []
            embeddings = []
            current_count = 0
            failed_attempts = 0
            last_captured_frame = None
            preview_time = 4  # 4 seconds delay
            last_capture_time = time.time()
            
            # Load action instructions
            actions = self.get_action_instructions()
            
            # Main capture loop
            while current_count < 20:
                try:
                    ret, frame = cap.read()
                    if not ret:
                        logging.error("Failed to read frame during capture loop")
                        break

                    current_time = time.time()
                    time_since_last_capture = current_time - last_capture_time

                    # Get current instruction
                    current_action = actions[current_count] if current_count < len(actions) else ""
                    
                    # Update instruction window if available
                    if instruction_window and instruction_window.winfo_exists():
                        self.update_instruction_window(instruction_window, instruction_text, current_action)

                    # Process frame if enough time has passed
                    if time_since_last_capture >= preview_time:
                        face_detected, current_embedding = self.verify_face_detection(frame)
                        
                        if face_detected:
                            # Capture frame
                            last_captured_frame = frame.copy()
                            embeddings.append(current_embedding)
                            captured_frames.append(frame)
                            current_count += 1
                            last_capture_time = current_time
                        else:
                            failed_attempts += 1
                            if failed_attempts > 100:
                                if messagebox.askyesno("Detection Issues",
                                                    "Having trouble detecting face. Would you like to retry?"):
                                    failed_attempts = 0
                                else:
                                    break

                    # Update OpenCV display
                    status_copy = self.update_simple_status_display(
                        status_window.copy(), frame, last_captured_frame,
                        current_count, 20
                    )
                    
                    cv2.imshow('Face Capture Status', status_copy)

                    # Check for exit key
                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        break

                except Exception as e:
                    logging.error(f"Error in capture loop: {e}")
                    continue

            # Cleanup
            if cap is not None:
                cap.release()
            cv2.destroyAllWindows()
            if instruction_window is not None:
                instruction_window.destroy()
            
            return captured_frames, embeddings

        except Exception as e:
            logging.error(f"Critical error in capture_faces_continuous: {e}")
            if cap is not None:
                cap.release()
            cv2.destroyAllWindows()
            if instruction_window is not None:
                instruction_window.destroy()
            return [], []
        
        
    def update_simple_status_display(self, status_image, frame, last_captured_frame, current_count, total_count):
        """
        Simplified status display that focuses just on showing the camera feed and progress
        since instructions are now shown in the dedicated Tkinter window.
        """
        try:
            h, w = status_image.shape[:2]
            half_w = w // 2

            # Display current frame
            if frame is not None:
                frame_resized = cv2.resize(frame, (half_w, h))
                status_image[0:h, 0:half_w] = frame_resized

            # Display last captured frame
            if last_captured_frame is not None:
                last_resized = cv2.resize(last_captured_frame, (half_w, h))
                status_image[0:h, half_w:w] = last_resized

            # Add labels
            font = cv2.FONT_HERSHEY_SIMPLEX
            font_scale = 0.8  # Larger font
            thickness = 2
            color = (255, 255, 255)
            
            # Add frame labels
            cv2.putText(status_image, "Live Camera", (20, 30), font, font_scale, color, thickness)
            cv2.putText(status_image, "Last Captured", (half_w + 20, 30), font, font_scale, color, thickness)
            
            # Add progress
            progress_text = f"Progress: {current_count}/{total_count}"
            cv2.putText(status_image, progress_text, (20, h - 20), font, font_scale, color, thickness)
            
            return status_image
            
        except Exception as e:
            logging.error(f"Error in update_simple_status_display: {e}")
            return status_image
        
        
    def get_action_instructions(self):
        """Return the action instructions list - separated for optimization."""
                        
                    
        base_actions = [
            "1. Frontal:\nPlease look directly at the camera, face forward.\nসামনের দিকে তাকান, সরাসরি ক্যামেরার দিকে মুখ করুন।\n(Shamner dike takan, shorashori camera dike mukh korun.)",
            
            "2. Left profile:\nTurn your head to the left, showing your left side.\nআপনার মাথা বাম দিকে ঘুরান, আপনার বাম পাশ দেখান।\n(Apnar matha bam dike ghuran, apnar bam pash dekhan.)",
            
            "3. Right profile:\nTurn your head to the right, showing your right side.\nআপনার মাথা ডান দিকে ঘুরান, আপনার ডান পাশ দেখান।\n(Apnar matha dan dike ghuran, apnar dan pash dekhan.)",
            
            "4. 45 degree angle:\nTurn your body slightly, about 45 degrees to the left or right.\nআপনার শরীর সামান্য ঘোরান, প্রায় ৪৫ ডিগ্রি বাম বা ডানে।\n(Apnar shorir shamanno ghoran, pray 45 degree bam ba dane.)",
            
            "5. Upper view:\nTilt your chin down slightly, so we're looking down at you.\nআপনার চিবুক সামান্য নিচে নামান, যাতে আমরা আপনার দিকে নিচে থেকে তাকাতে পারি।\n(Apnar chibuk shamanno niche naman, jate amra apnar dike niche theke takate pari.)",
            
            "6. Lower view:\nTilt your chin up slightly, so we're looking up at you.\nআপনার চিবুক সামান্য উপরে তুলুন, যাতে আমরা আপনার দিকে উপর থেকে তাকাতে পারি।\n(Apnar chibuk shamanno upore tulun, jate amra apnar dike upor theke takate pari.)",

            "7. Quarter turn left:\nTurn your body a quarter turn to the left.\nআপনার শরীর এক চতুর্থাংশ বাম দিকে ঘোরান।\n(Apnar shorir ek choturthangsho bam dike ghoran.)",
            "8. Quarter turn right:\nTurn your body a quarter turn to the right.\nআপনার শরীর এক চতুর্থাংশ ডান দিকে ঘোরান।\n(Apnar shorir ek choturthangsho dan dike ghoran.)",

            "9.  Below view:\nPlease look directly down, as if looking at something on the floor.\nদয়া করে সরাসরি নিচে তাকান, যেন মেঝেতে কিছু দেখছেন।\n(Doya kore shorashori niche takan, jeno mejhete kichu dekhchen.)",
            "10. Head straight:\nKeep your head upright and centered.\nআপনার মাথা সোজা এবং মাঝখানে রাখুন।\n(Apnar matha soja ebong majkhane rakhun.)",
            "11. Head tilted left:\nTilt your head gently to your left shoulder.\nআপনার মাথা আলতো করে বাম কাঁধের দিকে কাত করুন।\n(Apnar matha alto kore bam kandher dike kat korun.)",
            "12. Head tilted right:\nTilt your head gently to your right shoulder.\nআপনার মাথা আলতো করে ডান কাঁধের দিকে কাত করুন।\n(Apnar matha alto kore dan kandher dike kat korun.)",
            "13. Head tilted up:\nTilt your head back slightly, looking upwards.\nআপনার মাথা সামান্য পিছনের দিকে কাত করুন, উপরের দিকে তাকান।\n(Apnar matha shamanno pichoner dike kat korun, upper dike takan.)",
            "14. Head tilted down:\nTilt your head forward, looking downwards.\nআপনার মাথা সামনের দিকে কাত করুন, নিচের দিকে তাকান।\n(Apnar matha shamner dike kat korun, nicher dike takan.)",
            "15. Looking left:\nTurn your eyes to look to your left.\nআপনার চোখ বাম দিকে ঘুরান।\n(Apnar chokh bam dike ghuran.)",
            "16. Looking right:\nTurn your eyes to look to your right.\nআপনার চোখ ডান দিকে ঘুরান।\n(Apnar chokh dan dike ghuran.)",
            "17. Looking up:\nLook upwards with your eyes.\nআপনার চোখ উপরের দিকে তাকান।\n(Apnar chokh upper dike takan.)",
            "18. Looking down:\nLook downwards with your eyes.\nআপনার চোখ নিচের দিকে তাকান।\n(Apnar chokh nicher dike takan.)",
            "19. Looking into camera:\nLook directly at the camera lens.\nসরাসরি ক্যামেরার লেন্সের দিকে তাকান।\n(Shorashori camera lens er dike takan.)",
            "20. Looking past camera:\nLook slightly beyond the camera, as if looking at something in the distance.\nক্যামেরার সামান্য বাইরে তাকান, যেন দূরের কিছু দেখছেন।\n(Camera shamanno baire takan, jeno durer kichu dekhchen.)",
            "21. Cross eyed:\nCross your eyes, looking towards the bridge of your nose.\nআপনার চোখ ট্যারা করুন, নাকের ডগার দিকে তাকান।\n(Apnar chokh tera korun, naker dogar dike takan.)",
            "22. Eyes closed:\nClose your eyes gently.\nআলতো করে চোখ বন্ধ করুন।\n(Alto kore chokh bondho korun.)",
            "23. Eyes open:\nOpen your eyes wide.\nচোখ বড় করে খুলুন।\n(Chokh boro kore khulun.)",
            "24. Wide eyes:\nOpen your eyes very wide, as if surprised.\nখুব বড় করে চোখ খুলুন, যেন বিস্মিত হয়েছেন।\n(Khub boro kore chokh khulun, jeno bismito hoyechen.)",
            "25. Eyes narrowed:\nSquint your eyes slightly.\nচোখ সামান্য কুঁচকে নিন।\n(Chokh shamanno kuchke nin.)",
            "26. Eyebrows raised:\nRaise your eyebrows.\nআপনার ভ্রু তুলুন।\n(Apnar bhru tulun.)",
            "27. Mouth open:\nOpen your mouth slightly.\nআপনার মুখ সামান্য খুলুন।\n(Apnar mukh shamanno khulun.)",
            "28. Teeth showing:\nSmile and show your teeth.\nহাসুন এবং আপনার দাঁত দেখান।\n(Hasun ebong apnar dant dekhan.)",
            "29. Tight lipped:\nPress your lips together firmly.\nআপনার ঠোঁট শক্ত করে একসাথে চেপে ধরুন।\n(Apnar thont shokto kore ekshathe chepe dhorun.)",
            "30. Relaxed lips:\nKeep your lips relaxed and natural.\nআপনার ঠোঁট শিথিল এবং স্বাভাবিক রাখুন।\n(Apnar thont shithil ebong shavabik rakhun.)",
            "31. Sad:\nShow a sad expression.\nএকটি দুঃখিত অভিব্যক্তি দেখান।\n(Ekta dukkhito obhibyakti dekhan.)",
            "32. Surprised:\nShow a surprised expression.\nএকটি বিস্মিত অভিব্যক্তি দেখান।\n(Ekta bismito obhibyakti dekhan.)",
            "33. Angry:\nShow an angry expression.\nএকটি রাগান্বিত অভিব্যক্তি দেখান।\n(Ekta raganbito obhibyakti dekhan.)",
            "34. Thoughtful:\nShow a thoughtful or contemplative expression.\nএকটি চিন্তাশীল বা ধ্যানমগ্ন অভিব্যক্তি দেখান।\n(Ekta chintashil ba dhyanomogno obhibyakti dekhan.)",
            "35. Neutral:\nKeep your face relaxed and without strong emotion.\nআপনার মুখ শিথিল এবং শক্তিশালী আবেগ ছাড়া রাখুন।\n(Apnar mukh shithil ebong shoktishali abeg chara rakhun.)",
            "36. Smile:\nSmile naturally.\nস্বাভাবিকভাবে হাসুন।\n(Shavabikvabe hasun.)",
            "37. Slight smile:\nGive a very small, subtle smile.\nখুব ছোট, সূক্ষ্ম হাসি দিন।\n(Khub choto, shukshmo hashi din.)",
            "38. Broad smile:\nGive a large, wide smile.\nএকটি বড়, প্রশস্ত হাসি দিন।\n(Ekta boro, proshosto hashi din.)",
            "39. Full smile:\nShow a full, happy smile.\nএকটি পূর্ণ, সুখী হাসি দেখান।\n(Ekta purno, sukhi hashi dekhan.)",
            "40. Half smile:\nGive a smile with only one side of your mouth curving upwards.\nআপনার মুখের শুধুমাত্র এক পাশ উপরের দিকে বাঁকিয়ে হাসুন।\n(Apnar mukher shudhu ek pash upper dike bankiye hasun.)",
            "41. Light smile:\nA gentle, soft smile.\nএকটি মৃদু, নরম হাসি।\n(Ekta mridu, norom hashi.)",
            "42. Fake smile:\nShow a forced, unnatural smile.\nএকটি জোর করে, অস্বাভাবিক হাসি দেখান।\n(Ekta jor kore, oshavabik hashi dekhan.)",
            "43. Forced smile:\nShow a smile that feels unnatural.\nএমন একটি হাসি দেখান যা অস্বাভাবিক মনে হয়।\n(Emon ekta hashi dekhan ja oshavabik mone hoy.)",
            "44. Genuine smile:\nShow a real, happy smile.\nএকটি সত্যিকারের, সুখী হাসি দেখান।\n(Ekta shotyikarer, sukhi hashi dekhan.)",
            "45. Expression changing:\nLet your expression change slowly, from one emotion to another.\nআপনার অভিব্যক্তি ধীরে ধীরে পরিবর্তন করুন, এক আবেগ থেকে অন্য আবেগে।\n(Apnar obhibyakti dhire dhire poriborton korun, ek abeg theke onno abege.)",
            "46. Expression neutral:\nReturn to a neutral, expressionless face.\nএকটি নিরপেক্ষ, অভিব্যক্তিহীন মুখে ফিরে আসুন।\n(Ekta niropekkho, obhibyaktihin mukhe phire ashun.)",
            "47. Expression intense:\nShow a very strong, intense expression.\nএকটি খুব শক্তিশালী, তীব্র অভিব্যক্তি দেখান।\n(Ekta khub shoktishali, tibro obhibyakti dekhan.)",
            "48. Serious gaze:\nLook serious and focused.\nগুরুতর এবং মনোযোগী দেখুন।\n(Gurutoro ebong monojogi dekhun.)",
            "49. Subtle smile:\nAn almost invisible, very slight smile.\nপ্রায় অদৃশ্য, খুব সামান্য হাসি।\n(Pray odrishyo, khub shamanno hashi.)",
            "50. Focused:\nLook very focused and attentive.\nখুব মনোযোগী এবং সতর্ক দেখুন।\n(Khub monojogi ebong shotorko dekhun.)",]


        
        return (base_actions * (20 // len(base_actions) + 1))[:20]

    def verify_face_detection(self, frame):
        """Verify if a face can be detected in the frame - optimized version."""
        try:
            # Import DeepFace only when needed to avoid startup delay
            from deepface import DeepFace
            
            # Use a smaller image for faster processing
            small_frame = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
            
            obj = DeepFace.represent(
                img_path=small_frame,
                model_name='Facenet512',
                detector_backend='mtcnn',
                enforce_detection=True,
                align=True
            )
            return True, obj[0]['embedding']
        except Exception as e:
            return False, None

        
        

    def create_instruction_window(self):
        """Create a Tkinter window for displaying multilingual instructions at the bottom of the screen."""
        window = tk.Tk()
        window.title("Instructions")
        
        # Get screen width and height
        screen_width = window.winfo_screenwidth()
        screen_height = window.winfo_screenheight()
        
        # Set window dimensions - wider and shorter as requested
        window_width = int(screen_width * 0.5)  # 50% of screen width
        window_height = 140  # Reduced height
        
        # Calculate x and y coordinates for the window to be at the bottom center
        x = screen_width - window_width
        y = screen_height - window_height - 5  # Position at bottom with some margin
        
        # Set window geometry
        window.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # Set window to have black background
        window.configure(bg='black')
        
        # CRITICAL FIX: Keep window always on top and focused
        window.attributes('-topmost', True)
        window.attributes('-toolwindow', True)  # Makes it stay on top in Windows
        window.lift()
        window.focus_force()
        
        # Configure font that supports Bengali with larger size
        bengali_font = tkfont.Font(family="Nirmala UI", size=16)  # Increased font size
        
        # Text widget for instructions
        instruction_text = tk.Text(window, font=bengali_font, wrap=tk.WORD, 
                                bg='black', fg='white',  # Black background, white text
                                padx=20, pady=10)  # Add some padding
        instruction_text.pack(fill=tk.BOTH, expand=True)
        
        # Add a scrollbar with custom colors
        scrollbar = tk.Scrollbar(instruction_text, bg='gray', troughcolor='black')
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        instruction_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=instruction_text.yview)
        
        return window, instruction_text

    def update_instruction_window(self, window, text_widget, instruction):
        """Update the instruction window with the current instruction."""
        if window.winfo_exists():
            text_widget.delete(1.0, tk.END)
            
            # Parse the instruction and apply custom formatting
            if '\n' in instruction:
                parts = instruction.split('\n')
                
                # Make the first line (title/number) bold with yellow color
                text_widget.tag_configure("title", foreground="yellow", font=tkfont.Font(family="Nirmala UI", size=18, weight="bold"))
                text_widget.insert(tk.END, parts[0] + "\n", "title")
                
                # Add remaining content with white color
                remaining_text = "\n".join(parts[1:])
                text_widget.insert(tk.END, remaining_text)
            else:
                text_widget.insert(tk.END, instruction)
            
            window.update()




    def update_status_display(self, status_image, frame, last_captured_frame, action, current_count, total_count):
        try:
            h, w = status_image.shape[:2]
            half_w = w // 2

            # Display current frame
            if frame is not None:
                frame_resized = cv2.resize(frame, (half_w, h//2))
                status_image[0:h//2, 0:half_w] = frame_resized

            # Display last captured frame
            if last_captured_frame is not None:
                last_resized = cv2.resize(last_captured_frame, (half_w, h//2))
                status_image[0:h//2, half_w:w] = last_resized

            # Add labels
            font = cv2.FONT_HERSHEY_SIMPLEX
            font_scale = 0.6
            thickness = 1
            color = (255, 255, 255)
            
            # Add frame labels
            cv2.putText(status_image, "Live Camera", (20, h//2 - 10), font, font_scale, color, thickness)
            cv2.putText(status_image, "Last Captured", (half_w + 20, h//2 - 10), font, font_scale, color, thickness)
            
            # Add progress
            progress_text = f"Progress: {current_count}/{total_count}"
            cv2.putText(status_image, progress_text, (20, h - 20), font, font_scale, color, thickness)

            # Create a black background for instructions
            instruction_bg = np.zeros((h//2, w, 3), dtype=np.uint8)
            
            # Draw instruction text on the background
            if action:
                # Split text into parts (English, Bangla, Pronunciation)
                parts = action.split('\n')
                
                # Start position for text
                x_pos = 20
                y_pos = 30
                line_height = 30
                
                for i, part in enumerate(parts):
                    if i == 0:  # Title/number - make it bold and larger
                        cv2.putText(instruction_bg, part, (x_pos, y_pos), 
                                    font, font_scale * 1.2, (255, 255, 0), thickness + 1)
                        y_pos += line_height
                    else:
                        # Handle long lines by wrapping text
                        words = part.split()
                        if words:
                            current_line = words[0]
                            for word in words[1:]:
                                test_line = current_line + " " + word
                                # Get text size
                                (text_width, text_height), _ = cv2.getTextSize(test_line, font, font_scale, thickness)
                                
                                # If too wide, print current line and start new line
                                if text_width > w - 40:  # 40px margin
                                    cv2.putText(instruction_bg, current_line, (x_pos, y_pos), 
                                            font, font_scale, color, thickness)
                                    y_pos += line_height
                                    current_line = word
                                else:
                                    current_line = test_line
                                    
                            # Print the last line
                            cv2.putText(instruction_bg, current_line, (x_pos, y_pos), 
                                    font, font_scale, color, thickness)
                            y_pos += line_height
                
                # If there's Bengali text that can't be rendered with cv2.putText,
                # add a message explaining this
                cv2.putText(instruction_bg, "See instruction window for Bengali text", 
                        (x_pos, y_pos + line_height), font, font_scale, (0, 255, 255), thickness)
            
            # Place the instruction background on the bottom half of the status image
            status_image[h//2:h, :] = instruction_bg
            
            return status_image
            
        except Exception as e:
            logging.error(f"Error in update_status_display: {e}")
            return status_image

        

    def open_hr_window(self):
        """Open the HR window by delegating to the App class."""
        if self.app and hasattr(self.app, 'open_hr_window'):
            try:
                self.app.open_hr_window()
            except Exception as e:
                logging.error(f"Error opening HR window: {e}")
                messagebox.showerror("Error", "Failed to open HR window!")
        else:
            logging.error("App instance or open_hr_window method not found.")
            messagebox.showerror("Error", "HR functionality is unavailable.")
    

              
    # Add this method to the FaceRecognitionSystem class
    def analyze_distances(self, distances, indices):
        """Analyze the distances returned by FAISS search"""
        distance_analysis = defaultdict(list)
        
        # Map each distance to corresponding user
        for dist, idx in zip(distances[0], indices[0]):
            for company_user_uuid, user_indices in self.user_embeddings.items():
                if idx in user_indices:
                    distance_analysis[company_user_uuid].append(dist)
                    break
        
        # Calculate statistics for each user
        user_stats = {}
        for company_user_uuid, dists in distance_analysis.items():
            user_stats[company_user_uuid] = {
                'min_distance': min(dists),
                'max_distance': max(dists),
                'avg_distance': sum(dists) / len(dists),
                'num_matches': len(dists)
            }
        
        return user_stats

    

    def verify_identity(self, frame=None):
        # Check if verification window already exists and destroy it
        if hasattr(self, 'verification_window') and self.verification_window is not None:
            try:
                self.verification_window.destroy()
            except:
                pass
        
        try:
            captured_frame = None
            
            if frame is not None:
                captured_frame = frame
            elif hasattr(self, 'current_frame') and self.current_frame is not None:
                captured_frame = self.current_frame
                self.current_frame = None
            else:
                cap = cv2.VideoCapture(0)
                ret, captured_frame = cap.read()
                cap.release()
                if not ret:
                    messagebox.showerror("Error", "Failed to capture image")
                    return

            # Create a sleek, modern verification window

            # Change window background to black and reduce size to 75%
            self.verification_window = tk.Toplevel()
            self.verification_window.title("Identity Verification")
            # Change from 1100x700 to 825x525 (75% of original size)
            self.verification_window.geometry("825x925")  
            self.verification_window.configure(bg="black")  # Change from "#f8f9fa" to "black"


            # Center the window on the screen
            screen_width = self.verification_window.winfo_screenwidth()
            screen_height = self.verification_window.winfo_screenheight()
            x = (screen_width - 825) // 2
            y = (screen_height - 925) // 2
            self.verification_window.geometry(f"+{x}+{y}")
            self.verification_window.resizable(False, False)  # Disable resizing
            self.verification_window.attributes('-topmost', True)  # Keep on top
            self.verification_window.attributes('-fullscreen', False)  # Disable fullscreen
            self.verification_window.attributes('-alpha', 0.95)  # Set transparency
            self.verification_window.overrideredirect(True)  # Remove window decorations
            self.verification_window.focus_force()  # Bring to front
            self.verification_window.grab_set()  # Make it modal
            self.verification_window.bind("<Escape>", lambda e: self.verification_window.destroy())  # Close on Escape
            
            # Set window icon if available
            try:
                self.verification_window.iconbitmap("assets/icon.ico")
            except:
                pass
                
            # Add smooth rounded corners effect with a frame
            main_container = tk.Frame(self.verification_window, bg="black", padx=20, pady=20)
            main_container.pack(fill="both", expand=True)

            if captured_frame is None:
                error_frame = tk.Frame(main_container, bg="black")
                error_frame.pack(expand=True)
                
                # Error icon
                try:
                    error_icon = Image.open("assets/error.png")
                    error_icon = error_icon.resize((64, 64))
                    error_icon_img = ImageTk.PhotoImage(error_icon)
                    tk.Label(error_frame, image=error_icon_img, bg="#f8f9fa").pack(pady=(20, 10))
                    error_frame.image = error_icon_img
                except:
                    pass
                
                tk.Label(
                    error_frame,
                    text="Failed to capture frame",
                    font=('Helvetica', 14, 'bold'),
                    fg="#dc3545",
                    bg="#f8f9fa"
                ).pack(pady=10)

                btn_return = tk.Button(
                    error_frame,
                    text="Return to Main Window",
                    command=lambda: [self.verification_window.destroy(), self.root.deiconify()],
                    font=('Helvetica', 12),
                    bg="#0d6efd",
                    fg="white",
                    relief="flat",
                    padx=15,
                    pady=8,
                    cursor="hand2"
                )
                btn_return.pack(pady=20)
                return

            # Create two main sections with a modern split layout
            left_panel = tk.Frame(main_container, bg="#ffffff", padx=30, pady=30, relief="flat", bd=1)
            left_panel.grid(row=0, column=0, sticky="nsew", padx=(0, 10))
            
            right_panel = tk.Frame(main_container, bg="#ffffff", padx=30, pady=30, relief="flat", bd=1)
            right_panel.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
            
            # Configure the grid
            main_container.grid_columnconfigure(0, weight=3)
            main_container.grid_columnconfigure(1, weight=2)
            main_container.grid_rowconfigure(0, weight=1)
                        
            # Add a header to the left panel
            header_frame = tk.Frame(left_panel, bg="#ffffff")
            header_frame.pack(fill="x", pady=(0, 20))
            
            tk.Label(
                header_frame,
                text="IDENTITY VERIFICATION",
                font=('Helvetica', 18, 'bold'),
                fg="#212529",
                bg="#ffffff"
            ).pack(side="left")

            # Create an elegant image display area
            image_frame = tk.Frame(left_panel, bg="#ffffff", bd=0)
            image_frame.pack(fill="x", pady=(0, 20))
            
            # Process and display the captured image with a modern look
            captured_frame_rgb = cv2.cvtColor(captured_frame, cv2.COLOR_BGR2RGB)
            captured_frame_rgb = cv2.resize(captured_frame_rgb, (240, 240))
            
            # Add a subtle border to the image
            img = Image.fromarray(captured_frame_rgb)
            
            # Create a circular mask for the image (optional)
            # Uncomment the following to make the image circular
            """
            size = img.size
            mask = Image.new('L', size, 0)
            draw = ImageDraw.Draw(mask)
            draw.ellipse((0, 0) + size, fill=255)
            img.putalpha(mask)
            """
            
            img_tk = ImageTk.PhotoImage(img)
            
            # Create a canvas for the image with a subtle shadow effect
            img_canvas = tk.Canvas(image_frame, width=250, height=250, bg="#ffffff", bd=0, highlightthickness=0)
            img_canvas.pack(side="left", padx=15)
            
            # Create shadow effect
            img_canvas.create_rectangle(5, 5, 245, 245, fill="#f0f0f0", outline="")
            img_canvas.create_image(10, 10, anchor="nw", image=img_tk)
            img_canvas.image = img_tk

            embedding = self.get_embedding(captured_frame)
            if embedding is None:
                no_face_frame = tk.Frame(left_panel, bg="#ffffff")
                no_face_frame.pack(fill="x", pady=20)
                
                tk.Label(
                    no_face_frame,
                    text="No face detected",
                    font=('Helvetica', 14, 'bold'),
                    fg="#dc3545",
                    bg="#ffffff"
                ).pack(pady=10)
                
                btn_return = tk.Button(
                    no_face_frame,
                    text="Return to Main Window",
                    command=lambda: [self.verification_window.destroy(), self.root.deiconify()],
                    font=('Helvetica', 12),
                    bg="#0d6efd",
                    fg="white",
                    relief="flat",
                    padx=15,
                    pady=8,
                    cursor="hand2"
                )
                btn_return.pack(pady=10)
                return

            if self.index is None:
                no_db_frame = tk.Frame(left_panel, bg="#ffffff")
                no_db_frame.pack(fill="x", pady=20)
                
                tk.Label(
                    no_db_frame,
                    text="No face database available",
                    font=('Helvetica', 14, 'bold'),
                    fg="#dc3545",
                    bg="#ffffff"
                ).pack(pady=10)
                
                btn_return = tk.Button(
                    no_db_frame,
                    text="Return to Main Window",
                    command=lambda: [self.verification_window.destroy(), self.root.deiconify()],
                    font=('Helvetica', 12),
                    bg="#0d6efd",
                    fg="white",
                    relief="flat",
                    padx=15,
                    pady=8,
                    cursor="hand2"
                )
                btn_return.pack(pady=10)
                return

            norm = np.linalg.norm(embedding)
            if norm > 0:
                embedding = embedding / norm

            THRESHOLDS = {
                'EXACT_MATCH': 0.4,
                'STRONG_MATCH': 0.5,
                'POSSIBLE_MATCH': 0.6,
                'REJECT': 0.7
            }

            MIN_MATCHES = {
                'EXACT_MATCH': 8,
                'STRONG_MATCH': 7,
                'POSSIBLE_MATCH': 6
            }

            D, I = self.index.search(embedding.reshape(1, -1), k=10)

            user_matches = {}
            min_distances = {}
            quality_matches = {}

            for idx, dist in zip(I[0], D[0]):
                if dist > THRESHOLDS['REJECT']:
                    continue

                for company_user_uuid, indices in self.user_embeddings.items():
                    if idx in indices:
                        if company_user_uuid not in user_matches:
                            user_matches[company_user_uuid] = 0
                            min_distances[company_user_uuid] = float('inf')
                            quality_matches[company_user_uuid] = {
                                'exact': 0,
                                'strong': 0,
                                'possible': 0
                            }
                        user_matches[company_user_uuid] += 1
                        min_distances[company_user_uuid] = min(min_distances[company_user_uuid], dist)

                        if dist <= THRESHOLDS['EXACT_MATCH']:
                            quality_matches[company_user_uuid]['exact'] += 1
                        elif dist <= THRESHOLDS['STRONG_MATCH']:
                            quality_matches[company_user_uuid]['strong'] += 1
                        elif dist <= THRESHOLDS['POSSIBLE_MATCH']:
                            quality_matches[company_user_uuid]['possible'] += 1

            if not user_matches:
                no_match_frame = tk.Frame(left_panel, bg="#ffffff")
                no_match_frame.pack(fill="x", pady=20)
                
                tk.Label(
                    no_match_frame,
                    text="No valid matches found",
                    font=('Helvetica', 14, 'bold'),
                    fg="#dc3545",
                    bg="#ffffff"
                ).pack(pady=10)
                
                btn_return = tk.Button(
                    no_match_frame,
                    text="Return to Main Window",
                    command=lambda: [self.verification_window.destroy(), self.root.deiconify()],
                    font=('Helvetica', 12),
                    bg="#0d6efd",
                    fg="white",
                    relief="flat",
                    padx=15,
                    pady=8,
                    cursor="hand2"
                )
                btn_return.pack(pady=10)
                return

            best_match_user = None
            best_match_quality = None
            best_distance = float('inf')
            best_confidence = 0

            for company_user_uuid in user_matches:
                total_matches = user_matches[company_user_uuid]
                user_min_distance = min_distances[company_user_uuid]
                quality = quality_matches[company_user_uuid]

                if (quality['exact'] >= MIN_MATCHES['EXACT_MATCH'] and 
                    user_min_distance <= THRESHOLDS['EXACT_MATCH']):
                    match_quality = "EXACT"
                    confidence = 95.0 + (quality['exact'] / 10) * 5
                elif (quality['exact'] + quality['strong'] >= MIN_MATCHES['STRONG_MATCH'] and 
                    user_min_distance <= THRESHOLDS['STRONG_MATCH']):
                    match_quality = "STRONG"
                    confidence = 80.0 + (total_matches / 10) * 15
                elif (quality['exact'] + quality['strong'] + quality['possible'] >= MIN_MATCHES['POSSIBLE_MATCH'] and 
                    user_min_distance <= THRESHOLDS['POSSIBLE_MATCH']):
                    match_quality = "POSSIBLE"
                    confidence = 60.0 + (total_matches / 10) * 20
                else:
                    continue

                if confidence > best_confidence:
                    best_match_user = company_user_uuid
                    best_match_quality = match_quality
                    best_distance = user_min_distance
                    best_confidence = confidence

            if best_match_user is None:
                no_conf_frame = tk.Frame(left_panel, bg="#ffffff")
                no_conf_frame.pack(fill="x", pady=20)
                
                tk.Label(
                    no_conf_frame,
                    text="No confident match found",
                    font=('Helvetica', 14, 'bold'),
                    fg="#dc3545",
                    bg="#ffffff"
                ).pack(pady=10)
                
                btn_return = tk.Button(
                    no_conf_frame,
                    text="Return to Main Window",
                    command=lambda: [self.verification_window.destroy(), self.root.deiconify()],
                    font=('Helvetica', 12),
                    bg="#0d6efd",
                    fg="white",
                    relief="flat",
                    padx=15,
                    pady=8,
                    cursor="hand2"
                )
                btn_return.pack(pady=10)
                return

            conn = self.db_pool.get_local_connection()
            try:
                cur = conn.cursor()
              
                cur.execute("SELECT username FROM users WHERE company_user_uuid = ?", (best_match_user,))
                result = cur.fetchone()
                matched_username = result[0] if result else None
            finally:
                cur.close()

            if not matched_username:
                no_user_info_frame = tk.Frame(left_panel, bg="#ffffff")
                no_user_info_frame.pack(fill="x", pady=20)
                
                tk.Label(
                    no_user_info_frame,
                    text="Could not retrieve user information",
                    font=('Helvetica', 14, 'bold'),
                    fg="#dc3545",
                    bg="#ffffff"
                ).pack(pady=10)
                
                btn_return = tk.Button(
                    no_user_info_frame,
                    text="Return to Main Window",
                    command=lambda: [self.verification_window.destroy(), self.root.deiconify()],
                    font=('Helvetica', 12),
                    bg="#0d6efd",
                    fg="white",
                    relief="flat",
                    padx=15,
                    pady=8,
                    cursor="hand2"
                )
                btn_return.pack(pady=10)
                return

            timestamp = datetime.now()
            
            try:
                login_type = self.db_pool.determine_login_type(matched_username)
            except Exception as login_err:
                logging.error(f"Error determining login type: {login_err}")
                login_type = "Unknown"

            # Add the match analysis with a modern card layout
            match_card = tk.Frame(left_panel, bg="#ffffff", bd=1, relief="solid")
            match_card.pack(fill="x", pady=10)
            
            match_header = tk.Frame(match_card, bg="#f1f8ff", padx=15, pady=10)
            match_header.pack(fill="x")
            
            tk.Label(
                match_header,
                text="MATCH ANALYSIS",
                font=('Helvetica', 12, 'bold'),
                fg="#0d6efd",
                bg="#f1f8ff"
            ).pack(anchor="w")
            
            match_body = tk.Frame(match_card, bg="#ffffff", padx=15, pady=15)
            match_body.pack(fill="x")
            
            # Use grid for more precise control
            row = 0
            
            # Match quality with color indicator
            quality_color = "#198754" if best_match_quality == "EXACT" else "#fd7e14" if best_match_quality == "STRONG" else "#6c757d"
            quality_frame = tk.Frame(match_body, bg="#ffffff")
            quality_frame.grid(row=row, column=0, sticky="w", pady=5)
            
            tk.Label(
                quality_frame,
                text="Match Quality:",
                font=('Helvetica', 10, 'bold'),
                fg="#212529",
                bg="#ffffff"
            ).pack(side="left", padx=(0, 5))
            
            tk.Label(
                quality_frame,
                text=best_match_quality,
                font=('Helvetica', 10, 'bold'),
                fg=quality_color,
                bg="#ffffff"
            ).pack(side="left")
            
            row += 1
            
            # Confidence score with progress bar
            conf_frame = tk.Frame(match_body, bg="#ffffff")
            conf_frame.grid(row=row, column=0, sticky="w", pady=5)
            
            tk.Label(
                conf_frame,
                text="Confidence Score:",
                font=('Helvetica', 10, 'bold'),
                fg="#212529",
                bg="#ffffff"
            ).pack(side="left", padx=(0, 5))
            
            tk.Label(
                conf_frame,
                text=f"{best_confidence:.2f}%",
                font=('Helvetica', 10),
                fg="#212529",
                bg="#ffffff"
            ).pack(side="left")
            
            # Progress bar frame
            bar_frame = tk.Frame(match_body, bg="#ffffff")
            bar_frame.grid(row=row+1, column=0, sticky="ew", pady=(0, 5))
            
            # Progress bar background
            bar_bg = tk.Frame(bar_frame, bg="#e9ecef", height=8, width=300)
            bar_bg.pack(fill="x")
            
            # Progress bar fill - determine color based on confidence
            bar_color = "#198754" if best_confidence >= 90 else "#0d6efd" if best_confidence >= 75 else "#fd7e14" if best_confidence >= 60 else "#dc3545"
            bar_width = int(3 * best_confidence)
            bar_fill = tk.Frame(bar_bg, bg=bar_color, height=8, width=bar_width)
            bar_fill.place(x=0, y=0)
            
            row += 2
            
            # Technical details in a grid layout
            details = [
                ("Minimum Distance", f"{best_distance:.4f}"),
                ("Exact Matches", f"{quality_matches[best_match_user]['exact']}"),
                ("Strong Matches", f"{quality_matches[best_match_user]['strong']}"),
                ("Possible Matches", f"{quality_matches[best_match_user]['possible']}"),
                ("Total Matches", f"{user_matches[best_match_user]}")
            ]
            
            for label, value in details:
                label_widget = tk.Label(
                    match_body,
                    text=f"{label}:",
                    font=('Helvetica', 10, 'bold'),
                    fg="#212529",
                    bg="#ffffff",
                    anchor="w"
                )
                label_widget.grid(row=row, column=0, sticky="w", pady=2)
                
                value_widget = tk.Label(
                    match_body,
                    text=value,
                    font=('Helvetica', 10),
                    fg="#6c757d",
                    bg="#ffffff",
                    anchor="e"
                )
                value_widget.grid(row=row, column=1, sticky="e", pady=2)
                
                row += 1
                
            # Configure columns
            match_body.grid_columnconfigure(0, weight=1)
            match_body.grid_columnconfigure(1, weight=1)
            
            # Add the attendance record with a stylish card
            attendance_card = tk.Frame(left_panel, bg="#ffffff", bd=1, relief="solid")
            attendance_card.pack(fill="x", pady=10)
            
            attendance_header = tk.Frame(attendance_card, bg="#f8f9fa", padx=15, pady=10)
            attendance_header.pack(fill="x")
            
            tk.Label(
                attendance_header,
                text="ATTENDANCE RECORD",
                font=('Helvetica', 12, 'bold'),
                fg="#212529",
                bg="#f8f9fa"
            ).pack(anchor="w")
            
            attendance_body = tk.Frame(attendance_card, bg="#ffffff", padx=15, pady=15)
            attendance_body.pack(fill="x")
            
            # Create a modern user info display
            user_info_frame = tk.Frame(attendance_body, bg="#ffffff")
            user_info_frame.pack(fill="x", pady=5)
            
            # Username display with highlight
            username_frame = tk.Frame(user_info_frame, bg="#ffffff")
            username_frame.pack(fill="x", pady=5)
            
            tk.Label(
                username_frame, 
                text="USER", 
                font=('Helvetica', 9),
                fg="#6c757d",
                bg="#ffffff"
            ).pack(anchor="w")
            
            tk.Label(
                username_frame,
                text=matched_username.upper(),
                font=('Helvetica', 16, 'bold'),
                fg="#212529",
                bg="#ffffff"
            ).pack(anchor="w")
            
            # Horizontal divider
            tk.Frame(user_info_frame, height=1, bg="#dee2e6").pack(fill="x", pady=10)
            
            # Type and timestamp in a grid
            details_frame = tk.Frame(user_info_frame, bg="#ffffff")
            details_frame.pack(fill="x")
            
            # Two column layout for details
            details_frame.columnconfigure(0, weight=1)
            details_frame.columnconfigure(1, weight=1)
            
            # Type info
            type_frame = tk.Frame(details_frame, bg="#ffffff")
            type_frame.grid(row=0, column=0, sticky="w", pady=5)
            
            tk.Label(
                type_frame,
                text="TYPE",
                font=('Helvetica', 9),
                fg="#6c757d",
                bg="#ffffff"
            ).pack(anchor="w")
            
            tk.Label(
                type_frame,
                text=login_type,
                font=('Helvetica', 12),
                fg="#212529",
                bg="#ffffff"
            ).pack(anchor="w")
            
            # Time info
            time_frame = tk.Frame(details_frame, bg="#ffffff")
            time_frame.grid(row=0, column=1, sticky="w", pady=5)
            
            tk.Label(
                time_frame,
                text="TIME",
                font=('Helvetica', 9),
                fg="#6c757d",
                bg="#ffffff"
            ).pack(anchor="w")
            
            tk.Label(
                time_frame,
                text=timestamp.strftime('%I:%M:%S %p'),
                font=('Helvetica', 12),
                fg="#212529",
                bg="#ffffff"
            ).pack(anchor="w")
            
            # Date info
            date_frame = tk.Frame(details_frame, bg="#ffffff")
            date_frame.grid(row=1, column=0, columnspan=2, sticky="w", pady=5)
            
            tk.Label(
                date_frame,
                text="DATE",
                font=('Helvetica', 9),
                fg="#6c757d",
                bg="#ffffff"
            ).pack(anchor="w")
            
            tk.Label(
                date_frame,
                text=timestamp.strftime('%d %B %Y'),
                font=('Helvetica', 12),
                fg="#212529",
                bg="#ffffff"
            ).pack(anchor="w")
            
            # Create the right panel with verification actions
            verification_header = tk.Frame(right_panel, bg="#ffffff")
            verification_header.pack(fill="x", pady=(0, 30))
            
            tk.Label(
                verification_header,
                text="CONFIRM IDENTITY",
                font=('Helvetica', 18, 'bold'),
                fg="#212529",
                bg="#ffffff"
            ).pack(anchor="w")
            
            # Add countdown widget with modern styling
            countdown_frame = tk.Frame(right_panel, bg="#ffffff", pady=15)
            countdown_frame.pack(fill="x")
            
            countdown_label = tk.Label(
                countdown_frame,
                text="",
                font=('Helvetica', 14),
                fg="#0d6efd",
                bg="#ffffff"
            )
            countdown_label.pack(pady=10)
            
            # Question
            question_frame = tk.Frame(right_panel, bg="#ffffff", pady=15)
            question_frame.pack(fill="x")
            
            tk.Label(
                question_frame,
                text=f"Are you {matched_username}?",
                font=('Helvetica', 16, 'bold'),
                fg="#212529",
                bg="#ffffff"
            ).pack(anchor="center")
            
            # Space for buttons
            button_frame = tk.Frame(right_panel, bg="#ffffff", pady=15)
            button_frame.pack(fill="x")
            
            # Actions section with modern buttons
            countdown_remaining = 5

            def save_attendance(username):
                try:
                    year_str = str(timestamp.year)
                    month_str = timestamp.strftime("%B")
                    user_dir = os.path.join(self.db_pool.db_dir, username)
                    year_dir = os.path.join(user_dir, year_str)
                    month_dir = os.path.join(year_dir, month_str)
                    os.makedirs(month_dir, exist_ok=True)

                    filename = f"{username.upper()}_{timestamp.strftime('%d_%b_%Y')}_{login_type}_{timestamp.strftime('%I_%M_%p')}.jpg"
                    filepath = os.path.join(month_dir, filename)
                    cv2.imwrite(filepath, captured_frame)

                    try:
                        local_success, aws_success = self.db_pool.save_login_to_db(
                            username,
                            login_type,
                            timestamp,
                            filepath,
                            with_camera=True
                        )

                        if local_success:
                            success_window = tk.Toplevel()
                            success_window.title("Success")
                            success_window.configure(bg="#ffffff")
                            
                            success_frame = tk.Frame(success_window, bg="#ffffff", padx=25, pady=20)
                            success_frame.pack(fill="both", expand=True)
                            
                            try:
                                check_icon = Image.open("assets/checkmark.png")
                                check_icon = check_icon.resize((48, 48))
                                check_icon_img = ImageTk.PhotoImage(check_icon)
                                tk.Label(success_frame, image=check_icon_img, bg="#ffffff").pack(pady=(0, 10))
                                success_frame.image = check_icon_img
                            except:
                                pass
                                
                            tk.Label(
                                success_frame,
                                text=f"{login_type} recorded successfully!",
                                font=('Helvetica', 12, 'bold'),
                                fg="#198754",
                                bg="#ffffff"
                            ).pack(padx=20, pady=10)
                            
                            success_window.after(3000, success_window.destroy)
                        else:
                            messagebox.showerror("Error", "Failed to record attendance")
                    except Exception as db_err:
                        logging.error(f"Error in save_login_to_db: {db_err}")
                        messagebox.showerror("Error", "Failed to record attendance: Database error")
                except Exception as save_err:
                    logging.error(f"Error saving attendance: {save_err}")
                    messagebox.showerror("Error", "Failed to save attendance")

            def save_and_close():
                save_attendance(matched_username)
                # Trigger sync and closing BEFORE destroying windows
                if self.app and hasattr(self.app, 'main_window'):
                    self.app.main_window.after(500, lambda: self.app.on_closing())
                # Now destroy verification window
                if self.verification_window and self.verification_window.winfo_exists():
                    self.verification_window.destroy()
                self.verification_window = None
            
            
               
            def confirm_user():
                nonlocal countdown_remaining
                countdown_remaining = -1
                save_and_close()

            def deny_user():
                nonlocal countdown_remaining
                countdown_remaining = -1
                self.verification_window.destroy()
                self.verification_window = None
                self.verify_identity()

            def done_action():
                nonlocal countdown_remaining
                countdown_remaining = -1
                save_and_close()

            def return_to_main():
                save_and_close()
                self.root.deiconify()

            def update_countdown():
                nonlocal countdown_remaining
                if countdown_remaining > 0:
                    countdown_label.config(text=f"Auto-verification in: {countdown_remaining} seconds")
                    countdown_remaining -= 1
                    self.verification_window.after(1000, update_countdown)
                elif countdown_remaining == 0:
                    save_and_close()  # This will now trigger main window close
            
            # Button styles using modern design principles
            btn_yes = tk.Button(
                button_frame,
                text=f"Yes, I am {matched_username}",
                command=confirm_user,
                font=('Helvetica', 12),
                bg="#198754",
                fg="white",
                relief="flat",
                padx=15,
                pady=10,
                cursor="hand2",
                width=25
            )
            btn_yes.pack(pady=10, fill="x")
            
            btn_no = tk.Button(
                button_frame,
                text="No, I'm someone else",
                command=deny_user,
                font=('Helvetica', 12),
                bg="#dc3545",
                fg="white",
                relief="flat",
                padx=15,
                pady=10,
                cursor="hand2",
                width=25
            )
            btn_no.pack(pady=10, fill="x")
            
            btn_done = tk.Button(
                button_frame,
                text="Done",
                command=done_action,
                font=('Helvetica', 12),
                bg="#0d6efd",
                fg="white",
                relief="flat",
                padx=15,
                pady=10,
                cursor="hand2",
                width=25
            )
            btn_done.pack(pady=10, fill="x")
            
            btn_return = tk.Button(
                button_frame,
                text="Return to Main Window",
                command=return_to_main,
                font=('Helvetica', 12),
                bg="#6c757d",
                fg="white",
                relief="flat",
                padx=15,
                pady=10,
                cursor="hand2",
                width=25
            )
            btn_return.pack(pady=10, fill="x")
            
            # Start the countdown
            update_countdown()

        except Exception as e:
            logging.error(f"Error during verification: {e}")
            logging.error("Traceback: %s", traceback.format_exc())
            
            if hasattr(self, 'verification_window'):
                self.verification_window.destroy()
            
            # Create a modern error window
            self.verification_window = tk.Toplevel()
            self.verification_window.title("Verification Error")
            self.verification_window.geometry("500x350")
            self.verification_window.configure(bg="#ffffff")
            
            error_container = tk.Frame(self.verification_window, bg="#ffffff", padx=30, pady=30)
            error_container.pack(fill="both", expand=True)
            
            # Add error icon if available
            try:
                error_icon = Image.open("assets/error.png")
                error_icon = error_icon.resize((64, 64))
                error_icon_img = ImageTk.PhotoImage(error_icon)
                tk.Label(error_container, image=error_icon_img, bg="#ffffff").pack(pady=(10, 20))
                error_container.image = error_icon_img
            except:
                pass
            
            # Title
            tk.Label(
                error_container,
                text="Verification Error",
                font=('Helvetica', 16, 'bold'),
                fg="#dc3545",
                bg="#ffffff"
            ).pack(pady=(0, 20))
            
            # Error message in a scrollable text area with a border
            message_frame = tk.Frame(error_container, bg="#ffffff", bd=1, relief="solid")
            message_frame.pack(fill="both", expand=True, pady=(0, 20))
            
            error_text = tk.Text(
                message_frame,
                wrap="word",
                height=5,
                font=('Helvetica', 10),
                bg="#f8f9fa",
                relief="flat",
                padx=10,
                pady=10
            )
            error_text.pack(fill="both", expand=True)
            error_text.insert("1.0", f"Error details: {str(e)}")
            error_text.config(state="disabled")
            
            # Return button with modern styling
            btn_return = tk.Button(
                error_container,
                text="Return to Main Window",
                command=lambda: [self.verification_window.destroy(), self.root.deiconify()],
                font=('Helvetica', 12),
                bg="#0d6efd",
                fg="white",
                relief="flat",
                padx=15,
                pady=10,
                cursor="hand2"
            )
            btn_return.pack(pady=10)
            
            self.verification_window = None
  


    def monitor_sync_status(self, cloud_status_label, cloud_update_label):
        def check_and_update():
            try:
                if self.db_pool.is_online:
                    conn = self.db_pool.get_aws_connection()
                    if conn:
                        cur = conn.cursor()
                        cur.execute("SELECT COUNT(*) FROM sync_status WHERE synced = 0")
                        pending_count = cur.fetchone()[0]
                        
                        if pending_count == 0:
                            cloud_status_label.config(text="Cloud Database: Connected", fg='green')
                            cloud_update_label.config(text="Cloud Sync: Complete", fg='green')
                        else:
                            cloud_status_label.config(text="Cloud Database: Connected", fg='green')
                            cloud_update_label.config(text="Cloud Sync: In Progress", fg='orange')
                        
                        self.db_pool.return_aws_connection(conn)
                    else:
                        cloud_status_label.config(text="Cloud Database: Connection Error", fg='red')
                        cloud_update_label.config(text="Cloud Sync: Failed", fg='red')
                else:
                    cloud_status_label.config(text="Cloud Database: Offline", fg='red')
                    cloud_update_label.config(text="Cloud Sync: Offline", fg='red')
                    
            except Exception as e:
                logging.error(f"Error monitoring sync status: {e}")
                cloud_status_label.config(text="Cloud Database: Error", fg='red')
                cloud_update_label.config(text="Cloud Sync: Error", fg='red')
                
            # Schedule next update
            if cloud_status_label.winfo_exists():
                cloud_status_label.after(5000, check_and_update)  # Update every 5 seconds
            
        check_and_update()

    def handle_sync_completion(self, cloud_status_label, cloud_update_label):
        try:
            if self.db_pool.is_online:
                conn = self.db_pool.get_aws_connection()
                if conn:
                    cur = conn.cursor()
                    cur.execute("""
                        UPDATE sync_status 
                        SET synced = 1, 
                            last_sync_attempt = CURRENT_TIMESTAMP 
                        WHERE synced = 0
                    """)
                    conn.commit()
                    
                    cloud_status_label.config(text="Cloud Database: Connected", fg='green')
                    cloud_update_label.config(text="Cloud Sync: Complete", fg='green')
                    
                    self.db_pool.return_aws_connection(conn)
                    
        except Exception as e:
            logging.error(f"Error handling sync completion: {e}")
            cloud_status_label.config(text="Cloud Database: Error", fg='red')
            cloud_update_label.config(text="Cloud Sync: Error", fg='red')
            
            

    def _save_frame(self, frame, username, login_type, timestamp):
        """Save a single frame as an image file"""
        try:
            # Create directory structure
            year_folder = os.path.join(self.db_pool.db_dir, username, str(timestamp.year))
            month_folder = os.path.join(year_folder, timestamp.strftime("%B"))
            os.makedirs(month_folder, exist_ok=True)
            logging.debug(f"Directory created/exists: {month_folder}")
            
            # Create filename
            filename = f"{username.upper()}_{timestamp.strftime('%d_%b_%Y')}_{login_type}_{timestamp.strftime('%I_%M_%p')}.jpg"
            filepath = os.path.join(month_folder, filename)

                       # Validate the file path

            try:
                validate_filepath(filepath)
            except ValueError as e:
                logging.error(f"Invalid file path: {e}")
                messagebox.showerror("Error", "The file path is invalid. Please check the directory structure.")
                return None
            
            # Save frame
            cv2.imwrite(filepath, frame)
            return filepath

        except Exception as e:
            logging.error(f"Error saving frame: {e}")
            return None



# Add this method to the FaceRecognitionSystem class
    def generate_embeddings_report(self):
        if self.index is None or self.index.ntotal == 0:
            messagebox.showerror("Error", "No embeddings in database")
            return

        report_data = []
        
        for company_user_uuid, indices in self.user_embeddings.items():
            # Get all embeddings for this user
            user_embeddings = self.index.reconstruct_n(indices[0], len(indices))
            
            for i, embedding in enumerate(user_embeddings):
                image_name = f"image_{i+1}"
                
                # Basic counts
                nan_count = np.isnan(embedding).sum()
                zero_count = np.count_nonzero(embedding == 0)
                negative_count = np.count_nonzero(embedding < 0)
                positive_count = np.count_nonzero(embedding > 0)
                
                # Sort values for analysis
                sorted_vals = np.sort(embedding)
                non_nan_vals = sorted_vals[~np.isnan(sorted_vals)]
                
                # Get top/middle/bottom values
                if len(non_nan_vals) >= 5:
                    lowest_5 = ','.join(f'{x:.6f}' for x in non_nan_vals[:5])
                    highest_5 = ','.join(f'{x:.6f}' for x in non_nan_vals[-5:])
                    mid_idx = len(non_nan_vals) // 2
                    middle_5 = ','.join(f'{x:.6f}' for x in non_nan_vals[mid_idx-2:mid_idx+3])
                else:
                    lowest_5 = highest_5 = middle_5 = "Insufficient values"

                # Count unusable embeddings (all zero or all nan)
                unusable = 1 if (nan_count == len(embedding) or zero_count == len(embedding)) else 0
                
                # Create row
                row = {
                    'Username': company_user_uuid,
                    'Image_Name': image_name,
                    'Number_of_Embeddings_Generated': len(embedding),
                    'Number_of_Errors': 0,  # This would need error tracking during generation
                    'Number_of_Unusable_Embeddings': unusable,
                    'Number_of_Nan_Values': nan_count,
                    'Number_of_Zero_Values': zero_count,
                    'Number_of_Negative_Values': negative_count,
                    'Number_of_Positive_Values': positive_count,
                    'Top_5_Lower_Values': lowest_5,
                    'Middle_5_Values': middle_5,
                    'Top_5_Highest_Values': highest_5
                }
                report_data.append(row)
        
        # Create DataFrame and save to CSV
        df = pd.DataFrame(report_data)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'embeddings_analysis_{timestamp}.csv'
        df.to_csv(filename, index=False)
        
        # Also save summary statistics
        summary_df = df.groupby('Username').agg({
            'Number_of_Embeddings_Generated': 'sum',
            'Number_of_Errors': 'sum',
            'Number_of_Unusable_Embeddings': 'sum',
            'Number_of_Nan_Values': 'mean',
            'Number_of_Zero_Values': 'mean',
            'Number_of_Negative_Values': 'mean',
            'Number_of_Positive_Values': 'mean'
        }).round(2)
        
        summary_filename = f'embeddings_summary_{timestamp}.csv'
        summary_df.to_csv(summary_filename)
        
        messagebox.showinfo("Success", 
                        f"Analysis complete!\n"
                        f"Detailed report saved as: {filename}\n"
                        f"Summary report saved as: {summary_filename}")

        
   
        
    
    def update_status(self, message):
        """Update the status label in the UI."""
        self.status_label.config(text=message)
        
            

    def create_status_window(self, width=800, height=600):
        status_image = np.zeros((height, width, 3), np.uint8)
        return status_image

    def update_status_display(self, status_image, frame, captured_frame, action, current_count, total_count):
        h, w = status_image.shape[:2]
        half_w = w // 2

        # Resize frames to fit in the window
        frame_height = h // 2
        if frame is not None:
            frame_resized = cv2.resize(frame, (half_w, frame_height))
            status_image[0:frame_height, 0:half_w] = frame_resized

        if captured_frame is not None:
            captured_resized = cv2.resize(captured_frame, (half_w, frame_height))
            status_image[0:frame_height, half_w:w] = captured_resized

        # Add text information
        font = cv2.FONT_HERSHEY_SIMPLEX
        cv2.putText(status_image, f"Current Action: {action}", (20, h//2 + 40), font, 0.7, (255, 255, 255), 2)
        cv2.putText(status_image, f"Progress: {current_count}/{total_count} images", (20, h//2 + 80), font, 0.7, (255, 255, 255), 2)
        cv2.putText(status_image, "Live Camera Feed", (20, 30), font, 0.7, (255, 255, 255), 2)
        cv2.putText(status_image, "Last Captured Image", (half_w + 20, 30), font, 0.7, (255, 255, 255), 2)
        
        return status_image

   


    def get_embedding(self, frame):
        try:
            from deepface import DeepFace
            
            
            obj = DeepFace.represent(
                img_path=frame,
                model_name='Facenet512', 
                detector_backend='mtcnn',
                enforce_detection=True
            )
            if len(obj) != 1:
                return None
            embedding = np.array(obj[0]['embedding'], dtype=np.float32)
            # Normalize the embedding
            embedding = embedding / np.linalg.norm(embedding)
            return embedding
        except Exception as e:
            print(f"Error in getting embedding: {e}")
            return None

    def load_faiss_index(self):
        # Get database directory path from the database pool
        db_dir = self.db_pool.db_dir
        faiss_index_path = os.path.join(db_dir, 'face_index.faiss')
        user_mappings_path = os.path.join(db_dir, 'user_mappings.txt')
        
        if os.path.exists(faiss_index_path):
            try:
                self.index = faiss.read_index(faiss_index_path)
                self.load_user_mappings(user_mappings_path)
                logging.info("Successfully loaded FAISS index")
            except Exception as e:
                logging.error(f"Error loading index: {e}")
                self.index = None
        else:
            logging.info("No FAISS Index Found")
            self.index = None

    def load_user_mappings(self, user_mappings_path=None):
        if user_mappings_path is None:
            user_mappings_path = os.path.join(self.db_pool.db_dir, 'user_mappings.txt')
            
        if os.path.exists(user_mappings_path):
            with open(user_mappings_path, "r") as f:
                lines = f.readlines()
                for line in lines:
                    company_user_uuid, indices = line.strip().split(":")
                    self.user_embeddings[company_user_uuid] = [int(idx) for idx in indices.split(",")]

    def save_user_mappings(self):
        user_mappings_path = os.path.join(self.db_pool.db_dir, 'user_mappings.txt')
        with open(user_mappings_path, "w") as f:
            for company_user_uuid, indices in self.user_embeddings.items():
                f.write(f"{company_user_uuid}:{','.join(map(str, indices))}\n")

    

    def reset_database(self):
        if messagebox.askyesno("Confirm Reset", "Are you sure you want to reset the database?"):
            faiss_index_path = os.path.join(self.db_pool.db_dir, 'face_index.faiss')
            user_mappings_path = os.path.join(self.db_pool.db_dir, 'user_mappings.txt')
            
            if os.path.exists(faiss_index_path):
                os.remove(faiss_index_path)
            if os.path.exists(user_mappings_path):
                os.remove(user_mappings_path)
                
            self.index = None
            self.user_embeddings = {}
            messagebox.showinfo("Success", "Database reset successfully")



class NetworkSyncManager:
    def __init__(self, db_pool):
        self.db_pool = db_pool
        self.is_online = True
        self.last_sync_time = None
        self.sync_thread = None
        self.stop_thread = threading.Event()  # Use Event instead of boolean
        self.lock = threading.Lock()
        self.max_retries = 3
        self.retry_delay = 1


    
    def _check_connectivity(self) -> bool:
        """Check internet connectivity with proper separation of internet vs AWS failures"""
        retry_count = 0
        while retry_count < self.max_retries:
            try:
                # Check basic internet first
                for dns in ["8.8.8.8", "1.1.1.1", "208.67.222.222"]:
                    try:
                        socket.create_connection((dns, 53), timeout=3)
                        # We have internet, now check if AWS services work
                        aws_working = self._verify_aws_services()
                        
                        # Update online status based on INTERNET, not AWS
                        self.is_online = True
                        
                        if not aws_working:
                            logging.warning("Internet available but AWS services unavailable")
                        else:
                            logging.info("Internet and AWS services both available")
                            
                        return True  # We have internet regardless of AWS status
                    except:
                        continue
                        
                # No internet connection
                logging.warning("No internet connection detected")
                self.is_online = False
                return False
                
            except Exception as e:
                retry_count += 1
                if retry_count == self.max_retries:
                    self.is_online = False
                    return False
                time.sleep(self.retry_delay * (2 ** (retry_count - 1)))
        
        self.is_online = False
        return False


        
    def _verify_aws_services(self) -> bool:
        """Verify AWS services are accessible - separate from internet connectivity"""
        try:
            # Check S3
            self.db_pool.s3_client.list_buckets()
            
            # Check RDS
            conn = self.db_pool.get_aws_connection()
            if conn:
                cur = conn.cursor()
                cur.execute("SELECT 1")
                cur.fetchone()
                self.db_pool.return_aws_connection(conn)
                return True
            return False
        except Exception as e:
            logging.error(f"AWS service verification failed: {e}")
            return False
        
        
    def start_monitoring(self):
        """Start network monitoring in a separate thread"""
        self.sync_thread = threading.Thread(target=self._monitor_network, daemon=True)
        self.sync_thread.start()



    def _has_pending_syncs(self):
        """Check if there are any pending syncs"""
        try:
            conn = sqlite3.connect(self.db_pool.local_db_path)
            cur = conn.cursor()
            cur.execute("SELECT COUNT(*) FROM sync_status WHERE synced = 0")
            count = cur.fetchone()[0]
            return count > 0
        except Exception as e:
            logging.error(f"Error checking pending syncs: {e}")
            return False
        finally:
            if conn:
                conn.close()



    def _sync_all_pending_data(self, days_filter=None):
        """Sync data between local SQLite and AWS RDS databases."""
        try:
            # Add this date filter logic at the beginning of the method
            date_filter = ""
            if days_filter:
                filter_date = (datetime.now() - timedelta(days=days_filter)).strftime('%Y-%m-%d')
                date_filter = f"WHERE DATE(date) >= '{filter_date}'"
                logging.info(f"Filtering records from {filter_date}")

            logging.info(f"Starting sync process{' for last ' + str(days_filter) + ' days' if days_filter else ''}...")
            local_conn = sqlite3.connect(self.db_pool.local_db_path)
            aws_conn = self.db_pool.get_aws_connection()
            
            
            if not aws_conn:
                logging.error("Failed to get AWS connection for sync")
                return

            local_conn.row_factory = sqlite3.Row
            local_cur = local_conn.cursor()
            aws_cur = aws_conn.cursor()

            # First ensure all users exist in AWS
            try:
                local_cur.execute("SELECT company_user_uuid, username, face_encoding FROM users")
                local_users = local_cur.fetchall()
                
                for user in local_users:
                    if user is None:  # Skip if user is None
                        continue
                    try:
                        company_user_uuid = user['company_user_uuid']
                        if not company_user_uuid:  # Skip if no company_user_uuid
                            continue
                            
                        aws_cur.execute("SELECT 1 FROM users WHERE company_user_uuid = %s", (company_user_uuid,))
                        user_exists = aws_cur.fetchone()
                        
                        if not user_exists:
                            aws_cur.execute("""
                                INSERT INTO users 
                                (company_user_uuid, username, face_encoding, created_at, last_updated)
                                VALUES (%s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                            """, (company_user_uuid, user['username'], user['face_encoding']))
                            aws_conn.commit()
                            logging.info(f"Added missing user {company_user_uuid} to AWS")
                    except Exception as e:
                        logging.error(f"Error syncing user {company_user_uuid if 'company_user_uuid' in locals() else 'unknown'}: {str(e)}")
                        aws_conn.rollback()
                        continue
                        
            except Exception as e:
                logging.error(f"Error in users sync: {str(e)}")
                aws_conn.rollback()

            def parse_datetime(dt_str):
                if not dt_str:
                    return None
                try:
                    dt_str = str(dt_str).strip()
                    if '+' in dt_str:
                        dt_str = dt_str.split('+')[0]
                    elif '-' in dt_str and dt_str.count('-') > 2:
                        dt_str = dt_str.rsplit('-', 1)[0]
                    return datetime.strptime(dt_str, '%Y-%m-%d %H:%M:%S')
                except Exception:
                    return None

            def time_diff_seconds(time1, time2):
                if not time1 or not time2:
                    return float('inf')  # Return infinity if either time is None
                t1 = parse_datetime(time1)
                t2 = parse_datetime(time2)
                if not t1 or not t2:
                    return float('inf')
                return abs((t1 - t2).total_seconds())

            try:
                # Build query based on days filter
                if days_filter:
                    filter_date = (datetime.now() - timedelta(days=days_filter)).strftime('%Y-%m-%d')
                    query = f"""
                        SELECT ar.*, u.company_user_uuid as verified_company_user_uuid 
                        FROM attendance_records ar
                        LEFT JOIN users u ON ar.company_user_uuid = u.company_user_uuid
                        {date_filter}
                    """
                else:
                    query = """
                        SELECT ar.*, u.company_user_uuid as verified_company_user_uuid 
                        FROM attendance_records ar
                        LEFT JOIN users u ON ar.company_user_uuid = u.company_user_uuid
                    """
                    
                local_cur.execute(query)
                local_records = local_cur.fetchall()
                
                if not local_records:  # Handle empty result
                    logging.info("No attendance records found to sync")
                    return
                    
                for record in local_records:
                    if record is None:  # Skip None records
                        continue
                        
                    try:
                        # Start a new transaction for each record
                        aws_conn.rollback()  # Clear any previous failed transaction
                        
                        record_id = record['record_id']
                        if not record_id:  # Skip if no record_id
                            continue
                            
                        company_user_uuid = record['company_user_uuid']
                        if not company_user_uuid:  # Skip if no company_user_uuid
                            continue

                        # Get AWS record if exists
                        try:
                            aws_cur.execute("SELECT * FROM attendance_records WHERE record_id = %s", (record_id,))
                            aws_record = aws_cur.fetchone()
                        except Exception as e:
                            logging.error(f"Error fetching AWS record {record_id}: {str(e)}")
                            aws_record = None

                        # Check if record exists in AWS and data needs to be synced
                        needs_sync = False
                        if not aws_record:
                            needs_sync = True  # New record, needs sync
                        else:
                            # Convert aws_record to dict for easier access
                            aws_column_names = [desc[0] for desc in aws_cur.description]
                            aws_dict = dict(zip(aws_column_names, aws_record))
                            
                            # Check specific fields that matter for sync decision
                            critical_fields = ['in_time', 'out_time']
                            
                            for field in critical_fields:
                                local_val = record[field]
                                aws_val = aws_dict.get(field)
                                
                                # If in_time/out_time differ by more than 60 seconds, sync is needed
                                if time_diff_seconds(local_val, aws_val) > 60:
                                    needs_sync = True
                                    break
                        
                        if not needs_sync:
                            logging.info(f"Record {record_id} matches within tolerance - skipping sync")
                            continue

                        # Verify user exists in AWS before proceeding
                        aws_cur.execute("SELECT 1 FROM users WHERE company_user_uuid = %s", (company_user_uuid,))
                        if not aws_cur.fetchone():
                            logging.error(f"User {company_user_uuid} not found in AWS for record {record_id}")
                            continue

                        # Prepare final values for sync
                        final_values = {}
                        for field in record.keys():
                            if field == 'verified_company_user_uuid':  # Skip join field
                                continue
                                
                            try:
                                local_val = record[field]
                                
                                if aws_record:
                                    aws_column_names = [desc[0] for desc in aws_cur.description]
                                    aws_dict = dict(zip(aws_column_names, aws_record))
                                    aws_val = aws_dict.get(field)
                                else:
                                    aws_val = None
                                
                                if field in ['in_time', 'out_time']:
                                    local_time = parse_datetime(local_val)
                                    aws_time = parse_datetime(aws_val)
                                    
                                    if field == 'in_time':
                                        if local_time and aws_time:
                                            final_values[field] = min(local_time, aws_time).strftime('%Y-%m-%d %H:%M:%S')
                                        elif local_time or aws_time:
                                            final_values[field] = (local_time or aws_time).strftime('%Y-%m-%d %H:%M:%S')
                                        else:
                                            final_values[field] = None
                                    else:  # out_time
                                        if local_time and aws_time:
                                            final_values[field] = max(local_time, aws_time).strftime('%Y-%m-%d %H:%M:%S')
                                        elif local_time or aws_time:
                                            final_values[field] = (local_time or aws_time).strftime('%Y-%m-%d %H:%M:%S')
                                        else:
                                            final_values[field] = None
                                else:
                                    final_values[field] = local_val if local_val not in [None, '', 0] else aws_val
                            except Exception as field_error:
                                logging.error(f"Error processing field {field} for record {record_id}: {str(field_error)}")
                                final_values[field] = local_val  # Use local value as fallback
                                continue

                        try:
                            if aws_record:
                                # Update existing record
                                update_fields = [f for f in final_values.keys() if f != 'record_id']
                                if not update_fields:  # Skip if no fields to update
                                    continue
                                    
                                query = """
                                    UPDATE attendance_records 
                                    SET """ + ", ".join(f"{field} = %s" for field in update_fields) + """
                                    WHERE record_id = %s
                                """
                                values = [final_values[field] for field in update_fields] + [record_id]
                                aws_cur.execute(query, tuple(values))
                            else:
                                # Insert new record
                                fields = [f for f in final_values.keys() if final_values[f] is not None]
                                if not fields:  # Skip if no fields to insert
                                    continue
                                    
                                query = f"""
                                    INSERT INTO attendance_records 
                                    ({', '.join(fields)})
                                    VALUES ({', '.join(['%s'] * len(fields))})
                                """
                                values = [final_values[field] for field in fields]
                                aws_cur.execute(query, tuple(values))
                            
                            aws_conn.commit()
                            
                            # Update sync status
                            local_cur.execute("""
                                INSERT OR REPLACE INTO sync_status 
                                (record_id, synced, last_sync_attempt)
                                VALUES (?, 1, datetime('now'))
                            """, (record_id,))
                            local_conn.commit()
                            
                            logging.info(f"Successfully synced record {record_id}")
                            
                        except Exception as db_error:
                            logging.error(f"Database operation error for record {record_id}: {str(db_error)}")
                            aws_conn.rollback()
                            continue

                    except Exception as record_error:
                        logging.error(f"Error processing record {record_id if 'record_id' in locals() else 'unknown'}: {str(record_error)}")
                        if 'aws_conn' in locals():
                            aws_conn.rollback()
                        continue
                        
            except Exception as e:
                logging.error(f"Error in attendance records sync: {str(e)}")
                if 'aws_conn' in locals():
                    aws_conn.rollback()
                
        except Exception as e:
            logging.error(f"Error in sync_all_pending_data: {str(e)}")
        finally:
            try:
                if 'local_conn' in locals() and local_conn:
                    local_conn.close()
                if 'aws_conn' in locals() and aws_conn:
                    self.db_pool.return_aws_connection(aws_conn)
            except Exception as e:
                logging.error(f"Error closing connections: {str(e)}")
            logging.info("Sync process completed")



    def sync_recent_data(self):
        """Sync both AWS RDS and S3 with proper validation"""
        try:
            # Create progress window
            progress_window = tk.Toplevel()
            progress_window.title("Sync Status")
            progress_window.geometry("400x250")
            
            status_label = tk.Label(progress_window, text="Checking AWS RDS and S3 status...", pady=10)
            status_label.pack()
            
            rds_label = tk.Label(progress_window, text="", pady=5)
            rds_label.pack()
            
            s3_label = tk.Label(progress_window, text="", pady=5)
            s3_label.pack()
            
            details_label = tk.Label(progress_window, text="", pady=5)
            details_label.pack()
            
            # First validate AWS connectivity
            aws_conn = self.db_pool.get_aws_connection()
            if not aws_conn:
                status_label.config(text="Cannot connect to AWS. Sync aborted.")
                return
                
            try:
                aws_cur = aws_conn.cursor()
                
                # Calculate date 7 days ago
                seven_days_ago = (datetime.now() - timedelta(days=7)).date()
                
                # Get local records that need syncing
                local_conn = self.db_pool.get_local_connection()
                local_cur = local_conn.cursor()
                
                # First check RDS records
                local_cur.execute("""
                    SELECT ar.record_id, ar.company_user_uuid, ar.username, ar.date, 
                        ar.in_time, ar.out_time, ar.video_path_in, ar.video_path_out,
                        ar.in_time_with_camera, ar.out_time_with_camera
                    FROM attendance_records ar
                    WHERE date(ar.date) >= ?
                """, (seven_days_ago.isoformat(),))
                
                local_records = local_cur.fetchall()
                
                if not local_records:
                    status_label.config(text="No records found for the last 7 days!")
                    progress_window.after(2000, progress_window.destroy)
                    return
                
                # Check each record in AWS RDS
                rds_sync_needed = []
                s3_sync_needed = []
                
                for record in local_records:
                    record_id = record[0]
                    
                    # Check RDS
                    aws_cur.execute("""
                        SELECT record_id, in_time, out_time, video_path_in, video_path_out
                        FROM attendance_records 
                        WHERE record_id = %s
                    """, (record_id,))
                    aws_record = aws_cur.fetchone()
                    
                    if not aws_record:
                        rds_sync_needed.append(record)
                    else:
                        # Compare timestamps and paths
                        local_in_time = record[4]
                        local_out_time = record[5]
                        aws_in_time = aws_record[1]
                        aws_out_time = aws_record[2]
                        
                        if (local_in_time != aws_in_time or 
                            local_out_time != aws_out_time):
                            rds_sync_needed.append(record)
                    
                    # Check S3 paths
                    video_in = record[6]
                    video_out = record[7]
                    
                    if video_in and not video_in.startswith('s3://'):
                        s3_sync_needed.append((record_id, video_in, 'IN_TIME'))
                    if video_out and not video_out.startswith('s3://'):
                        s3_sync_needed.append((record_id, video_out, 'OUT_TIME'))
                
                # Update status
                rds_label.config(text=f"RDS Records to sync: {len(rds_sync_needed)}")
                s3_label.config(text=f"S3 Files to sync: {len(s3_sync_needed)}")
                
                if not rds_sync_needed and not s3_sync_needed:
                    status_label.config(text="All data is already synced!")
                    progress_window.after(2000, progress_window.destroy)
                    return
                
                # First sync RDS records
                if rds_sync_needed:
                    status_label.config(text="Syncing RDS records...")
                    for i, record in enumerate(rds_sync_needed, 1):
                        record_id, company_user_uuid, username, date, in_time, out_time = record[:6]
                        in_camera, out_camera = record[8:10]
                        
                        details_label.config(
                            text=f"Syncing RDS record {i}/{len(rds_sync_needed)}\nUser: {username}"
                        )
                        progress_window.update()
                        
                        # Insert or update AWS RDS record
                        aws_cur.execute("""
                            INSERT INTO attendance_records 
                            (record_id, company_user_uuid, username, date, in_time, out_time,
                            in_time_with_camera, out_time_with_camera)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            ON CONFLICT (record_id) 
                            DO UPDATE SET
                                in_time = EXCLUDED.in_time,
                                out_time = EXCLUDED.out_time,
                                in_time_with_camera = EXCLUDED.in_time_with_camera,
                                out_time_with_camera = EXCLUDED.out_time_with_camera
                        """, (record_id, company_user_uuid, username, date, in_time, out_time, 
                            in_camera, out_camera))
                        
                        # Update sync status
                        local_cur.execute("""
                            INSERT OR REPLACE INTO sync_status 
                            (record_id, synced, last_sync_attempt)
                            VALUES (?, 1, datetime('now'))
                        """, (record_id,))
                        
                        local_conn.commit()
                        aws_conn.commit()
                
                # Then sync S3 files
                if s3_sync_needed:
                    status_label.config(text="Syncing S3 files...")
                    for i, (record_id, video_path, login_type) in enumerate(s3_sync_needed, 1):
                        details_label.config(
                            text=f"Syncing S3 file {i}/{len(s3_sync_needed)}\nFile: {os.path.basename(video_path)}"
                        )
                        progress_window.update()
                        
                        # Get username for the record
                        local_cur.execute("SELECT username FROM attendance_records WHERE record_id = ?", 
                                        (record_id,))
                        username = local_cur.fetchone()[0]
                        
                        # Upload to S3
                        s3_url = self.db_pool.upload_video_to_s3(video_path, username, login_type)
                        if s3_url:
                            # Update both local and AWS records with S3 URL
                            if login_type == 'IN_TIME':
                                local_cur.execute("""
                                    UPDATE attendance_records 
                                    SET video_path_in = ? 
                                    WHERE record_id = ?
                                """, (s3_url, record_id))
                                aws_cur.execute("""
                                    UPDATE attendance_records 
                                    SET video_path_in = %s 
                                    WHERE record_id = %s
                                """, (s3_url, record_id))
                            else:
                                local_cur.execute("""
                                    UPDATE attendance_records 
                                    SET video_path_out = ? 
                                    WHERE record_id = ?
                                """, (s3_url, record_id))
                                aws_cur.execute("""
                                    UPDATE attendance_records 
                                    SET video_path_out = %s 
                                    WHERE record_id = %s
                                """, (s3_url, record_id))
                            
                            local_conn.commit()
                            aws_conn.commit()
                
                status_label.config(text="Sync completed successfully!")
                details_label.config(text=f"""
                    RDS Records synced: {len(rds_sync_needed)}
                    S3 Files synced: {len(s3_sync_needed)}
                """)
                progress_window.after(3000, progress_window.destroy)
                
            except Exception as e:
                logging.error(f"Error during sync: {e}")
                status_label.config(text=f"Error during sync: {str(e)}")
            finally:
                if aws_conn:
                    self.db_pool.return_aws_connection(aws_conn)
                    
        except Exception as e:
            logging.error(f"Error in sync process: {e}")
            messagebox.showerror("Error", "Failed to start sync process")
            
            
            
    
    def verify_sync_status(self, record_id):
        """Verify if a record is properly synced between local and AWS"""
        try:
            local_conn = sqlite3.connect(self.db_pool.local_db_path)
            aws_conn = self.db_pool.get_aws_connection()
            
            if not aws_conn:
                return False, "No AWS connection"

            local_cur = local_conn.cursor()
            aws_cur = aws_conn.cursor()
            
            # Get local record
            local_cur.execute("""
                SELECT out_time 
                FROM attendance_records 
                WHERE record_id = ?
            """, (record_id,))
            local_record = local_cur.fetchone()
            
            # Get AWS record
            aws_cur.execute("""
                SELECT out_time 
                FROM attendance_records 
                WHERE record_id = %s
            """, (record_id,))
            aws_record = aws_cur.fetchone()
            
            if not local_record or not aws_record:
                return False, "Record not found in both databases"
                
            if local_record[0] == aws_record[0]:
                return True, "Sync successful"
            else:
                return False, f"Out times don't match: Local={local_record[0]}, AWS={aws_record[0]}"
                
        except Exception as e:
            return False, f"Error verifying sync: {e}"
        finally:
            if local_conn:
                local_conn.close()
            if aws_conn:
                self.db_pool.return_aws_connection(aws_conn)
                
                          

    def _verify_aws_connection(self) -> bool:
        """Verify AWS services are accessible"""
        try:
            # Check S3
            self.db_pool.s3_client.list_buckets()
            
            # Check RDS
            conn = self.db_pool.get_aws_connection()
            if conn:
                cur = conn.cursor()
                cur.execute("SELECT 1")
                cur.fetchone()
                self.db_pool.return_aws_connection(conn)
                return True
            return False
        except Exception as e:
            logging.error(f"AWS service verification failed: {e}")
            return False

    
    def _sync_databases(self):
        try:
            # Get local database path from DatabasePool
            local_conn = sqlite3.connect(self.db_pool.local_db_path)  # Fixed line
            aws_conn = None
            retry_count = 0
            max_retries = 3
            while retry_count < max_retries:
                try:
                    aws_conn = self.get_aws_connection()
                    if not aws_conn:
                        raise Exception("Could not establish AWS connection for sync")
                    break
                except Exception as e:
                    retry_count += 1
                    if retry_count == max_retries:
                        logging.error(f"Failed to get AWS connection after {max_retries} attempts")
                        return
                    time.sleep(1 * (2 ** (retry_count - 1)))
                    logging.warning(f"Retry {retry_count}: Getting AWS connection")
            try:
                # Sync users first
                self._sync_table(local_conn, aws_conn, "users")
                # Then sync attendance records
                self._sync_table(local_conn, aws_conn, "attendance_records")
                logging.info("Initial database sync completed successfully.")
            finally:
                local_conn.close()
                if aws_conn:
                    self.return_aws_connection(aws_conn)
        except Exception as e:
            logging.error(f"Error during initial database sync: {e}")
            raise



    def verify_database_times(self) -> bool:

        """Verifies that SQLite and AWS RDS times are synchronized."""
        return self.db_pool.verify_database_times()

        
        
            
    def _get_aws_time_with_retry(self):
        """Get AWS time with exponential backoff retry."""
        retry_count = 0
        max_retries = 3
        base_delay = 1  # Initial delay in seconds
        
        while retry_count < max_retries:
            try:
                aws_conn = self.db_pool.aws_pool.getconn()
                aws_cur = aws_conn.cursor()
                aws_cur.execute("SELECT CURRENT_TIMESTAMP AT TIME ZONE 'UTC'")
                aws_time = aws_cur.fetchone()[0]
                self.db_pool.aws_pool.putconn(aws_conn)
                return aws_time
                
            except Exception as e:
                retry_count += 1
                if retry_count == max_retries:
                    logging.error(f"Failed to get AWS time after {max_retries} attempts")
                    return None
                    
                delay = base_delay * (2 ** (retry_count - 1))  # Exponential backoff
                logging.warning(f"Retry {retry_count}: Getting AWS time in {delay} seconds")
                time.sleep(delay)
    
    def _attempt_time_sync(self):
        """Attempt to synchronize system time with NTP."""
        try:
            ntp_client = ntplib.NTPClient()
            response = ntp_client.request('pool.ntp.org', version=3)
            system_time = datetime.fromtimestamp(response.tx_time)
            
            if os.name == 'nt':  # Windows
                os.system(f'time {system_time.strftime("%H:%M:%S")}')
            else:  # Linux/Unix
                os.system(f'date -s "{system_time.strftime("%Y-%m-%d %H:%M:%S")}"')
                
            logging.info("System time synchronized with NTP server")
            
        except Exception as e:
            logging.error(f"Failed to synchronize system time: {e}")
    
    def check_internet_connection(self) -> Tuple[bool, Optional[str]]:
        """Checks internet connectivity with retry mechanism."""
        retry_count = 0
        
        while retry_count < self.max_retries:
            try:
                # Check multiple DNS servers
                for dns in ["8.8.8.8", "1.1.1.1", "208.67.222.222"]:
                    try:
                        socket.create_connection((dns, 53), timeout=3)
                        return self._verify_aws_connection()
                    except:
                        continue
                        
                raise ConnectionError("No DNS servers reachable")
                
            except Exception as e:
                retry_count += 1
                if retry_count == self.max_retries:
                    return False, f"Connection failed after {self.max_retries} attempts: {str(e)}"
                    
                delay = self.retry_delay * (2 ** (retry_count - 1))
                logging.warning(f"Retry {retry_count}: Checking connection in {delay} seconds")
                time.sleep(delay)
        
        return False, "Maximum retries exceeded"
    
    
    
   
    
    def verify_aws_skip(self, service: str) -> bool:
        """
        Verifies that AWS services are properly skipped when offline.
        Returns True if skip was successful.
        """
        with self.lock:
            if not self.is_online:
                logging.info(f"Offline mode: Skipping AWS {service} operations")
                # Add to sync queue for later processing
                self.sync_queue.put(('verify_skip', service))
                return True
            return False
    
    def start_network_monitoring(self):
        """Starts the network monitoring thread."""
        self.sync_thread = threading.Thread(target=self._monitor_network, daemon=True)
        self.sync_thread.start()
        
    
    
                
    # In NetworkSyncManager class, modify _sync_table method:

    def _sync_table(self, local_conn, aws_conn, table_name: str):
        try:
            local_cur = local_conn.cursor()
            aws_cur = aws_conn.cursor()
            
            # Get all records from both databases
            local_cur.execute(f"SELECT * FROM {table_name}")
            aws_cur.execute(f"SELECT * FROM {table_name}")
            
            local_records = {row['company_user_uuid']: dict(row) for row in local_cur.fetchall()}
            aws_records = {row['company_user_uuid']: dict(row) for row in aws_cur.fetchall()}
            
            logging.info(f"Local records for {table_name}: {local_records}")
            logging.info(f"AWS records for {table_name}: {aws_records}")
            
            # Sync missing or updated records in both directions
            self._sync_records(local_conn, aws_conn, table_name, local_records, aws_records)
        except Exception as e:
            logging.error(f"Error syncing table {table_name}: {e}")
            raise
    
    def _monitor_network(self):
        """Continuously monitor network connectivity with proper thread termination"""
        while not self.stop_thread.is_set():  # Use Event.is_set()
            try:
                is_connected = self._check_connectivity()
                with self.lock:
                    old_status = self.is_online
                    self.is_online = is_connected
                    
                    if self.is_online and (not old_status or self._has_pending_syncs()):
                        logging.info("Network connection available - checking for pending syncs")
                        self._sync_all_pending_data()
                    elif not self.is_online:
                        logging.warning("Network connection lost - switching to offline mode")
                
                # Check for stop event with timeout instead of sleep
                if self.stop_thread.wait(300):  # 5 minutes with early exit capability
                    break
                    
            except Exception as e:
                logging.error(f"Error in network monitoring: {e}")
                if self.stop_thread.wait(5):  # 5 seconds with early exit
                    break


    def force_stop_monitoring(self):
        """Force stop network monitoring thread"""
        try:
            self.stop_thread.set()
            
            # Give thread a chance to stop gracefully
            if self.sync_thread and self.sync_thread.is_alive():
                self.sync_thread.join(timeout=5)
                
                if self.sync_thread.is_alive():
                    logging.warning("Network monitoring thread did not stop gracefully")
                    # In Python, we can't force kill threads, so we just log and continue
                    
        except Exception as e:
            logging.error(f"Error force stopping network monitoring: {e}")
            
            
        logging.info("Network monitoring thread stopped successfully")
        
        
        
    def cleanup(self):
        """FIXED: Safe cleanup with proper thread termination"""
        try:
            logging.info("Starting NetworkSyncManager cleanup...")
            
            # Set stop flag using Event
            self.stop_thread.set()
            
            # Wait for thread to finish with longer timeout
            if hasattr(self, 'sync_thread') and self.sync_thread and self.sync_thread.is_alive():
                try:
                    self.sync_thread.join(timeout=10)  # Increased timeout to 10 seconds
                    if self.sync_thread.is_alive():
                        logging.warning("Sync thread did not terminate within 10 seconds - forcing cleanup")
                        # Don't try to force terminate, just log and continue
                    else:
                        logging.info("Sync thread terminated successfully")
                except Exception as e:
                    logging.error(f"Error joining sync thread: {e}")
            
            logging.info("Network sync manager cleanup completed")
            
        except Exception as e:
            logging.error(f"Error during network sync cleanup: {e}")
            
            

class TimeManager:
    def __init__(self, db_pool):
        self.ntp_servers = [
            'pool.ntp.org',
            'time.google.com',
            'time.cloudflare.com'
        ]
        self.db_pool = db_pool
        self.time_offset = 0
        self.last_ntp_sync = None
        self.ntp_sync_interval = 3600  # 1 hour

    def get_accurate_time(self) -> datetime:
        """Get accurate time with proper timezone handling"""
        if self._should_sync_ntp():
            self._sync_with_ntp()
        
        current_time = datetime.now()
        if current_time.tzinfo is None:
            current_time = self.db_pool.dhaka_tz.localize(current_time)
        
        return current_time + timedelta(seconds=self.time_offset)

    
    
    def _convert_to_utc(self, dt):
        """Ensure consistent UTC time handling"""
        if dt.tzinfo is None:
            dt = pytz.timezone('Asia/Dhaka').localize(dt)
        return dt.astimezone(pytz.UTC).replace(microsecond=0)



    def _should_sync_ntp(self) -> bool:
        if not self.last_ntp_sync:
            return True
        time_since_sync = (datetime.now() - self.last_ntp_sync).total_seconds()
        return time_since_sync >= self.ntp_sync_interval

    def _sync_with_ntp(self):
        """Sync with NTP servers with fallback"""
        for server in self.ntp_servers:
            try:
                client = ntplib.NTPClient()
                response = client.request(server, timeout=5)
                ntp_time = datetime.fromtimestamp(response.tx_time)
                system_time = datetime.now()
                self.time_offset = (ntp_time - system_time).total_seconds()
                self.last_ntp_sync = datetime.now()
                logging.info(f"NTP sync successful with {server}")
                return
            except Exception as e:
                continue
        logging.error("Failed to sync with any NTP server")
        
        
        
    def _adjust_system_time(self, ntp_time):
        """Adjust system time with proper privileges handling"""
        try:
            if os.name == 'nt':
                os.system(f'time {ntp_time.strftime("%H:%M:%S")}')
                os.system(f'date {ntp_time.strftime("%Y-%m-%d")}')
            else:
                os.system(f'sudo date -s "{ntp_time.isoformat()}"')
        except Exception as e:
            logging.error(f"Time adjustment failed: {e}")
            
            
    
        
        
    def validate_attendance_time(self, username: str, timestamp: datetime, 
                               is_in_time: bool) -> datetime:
        """
        Validates and corrects attendance time based on existing records.
        For in_time: Uses earlier time between local and AWS
        For out_time: Uses later time between local and AWS
        """
        try:
            # Get local and AWS times for comparison
            local_time = self._get_local_time(username, is_in_time)
            aws_time = self._get_aws_time(username, is_in_time)
            
            if local_time is None and aws_time is None:
                # No existing records, use provided timestamp
                return timestamp
                
            if is_in_time:
                # For in_time, use the earlier time
                valid_times = [t for t in [local_time, aws_time, timestamp] if t is not None]
                return min(valid_times)
            else:
                # For out_time, use the later time
                valid_times = [t for t in [local_time, aws_time, timestamp] if t is not None]
                return max(valid_times)
                
        except Exception as e:
            logging.error(f"Error validating attendance time: {e}")
            return timestamp
            
    
            
    def _get_local_time(self, username: str, is_in_time: bool) -> Optional[datetime]:
        """Get attendance time from local SQLite database."""
        try:
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            
            today = datetime.now().date().isoformat()
            time_field = "in_time" if is_in_time else "out_time"
            
            cur.execute(f"""
                SELECT {time_field} 
                FROM attendance_records 
                WHERE username = ? AND date = ?
            """, (username, today))
            
            result = cur.fetchone()
            if result and result[0]:
                return datetime.strptime(result[0], '%Y-%m-%d %H:%M:%S')
            return None
            
        except Exception as e:
            logging.error(f"Error getting local time: {e}")
            return None
            
    def _get_aws_time(self, username: str, is_in_time: bool) -> Optional[datetime]:
        """Get attendance time from AWS RDS database."""
        if not self.db_pool.is_online:
            return None
            
        try:
            conn = self.db_pool.get_aws_connection()
            if not conn:
                return None
                
            cur = conn.cursor()
            today = datetime.now().date()
            time_field = "in_time" if is_in_time else "out_time"
            
            cur.execute(f"""
                SELECT {time_field} 
                FROM attendance_records 
                WHERE username = %s AND date = %s
            """, (username, today))
            
            result = cur.fetchone()
            return result[0] if result and result[0] else None
            
        except Exception as e:
            logging.error(f"Error getting AWS time: {e}")
            return None
        finally:
            if conn:
                self.db_pool.return_aws_connection(conn)
                        
        
                
class AWSConfig:
    def __init__(self):
        # AWS Credentials - Better to use AWS IAM roles for EC2
        self.aws_access_key = aws_access_key
        self.aws_secret_key = aws_secret_key
        self.region = 'ap-southeast-1'  # e.g., 'us-east-1'
        
        # S3 Configuration
        self.s3_bucket = 'my-attendance-bucket-s3'
        self.s3_video_prefix = 'attendance_videos/'
        
        # RDS Configuration
        self.rds_host = rds_host
        self.rds_port = rds_port
        self.rds_database = rds_database
        self.rds_user = rds_user
        self.rds_password = rds_password

    def get_s3_client(self):
        return boto3.client(
            's3',
            aws_access_key_id=self.aws_access_key,
            aws_secret_access_key=self.aws_secret_key,
            region_name=self.region
        )






def serialize_face_embedding(embedding):
    """Serialize face embedding to JSON string"""
    return json.dumps(embedding.tolist())

def deserialize_face_embedding(embedding_json):
    """Deserialize face embedding from JSON string"""
    return np.array(json.loads(embedding_json))





class DatabasePool:
    _instance = None
    max_retries = 3
    retry_delay = 1

    def __init__(self):
        if DatabasePool._instance:
            raise Exception("DatabasePool is a singleton. Use get_instance() instead.")
        try:
            self.db_dir = os.path.join(os.path.expanduser('~'), 'db')
            os.makedirs(self.db_dir, exist_ok=True)
            logging.info(f"Ensured directory exists: {self.db_dir}")

            self.local_db_path = os.path.join(self.db_dir, 'local_attendance.db')
            logging.info(f"Local database path: {self.local_db_path}")
            
            self._local_conn_lock = threading.Lock()
            # Initialize basic attributes
            self.is_online = False
            self.local_conn = None
            self._local_conn_count = 0
            self.db_lock = threading.Lock()
            self.aws_pool = None
            self.s3_client = None
            
            # Set up timezone info
            self.dhaka_tz = pytz.timezone('Asia/Dhaka')
            self.utc_tz = pytz.utc
            
            # Check for .h5 model first
            home_dir = os.path.expanduser('~')
            deepface_dir = os.path.join(home_dir, '.deepface')
            weights_dir = os.path.join(deepface_dir, 'weights')
            model_path = os.path.join(weights_dir, 'facenet512_weights.h5')
            
            self.model_exists = os.path.exists(model_path)
            logging.info(f"Face recognition model exists: {self.model_exists}")
            
            # ALWAYS initialize local database first
            self._initialize_local_tables()
            
            # Try AWS connections, but don't fail if they're not available
            try:
                logging.info("Attempting to connect to AWS services...")
                self.aws_config = AWSConfig()
                
                # Initialize S3 client
                self.s3_client = self.aws_config.get_s3_client()
                if not self.s3_client:
                    logging.warning("Failed to initialize S3 client, continuing in offline mode")
                    return
                    
                # Initialize AWS RDS pool - Use ThreadedConnectionPool instead of SimpleConnectionPool
                self.aws_pool = psycopg2.pool.ThreadedConnectionPool(
                    minconn=1,
                    maxconn=5,  # Reduced for free tier
                    host=self.aws_config.rds_host,
                    database=self.aws_config.rds_database,
                    user=self.aws_config.rds_user,
                    password=self.aws_config.rds_password,
                    port=self.aws_config.rds_port,
                    connect_timeout=30,
                    keepalives=1,
                    keepalives_idle=30,
                    keepalives_interval=10,
                    keepalives_count=5
                )
                                
                # Test AWS connection
                test_conn = self.aws_pool.getconn()
                if test_conn and not test_conn.closed:
                    self.aws_pool.putconn(test_conn)
                    self.is_online = True
                    logging.info("Successfully connected to AWS services")
                    # Initialize AWS tables only if connection is successful
                    try:
                        self._initialize_aws_tables()
                    except Exception as e:
                        logging.warning(f"Non-critical: AWS tables initialization failed: {e}")
                else:
                    logging.warning("AWS RDS connection test failed, continuing in offline mode")
                    
            except Exception as e:
                logging.warning(f"AWS services unavailable: {e}")
                self.aws_pool = None
                self.s3_client = None
                
            # Set online status
            if not self.is_online:
                logging.info("Starting in offline mode, will sync when online")
                
            # Initialize FAISS index
            self.index = None
            self.faiss_index_path = os.path.join(self.db_dir, 'face_index.faiss')
            if os.path.exists(self.faiss_index_path):
                try:
                    self.index = faiss.read_index(self.faiss_index_path)
                    logging.info("Successfully loaded FAISS index")
                except Exception as e:
                    logging.error(f"Error loading FAISS index: {e}")
                    self.index = faiss.IndexFlatL2(512)
                    logging.info("Created new FAISS index")
            else:
                self.index = faiss.IndexFlatL2(512)
                logging.info("No FAISS Index Found")
        
        except Exception as e:
            logging.error(f"Error initializing DatabasePool: {e}")
            raise        
            
    @staticmethod
    def get_instance():
        if DatabasePool._instance is None:
            DatabasePool._instance = DatabasePool()
        return DatabasePool._instance



    def get_aws_connection(self):
        """Get AWS connection with improved error handling - separate AWS failures from internet failures"""
        retry_count = 0
        while retry_count < self.max_retries:
            try:
                # First check if we have internet
                if not self._check_basic_internet():
                    logging.warning("No internet connection available")
                    self.is_online = False
                    return None
                    
                # We have internet, try AWS connection
                if self.aws_pool:
                    conn = self.aws_pool.getconn()
                    
                    if conn and not conn.closed:
                        try:
                            # Test connection with a simple query
                            with conn.cursor() as cur:
                                cur.execute("SELECT 1")
                            logging.info("Successfully connected to AWS")
                            # AWS is working, but don't change is_online status here
                            return conn
                        except Exception as test_error:
                            logging.warning(f"AWS connection test failed: {test_error}")
                            try:
                                conn.close()
                            except:
                                pass
                            # AWS failed but we have internet - don't set offline
                            logging.error("AWS database is unavailable but internet is working")
                            return None
                    else:
                        logging.warning("Received closed or invalid AWS connection from pool")
                        return None
                else:
                    logging.warning("No AWS connection pool available")
                    return None
                    
            except Exception as e:
                retry_count += 1
                if retry_count == self.max_retries:
                    logging.error(f"Failed to get AWS connection after {self.max_retries} attempts: {e}")
                    # Don't set offline here - AWS failure doesn't mean no internet
                    return None
                sleep_time = self.retry_delay * (2 ** (retry_count - 1))
                logging.warning(f"Retrying AWS connection attempt {retry_count} after {sleep_time} seconds")
                time.sleep(sleep_time)
        
        return None


    def _check_basic_internet(self):
        """Check basic internet connectivity using DNS servers"""
        try:
            test_hosts = [
                ("8.8.8.8", 53),
                ("1.1.1.1", 53),
                ("208.67.222.222", 53)
            ]
            
            for host, port in test_hosts:
                try:
                    socket.create_connection((host, port), timeout=3)
                    return True
                except:
                    continue
            return False
        except:
            return False




    def return_aws_connection(self, conn):
        """Return AWS connection with proper error handling - NO key parameter for ThreadedConnectionPool"""
        if not conn:
            logging.warning("Attempted to return None connection")
            return
            
        # Check if pool exists and is not closed
        if not hasattr(self, 'aws_pool') or not self.aws_pool:
            logging.warning("No AWS pool available, closing connection directly")
            try:
                if not conn.closed:
                    conn.close()
            except:
                pass
            return
            
        # Check if pool is closed before attempting to return connection
        if self.aws_pool.closed:
            logging.warning("AWS pool is closed, closing connection directly")
            try:
                if not conn.closed:
                    conn.close()
            except:
                pass
            return
            
        try:
            # Only return connection if it's not closed and pool is open
            if not conn.closed:
                try:
                    # Quick connection test
                    with conn.cursor() as cur:
                        cur.execute("SELECT 1")
                    # Connection is good, return to pool
                    self.aws_pool.putconn(conn)
                    logging.debug("Successfully returned connection to pool")
                except Exception as test_error:
                    logging.warning(f"Connection test failed before return: {test_error}")
                    # Connection is bad, close it
                    try:
                        conn.close()
                    except:
                        pass
                    logging.debug("Closed bad connection instead of returning to pool")
            else:
                logging.warning("Connection is already closed, not returning to pool")
                
        except psycopg2.pool.PoolError as pool_error:
            logging.error(f"Pool error when returning connection: {pool_error}")
            # If we can't return to pool, close the connection
            try:
                if not conn.closed:
                    conn.close()
            except:
                pass
        except Exception as e:
            logging.error(f"Error returning AWS connection: {e}")
            # If we can't return to pool, try to close it
            try:
                if not conn.closed:
                    conn.close()
            except:
                pass





    def close_pool(self):
        """Properly close the connection pool"""
        if self.aws_pool:
            try:
                self.aws_pool.closeall()
                logging.info("All AWS connections closed")
            except Exception as e:
                logging.error(f"Error closing AWS pool: {e}")
                
    def get_connection_context(self):
        """Context manager for database connections to ensure proper cleanup"""
        class ConnectionContext:
            def __init__(self, pool):
                self.pool = pool
                self.conn = None
                self.cur = None
                
            def __enter__(self):
                self.conn = self.pool.get_local_connection()
                self.cur = self.conn.cursor()
                return self.conn, self.cur
                
            def __exit__(self, exc_type, exc_val, exc_tb):
                if self.cur:
                    self.cur.close()
                if self.conn:
                    self.conn.close()
                    
        return ConnectionContext(self)

    def get_aws_connection_context(self):
        """Context manager for AWS connections to ensure proper cleanup"""
        class AWSConnectionContext:
            def __init__(self, pool):
                self.pool = pool
                self.conn = None
                self.cur = None
                
            def __enter__(self):
                self.conn = self.pool.get_aws_connection()
                if self.conn:
                    self.cur = self.conn.cursor()
                    return self.conn, self.cur
                return None, None
                
            def __exit__(self, exc_type, exc_val, exc_tb):
                if self.cur:
                    try:
                        self.cur.close()
                    except:
                        pass
                if self.conn:
                    self.pool.return_aws_connection(self.conn)
                    
        return AWSConnectionContext(self)


           
       
    def initialize_database(self):
        """Initialize both local SQLite and AWS RDS databases."""
        try:
            # Initialize local SQLite database
            self._initialize_local_tables()
            logging.info("Local SQLite database initialized successfully")

            # Initialize AWS RDS database only if online
            if self.is_online and self.aws_pool:
                try:
                    self._initialize_aws_tables()
                    logging.info("AWS PostgreSQL database initialized successfully")
                except Exception as e:
                    logging.error(f"Failed to initialize AWS database: {e}")
                    self.is_online = False
        except Exception as e:
            logging.error(f"Database initialization error: {e}")
            raise
        
     



     
    
    def _initialize_local_tables(self):
        """Initialize local SQLite database tables."""
        conn = sqlite3.connect(self.local_db_path)
        try:
            cur = conn.cursor()
            
            # Users table (unchanged)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                company_user_uuid TEXT PRIMARY KEY,
                company_name TEXT,
                company_uuid TEXT,
                username TEXT UNIQUE,
                user_uuid TEXT UNIQUE,
                user_designation TEXT,
                user_android_id TEXT,
                face_encoding TEXT,
                password TEXT,
                created_at TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc')),
                last_updated TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc'))
            )
            """)
            
            # Attendance records table (unchanged)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS attendance_records (
                record_id TEXT PRIMARY KEY,
                company_user_uuid TEXT REFERENCES users(company_user_uuid),
                username TEXT NOT NULL,
                date TEXT NOT NULL,
                in_time TEXT,
                out_time TEXT,
                video_path_in TEXT,
                video_path_out TEXT,
                in_time_with_camera TEXT CHECK(in_time_with_camera IN ('Y', 'N')),
                out_time_with_camera TEXT CHECK(out_time_with_camera IN ('Y', 'N')),
                in_time_with_qr TEXT CHECK(in_time_with_qr IN ('Y', 'N')),
                out_time_with_qr TEXT CHECK(out_time_with_qr IN ('Y', 'N')),
                created_at TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc'))
            )
            """)

            # Company info table (unchanged)
           
            cur.execute("""
            CREATE TABLE IF NOT EXISTS company_info (
                company_id TEXT PRIMARY KEY,
                company_name TEXT NOT NULL UNIQUE,
                password TEXT NOT NULL,
                address_road TEXT,
                address_city TEXT,
                contact_person_name TEXT,
                contact_person_designation TEXT,
                contact_person_number TEXT,
                created_at TEXT,
                last_updated TEXT
            )
            """)

            
            # Sync status table (unchanged)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS sync_status (
                record_id TEXT PRIMARY KEY,
                synced INTEGER DEFAULT 0,
                last_sync_attempt TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc')),
                retry_count INTEGER DEFAULT 0
            )
            """)

            # UPDATED: Company-Oriented HR Password table
            cur.execute("""
            CREATE TABLE IF NOT EXISTS hr_passwords (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_uuid TEXT NOT NULL,
                company_name TEXT NOT NULL,
                password TEXT NOT NULL,
                created_at TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc')),
                last_updated TEXT DEFAULT (strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc')),
                UNIQUE(company_uuid)  -- One HR password per company
            )
            """)
            
            conn.commit()
            
        except Exception as e:
            logging.error(f"Error initializing local database tables: {e}")
            conn.rollback()
            raise
        finally:
            conn.close()

    def _initialize_aws_tables(self):
        """Initialize AWS PostgreSQL database tables."""
        conn = None
        try:
            conn = self.get_aws_connection()
            if not conn:
                raise Exception("Could not get AWS connection")
            cur = conn.cursor()

            # Users table (unchanged)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS users (
                company_user_uuid TEXT PRIMARY KEY,
                company_name TEXT,
                company_uuid TEXT,
                username TEXT UNIQUE,
                user_uuid TEXT UNIQUE,
                user_designation TEXT,
                user_android_id TEXT,
                face_encoding TEXT,
                password TEXT,
                created_at TEXT DEFAULT TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'),
                last_updated TEXT DEFAULT TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS')
            )
            """)

            # Attendance records table (unchanged)
            cur.execute("""
            CREATE TABLE IF NOT EXISTS attendance_records (
                record_id TEXT PRIMARY KEY,
                company_user_uuid TEXT REFERENCES users(company_user_uuid),
                username TEXT NOT NULL,
                date TEXT NOT NULL, 
                in_time TEXT,
                out_time TEXT,
                video_path_in TEXT,
                video_path_out TEXT,
                in_time_with_camera TEXT CHECK(in_time_with_camera IN ('Y', 'N')),
                out_time_with_camera TEXT CHECK(out_time_with_camera IN ('Y', 'N')),
                in_time_with_qr TEXT CHECK(in_time_with_qr IN ('Y', 'N')),
                out_time_with_qr TEXT CHECK(out_time_with_qr IN ('Y', 'N')),
                created_at TEXT DEFAULT TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS')
            )
            """)

            # Company info table (unchanged)

        
            # Updated Company info table with new fields
            cur.execute("""
            CREATE TABLE IF NOT EXISTS company_info (
                company_id TEXT PRIMARY KEY,
                company_name VARCHAR(255) NOT NULL UNIQUE,
                password VARCHAR(255) NOT NULL,
                address_road VARCHAR(255),
                address_city VARCHAR(255),
                contact_person_name VARCHAR(255),
                contact_person_designation VARCHAR(255),
                contact_person_number VARCHAR(50),
                created_at TEXT,
                last_updated TEXT
            )
            """)
            
                    
            # UPDATED: Company-Oriented HR Password table for AWS
            cur.execute("""
            CREATE TABLE IF NOT EXISTS hr_passwords (
                id SERIAL PRIMARY KEY,
                company_uuid TEXT NOT NULL,
                company_name TEXT NOT NULL,
                password TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                last_updated TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(company_uuid)  -- One HR password per company
            )
            """)

            conn.commit()

        except Exception as e:
            logging.error(f"Error initializing AWS database tables: {e}")
            if conn:
                conn.rollback()
            raise
        finally:
            if conn:
                self.return_aws_connection(conn)








    # REPLACE the existing determine_login_type method in DatabasePool class
    def determine_login_type(self, username):
        """
        FIXED: Determine login type by checking BOTH local and AWS databases
        - If no in-time exists in EITHER database for today -> IN_TIME  
        - If in-time exists in EITHER database -> OUT_TIME
        """
        try:
            # Convert username to uppercase
            username = username.upper()
            
            conn = self.get_local_connection()
            cur = conn.cursor()
            current_date = datetime.now().date().isoformat()
            
            # Check local database for today's in_time
            cur.execute("""
                SELECT in_time 
                FROM attendance_records 
                WHERE username = ? AND date = ?
            """, (username, current_date))
            
            local_result = cur.fetchone()
            local_has_in_time = local_result and local_result[0] is not None
            
            # FIXED: Also check AWS database if online
            aws_has_in_time = False
            if self.is_online:
                aws_conn = self.get_aws_connection()
                if aws_conn:
                    try:
                        aws_cur = aws_conn.cursor()
                        aws_cur.execute("""
                            SELECT in_time 
                            FROM attendance_records 
                            WHERE username = %s AND date::date = %s::date
                        """, (username, current_date))
                        
                        aws_result = aws_cur.fetchone()
                        aws_has_in_time = aws_result and aws_result[0] is not None
                        
                    except Exception as e:
                        logging.error(f"Error checking AWS for in_time: {e}")
                    finally:
                        self.return_aws_connection(aws_conn)
            
            # FIXED: If in_time exists in EITHER database, it's OUT_TIME
            if local_has_in_time or aws_has_in_time:
                logging.info(f"Found in_time for {username} - returning OUT_TIME")
                return "OUT_TIME"
            else:
                logging.info(f"No in_time found for {username} - returning IN_TIME")
                return "IN_TIME"
                
        except Exception as e:
            logging.error(f"Error determining login type: {e}")
            return "IN_TIME"  # Default to IN_TIME if error occurs


        
        
    
    def _convert_to_local(self, dt):
        """Convert datetime to local timezone (Asia/Dhaka)."""
        if dt.tzinfo is None:
            dt = self.utc_tz.localize(dt)  # Assume UTC if no timezone info
        return dt.astimezone(self.dhaka_tz)

    def _convert_to_text(self, dt):
        """Convert datetime to TEXT format YYYY-MM-DD HH:MM:SS."""
        return dt.strftime('%Y-%m-%d %H:%M:%S')


    def save_timestamps(self):
        """Save timestamps to both SQLite and AWS PostgreSQL."""
        # Get current time in UTC
        current_time = datetime.now(pytz.utc)
        
        # Convert to local timezone (Asia/Dhaka)
        local_time = self._convert_to_local(current_time)
        
        # Format as TEXT (YYYY-MM-DD HH:MM:SS)
        sqlite_time = self._convert_to_text(local_time)
        postgres_time = self._convert_to_text(local_time)
        
        logging.info(f"Local Time: {sqlite_time}")
        logging.info(f"AWS Time: {postgres_time}")
        
        # Save to SQLite
        conn = self.get_local_connection()
        try:
            cur = conn.cursor()
            cur.execute("""
            INSERT INTO some_table (timestamp_column)
            VALUES (?)
            """, (sqlite_time,))
            conn.commit()
        except Exception as e:
            logging.error(f"Error saving timestamp to SQLite: {e}")
            conn.rollback()
        finally:
            conn.close()
        
        # Save to AWS PostgreSQL
        aws_conn = self.get_aws_connection()
        if aws_conn:
            try:
                aws_cur = aws_conn.cursor()
                aws_cur.execute("""
                INSERT INTO some_table (timestamp_column)
                VALUES (%s)
                """, (postgres_time,))
                aws_conn.commit()
            except Exception as e:
                logging.error(f"Error saving timestamp to AWS PostgreSQL: {e}")
                aws_conn.rollback()
            finally:
                self.return_aws_connection(aws_conn)
                
                

    def sync_database_times(self):
        """Synchronize local and AWS database times."""
        try:
            # Get local database time
            local_conn = sqlite3.connect(self.local_db_path)
            local_cur = local_conn.cursor()
            local_cur.execute("SELECT MAX(in_time), MAX(out_time) FROM attendance_records")
            local_max_in_time, local_max_out_time = local_cur.fetchone() or (None, None)
            local_conn.close()

            # Get AWS database time
            aws_conn = self.get_aws_connection()
            if not aws_conn:
                logging.warning("Cannot synchronize database times: No AWS connection")
                return
            aws_cur = aws_conn.cursor()
            aws_cur.execute("SELECT MAX(in_time), MAX(out_time) FROM attendance_records")
            aws_max_in_time, aws_max_out_time = aws_cur.fetchone() or (None, None)
            self.return_aws_connection(aws_conn)

            # Calculate time offsets
            in_time_offset = timedelta(seconds=0)
            out_time_offset = timedelta(seconds=0)
            if local_max_in_time and aws_max_in_time:
                in_time_offset = local_max_in_time - aws_max_in_time
            if local_max_out_time and aws_max_out_time:
                out_time_offset = local_max_out_time - aws_max_out_time

            # Log time offsets
            logging.info(f"Local database time offset: {in_time_offset}, {out_time_offset}")
            logging.info(f"AWS database time offset: {-in_time_offset}, {-out_time_offset}")
        except Exception as e:
            logging.error(f"Error synchronizing database times: {e}")
    
    
   
   
    def sync_database_times(self):
        """Synchronize local and AWS database times."""
        try:
            # Get local database time
            local_conn = sqlite3.connect(self.local_db_path)
            local_cur = local_conn.cursor()
            local_cur.execute("SELECT MAX(in_time), MAX(out_time) FROM attendance_records")
            local_max_in_time, local_max_out_time = local_cur.fetchone()
            local_conn.close()

            # Get AWS database time
            aws_conn = self.get_aws_connection()
            aws_cur = aws_conn.cursor()
            aws_cur.execute("SELECT MAX(in_time), MAX(out_time) FROM attendance_records")
            aws_max_in_time, aws_max_out_time = aws_cur.fetchone()
            self.return_aws_connection(aws_conn)

            # Calculate time offsets
            in_time_offset = timedelta(seconds=0)
            out_time_offset = timedelta(seconds=0)

            if local_max_in_time and aws_max_in_time:
                in_time_offset = local_max_in_time - aws_max_in_time
            if local_max_out_time and aws_max_out_time:
                out_time_offset = local_max_out_time - aws_max_out_time

            # Update local and AWS database times
            self._update_local_database_time(in_time_offset, out_time_offset)
            self._update_aws_database_time(in_time_offset, out_time_offset)

            logging.info(f"Local database time offset: {in_time_offset}, {out_time_offset}")
            logging.info(f"AWS database time offset: {-in_time_offset}, {-out_time_offset}")

        except Exception as e:
            logging.error(f"Error synchronizing database times: {e}")
   
    def _get_local_records(self):
        """Retrieve local database records."""
        local_records = {}
        try:
            local_conn = sqlite3.connect(self.local_db_path)
            local_cur = local_conn.cursor()
            local_cur.execute("SELECT company_user_uuid, date, in_time, out_time FROM attendance_records")
            for row in local_cur:
                company_user_uuid, date, in_time, out_time = row
                local_records[company_user_uuid] = (in_time, out_time, date)
            local_conn.close()
            return local_records
        except Exception as e:
            logging.error(f"Error retrieving local database records: {e}")
            raise
        
    def _get_aws_records(self):
        """Retrieve AWS database records."""
        aws_records = {}
        try:
            aws_conn = self.get_aws_connection()
            aws_cur = aws_conn.cursor()
            aws_cur.execute("SELECT company_user_uuid, date, in_time, out_time FROM attendance_records")
            for row in aws_cur:
                company_user_uuid, date, in_time, out_time = row
                aws_records[company_user_uuid] = (in_time, out_time, date)
            self.return_aws_connection(aws_conn)
            return aws_records
        except Exception as e:
            logging.error(f"Error retrieving AWS database records: {e}")
            if aws_conn:
                self.return_aws_connection(aws_conn)
            raise
        
            
    def _sync_databases(self):
        """Sync data between local SQLite and AWS RDS databases."""
        try:
            # Get local and AWS database connections
            local_conn = sqlite3.connect(self.local_db_path)
            aws_conn = self.get_aws_connection()
            if not aws_conn:
                raise Exception("Could not establish AWS connection for sync")
            try:
                # Sync users first
                self._sync_table(local_conn, aws_conn, "users")
                # Then sync attendance records
                self._sync_table(local_conn, aws_conn, "attendance_records")
                logging.info("Initial database sync completed successfully.")
            finally:
                local_conn.close()
                self.return_aws_connection(aws_conn)
        except Exception as e:
            logging.error(f"Error during initial database sync: {e}")
            raise
   
   
    
    
    

        
        
    def get_local_connection(self):
        """Get local SQLite connection with dictionary row factory."""
        if not hasattr(self, '_local_conn_lock'):
            self._local_conn_lock = threading.Lock()
            
        with self._local_conn_lock:  # Add a lock to prevent race conditions
            if not self.local_conn or self.local_conn is None:
                try:
                    # Test if connection is valid
                    if self.local_conn:
                        try:
                            self.local_conn.execute("SELECT 1").fetchone()
                        except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                            # Connection is closed, create a new one
                            self.local_conn = None
                    
                    # Create new connection if needed
                    if not self.local_conn:
                        self.local_conn = sqlite3.connect(self.local_db_path)
                        self.local_conn.row_factory = sqlite3.Row  # Enable dictionary-style access
                    
                    self._local_conn_count += 1  # Increment counter
                except Exception as e:
                    logging.error(f"Error getting local connection: {e}")
                    # Create a fresh connection as fallback
                    try:
                        self.local_conn = sqlite3.connect(self.local_db_path)
                        self.local_conn.row_factory = sqlite3.Row
                        self._local_conn_count = 1
                    except Exception as e2:
                        logging.error(f"Critical error creating fallback connection: {e2}")
                        raise  # Re-raise the exception if we can't create a connection
            else:
                # Verify the connection is still valid before returning it
                try:
                    self.local_conn.execute("SELECT 1").fetchone()
                    self._local_conn_count += 1  # Increment counter for existing connection
                except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                    # Connection is invalid, create a new one
                    try:
                        self.local_conn = sqlite3.connect(self.local_db_path)
                        self.local_conn.row_factory = sqlite3.Row
                        self._local_conn_count = 1
                    except Exception as e:
                        logging.error(f"Critical error recreating invalid connection: {e}")
                        raise  # Re-raise the exception if we can't create a connection
                
            return self.local_conn




    def close_local_connection(self):
        self._local_conn_count -= 1  # Decrement counter
        if self._local_conn_count <= 0:  # Only close if no active uses
            if self.local_conn:
                self.local_conn.close()
                self.local_conn = None
            self._local_conn_count = 0    
            
 

     
   
    def upload_video_to_s3(self, video_path, username, video_type):
        """Upload video with offline handling."""
        if not self.is_online:
            logging.info("Offline mode: Skipping S3 upload")
            return None
        try:
            local_path_parts = os.path.normpath(video_path).split(os.sep)
            username_index = local_path_parts.index(username)
            s3_key = '/'.join(local_path_parts[username_index:])
            self.s3_client.upload_file(video_path, self.aws_config.s3_bucket, s3_key)
            url = f"s3://{self.aws_config.s3_bucket}/{s3_key}"
            logging.info(f"Video uploaded to S3: {url}")
            return url
        except Exception as e:
            logging.error(f"Error uploading to S3: {e}")
            self.is_online = False
            return None
        
        
        

    def save_user_to_db(self, username, video_path):
        """
        Save or update user with face encoding from video.
        """
        try:
            # Get face encoding from video
            new_encoding = self._get_face_encoding_from_video(video_path)
            if new_encoding is None:
                logging.error(f"Failed to extract face encoding from video for user {username}")
                return False
            
            # Check if user exists
            company_user_uuid = None
            if self._is_username_exists(username):
                logging.info(f"Username {username} already exists. Updating face encoding...")
                # Get company_user_uuid for existing user
                conn = self.get_local_connection()
                cur = conn.cursor()
                cur.execute("SELECT company_user_uuid FROM users WHERE username = ?", (username,))
                result = cur.fetchone()
                if result:
                    company_user_uuid = result[0]
            
            # Save or update user
            return self.save_or_update_user(username, new_encoding, company_user_uuid)
        except Exception as e:
            logging.error(f"Error in save_user_to_db: {e}")
            return False



        
    def _is_duplicate_user(self, new_encoding):
        """
        Check if the face encoding already exists in the local database.
        """
        conn = self.get_local_connection()
        try:
            cur = conn.cursor()
            cur.execute("SELECT face_encoding FROM users")
            all_encodings = cur.fetchall()
            
            for encoding in all_encodings:
                existing_encoding = json.loads(encoding[0])
                if self._compare_face_encodings(existing_encoding, new_encoding):
                    return True  # Duplicate user found
            return False  # No duplicate user
        except Exception as e:
            logging.error(f"Error checking for duplicate user: {e}")
            return False
        finally:
            if conn:
                conn.close()


    def _compare_face_embeddings(self, embedding1, embedding2):
        """
        Compare two face embeddings using DeepFace.
        Returns True if the faces match, False otherwise.
        """
        try:
            # Deserialize JSON strings back to arrays
            embedding1_array = deserialize_face_embedding(embedding1)
            embedding2_array = deserialize_face_embedding(embedding2)
            
            # Convert to numpy arrays if they aren't already
            embedding1_array = np.array(embedding1_array)
            embedding2_array = np.array(embedding2_array)
            
            # Calculate cosine similarity
            cosine_similarity = np.dot(embedding1_array, embedding2_array) / (
                np.linalg.norm(embedding1_array) * np.linalg.norm(embedding2_array)
            )
            
            # Convert to distance (1 - similarity)
            distance = 1 - cosine_similarity
            
            # Use a threshold to determine if faces match (lower distance = better match)
            threshold = 0.4  # Adjust this threshold as needed for Facenet512
            return distance < threshold
            
        except Exception as e:
            logging.error(f"Error comparing face embeddings: {e}")
            return False


    def get_latest_video_path(self, username: str, login_type: str) -> Optional[str]:
        """Get the most recent video path for a user"""
        try:
            now = datetime.now()
            year_folder = os.path.join(self.db_dir, username, str(now.year))
            month_folder = os.path.join(year_folder, now.strftime("%B"))
            
            if not os.path.exists(month_folder):
                return None
                
            # Find the most recent video file for today
            today_str = now.strftime('%d %b %Y')
            matching_files = [
                f for f in os.listdir(month_folder)
                if today_str in f and login_type in f
            ]
            
            if matching_files:
                # Return the most recent file
                latest_file = max(matching_files)
                return os.path.join(month_folder, latest_file)
                
            return None
            
        except Exception as e:
            logging.error(f"Error getting video path: {e}")
            return None


    def _save_login_to_local_db(self, company_user_uuid, username, login_type, validated_time):
        """Convert to UTC before storage"""
        utc_time = validated_time.astimezone(pytz.utc)
        date_str = utc_time.date().isoformat()
        time_str = utc_time.strftime('%Y-%m-%d %H:%M:%S')
        
        """Save login to local SQLite database with validated time"""
        conn = self.get_local_connection()
        try:
            cur = conn.cursor()
            date_str = validated_time.date().isoformat()
            time_str = validated_time.strftime('%Y-%m-%d %H:%M:%S')
            video_path = self.get_latest_video_path(username, login_type)
            if login_type == "IN TIME":
                cur.execute("""
                INSERT OR REPLACE INTO attendance_records 
                (company_user_uuid, username, date, in_time, video_path_in)
                VALUES (?, ?, ?, ?, ?)
                """, (company_user_uuid, username, date_str, time_str, video_path))
            else:
                cur.execute("""
                UPDATE attendance_records 
                SET out_time = ?, video_path_out = ?
                WHERE company_user_uuid = ? AND date = ?
                """, (time_str, video_path, company_user_uuid, date_str))
            conn.commit()
            return True
        except Exception as e:
            logging.error(f"Error saving to local database: {e}")
            if conn:
                conn.rollback()
            return False


    def _save_login_to_aws_db(self, company_user_uuid, username, login_type, validated_time):
        """Save login to AWS RDS with validated time"""
        retry_count = 0
        max_retries = 3
        base_delay = 1  # Initial delay in seconds

        while retry_count < max_retries:
            try:
                conn = self.get_aws_connection()
                if not conn:
                    return False

                cur = conn.cursor()
                video_path = self.get_latest_video_path(username, login_type)
                s3_url = self.upload_video_to_s3(video_path, username, login_type.lower().replace(' ', '_'))

        
                if login_type == "IN TIME":
                    cur.execute("""
                        INSERT INTO attendance_records 
                        (company_user_uuid, username, date, in_time, video_path_in, in_time_with_camera)
                        VALUES (%s, %s, %s, %s, %s, %s)
                        ON CONFLICT (company_user_uuid, date) 
                        DO UPDATE SET 
                            in_time = EXCLUDED.in_time, 
                            video_path_in = EXCLUDED.video_path_in,
                            in_time_with_camera = EXCLUDED.in_time_with_camera
                    """, (company_user_uuid, username, validated_time.date(), validated_time, 
                          s3_url, 'Y' if s3_url else 'N'))
                else:
                    cur.execute("""
                        UPDATE attendance_records 
                        SET out_time = %s, 
                            video_path_out = %s,
                            out_time_with_camera = %s
                        WHERE company_user_uuid = %s AND date = %s
                    """, (validated_time, s3_url, 'Y' if s3_url else 'N', 
                          company_user_uuid, validated_time.date()))

                conn.commit()



                self.return_aws_connection(conn)
                return True
            except Exception as e:
                retry_count += 1
                if retry_count == max_retries:
                    logging.error(f"Failed to save login to AWS after {max_retries} attempts: {e}")
                    return False

                delay = base_delay * (2 ** (retry_count - 1))  # Exponential backoff
                logging.warning(f"Retry {retry_count}: Saving login to AWS in {delay} seconds")
                time.sleep(delay)
            finally:
                if conn:
                    self.return_aws_connection(conn)
            
    
    
            
    
    
    
    def save_or_update_user(self, username, face_embedding, company_user_uuid=None):
        try:
            # Convert time to UTC
            current_time = datetime.now(pytz.UTC)
            utc_timestamp = current_time.strftime('%Y-%m-%d %H:%M:%S')

            if not company_user_uuid:
                company_user_uuid = str(uuid.uuid4())

            # Normalize embedding for FAISS
            embedding_array = np.array(json.loads(face_embedding))
            normalized_embedding = embedding_array / np.linalg.norm(embedding_array)
            
            # Update FAISS index
            if self.index is None:
                self.index = faiss.IndexFlatL2(512)
            self.index.add(normalized_embedding.reshape(1, -1))
            
            # Save FAISS index to the db_dir location
            try:
                faiss.write_index(self.index, self.faiss_index_path)
                logging.info("Successfully saved FAISS index")
            except Exception as e:
                logging.error(f"Error saving FAISS index: {e}")

            # Save to local database with UTC time
            conn = self.get_local_connection()
            cur = conn.cursor()
            cur.execute("""
                INSERT OR REPLACE INTO users 
                (company_user_uuid, username, face_encoding, created_at, last_updated)
                VALUES (?, ?, ?, ?, ?)
            """, (company_user_uuid, username, face_embedding, utc_timestamp, utc_timestamp))
            conn.commit()

            # Save to AWS if online
            if self.is_online and self.aws_pool:
                aws_conn = self.get_aws_connection()
                if aws_conn:
                    try:
                        aws_cur = aws_conn.cursor()
                        aws_cur.execute("""
                            INSERT INTO users 
                            (company_user_uuid, username, face_encoding, created_at, last_updated)
                            VALUES (%s, %s, %s, %s, %s)
                            ON CONFLICT (username) 
                            DO UPDATE SET 
                                face_encoding = EXCLUDED.face_encoding,
                                last_updated = EXCLUDED.last_updated
                        """, (company_user_uuid, username, face_embedding, utc_timestamp, utc_timestamp))
                        aws_conn.commit()
                    finally:
                        self.return_aws_connection(aws_conn)

            return True
        except Exception as e:
            logging.error(f"Error saving user: {e}")
            return False



    def verify_database_times(self) -> bool:
        """Check UTC times between databases"""
        try:
            # Get local UTC time
            local_conn = sqlite3.connect(self.local_db_path)
            local_cur = local_conn.cursor()
            local_cur.execute("SELECT strftime('%Y-%m-%d %H:%M:%S','now')")
            local_time = pytz.utc.localize(
                datetime.strptime(local_cur.fetchone()[0], '%Y-%m-%d %H:%M:%S')
            )

            # Get AWS UTC time
            aws_conn = self.get_aws_connection()
            aws_cur = aws_conn.cursor()
            aws_cur.execute("SELECT CURRENT_TIMESTAMP AT TIME ZONE 'UTC'")
            aws_time = aws_cur.fetchone()[0].replace(tzinfo=pytz.utc)

            # Compare UTC times
            time_diff = abs((aws_time - local_time).total_seconds())
            if time_diff > 300:  # 5-minute threshold
                logging.warning(f"UTC time mismatch: {time_diff} seconds")
                return False
            return True
        finally:
            if aws_conn:
                self.return_aws_connection(aws_conn)         
  
    
    def generate_record_id(self, company_user_uuid: str, date: datetime.date) -> str:
        """Generate record ID using company_user_uuid instead of company_user_uuid"""
        return f"{company_user_uuid}_{date.day}_{date.strftime('%b')}_{date.year}"
    
    
    
    def _is_username_exists(self, username):
        """Check if username exists in either local or AWS database"""
        try:
            # Check local database
            conn = self.get_local_connection()
            cur = conn.cursor()
            cur.execute("SELECT 1 FROM users WHERE username = ?", (username,))
            if cur.fetchone():
                return True

            # Check AWS if online
            if self.is_online and self.aws_pool:
                aws_conn = self.get_aws_connection()
                if aws_conn:
                    try:
                        aws_cur = aws_conn.cursor()
                        aws_cur.execute("SELECT 1 FROM users WHERE username = %s", (username,))
                        if aws_cur.fetchone():
                            return True
                    finally:
                        self.return_aws_connection(aws_conn)
            
            return False
            
        except Exception as e:
            logging.error(f"Error checking username existence: {e}")
            return False
    
    

    def _convert_to_utc(self, dt):
        """
        Convert datetime to UTC with proper timezone handling.
        Ensures that all timestamps are stored in UTC in the database.
        """
        if dt.tzinfo is None:
            # Assume the input time is in Asia/Dhaka timezone
            dt = self.dhaka_tz.localize(dt)
        return dt.astimezone(self.utc_tz).replace(microsecond=0)  # Remove fractional seconds






    def save_login_to_db(self, username, login_type, timestamp, filepath=None, with_camera=False):
        """Save login with improved error handling"""
        try:
            # Convert timestamp to UTC if it's not already
            if timestamp.tzinfo is None:
                local_tz = pytz.timezone('Asia/Dhaka')
                timestamp = local_tz.localize(timestamp)
            utc_timestamp = timestamp.astimezone(pytz.UTC)
            utc_time_str = utc_timestamp.strftime('%Y-%m-%d %H:%M:%S')
            utc_date_str = utc_timestamp.date().isoformat()

            # FIXED: Get company_user_uuid instead of company_user_uuid
            conn = self.get_local_connection()
            cur = conn.cursor()
            cur.execute("SELECT company_user_uuid FROM users WHERE username = ?", (username,))
            result = cur.fetchone()

            if not result:
                logging.error(f"User {username} not found in database")
                return False, False

            company_user_uuid = result[0]
            record_id = self.generate_record_id(company_user_uuid, utc_timestamp.date())

            # Determine if camera was used based on the button clicked
            used_camera = 'Y' if with_camera else 'N'

            local_success = False
            # Save to local database
            try:
                if login_type == "IN_TIME":
                    cur.execute("""
                        INSERT OR REPLACE INTO attendance_records 
                        (record_id, company_user_uuid, username, date, in_time, video_path_in, in_time_with_camera)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    """, (record_id, company_user_uuid, username, utc_date_str, utc_time_str, filepath, used_camera))
                else:  # OUT_TIME
                    cur.execute("""
                        UPDATE attendance_records 
                        SET out_time = ?,
                            video_path_out = ?,
                            out_time_with_camera = ?
                        WHERE record_id = ?
                    """, (utc_time_str, filepath, used_camera, record_id))

                conn.commit()
                local_success = True
            except Exception as e:
                logging.error(f"Error saving to local database: {e}")
                conn.rollback()
                return False, False

            # Save to AWS if online
            aws_success = False
            if self.is_online:
                aws_conn = self.get_aws_connection()
                if aws_conn:
                    try:
                        s3_url = None
                        if filepath:
                            s3_url = self.upload_video_to_s3(filepath, username, login_type)

                        aws_cur = aws_conn.cursor()
                        if login_type == "IN_TIME":
                            aws_cur.execute("""
                                INSERT INTO attendance_records 
                                (record_id, company_user_uuid, username, date, in_time, video_path_in, in_time_with_camera)
                                VALUES (%s, %s, %s, %s, %s, %s, %s)
                                ON CONFLICT (record_id) DO UPDATE SET 
                                    in_time = EXCLUDED.in_time,
                                    video_path_in = EXCLUDED.video_path_in,
                                    in_time_with_camera = EXCLUDED.in_time_with_camera
                            """, (record_id, company_user_uuid, username, utc_date_str, utc_time_str, s3_url, used_camera))
                        else:  # OUT_TIME
                            aws_cur.execute("""
                                UPDATE attendance_records 
                                SET out_time = %s,
                                    video_path_out = %s,
                                    out_time_with_camera = %s
                                WHERE record_id = %s
                            """, (utc_time_str, s3_url, used_camera, record_id))

                        aws_conn.commit()
                        aws_success = True
                    except Exception as e:
                        logging.error(f"Error saving to AWS: {e}")
                        aws_conn.rollback()
                    finally:
                        self.return_aws_connection(aws_conn)
                        
            if local_success:
                # Signal successful login to any observers
                if hasattr(self, 'on_login_success'):
                    self.on_login_success(username, login_type, timestamp, with_camera)
                            
            return local_success, aws_success

        except Exception as e:
            logging.error(f"Error in save_login_to_db: {e}")
            return False, False

      
        

      
        
            
        
        
            
        
    def _sync_records(self, local_conn, aws_conn, table_name: str, local_records: Dict, aws_records: Dict):
        """
        Syncs records between local SQLite and AWS PostgreSQL databases with improved error handling.
        """
        local_cur = local_conn.cursor()
        aws_cur = aws_conn.cursor()
        
        try:
            # Get all records from both databases with null checks
            local_cur.execute(f"SELECT * FROM {table_name}")
            aws_cur.execute(f"SELECT * FROM {table_name}")
            
            # Safely process local records with null handling
            local_records = {}
            for row in local_cur.fetchall():
                if row:  # Check if row exists
                    record_dict = dict(zip([column[0] for column in local_cur.description], row))
                    if record_dict.get('company_user_uuid'):  # Ensure key field exists
                        local_records[record_dict['company_user_uuid']] = record_dict
            
            # Safely process AWS records with null handling
            aws_records = {}
            aws_rows = aws_cur.fetchall()
            if aws_rows:  # Check if any rows returned
                for row in aws_rows:
                    if row:  # Check if row exists
                        record_dict = dict(zip([column.name for column in aws_cur.description], row))
                        if record_dict.get('company_user_uuid'):  # Ensure key field exists
                            aws_records[record_dict['company_user_uuid']] = record_dict
            
            # Sync records with validation
            for aws_id, aws_record in aws_records.items():
                if aws_id and aws_record:  # Validate record before processing
                    try:
                        if table_name == "users":
                            # Handle user table sync
                            self._sync_user_record(local_cur, aws_record, aws_id)
                        elif table_name == "attendance_records":
                            # Handle attendance record sync
                            self._sync_attendance_record(local_cur, aws_record)
                        
                        local_conn.commit()
                    except Exception as e:
                        logging.error(f"Error syncing record {aws_id}: {e}")
                        local_conn.rollback()
                        continue
                        
            # Log success
            logging.info(f"Successfully completed sync for table {table_name}")
                        
        except Exception as e:
            logging.error(f"Error in _sync_records for table {table_name}: {e}")
            raise
            
    def _sync_user_record(self, local_cur, aws_record, aws_id):
        """Handle user record synchronization with validation"""
        if not all(key in aws_record for key in ['username', 'face_encoding', 'created_at', 'last_updated']):
            raise ValueError(f"Missing required fields in AWS record {aws_id}")
            
        local_cur.execute("SELECT 1 FROM users WHERE company_user_uuid = ?", (aws_id,))
        exists = local_cur.fetchone() is not None
        
        if exists:
            local_cur.execute("""
                UPDATE users 
                SET username = ?, 
                    face_encoding = ?, 
                    created_at = ?, 
                    last_updated = ?
                WHERE company_user_uuid = ?
            """, (
                aws_record['username'],
                aws_record['face_encoding'],
                aws_record['created_at'],
                aws_record['last_updated'],
                aws_id
            ))
        else:
            local_cur.execute("""
                INSERT INTO users 
                (company_user_uuid, username, face_encoding, created_at, last_updated)
                VALUES (?, ?, ?, ?, ?)
            """, (
                aws_id,
                aws_record['username'],
                aws_record['face_encoding'],
                aws_record['created_at'],
                aws_record['last_updated']
            ))

    def _sync_attendance_record(self, local_cur, aws_record):
        """Handle attendance record synchronization with validation"""
        if not all(key in aws_record for key in ['record_id', 'company_user_uuid', 'username', 'date']):
            raise ValueError(f"Missing required fields in attendance record {aws_record.get('record_id')}")
            
        local_cur.execute("SELECT 1 FROM attendance_records WHERE record_id = ?", 
                        (aws_record['record_id'],))
        exists = local_cur.fetchone() is not None
        
        if exists:
            local_cur.execute("""
                UPDATE attendance_records 
                SET company_user_uuid = ?,
                    username = ?,
                    date = ?,
                    in_time = ?,
                    out_time = ?,
                    video_path_in = ?,
                    video_path_out = ?,
                    in_time_with_camera = ?,
                    out_time_with_camera = ?,
                    created_at = ?
                WHERE record_id = ?
            """, (
                aws_record['company_user_uuid'],
                aws_record['username'],
                aws_record['date'],
                aws_record.get('in_time'),
                aws_record.get('out_time'),
                aws_record.get('video_path_in'),
                aws_record.get('video_path_out'),
                aws_record.get('in_time_with_camera'),
                aws_record.get('out_time_with_camera'),
                aws_record.get('created_at'),
                aws_record['record_id']
            ))
        else:
            local_cur.execute("""
                INSERT INTO attendance_records (
                    record_id, company_user_uuid, username, date,
                    in_time, out_time,
                    video_path_in, video_path_out,
                    in_time_with_camera, out_time_with_camera,
                    created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                aws_record['record_id'],
                aws_record['company_user_uuid'],
                aws_record['username'],
                aws_record['date'],
                aws_record.get('in_time'),
                aws_record.get('out_time'),
                aws_record.get('video_path_in'),
                aws_record.get('video_path_out'),
                aws_record.get('in_time_with_camera'),
                aws_record.get('out_time_with_camera'),
                aws_record.get('created_at')
            ))



    def retry_failed_syncs(self):
        """Retry failed synchronizations with exponential backoff"""
        max_retries = 3
        base_delay = 1  # seconds
        
        try:
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            
            # Get failed syncs
            cur.execute("""
                SELECT record_id, retry_count 
                FROM sync_status 
                WHERE synced = 0 AND retry_count < ?
            """, (max_retries,))
            
            failed_syncs = cur.fetchall()
            
            for record_id, retry_count in failed_syncs:
                try:
                    # Calculate delay with exponential backoff
                    delay = base_delay * (2 ** retry_count)
                    time.sleep(delay)
                    
                    # Attempt to sync
                    success = self._sync_single_record(record_id)
                    
                    if success:
                        # Update sync status
                        cur.execute("""
                            UPDATE sync_status 
                            SET synced = 1,
                                last_sync_attempt = datetime('now'),
                                retry_count = retry_count + 1
                            WHERE record_id = ?
                        """, (record_id,))
                    else:
                        # Update retry count
                        cur.execute("""
                            UPDATE sync_status 
                            SET retry_count = retry_count + 1,
                                last_sync_attempt = datetime('now')
                            WHERE record_id = ?
                        """, (record_id,))
                    
                    conn.commit()
                    
                except Exception as e:
                    logging.error(f"Error retrying sync for record {record_id}: {e}")
                    conn.rollback()
                    continue
                    
        except Exception as e:
            logging.error(f"Error in retry_failed_syncs: {e}")
        finally:
            if conn:
                conn.close()

    def validate_and_sync_times(self, local_time, aws_time, record_id):
        """Validate and synchronize timestamps between local and AWS"""
        try:
            # Convert times to datetime objects if they're strings
            if isinstance(local_time, str):
                local_time = datetime.strptime(local_time, '%Y-%m-%d %H:%M:%S')
            if isinstance(aws_time, str):
                aws_time = datetime.strptime(aws_time, '%Y-%m-%d %H:%M:%S')
                
            # Handle case where either time is None
            if local_time is None and aws_time is None:
                return None
            elif local_time is None:
                return aws_time
            elif aws_time is None:
                return local_time
                
            # Compare times and choose appropriate one based on type
            if 'in_time' in record_id.lower():
                # For in_time, use earlier time
                return min(local_time, aws_time)
            else:
                # For out_time, use later time
                return max(local_time, aws_time)
                
        except Exception as e:
            logging.error(f"Error validating times for record {record_id}: {e}")
            # Return local time as fallback
            return local_time
            
        
        
    def _sync_table(self, local_conn, aws_conn, table_name: str):
        """
        Syncs a specific table between SQLite and AWS RDS.
        Handles data type conversions, UUID casting, and transaction rollbacks.
        """
        try:
            local_cur = local_conn.cursor()
            aws_cur = aws_conn.cursor()

            # Get all records from both databases
            local_cur.execute(f"SELECT * FROM {table_name}")
            aws_cur.execute(f"SELECT * FROM {table_name}")

            # Convert rows to dictionaries for consistent access
            local_records = {}
            for row in local_cur.fetchall():
                record_dict = dict(zip([column[0] for column in local_cur.description], row))
                local_records[record_dict['company_user_uuid']] = record_dict

            aws_records = {}
            for row in aws_cur.fetchall():
                record_dict = dict(zip([column.name for column in aws_cur.description], row))
                aws_records[record_dict['company_user_uuid']] = record_dict

            # Sync missing or updated records in both directions
            self._sync_records(local_conn, aws_conn, table_name, local_records, aws_records)
        except Exception as e:
            logging.error(f"Error syncing table {table_name}: {e}")
            raise


        
        
            
    def cleanup(self):
        """FIXED: Safe database cleanup with proper connection handling"""
        try:
            logging.info("Starting database pool cleanup...")
            
            # Close local connection first
            if hasattr(self, 'local_conn') and self.local_conn:
                try:
                    # Check if connection is still valid before closing
                    try:
                        self.local_conn.execute("SELECT 1").fetchone()
                        self.local_conn.close()
                        logging.info("Local database connection closed successfully")
                    except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                        # Connection already closed
                        logging.info("Local database connection was already closed")
                    except Exception as e:
                        logging.warning(f"Error testing local connection before close: {e}")
                        try:
                            self.local_conn.close()
                        except:
                            pass
                    finally:
                        self.local_conn = None
                except Exception as e:
                    logging.error(f"Error closing local connection: {e}")
                    self.local_conn = None
            
            # Close AWS connection pool with enhanced checks
            if hasattr(self, 'aws_pool') and self.aws_pool:
                try:
                    # Check if pool is already closed before attempting to close
                    if not hasattr(self.aws_pool, 'closed') or not self.aws_pool.closed:
                        # Close all connections in the pool
                        self.aws_pool.closeall()
                        logging.info("AWS connection pool closed successfully")
                    else:
                        logging.info("AWS connection pool was already closed")
                except Exception as e:
                    logging.error(f"Error closing AWS connection pool: {e}")
                finally:
                    self.aws_pool = None
                    
            logging.info("Database pool cleanup completed successfully")
            
        except Exception as e:
            logging.error(f"Error during database pool cleanup: {e}")



    
    def verify_database_connection(self):
        """Verify database connectivity and schema"""
        try:
            # Check local database
            conn = self.get_local_connection()
            cur = conn.cursor()
            
            # Verify users table
            cur.execute("""
                SELECT COUNT(*) 
                FROM sqlite_master 
                WHERE type='table' AND name='users'
            """)
            if cur.fetchone()[0] == 0:
                self._initialize_local_tables()
                
            # Test query
            cur.execute("SELECT COUNT(*) FROM users")
            cur.fetchone()
            
            # Check AWS connection if online
            if self.is_online and self.aws_pool:
                aws_conn = self.get_aws_connection()
                if aws_conn:
                    aws_cur = aws_conn.cursor()
                    aws_cur.execute("SELECT COUNT(*) FROM users")
                    aws_cur.fetchone()
                    self.return_aws_connection(aws_conn)
                    
            return True
            
        except Exception as e:
            logging.error(f"Database verification failed: {e}")
            return False


# Setup logging with proper error handling and PyInstaller compatibility
def setup_logging():
    """Setup logging with proper error handling and PyInstaller compatibility"""
    try:
        # Create logs directory in user's home folder
        log_dir = os.path.join(os.path.expanduser('~'), '.attendance_logs')
        os.makedirs(log_dir, exist_ok=True)

        # Define log file paths
        debug_log = os.path.join(log_dir, 'attendance_debug.log')
        complete_log = os.path.join(log_dir, 'attendance_complete.log')
        terminal_log = os.path.join(log_dir, 'terminal_output.log')

        # Delete existing log files if they exist
        for log_file in [debug_log, complete_log, terminal_log]:
            if os.path.exists(log_file):
                os.remove(log_file)

        # Configure root logger with 'w' mode to overwrite instead of append
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(complete_log, encoding='utf-8', mode='w'),
                logging.StreamHandler()
            ]
        )

        # Add debug file handler in write mode
        debug_handler = logging.FileHandler(debug_log, encoding='utf-8', mode='w')
        debug_handler.setLevel(logging.DEBUG)
        debug_handler.setFormatter(
            logging.Formatter('%(asctime)s,%(msecs)03d - %(levelname)s - %(message)s', 
                           '%Y-%m-%d %H:%M:%S')
        )
        logging.getLogger().addHandler(debug_handler)

        # Create terminal logger wrapper with write mode
        class TerminalLogger:
            def __init__(self, stream, log_file):
                self.stream = stream
                self.log_file = log_file
                self.encoding = 'utf-8'
                # Open file in write mode initially to clear it
                with open(self.log_file, 'w', encoding=self.encoding) as f:
                    f.write(f"Terminal log started at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                
            def write(self, message):
                try:
                    if message.strip():
                        timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S,%f')[:-3]
                        log_message = f"{timestamp} - {message.strip()}\n"
                        # Use append mode after initial clearing
                        with open(self.log_file, 'a', encoding=self.encoding) as f:
                            f.write(log_message)
                    if self.stream:
                        self.stream.write(message)
                except Exception as e:
                    print(f"Error writing to terminal log: {e}")

            def flush(self):
                if self.stream:
                    self.stream.flush()

        # Redirect stdout and stderr
        sys.stdout = TerminalLogger(sys.stdout, terminal_log)
        sys.stderr = TerminalLogger(sys.stderr, terminal_log)

        logging.info("Logging system initialized successfully")
        return True

    except Exception as e:
        print(f"Failed to initialize logging: {e}")
        # Set up basic console logging as fallback
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[logging.StreamHandler()]
        )
        return False
    
    
  
# Initialize logging
if not setup_logging():
    print("Warning: Using fallback logging configuration")

# Setup logging with proper error handling
try:
    # Get log file paths
    debug_log = resource_path('attendance_debug.log')
    complete_log = resource_path('attendance_complete.log')
    terminal_log = resource_path('terminal_output.log')
    custom_log = resource_path('custom_handler.log')

    # Configure root logger first
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(complete_log, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )

    # Create and configure file handlers
    file_handler = logging.FileHandler(debug_log, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(
        logging.Formatter('%(asctime)s,%(msecs)03d - %(levelname)s - %(message)s', 
                         '%Y-%m-%d %H:%M:%S')
    )
    logging.getLogger().addHandler(file_handler)

    # Initialize custom handler
    class LogHandler(logging.Handler):
        def __init__(self, log_file):
            super().__init__()
            self.logs = []
            self.log_file = log_file
            
            # Create or clear the log file
            with open(self.log_file, 'w', encoding='utf-8') as f:
                f.write(f"Custom handler log started at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

        def emit(self, record):
            try:
                msg = self.format(record)
                timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S,%f')[:-3]
                formatted_msg = f"{timestamp} - {record.levelname} - {msg}"
                self.logs.append(formatted_msg)
                
                # Write to log file with error handling
                try:
                    with open(self.log_file, 'a', encoding='utf-8') as f:
                        f.write(f"{formatted_msg}\n")
                except Exception as e:
                    print(f"Error writing to log file: {e}")
                    
            except Exception:
                self.handleError(record)



    # Add custom handler
    custom_handler = LogHandler(custom_log)
    custom_handler.setLevel(logging.INFO)
    logging.getLogger().addHandler(custom_handler)


except Exception as e:
    print(f"Error setting up logging: {e}")
    # Fallback to basic logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[logging.StreamHandler(sys.stdout)]
    )

def handle_exception(exc_type, exc_value, exc_traceback):
    """Global exception handler"""
    if issubclass(exc_type, KeyboardInterrupt):
        # Don't log keyboard interrupt
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
        return
        
    logging.error("Uncaught exception", exc_info=(exc_type, exc_value, exc_traceback))

sys.excepthook = handle_exception





class AndroidIDReceiver:
    def __init__(self):
        self.root = None
        self.server = None
        self.port = 5000
        self.session_id = str(uuid.uuid4())[:8]
        self.public_url = None
        self.ngrok_process = None
        self.android_id_callback = None
        
    def setup_ui(self):
        if not self.root:
            return
            
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        ttk.Label(main_frame, text="📱 Android ID Receiver", 
                 font=('Arial', 16, 'bold')).pack(pady=(0, 15))
        
        # Status
        self.status_var = tk.StringVar(value="🟡 Starting...")
        ttk.Label(main_frame, textvariable=self.status_var, 
                 font=('Arial', 11)).pack(pady=(0, 10))
        
        # QR Code
        qr_frame = ttk.LabelFrame(main_frame, text="📱 SCAN QR CODE", padding="15")
        qr_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.qr_label = ttk.Label(qr_frame, text="⏳ Setting up...", font=('Arial', 12))
        self.qr_label.pack(expand=True)
        
        # Android ID Display
        self.id_var = tk.StringVar(value="🟡 Waiting for device...")
        ttk.Label(main_frame, text="📲 Received Android ID:", font=('Arial', 11, 'bold')).pack(anchor='w')
        ttk.Label(main_frame, textvariable=self.id_var, 
                 font=('Courier', 10), foreground='blue',
                 background='lightyellow', relief='solid', 
                 padding="8").pack(fill=tk.X, pady=5)

    def start_server(self):
        try:
            self.server = HTTPServer(('0.0.0.0', self.port), self.make_handler())
            threading.Thread(target=self.server.serve_forever, daemon=True).start()
            self.status_var.set("🟢 Server running...")
        except Exception as e:
            self.status_var.set(f"🔴 Server failed: {e}")

    def setup_ngrok(self):
        try:
            # Kill existing ngrok
            subprocess.run(['taskkill', '/f', '/im', 'ngrok.exe'], capture_output=True)
            time.sleep(0.5)
            
            # Start ngrok
            self.ngrok_process = subprocess.Popen([
                'ngrok', 'http', str(self.port)
            ], stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            
            # Quick polling for tunnel URL
            for attempt in range(15):
                try:
                    time.sleep(0.2)
                    response = requests.get('http://127.0.0.1:4040/api/tunnels', timeout=2)
                    if response.status_code == 200:
                        tunnels = response.json().get('tunnels', [])
                        if tunnels and len(tunnels) > 0:
                            self.public_url = tunnels[0]['public_url']
                            self.root.after(0, self.generate_qr)
                            self.root.after(0, lambda: self.status_var.set("🌍 Ready for scanning!"))
                            return
                except requests.exceptions.RequestException:
                    continue
            
            self.root.after(0, lambda: self.status_var.set("❌ Tunnel setup failed"))
        except Exception as e:
            error_msg = str(e)
            self.root.after(0, lambda: self.status_var.set(f"❌ Setup failed: {error_msg}"))

    def make_handler(self):
        app = self
        
        class Handler(BaseHTTPRequestHandler):
            def do_GET(self):
                try:
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.send_header('Connection', 'close')
                    self.end_headers()
                    
                    response = {"status": "ready", "session_id": self.path[1:]}
                    self.wfile.write(json.dumps(response).encode())
                    self.wfile.flush()
                except Exception:
                    pass

            def do_POST(self):
                try:
                    self.send_response(200)
                    self.send_header('Content-Type', 'application/json')
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.send_header('Connection', 'close')
                    self.end_headers()
                    
                    if self.path == '/receive_android_id':
                        try:
                            content_length = int(self.headers.get('Content-Length', 0))
                            post_data = self.rfile.read(content_length)
                            data = json.loads(post_data.decode('utf-8'))
                            
                            android_id = data.get('android_id', '').strip()
                            if android_id and app.android_id_callback:
                                app.root.after(0, lambda: app.android_id_callback(android_id))
                                response = {"status": "success", "message": "ID received!"}
                            else:
                                response = {"status": "error", "message": "No ID provided"}
                        except Exception as e:
                            response = {"status": "error", "message": str(e)}
                        
                        self.wfile.write(json.dumps(response).encode())
                        self.wfile.flush()
                except Exception:
                    pass

            def do_OPTIONS(self):
                try:
                    self.send_response(200)
                    self.send_header('Access-Control-Allow-Origin', '*')
                    self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
                    self.send_header('Connection', 'close')
                    self.end_headers()
                except Exception:
                    pass

            def log_message(self, format, *args):
                pass
        
        return Handler

    def generate_qr(self):
        try:
            qr_url = f"{self.public_url}/{self.session_id}"
            
            qr = qrcode.QRCode(version=1, box_size=6, border=2)
            qr.add_data(qr_url)
            qr.make(fit=True)
            
            qr_img = qr.make_image(fill_color="black", back_color="white")
            img = qr_img.resize((200, 200), Image.NEAREST)
            
            self.tk_img = ImageTk.PhotoImage(img)
            self.qr_label.config(image=self.tk_img, text="")
            
        except Exception as e:
            self.qr_label.config(text=f"❌ QR Failed: {e}")

    def display_android_id(self, android_id):
        timestamp = time.strftime("%H:%M:%S")
        self.id_var.set(f"✅ {android_id} ({timestamp})")
        self.status_var.set("✅ Android ID received!")

    def cleanup(self):
        try:
            if self.server:
                self.server.shutdown()
            if self.ngrok_process:
                self.ngrok_process.terminate()
        except:
            pass

    
        
class App:


    def __init__(self):
        try:
            # Initialize main window
            self.main_window = tk.Tk()
            self.main_window.title("Face Recognition System")
            self.main_window.attributes('-fullscreen', True)

            # Track open windows to prevent duplicates
            self.open_windows = {}
            
            self.db_pool = DatabasePool.get_instance()
            self.setup_ui()
            
            # MOVE THIS IMMEDIATELY AFTER setup_ui() - BEFORE tensorflow loading
            self.initialize_hr_button_protection()
            
            # Initialize face recognition system with dependencies
            self.face_recognition_system = FaceRecognitionSystem(
                main_window=self.main_window,
                db_pool=self.db_pool,
                app=self
            )

            self.load_initial_attendance_status()    

            # Initialize startup manager first
            self.startup_manager = StartupAndAdminManager()
            self.startup_manager.initialize()
            
            # Initialize core components
            self.dhaka_tz = pytz.timezone('Asia/Dhaka')
            self.network_manager = None
            self.face_recognition_system = None
            self.tensorflow_ready = False
            self.file_lock = Lock()
            self.captured_frame = None
            self.closing = False
            self.time_manager = TimeManager(self)
            
            # Start component initialization in background
            threading.Thread(target=self.initialize_components, daemon=True).start()
            self.db_pool.on_login_success = self.update_attendance_display_after_login
            
            self.main_window.protocol("WM_DELETE_WINDOW", self.on_closing)
            
            # Initialize error handling and logging
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(levelname)s - %(message)s'
            )
            
            # Monitor sync status
            if hasattr(self, 'cloud_status_label') and hasattr(self, 'cloud_update_label'):
                self.monitor_sync_status(self.cloud_status_label, self.cloud_update_label)
            
            self.setup_global_hotkeys()
            
        except Exception as e:
            logging.error(f"Error in App initialization: {e}")
            messagebox.showerror("Initialization Error", 
                            "Failed to start application. Please check logs.")



    def start_background_init(self):
        """Start background initialization in separate threads"""
        # Initialize database
        threading.Thread(target=self.init_database, daemon=True).start()
        
        # Load TensorFlow
        threading.Thread(target=self.init_tensorflow, daemon=True).start()
        
        # Initialize network
        threading.Thread(target=self.init_network, daemon=True).start()

    def init_database(self):
        """Initialize database in background"""
        try:
            self.db_pool = DatabasePool.get_instance()
            self.db_pool.initialize_database()
            self.main_window.after(0, lambda: self.status_labels['database'].config(
                text="Database Ready ✓", fg="green"))
        except Exception as e:
            logging.error(f"Database initialization error: {e}")
            self.main_window.after(0, lambda: self.status_labels['database'].config(
                text="Database Error!", fg="red"))

    def init_network(self):
        """Initialize network in background"""
        try:
            self.network_manager = NetworkSyncManager(self.db_pool)
            self.network_manager.start_monitoring()
            self.main_window.after(0, lambda: self.status_labels['network'].config(
                text="Network Ready ✓", fg="green"))
        except Exception as e:
            logging.error(f"Network initialization error: {e}")
            self.main_window.after(0, lambda: self.status_labels['network'].config(
                text="Network Error!", fg="red"))

    def start(self):
        """Start the application"""
        try:
            self.main_window.mainloop()
        except Exception as e:
            logging.error(f"Error in main loop: {e}")
        finally:
            if not self.closing:
                self.cleanup()



    def setup_ui(self):
        """Creates an extraordinary, modern-designed UI for the attendance system"""
        # Apply a dark-themed modern style with gradient background
        self.main_window.configure(bg="#171721")
        
        # Create base containers with dynamic layout
        self.main_container = tk.Frame(self.main_window, bg="#171721")
        self.main_container.pack(fill='both', expand=True)
        
        # Create a background canvas for advanced gradient effect
        self.bg_canvas = tk.Canvas(self.main_container, highlightthickness=0, bg="#171721")
        self.bg_canvas.pack(fill='both', expand=True)
        self.bg_canvas.bind("<Configure>", self._draw_advanced_background)
        
        # Create content frame on top of gradient
        self.content_frame = tk.Frame(self.bg_canvas, bg="#171721")  # Use theme color instead of empty string
        self.bg_canvas.create_window(0, 0, anchor="nw", window=self.content_frame)
        self.content_frame_id = self.bg_canvas.create_window(0, 0, anchor="nw", window=self.content_frame)
        
        # Lock Windows key (unchanged functionality)
        try:
            import keyboard
            keyboard.block_key('win')
            self.main_window.protocol("WM_DELETE_WINDOW", lambda: [keyboard.unblock_key('win'), self.main_window.destroy()])
            logging.info("Windows key blocked successfully")
        except ImportError:
            logging.warning("keyboard module not found, Windows key not blocked")
        
        # Create a modern floating header with logo and title
        self._create_modern_header()
        
        # Create main content area with glass-morphism layout
        main_area = tk.Frame(self.content_frame, bg="#171721")  # Use theme color
        main_area.pack(fill='both', expand=True, padx=40, pady=(0, 30))
        
        # Create two-column layout with proper spacing
        left_panel = tk.Frame(main_area, bg="#171721", width=550)  # Increased width from 520 to 550
        left_panel.pack(side='left', fill='both', expand=True, padx=(0, 20))
        left_panel.pack_propagate(False)  # Maintain fixed width
        
        right_panel = tk.Frame(main_area, bg="#171721")  # Use theme color
        right_panel.pack(side='right', fill='both', expand=True, padx=(20, 0))
        
        # Reorganize left panel for better space usage
        # -- First add action buttons (moved up) with modern design
        self._create_modern_action_panel(left_panel)
        
        # -- Add modern status dashboard to left panel with animations (reduced vertical space)
        self._create_glass_status_dashboard(left_panel)
        
        # -- Add HR button below action panel with reduced spacing
        self._create_floating_hr_button(left_panel)
        
        # Add log console to right panel with terminal-inspired design
        self._create_modern_console(right_panel)
        
        # Footer removed to save space
        
        # Initialize custom animations
        self._setup_animations()
        
            # Force layout update
        left_panel.update_idletasks()
        

        self.main_window.after(500, self.create_simple_hr_button)

    def _create_glass_status_dashboard(self, parent):
        """Create a glass morphism status dashboard with modern metrics display"""
        status_frame = tk.Frame(parent, bg="#171721")  # Use theme color
        status_frame.pack(fill='x', pady=(0, 15))  # Reduced bottom padding from 25 to 15
        
        # Add glass effect
        self._add_glass_effect(status_frame, alpha=10)
        
        # Title for status section with glowing accent - more compact
        title_container = tk.Frame(status_frame, bg="#171721")  # Use theme color
        title_container.pack(fill='x', padx=25, pady=(15, 10))  # Reduced padding
        
        # Add accent line
        accent = tk.Frame(title_container, width=30, height=3, bg="#6c5ce7")
        accent.pack(side='left')
        
        status_title = tk.Label(title_container, text="SYSTEM STATUS", 
                            font=("Montserrat", 12, "bold"),  # Reduced font size from 14 to 12
                            bg="#171721", fg="#ffffff")
        status_title.pack(side='left', padx=(10, 0))
        
        # Create status cards container
        status_grid = tk.Frame(status_frame, bg="#171721")  # Use theme color
        status_grid.pack(fill='x', padx=20, pady=(0, 15))  # Reduced bottom padding
        
        # Status items with better visual design - rearranged per requirement
        status_items = {
            "system": {"title": "System Status", "icon": "💻", "initial": "Loaded"},
            "network": {"title": "Network", "icon": "🌐", "initial": "Initializing..."},
            "tensorflow": {"title": "AI Models", "icon": "🧠", "initial": "Loading..."},
            "database": {"title": "Database", "icon": "🗃️", "initial": "Connecting..."}
        }
        
        self.status_labels = {}
        self.status_indicators = {}
        self.status_progress = {}
        
        # Create modern status cards in grid - 4 in a row to save vertical space
        for idx, (key, data) in enumerate(status_items.items()):
            col = idx % 4  # 4 columns instead of 2
            
            # Create card container with glass effect - make it more compact
            card = tk.Frame(status_grid, bg="#171721", padx=10, pady=10)  # Reduced padding
            card.grid(row=0, column=col, padx=4, pady=4, sticky="nsew")  # Reduced padding
            
            # Add glass effect to card
            self._add_glass_effect(card, alpha=15, highlight_top=True)
            
            # Status icon and title in card header
            header = tk.Frame(card, bg="#171721")  # Use theme color
            header.pack(fill='x')
            
            # Icon with glow effect - smaller
            icon_frame = tk.Frame(header, bg="#171721", width=25, height=25)  # Reduced size
            icon_frame.pack(side='left')
            icon_frame.pack_propagate(False)
            
            icon_label = tk.Label(icon_frame, text=data["icon"], 
                                font=("Segoe UI", 14),  # Reduced font size
                                bg="#171721", fg="#a29bfe")
            icon_label.place(relx=0.5, rely=0.5, anchor="center")
            
            title_label = tk.Label(header, text=data["title"], 
                                font=("Montserrat", 10, "bold"),  # Reduced font size
                                bg="#171721", fg="#ffffff")
            title_label.pack(side='left', padx=(8, 0))  # Reduced padding
            
            # Modern progress bar instead of indicator light
            progress_frame = tk.Frame(card, height=3, bg="#2d2d4a",  # Reduced height
                                relief="flat", bd=0, pady=6)  # Reduced padding
            progress_frame.pack(fill='x', pady=(8, 4))  # Reduced padding
            
            # Progress indicator that will animate
            progress_bar = tk.Frame(progress_frame, width=0, height=3, bg="#6c5ce7")  # Reduced height
            progress_bar.place(x=0, y=0)
            self.status_progress[key] = progress_bar
            
            # Status text with modern font - smaller
            self.status_labels[key] = tk.Label(card, text=data["initial"], 
                                            font=("Montserrat", 9),  # Reduced font size
                                            bg="#171721", fg="#dfe6e9")
            self.status_labels[key].pack(anchor='w', pady=(4, 0))  # Reduced padding
            
            # Hidden status indicator for compatibility
            self.status_indicators[key] = tk.Frame(card, width=0, height=0)
        
        # Make grid columns equal width
        for i in range(4):  # 4 columns
            status_grid.columnconfigure(i, weight=1)

    def _create_modern_action_panel(self, parent):
        """Create a panel with modern, animated action buttons"""
        action_frame = tk.Frame(parent, bg="#171721")
        action_frame.pack(fill='x', pady=(5, 10))  # Reduced bottom padding
        
        # Add glass effect
        self._add_glass_effect(action_frame, alpha=10)
        
        # Title container with accent line
        title_container = tk.Frame(action_frame, bg="#171721")
        title_container.pack(fill='x', padx=25, pady=(15, 10))  # Reduced padding
        
        # Add accent line
        accent = tk.Frame(title_container, width=30, height=3, bg="#00b894")
        accent.pack(side='left')
        
        action_title = tk.Label(title_container, text="ACTIONS", 
                            font=("Montserrat", 12, "bold"),  # Reduced font size
                            bg="#171721", fg="#ffffff")
        action_title.pack(side='left', padx=(10, 0))
        
        # Button container with proper spacing
        buttons_container = tk.Frame(action_frame, bg="#171721")
        buttons_container.pack(fill='x', padx=20, pady=(0, 15))  # Reduced bottom padding
        
        # Define action buttons with vibrant gradients
        button_data = [
            {
                "text": "In & Out with Camera", 
                "icon": "📷", 
                "command": self.handle_camera_login,
                "gradient": ["#00b894", "#00d2d3"]  # Teal gradient
            },
            {
                "text": "In & Out without Camera", 
                "icon": "🚪", 
                "command": self.handle_no_camera_login,
                "gradient": ["#fdcb6e", "#ffeaa7"]  # Yellow gradient
            },
            {
                "text": "In & Out with QR Code (Online)", 
                "icon": "📲", 
                "command": self.handle_qr_login,
                "gradient": ["#74b9ff", "#81ecec"]  # Blue gradient
            }
        ]
        
        # Create gradient buttons with hover animation - slightly smaller
        for btn_data in button_data:
            self._create_gradient_button(
                buttons_container, 
                btn_data["text"], 
                btn_data["icon"], 
                btn_data["command"],
                btn_data["gradient"]
            )

    def _create_gradient_button(self, parent, text, icon, command, gradient_colors):
        """Creates a button with gradient background and hover animation - taller for important buttons"""
        # Button container with increased height for important buttons (1.5x taller)
        # Increased from 50 to 75 for important buttons (50 * 1.5 = 75)
        btn_container = tk.Frame(parent, bg="#171721", height=75)
        btn_container.pack(fill='x', pady=6)  # Slightly increased padding
        btn_container.pack_propagate(False)  # Prevent inner contents from affecting frame size
        
        # Create gradient background
        gradient_canvas = tk.Canvas(btn_container, highlightthickness=0, bg="#171721")
        gradient_canvas.place(x=0, y=0, relwidth=1, relheight=1)
        
        # Draw gradient
        def draw_gradient(width, height):
            gradient_canvas.delete("gradient")
            for i in range(width):
                # Calculate color at this position
                ratio = i / width
                r1, g1, b1 = self._hex_to_rgb(gradient_colors[0])
                r2, g2, b2 = self._hex_to_rgb(gradient_colors[1])
                r = int(r1 + (r2 - r1) * ratio)
                g = int(g1 + (g2 - g1) * ratio)
                b = int(b1 + (b2 - b1) * ratio)
                color = f"#{r:02x}{g:02x}{b:02x}"
                gradient_canvas.create_line(i, 0, i, height, fill=color, tags="gradient")
        
        # Update gradient on resize
        gradient_canvas.bind("<Configure>", lambda e: draw_gradient(e.width, e.height))
        
        # Create button content - modify to use relative width
        content_frame = tk.Frame(btn_container, bg="#171721")
        content_frame.place(relx=0.5, rely=0.5, anchor="center", relwidth=0.98, relheight=0.9)
        
        # Icon with larger size since buttons are taller
        icon_label = tk.Label(content_frame, text=icon, font=("Segoe UI", 22), bg="#171721", fg="white")  # Increased font size
        icon_label.pack(side='left', padx=(20, 10))
        
        # Text with larger font
        text_label = tk.Label(content_frame, text=text, font=("Montserrat", 13, "bold"), bg="#171721", fg="white")  # Increased font size
        text_label.pack(side='left', fill='x', expand=True)
        
        # Add hover effect with hand cursor
        def on_enter(e):
            # Change cursor to hand
            widget = e.widget
            widget.config(cursor="hand2")  # Set cursor to hand/finger pointer
            
            # Scale up slightly
            btn_container.config(height=78)  # Slight increase on hover
            # Brighten text
            text_label.config(fg="#ffffff")
            icon_label.config(fg="#ffffff")
        
        def on_leave(e):
            # Restore cursor
            widget = e.widget
            widget.config(cursor="")
            
            # Scale back to normal
            btn_container.config(height=75)  # Return to base height
            # Normal text color
            text_label.config(fg="#f5f5f5")
            icon_label.config(fg="#f5f5f5")
        
        # Add binding for hover effect
        for widget in [btn_container, gradient_canvas, content_frame, icon_label, text_label]:
            widget.bind("<Enter>", on_enter)
            widget.bind("<Leave>", on_leave)
            widget.bind("<Button-1>", lambda e: command())
        
        return btn_container


    
    
    
    def create_simple_hr_button(self):
        """Create a simple, always-visible HR button at bottom-left - ONLY ONCE"""
        try:
            # Check if button already exists and is valid
            if hasattr(self, 'simple_hr_button'):
                try:
                    if (self.simple_hr_button.winfo_exists() and 
                        self.simple_hr_button.winfo_viewable()):
                        # Button already exists and is visible, don't create another
                        logging.info("HR button already exists and is visible")
                        return self.simple_hr_button
                    else:
                        # Button exists but not visible, destroy it first
                        self.simple_hr_button.destroy()
                except:
                    # Button reference is invalid, clear it
                    self.simple_hr_button = None
            
            # Create a simple HR button directly on the main window
            self.simple_hr_button = tk.Button(
                self.main_window,
                text="👥 HR MANAGEMENT",
                command=self.handle_hr_window,  # This calls the authentication window
                bg="#2C3E50",           # Dark blue-gray background
                fg="white",             # White text
                font=("Arial", 11, "bold"),
                relief="raised",
                bd=3,
                padx=15,
                pady=8,
                cursor="hand2",
                activebackground="#34495E",  # Slightly lighter when clicked
                activeforeground="white"
            )
            
            # Position at bottom-left corner - this is ALWAYS visible
            # Make sure window is updated first
            self.main_window.update_idletasks()
            window_height = self.main_window.winfo_height()
            if window_height <= 1:  # Window not yet rendered
                window_height = 800  # Default height
                
            self.simple_hr_button.place(x=20, y=window_height-60, width=200, height=40)
            
            # Bind resize event to keep button positioned correctly
            def reposition_hr_button(event=None):
                try:
                    if hasattr(self, 'simple_hr_button') and self.simple_hr_button.winfo_exists():
                        window_height = self.main_window.winfo_height()
                        if window_height > 100:  # Valid window height
                            self.simple_hr_button.place(x=20, y=window_height-60, width=200, height=40)
                except Exception as e:
                    logging.error(f"Error repositioning HR button: {e}")
            
            # Bind to window resize
            self.main_window.bind('<Configure>', reposition_hr_button)
            
            # Store positioning function for later use
            self.reposition_hr_button = reposition_hr_button
            
            # Add hover effects
            def on_enter(event):
                self.simple_hr_button.config(bg="#E74C3C", fg="white")  # Red on hover
            
            def on_leave(event):
                self.simple_hr_button.config(bg="#2C3E50", fg="white")  # Back to original
            
            self.simple_hr_button.bind("<Enter>", on_enter)
            self.simple_hr_button.bind("<Leave>", on_leave)
            
            # Force immediate positioning after a brief delay to ensure window is rendered
            self.main_window.after(100, reposition_hr_button)
            
            logging.info("Simple HR button created immediately at bottom-left")
            
            return self.simple_hr_button
            
        except Exception as e:
            logging.error(f"Error creating simple HR button: {e}")
            return None
        
        
    def recreate_hr_button_if_missing(self):
        """Simple check and recreate if needed - PREVENT MULTIPLE CREATION"""
        try:
            hr_exists = False
            if hasattr(self, 'simple_hr_button'):
                try:
                    if (self.simple_hr_button.winfo_exists() and 
                        self.simple_hr_button.winfo_viewable()):
                        hr_exists = True
                except:
                    hr_exists = False
            
            if not hr_exists:
                # Only log if we're actually creating a new button
                logging.info("Creating simple HR button...")
                self.create_simple_hr_button()
        
        except Exception as e:
            logging.error(f"Error in recreate_hr_button_if_missing: {e}")


    def _create_floating_hr_button(self, parent):
        """Simplified - just ensure simple HR button exists"""
        # Don't create another button here, just ensure the simple one exists
        if not hasattr(self, 'simple_hr_button') or not self.simple_hr_button.winfo_exists():
            return self.create_simple_hr_button()
        else:
            return self.simple_hr_button

    # Update the force_create_hr_button method:

    def force_create_hr_button(self):
        """Force create HR button - ONLY IF IT DOESN'T EXIST"""
        try:
            # Check if button already exists
            if hasattr(self, 'simple_hr_button'):
                try:
                    if self.simple_hr_button.winfo_exists() and self.simple_hr_button.winfo_viewable():
                        # Button already exists and is visible
                        logging.info("HR button already exists, skipping creation")
                        return
                except:
                    pass
            
            # Wait for window to be fully rendered
            self.main_window.update_idletasks()
            
            # Create the HR button
            self.create_simple_hr_button()
            
            # Force reposition after a moment
            self.main_window.after(200, lambda: self.reposition_hr_button())
            
            logging.info("Forced HR button creation completed")
            
        except Exception as e:
            logging.error(f"Error in force_create_hr_button: {e}")


    def start_hr_button_monitor(self):
        """Simple monitoring for HR button - REDUCED FREQUENCY"""
        def monitor():
            try:
                self.recreate_hr_button_if_missing()
                if hasattr(self, 'main_window') and self.main_window.winfo_exists():
                    # Check less frequently to reduce log spam
                    self.main_window.after(10000, monitor)  # Check every 10 seconds instead of 5
            except:
                pass
        
        # Start monitoring with initial delay
        if hasattr(self, 'main_window') and self.main_window.winfo_exists():
            self.main_window.after(3000, monitor)  # Start after 3 seconds

    def initialize_hr_button_protection(self):
        """Initialize HR button protection system - SINGLE CREATION"""
        try:
            # Create button immediately - ONLY ONCE
            self.main_window.after(100, self.create_simple_hr_button)
            
            # Start monitoring with reduced frequency
            self.main_window.after(2000, self.start_hr_button_monitor)
            
            # Bind window events for repositioning only (not recreation)
            def on_window_event(event=None):
                # Only reposition, don't recreate
                if hasattr(self, 'reposition_hr_button'):
                    self.main_window.after(100, self.reposition_hr_button)
            
            self.main_window.bind('<Map>', on_window_event)
            self.main_window.bind('<Visibility>', on_window_event)
            
            logging.info("Simple HR button protection system initialized")
            
        except Exception as e:
            logging.error(f"Error initializing HR button protection: {e}")


    def on_tensorflow_loaded(self):
        """Call this when TensorFlow loading is complete"""
        try:
            # Your existing tensorflow completion code...
            
            # Ensure HR button is created after TensorFlow loads
            #self.main_window.after(500, self.create_simple_hr_button)
            
            logging.info("TensorFlow loaded - Simple HR button activated")
            
        except Exception as e:
            logging.error(f"Error in tensorflow completion: {e}")

    # 9. Optional: Force HR button creation right after UI is ready:



    def initialize_hr_button_protection(self):
        """Initialize HR button protection system - CREATE IMMEDIATELY"""
        try:
            # Create button IMMEDIATELY - no delay
            self.create_simple_hr_button()
            
            # Start monitoring with reduced frequency
            self.main_window.after(2000, self.start_hr_button_monitor)
            
            # Bind window events for repositioning only (not recreation)
            def on_window_event(event=None):
                # Only reposition, don't recreate
                if hasattr(self, 'reposition_hr_button'):
                    self.main_window.after(100, self.reposition_hr_button)
            
            self.main_window.bind('<Map>', on_window_event)
            self.main_window.bind('<Visibility>', on_window_event)
            
            logging.info("HR button protection system initialized with immediate button creation")
            
        except Exception as e:
            logging.error(f"Error initializing HR button protection: {e}")

    

    
    def _create_modern_header(self):
        """Create a floating header with modern design elements"""
        header = tk.Frame(self.content_frame, bg="#171721")  # Use theme color
        header.pack(fill='x', pady=(20, 25))  # Reduced padding
        
        # Add glass effect to header
        self._add_glass_effect(header, alpha=15)
        
        # Left side with logo and interactive particles
        logo_frame = tk.Frame(header, bg="#171721")  # Use theme color
        logo_frame.pack(side='left', padx=(40, 0))
        
        # Create an interactive logo using Canvas - smaller
        logo_canvas = tk.Canvas(logo_frame, width=55, height=55, bg="#171721", 
                            highlightthickness=0)  # Reduced size
        logo_canvas.pack(side='left')
        
        # Draw modern logo with glowing effect
        self._draw_glowing_logo(logo_canvas)
        
        # App title with modern typography - smaller
        title_frame = tk.Frame(header, bg="#171721")  # Use theme color
        title_frame.pack(side='left', padx=15)
        
        main_title = tk.Label(title_frame, text="ATTENDANCE", 
                            font=("Montserrat", 24, "bold"),  # Reduced from 28 to 24
                            bg="#171721", fg="#ffffff")
        main_title.pack(anchor='w')
        
        sub_title = tk.Label(title_frame, text="MANAGEMENT SYSTEM", 
                            font=("Montserrat", 11),  # Reduced from 12 to 11
                            bg="#171721", fg="#a29bfe")
        sub_title.pack(anchor='w')
        
        # Right side with digital time display
        time_frame = tk.Frame(header, bg="#171721")  # Use theme color
        time_frame.pack(side='right', padx=(0, 40))
        
        self.time_label = tk.Label(time_frame, text="00:00:00", 
                                font=("Consolas", 20, "bold"),  # Reduced from 22 to 20
                                bg="#171721", fg="#ffffff")
        self.time_label.pack(anchor='e')
        
        self.date_label = tk.Label(time_frame, text="Wednesday, April 23", 
                                font=("Consolas", 9),  # Reduced from 10 to 9
                                bg="#171721", fg="#a29bfe")
        self.date_label.pack(anchor='e')
        
        # Update time with animation
        self._update_animated_time()
    
    

    def _create_modern_console(self, parent):
        """Create a modern terminal-inspired console panel with attendance info"""
        console_frame = tk.Frame(parent, bg="#171721")
        console_frame.pack(fill='both', expand=True)
        
        # Add glass effect
        self._add_glass_effect(console_frame, alpha=10)
        
        # Create two sections with equal height
        upper_section = tk.Frame(console_frame, bg="#171721")
        upper_section.pack(fill='both', expand=True)
        
        lower_section = tk.Frame(console_frame, bg="#171721")
        lower_section.pack(fill='both', expand=True)
        
        # Upper section (System Console)
        console_header = tk.Frame(upper_section, bg="#171721")
        console_header.pack(fill='x', padx=20, pady=(20, 10))
        
        accent = tk.Frame(console_header, width=30, height=3, bg="#6c5ce7")
        accent.pack(side='left')
        
        title_label = tk.Label(console_header, text="SYSTEM CONSOLE", 
                            font=("Montserrat", 14, "bold"), 
                            bg="#171721", fg="#ffffff")
        title_label.pack(side='left', padx=(10, 0))
        
        # RESTORE: Control buttons with modern design
        buttons_frame = tk.Frame(console_header, bg="#171721")
        buttons_frame.pack(side='right')
       
        # Download button in green box, emoji centered
        export_btn = tk.Frame(buttons_frame, width=30, height=30, bg="#2ecc71")
        export_btn.pack(side='right', padx=5)
        export_icon = tk.Label(export_btn, text="🡇", font=("Segoe UI Emoji", 16), bg="#2ecc71", fg="white", cursor="hand2")
        export_icon.place(relx=0.5, rely=0.5, anchor='center')
        export_icon.bind("<Button-1>", lambda e: self.collect_all_logs())

        # Sync button in blue box, emoji centered
        sync_btn = tk.Frame(buttons_frame, width=30, height=30, bg="#3498db")
        sync_btn.pack(side='right', padx=5)
        sync_icon = tk.Label(sync_btn, text="🔄", font=("Segoe UI Emoji", 16), bg="#3498db", fg="white", cursor="hand2")
        sync_icon.place(relx=0.5, rely=0.5, anchor='center')
        sync_icon.bind("<Button-1>", lambda e: self.force_sync())

        # Recycle bin button in red box, emoji centered
        clear_btn = tk.Frame(buttons_frame, width=30, height=30, bg="#e74c3c")
        clear_btn.pack(side='right', padx=5)
        clear_icon = tk.Label(clear_btn, text="🗑️", font=("Segoe UI Emoji", 16), bg="#e74c3c", fg="white", cursor="hand2")
        clear_icon.place(relx=0.5, rely=0.5, anchor='center')
        clear_icon.bind("<Button-1>", lambda e: self.log_text.delete(1.0, tk.END))

        # RESTORE: Console body with terminal design
        console_body = tk.Frame(upper_section, bg="#10101c", padx=0, pady=0)
        console_body.pack(fill='both', expand=True, padx=20, pady=(10, 20))
        
        # RESTORE: Add terminal header bar for realistic look
        term_header = tk.Frame(console_body, bg="#1e1e2e", height=30)
        term_header.pack(fill='x')
        term_header.pack_propagate(False)
        
        # RESTORE: Terminal controls (decorative)
        controls_frame = tk.Frame(term_header, bg="#1e1e2e")
        controls_frame.pack(side='left', padx=10)
        
        for color in ["#ff5f56", "#ffbd2e", "#27c93f"]:  # macOS style buttons
            dot = tk.Frame(controls_frame, width=12, height=12, bg=color, bd=0)
            dot.pack(side='left', padx=2)
            # Make round
            dot.bind("<Configure>", lambda e, d=dot: self._make_round(d))
        
        # RESTORE: Terminal title
        term_title = tk.Label(term_header, text="system_logs", font=("Consolas", 10), 
                        bg="#1e1e2e", fg="#a29bfe")
        term_title.pack(side='left', padx=10)
        
        # Create log area with dark terminal appearance
        log_area = tk.Frame(console_body, bg="#10101c")
        log_area.pack(fill='both', expand=True)
        
        # Create text widget for logs with terminal appearance
        self.log_text = tk.Text(log_area, bg="#10101c", fg="#dfe6e9", 
                            font=("Consolas", 10), relief="flat", bd=0,
                            insertbackground="#a29bfe", selectbackground="#6c5ce7")
        self.log_text.pack(fill='both', expand=True, padx=10, pady=10)
        
        # RESTORE: Add scrollbar with modern styling
        scrollbar = tk.Scrollbar(self.log_text, command=self.log_text.yview)
        scrollbar.pack(side='right', fill='y')
        
        # Custom styling for scrollbar
        self._style_modern_scrollbar(scrollbar)
        
        # Configure text widget to use scrollbar
        self.log_text.config(yscrollcommand=scrollbar.set)
        
        # RESTORE: Add welcome message with typing animation
        self._type_welcome_message()
        
        # Lower section (Attendance Info)
        attendance_header = tk.Frame(lower_section, bg="#171721")
        attendance_header.pack(fill='x', padx=20, pady=(0, 5))  # Reduced padding

        # accent = tk.Frame(attendance_header, width=30, height=3, bg="#00b894")
        # accent.pack(side='left')

        # title_label = tk.Label(attendance_header, text="TODAY'S ATTENDANCE", 
        #                     font=("Montserrat", 14, "bold"), 
        #                     bg="#171721", fg="#ffffff")
        # title_label.pack(side='left', padx=(10, 0))
        # # Create the attendance section with the new design
        self._create_attendance_section(lower_section)
        
    
    def _create_attendance_section(self, lower_section):
        """Create a modern card-based attendance info section"""
        # Header with title and accent
        attendance_header = tk.Frame(lower_section, bg="#171721")
        attendance_header.pack(fill='x', padx=20, pady=(5, 5))

        # Left side accent with animation
        accent_frame = tk.Frame(attendance_header, bg="#171721")
        accent_frame.pack(side='left')
        
        # Create animated accent bar
        accent = tk.Canvas(accent_frame, width=40, height=10, bg="#171721", highlightthickness=0)
        accent.pack(side='left')
        accent.create_rectangle(0, 3, 30, 7, fill="#00b894", outline="")
        
        # Add title back as requested
        title_frame = tk.Frame(attendance_header, bg="#171721")
        title_frame.pack(side='left', padx=(5, 0))
        
        # Add attendance icon (calendar symbol)
        icon_label = tk.Label(title_frame, text="", font=("Segoe UI Emoji", 14), 
                            bg="#171721", fg="#ffffff")
        icon_label.pack(side='left', padx=(0, 5))
        
        title_label = tk.Label(title_frame, text="TODAY'S ATTENDANCE", 
                            font=("Montserrat", 14, "bold"), 
                            bg="#171721", fg="#ffffff")
        title_label.pack(side='left')

        # Add refresh button on the right
        buttons_frame = tk.Frame(attendance_header, bg="#171721")
        buttons_frame.pack(side='right')
        

        
        # Create a 7-day data icon button with chart emoji, hand cursor, and tooltip
        seven_day_data_btn = self._create_icon_button(
            buttons_frame,
            "📈",
            lambda: self._update_attendance_data(True),
            "#5E0037"
        )
        seven_day_data_btn.pack(side='left', padx=5)
        seven_day_data_btn.config(cursor="hand2")  # Set hand cursor

        # Safe tooltip using Label widget, not Toplevel
        tooltip_label = tk.Label(
            buttons_frame,
            text="7 day's attendance",
            bg="#340263",
            relief='solid',
            borderwidth=1,
            font=("tahoma", 9)
        )
        tooltip_label.place_forget()

        seven_day_data_btn.bind("<Enter>", lambda e: tooltip_label.place(x=e.x_root - seven_day_data_btn.winfo_rootx() + 10,
                                                                        y=e.y_root - seven_day_data_btn.winfo_rooty() + 10))
        seven_day_data_btn.bind("<Leave>", lambda e: tooltip_label.place_forget())

        
        
        
        

        # Main container for attendance cards with scrollbar
        main_container = tk.Frame(lower_section, bg="#10101c")
        main_container.pack(fill='both', expand=True, padx=20, pady=(0, 10))
        
        # Create a canvas for smooth scrolling that doesn't cause disappearing
        canvas = tk.Canvas(main_container, bg="#10101c", highlightthickness=0)
        canvas.pack(side='left', fill='both', expand=True)
        
        # Add scrollbar with fixed styling
        scrollbar = tk.Scrollbar(main_container, command=canvas.yview)
        scrollbar.pack(side='right', fill='y')
        
        # Configure canvas
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Create frame inside canvas to hold cards
        cards_frame = tk.Frame(canvas, bg="#10101c")
        cards_inner_id = canvas.create_window((0, 0), window=cards_frame, anchor="nw", width=canvas.winfo_width())
        
        # Make sure the cards frame expands to full width
        def configure_scroll_region(event):
            # Update the width and the scroll region
            canvas.itemconfig(cards_inner_id, width=canvas.winfo_width())
            canvas.configure(scrollregion=canvas.bbox("all"))
        
        # Bind to canvas size changes
        canvas.bind('<Configure>', configure_scroll_region)
        cards_frame.bind('<Configure>', configure_scroll_region)
        
        # Define the attendance info items with icons in the requested order with correct descriptions
        info_items = [
            ("username", "👤 User Name", "#6c5ce7"),
            ("record_time", "⏱️ Record Time", "#fdcb6e"),
            ("login_type", "🔐 In/Out", "#00b894"),  # Corrected label
            ("attendance_type", "📋 Attendance Status", "#dfe6e9"),  # Corrected label
            ("date", "📅 Today's Date", "#e17055"),
            ("attendance_method", "📱 Camera Status", "#a29bfe"),  # Corrected label
            ("total_lates", "⏰ Total Lates", "#ff7675"),
            ("total_late_absents", "🚫 Total Late Absents", "#d63031"),
            ("total_early_leaves", "🏃 Total Early Leaves", "#e84393")
        ]
        
        # Create attendance cards in single column layout
        self.attendance_labels = {}
        
        for key, text, color in info_items:
            # Create card with minimal height and full width
            self._create_attendance_card(cards_frame, key, text, color)
        
        # First data load
        self._update_attendance_data(False)

    def _create_attendance_card(self, parent, key, text, accent_color):
        """Create a modern card for attendance data with reduced height"""
        # Card container
        card_container = tk.Frame(parent, bg="#10101c")
        card_container.pack(fill='x', pady=2)  # Reduced vertical spacing
        
        # Card with reduced padding for less height
        card = tk.Frame(card_container, bg="#1e1e2e", padx=12, pady=5)  # Reduced padding
        card.pack(fill='x')
        
        # Add subtle border on left with accent color
        accent = tk.Frame(card, width=4, bg=accent_color)
        accent.pack(side='left', fill='y', padx=(0, 10))
        
        # Content frame
        content = tk.Frame(card, bg="#1e1e2e")
        content.pack(side='left', fill='both', expand=True)
        
        # Split label text at icon
        label_parts = text.split(' ', 1)
        icon = label_parts[0]
        label_text = label_parts[1] if len(label_parts) > 1 else ""
        
        # Icon and title in same row as value for compact display
        row_frame = tk.Frame(content, bg="#1e1e2e")
        row_frame.pack(fill='x')
        
        # Left side with icon and title
        left_frame = tk.Frame(row_frame, bg="#1e1e2e")
        left_frame.pack(side='left', fill='y')
        
        # Icon and title in same line
        title_frame = tk.Frame(left_frame, bg="#1e1e2e")
        title_frame.pack(fill='x', anchor='w')
        
        # Icon label
        icon_label = tk.Label(title_frame, text=icon, font=("Segoe UI Emoji", 11),
                            bg="#1e1e2e", fg=accent_color)
        icon_label.pack(side='left')
        
        # Title label with smaller font
        title_label = tk.Label(title_frame, text=f" {label_text}", 
                            font=("Montserrat", 9),  # Smaller font
                            bg="#1e1e2e", fg="#ffffff")
        title_label.pack(side='left')
        
        # Value label on the right side with bold font
        self.attendance_labels[key] = tk.Label(
            row_frame, 
            text="---", 
            font=("Montserrat", 11, "bold"),  # Slightly smaller than before
            bg="#1e1e2e", 
            fg="#ffffff",
            anchor='e',  # Right-aligned
            width=20     # Fixed width to align values
        )
        self.attendance_labels[key].pack(side='right', padx=(0, 5))
        
        # Simple hover effect without animations
        def on_enter(e):
            card.configure(bg="#252538")
            content.configure(bg="#252538")
            row_frame.configure(bg="#252538")
            left_frame.configure(bg="#252538")
            title_frame.configure(bg="#252538")
            icon_label.configure(bg="#252538")
            title_label.configure(bg="#252538")
            self.attendance_labels[key].configure(bg="#252538")
        
        def on_leave(e):
            card.configure(bg="#1e1e2e")
            content.configure(bg="#1e1e2e")
            row_frame.configure(bg="#1e1e2e")
            left_frame.configure(bg="#1e1e2e")
            title_frame.configure(bg="#1e1e2e")
            icon_label.configure(bg="#1e1e2e")
            title_label.configure(bg="#1e1e2e")
            self.attendance_labels[key].configure(bg="#1e1e2e")
        
        # Bind hover events to card elements
        for element in [card, content, row_frame, title_frame, icon_label, title_label]:
            element.bind('<Enter>', on_enter)
            element.bind('<Leave>', on_leave)
            
    
    
    
    def _update_attendance_data(self, show_history=False):
        """Update attendance data with 7-day history option"""
        try:
            if show_history:
                # Get local connection
                local_conn = self.db_pool.get_local_connection()
                local_cur = local_conn.cursor()
                
                # Find the last updated user (either IN_TIME or OUT_TIME)
                local_cur.execute("""
                    SELECT username, date, in_time, out_time
                    FROM attendance_records
                    WHERE (in_time IS NOT NULL OR out_time IS NOT NULL)
                    ORDER BY 
                        CASE 
                            WHEN out_time IS NOT NULL THEN out_time
                            ELSE in_time
                        END DESC
                    LIMIT 1
                """)
                
                result = local_cur.fetchone()
                if not result:
                    messagebox.showinfo("No Data", "No attendance records found")
                    return
                    
                username = result[0]
                
                # Create popup window
                history_window = tk.Toplevel()
                history_window.title(f"Attendance History")
                history_window.geometry("650x650")
                history_window.configure(bg="#b8b8f3")  # Light background color
                
                # Make window non-resizable for consistent layout
                history_window.resizable(False, False)  # Disable resizing
                
                # Create header frame
                header_frame = tk.Frame(history_window, bg="#4a6fa5", height=70)
                header_frame.pack(fill="x", pady=(0, 20))
                
                # Add user info in header
                user_label = tk.Label(
                    header_frame, 
                    text=f"{username.upper()} - 7 Day Attendance History",
                    font=("Helvetica", 16, "bold"),
                    fg="white",
                    bg="#4a6fa5",
                    padx=20
                )
                user_label.pack(side="left", pady=20)
                
                # Add current time in header
                dhaka_tz = pytz.timezone('Asia/Dhaka')
                now = datetime.now(dhaka_tz)
                today_str = now.strftime('%d %b %Y')
                date_label = tk.Label(
                    header_frame, 
                    text=f"Generated on: {today_str}",
                    font=("Helvetica", 10),
                    fg="white",
                    bg="#4a6fa5",
                    padx=20
                )
                date_label.pack(side="right", pady=20)
                
                # Create main content frame
                content_frame = tk.Frame(history_window, bg="#f5f5f7")
                content_frame.pack(fill="both", expand=True, padx=30, pady=10)
                
                # Create card effect for table
                table_frame = tk.Frame(
                    content_frame, 
                    bg="white",
                    highlightbackground="#e0e0e0",
                    highlightthickness=1,
                )
                table_frame.pack(fill="both", expand=True, padx=10, pady=10)
                
                # Create headers with modern styling
                header_bg = "#4a6fa5"
                header_font = ('Helvetica', 12, 'bold')
                
                headers = ["Date", "In Time", "Out Time", "Duration"]
                
                for col, header_text in enumerate(headers):
                    header = tk.Frame(table_frame, bg=header_bg)
                    header.grid(row=0, column=col, sticky="nsew", padx=1, pady=1)
                    
                    label = tk.Label(
                        header, 
                        text=header_text, 
                        font=header_font, 
                        bg=header_bg, 
                        fg='white',
                        padx=15,
                        pady=12
                    )
                    label.pack(fill="both")

                # Add footer with exit button
                footer_frame = tk.Frame(history_window, bg="#f5f5f7", height=50)
                footer_frame.pack(fill="x", side="bottom", pady=15)

                def handle_exit():
                    # Call on_closing method before destroying the window
                    self.on_closing()
                    history_window.destroy()

                exit_button = tk.Button(
                    footer_frame,
                    text="Exit",
                    font=("Helvetica", 10, "bold"),
                    bg="#dc3545",  # Red color for exit button
                    fg="white",
                    padx=20,
                    pady=8,
                    borderwidth=0,
                    cursor="hand2",
                    command=handle_exit
                )
                exit_button.pack(side="right", padx=30)
                # Get AWS connection for fetching 7-day history
                aws_conn = self.db_pool.get_aws_connection()
                if not aws_conn:
                    messagebox.showerror("Error", "Cannot connect to AWS database")
                    return
                    
                try:
                    aws_cur = aws_conn.cursor()
                    
                    # Calculate date range
                    today = datetime.now(dhaka_tz).date()
                    seven_days_ago = today - timedelta(days=7)
                    
                    # Fetch attendance records for last 7 days
                    aws_cur.execute("""
                        SELECT 
                            date::date as attendance_date,
                            in_time::timestamp as in_time,
                            out_time::timestamp as out_time
                        FROM attendance_records 
                        WHERE username = %s 
                        AND date::date BETWEEN %s AND %s
                        ORDER BY date::date DESC
                    """, (username, seven_days_ago, today))
                    
                    records = aws_cur.fetchall()
                    
                    # Create dictionary to store records by date
                    attendance_dict = {}
                    for record in records:
                        date_str = record[0].strftime('%d %b %Y')
                        
                        # Convert string timestamps to datetime objects with timezone
                        try:
                            in_time = None
                            out_time = None
                            duration = "---"
                            
                            if record[1]:  # in_time
                                in_time = datetime.strptime(str(record[1]), '%Y-%m-%d %H:%M:%S')
                                in_time = pytz.utc.localize(in_time).astimezone(dhaka_tz)
                                
                            if record[2]:  # out_time
                                out_time = datetime.strptime(str(record[2]), '%Y-%m-%d %H:%M:%S')
                                out_time = pytz.utc.localize(out_time).astimezone(dhaka_tz)
                            
                            # Calculate duration if both in and out times exist
                            if in_time and out_time:
                                time_diff = out_time - in_time
                                hours, remainder = divmod(time_diff.seconds, 3600)
                                minutes, _ = divmod(remainder, 60)
                                duration = f"{hours}h {minutes}m"
                            
                            in_time_str = in_time.strftime('%I:%M %p') if in_time else '---'
                            out_time_str = out_time.strftime('%I:%M %p') if out_time else '---'
                            
                            attendance_dict[date_str] = (in_time_str, out_time_str, duration)
                            
                        except Exception as e:
                            logging.error(f"Error processing timestamp for {date_str}: {e}")
                            attendance_dict[date_str] = ('---', '---', '---')
                    
                    # Display records in table
                    row = 1
                    for i in range(7):
                        date = today - timedelta(days=i)
                        date_str = date.strftime('%d %b %Y')
                        
                        # Get attendance data for this date
                        in_time, out_time, duration = attendance_dict.get(date_str, ('---', '---', '---'))
                        
                        # Add row with alternating color for better readability
                        row_bg = "#f8f9fa" if row % 2 == 0 else "white"
                        
                        # Determine cell styling based on attendance status
                        if in_time != "---" and out_time != "---":
                            status_color = "#e8f5e9"  # Light green for complete attendance
                        elif in_time != "---" and out_time == "---":
                            status_color = "#fff3e0"  # Light orange for check-in only
                        else:
                            status_color = row_bg  # Default row color for no attendance
                        
                        # Date cell
                        date_cell = tk.Frame(table_frame, bg=row_bg)
                        date_cell.grid(row=row, column=0, sticky="nsew", padx=1, pady=1)
                        
                        # Check if this is today's date
                        if i == 0:
                            date_label = tk.Label(
                                date_cell, 
                                text=date_str + " (Today)", 
                                font=('Helvetica', 10),
                                bg=row_bg, 
                                fg='#1976d2',
                                padx=15, 
                                pady=10
                            )
                        else:
                            date_label = tk.Label(
                                date_cell, 
                                text=date_str, 
                                font=('Helvetica', 10),
                                bg=row_bg, 
                                fg='black',
                                padx=15, 
                                pady=10
                            )
                        date_label.pack(fill="both")
                        
                        # In time cell
                        in_cell = tk.Frame(table_frame, bg=status_color)
                        in_cell.grid(row=row, column=1, sticky="nsew", padx=1, pady=1)
                        
                        in_label = tk.Label(
                            in_cell, 
                            text=in_time, 
                            font=('Helvetica', 10),
                            bg=status_color, 
                            fg='#333333',
                            padx=15, 
                            pady=10
                        )
                        in_label.pack(fill="both")
                        
                        # Out time cell
                        out_cell = tk.Frame(table_frame, bg=status_color)
                        out_cell.grid(row=row, column=2, sticky="nsew", padx=1, pady=1)
                        
                        out_label = tk.Label(
                            out_cell, 
                            text=out_time, 
                            font=('Helvetica', 10),
                            bg=status_color, 
                            fg='#333333',
                            padx=15, 
                            pady=10
                        )
                        out_label.pack(fill="both")
                        
                        # Duration cell
                        duration_cell = tk.Frame(table_frame, bg=status_color)
                        duration_cell.grid(row=row, column=3, sticky="nsew", padx=1, pady=1)
                        
                        duration_label = tk.Label(
                            duration_cell, 
                            text=duration, 
                            font=('Helvetica', 10),
                            bg=status_color, 
                            fg='#333333',
                            padx=15, 
                            pady=10
                        )
                        duration_label.pack(fill="both")
                        
                        row += 1
                    
                    # Configure grid weights for proper expansion
                    for i in range(4):
                        table_frame.columnconfigure(i, weight=1)
                    
                    # Add footer with back button
                    footer_frame = tk.Frame(history_window, bg="#f5f5f7", height=50)
                    footer_frame.pack(fill="x", side="bottom", pady=15)
                    
                    close_button = tk.Button(
                        footer_frame,
                        text="Close",
                        font=("Helvetica", 10),
                        bg="#4a6fa5",
                        fg="white",
                        padx=15,
                        pady=5,
                        borderwidth=0,
                        command=history_window.destroy
                    )
                    close_button.pack(side="right", padx=30)
                    
                    # Add legend for status colors
                    legend_frame = tk.Frame(footer_frame, bg="#f5f5f7")
                    legend_frame.pack(side="left", padx=30)
                    
                    # Complete attendance legend
                    complete_frame = tk.Frame(legend_frame, bg="#f5f5f7")
                    complete_frame.pack(side="left", padx=(0, 15))
                    
                    complete_color = tk.Label(
                        complete_frame,
                        text="  ",
                        bg="#e8f5e9",
                        width=2,
                        height=1
                    )
                    complete_color.pack(side="left", padx=(0, 5))
                    
                    complete_text = tk.Label(
                        complete_frame,
                        text="Close: Report will Close",
                        font=("Helvetica", 9),
                        bg="#f5f5f7",
                        fg="#333333"
                    )
                    complete_text.pack(side="left")
                    
                    # Check-in only legend
                    checkin_frame = tk.Frame(legend_frame, bg="#f5f5f7")
                    checkin_frame.pack(side="left")
                    
                    checkin_color = tk.Label(
                        checkin_frame,
                        text="  ",
                        bg="#fff3e0",
                        width=2,
                        height=1
                    )
                    checkin_color.pack(side="left", padx=(0, 5))
                    
                    checkin_text = tk.Label(
                        checkin_frame,
                        text="Exit: Application will Exit after syncing.",
                        font=("Helvetica", 9),
                        bg="#f5f5f7",
                        fg="#333333"
                    )
                    checkin_text.pack(side="left")
                    
                finally:
                    self.db_pool.return_aws_connection(aws_conn)
                    
            else:
                # Original attendance update logic
                pass
                
        except Exception as e:
            logging.error(f"Error updating attendance data: {e}")
            messagebox.showerror("Error", "Failed to fetch attendance history")
        
        
    def _style_modern_scrollbar(self, scrollbar):
        """Apply simple styling to scrollbars"""
        scrollbar.configure(width=8, bd=0, highlightthickness=0)
        scrollbar.configure(troughcolor="#10101c", bg="#2d3436")

    

    def _make_round(self, widget):
        """Make a widget round by creating an oval in a canvas"""
        width = widget.winfo_width()
        height = widget.winfo_height()
        
        if width > 0 and height > 0:  # Check if widget is visible
            radius = min(width, height) // 2
            x = width // 2
            y = height // 2
            
            # Create a mask with an oval shape
            mask = tk.Canvas(widget, width=width, height=height, 
                        bg=widget["bg"], highlightthickness=0)
            mask.create_oval(x-radius, y-radius, x+radius, y+radius, fill=widget["bg"])
            mask.place(x=0, y=0)


                
    def update_attendance_display_after_login(self, username, login_type, timestamp, with_camera):
        """Update attendance display after new login"""
        try:
            if hasattr(self, 'attendance_labels'):
                # Get attendance type based on time
                hour = timestamp.hour
                if login_type == "IN_TIME":
                    if hour >= 10:
                        attendance_type = "Late Absent"
                    elif hour >= 9:
                        attendance_type = "Late"
                    else:
                        attendance_type = "On Time"
                else:  # OUT_TIME
                    if hour < 17:
                        attendance_type = "Early Leave"
                    else:
                        attendance_type = "Regular Leave"
                
                # Update display labels
                self.attendance_labels['username'].config(text=username.upper())
                self.attendance_labels['date'].config(text=timestamp.strftime('%d %b %Y'))
                self.attendance_labels['login_type'].config(text=login_type)
                self.attendance_labels['record_time'].config(text=timestamp.strftime('%I:%M %p'))
                self.attendance_labels['attendance_type'].config(text=attendance_type)
                self.attendance_labels['attendance_method'].config(text="Camera" if with_camera else "Without Camera")
                
                # Update statistics
                conn = self.db_pool.get_local_connection()
                cur = conn.cursor()
                cur.execute("""
                    SELECT 
                        COUNT(CASE WHEN strftime('%H', in_time) >= '09' THEN 1 END) as lates,
                        COUNT(CASE WHEN strftime('%H', in_time) >= '10' THEN 1 END) as late_absents,
                        COUNT(CASE WHEN strftime('%H', out_time) < '17' THEN 1 END) as early_leaves
                    FROM attendance_records 
                    WHERE username = ? 
                    AND strftime('%Y-%m', date) = ?
                """, (username, timestamp.strftime('%Y-%m')))
                
                stats = cur.fetchone()
                if stats:
                    self.attendance_labels['total_lates'].config(text=str(stats[0]))
                    self.attendance_labels['total_late_absents'].config(text=str(stats[1]))
                    self.attendance_labels['total_early_leaves'].config(text=str(stats[2]))
                
                #conn.close()
                
        except Exception as e:
            logging.error(f"Error updating attendance display: {e}")       
                    
    def load_initial_attendance_status(self):
        try:
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            
            # Get the most recent record (IN_TIME or OUT_TIME) from database
            cur.execute("""
                SELECT 
                    ar.username,
                    ar.date,
                    ar.in_time,
                    ar.out_time,
                    ar.in_time_with_camera,
                    ar.out_time_with_camera
                FROM attendance_records ar
                WHERE (ar.in_time IS NOT NULL OR ar.out_time IS NOT NULL)
                ORDER BY ar.date DESC, 
                    CASE 
                        WHEN ar.out_time IS NOT NULL THEN ar.out_time 
                        ELSE ar.in_time 
                    END DESC
                LIMIT 1
            """)
            
            result = cur.fetchone()
            
            if result and hasattr(self, 'attendance_labels'):
                username, date, in_time, out_time, in_camera, out_camera = result
                
                # Convert UTC timestamps to Dhaka time
                dhaka_tz = pytz.timezone('Asia/Dhaka')
                
                # Convert in_time to Dhaka time
                if in_time:
                    in_time_utc = datetime.strptime(in_time, '%Y-%m-%d %H:%M:%S')
                    in_time_utc = pytz.utc.localize(in_time_utc)
                    in_time = in_time_utc.astimezone(dhaka_tz)

                # Convert out_time to Dhaka time    
                if out_time:
                    out_time_utc = datetime.strptime(out_time, '%Y-%m-%d %H:%M:%S')
                    out_time_utc = pytz.utc.localize(out_time_utc)
                    out_time = out_time_utc.astimezone(dhaka_tz)

                # Determine which was the last record (IN or OUT)
                if out_time:
                    login_type = "OUT_TIME"
                    record_time = out_time
                    with_camera = out_camera == 'Y'
                else:
                    login_type = "IN_TIME"
                    record_time = in_time
                    with_camera = in_camera == 'Y'

                # Get hour for attendance type calculation
                hour = record_time.hour
                
                # Determine attendance type based on hour
                if login_type == "IN_TIME":
                    if hour >= 10:
                        attendance_type = "Late Absent"
                    elif hour >= 9:
                        attendance_type = "Late"
                    else:
                        attendance_type = "On Time"
                else:  # OUT_TIME
                    if hour < 17:
                        attendance_type = "Early Leave"
                    else:
                        attendance_type = "Regular Leave"
                        
                # Update all labels with the fetched data
                self.attendance_labels['username'].config(text=username.upper())
                self.attendance_labels['date'].config(text=datetime.strptime(date, '%Y-%m-%d').strftime('%d %b %Y'))
                self.attendance_labels['login_type'].config(text=login_type)
                self.attendance_labels['record_time'].config(text=record_time.strftime('%I:%M %p'))
                self.attendance_labels['attendance_type'].config(text=attendance_type)
                self.attendance_labels['attendance_method'].config(text="Camera" if with_camera else "Without Camera")

                # Get monthly statistics for this user's last attendance month
                cur.execute("""
                    SELECT 
                        COUNT(CASE WHEN strftime('%H', in_time) >= '09' THEN 1 END) as lates,
                        COUNT(CASE WHEN strftime('%H', in_time) >= '10' THEN 1 END) as late_absents,
                        COUNT(CASE WHEN strftime('%H', out_time) < '17' THEN 1 END) as early_leaves
                    FROM attendance_records 
                    WHERE username = ? 
                    AND strftime('%Y-%m', date) = strftime('%Y-%m', ?)
                """, (username, date))
                
                stats = cur.fetchone()
                if stats:
                    self.attendance_labels['total_lates'].config(text=str(stats[0]))
                    self.attendance_labels['total_late_absents'].config(text=str(stats[1]))
                    self.attendance_labels['total_early_leaves'].config(text=str(stats[2]))
            else:
                # Set default values if no record found
                for key in ['username', 'date', 'login_type', 'record_time', 
                        'attendance_type', 'attendance_method']:
                    self.attendance_labels[key].config(text="---")
                for key in ['total_lates', 'total_late_absents', 'total_early_leaves']:
                    self.attendance_labels[key].config(text="0")


        except Exception as e:
            logging.error(f"Error loading initial attendance status: {e}")
            logging.error(f"Traceback: {traceback.format_exc()}")
            # Set default values in case of error
            if hasattr(self, 'attendance_labels'):
                for key in ['username', 'date', 'login_type', 'record_time', 
                        'attendance_type', 'attendance_method']:
                    self.attendance_labels[key].config(text="---")
                for key in ['total_lates', 'total_late_absents', 'total_early_leaves']:
                    self.attendance_labels[key].config(text="0")
        
        # finally:
        #     # Safely close cursor and connection
        #     if cur:
        #         cur.close()
        #     if conn:
        #         conn.close()
                    
            
    def update_attendance_info(self, username=None):
        """Update attendance info display"""
        try:
            # Get current date
            today = datetime.now()
            
            # If no username provided, try to get last logged in user
            if not username:
                conn = self.db_pool.get_local_connection()
                cur = conn.cursor()
                cur.execute("""
                    SELECT username, in_time, out_time, in_time_with_camera
                    FROM attendance_records 
                    WHERE date = ? 
                    ORDER BY created_at DESC LIMIT 1
                """, (today.strftime('%Y-%m-%d'),))
                result = cur.fetchone()
                
                if result:
                    username = result[0]
                    login_time = result[1]
                    out_time = result[2]
                    with_camera = result[3] == 'Y'
                    
                    # Determine login type
                    login_type = "Out Time" if out_time else "In Time"
                    
                    # Determine attendance type
                    login_hour = datetime.strptime(login_time, '%H:%M:%S').hour
                    attendance_type = "On Time"
                    if login_hour >= 9:
                        attendance_type = "Late"
                    if login_hour >= 10:
                        attendance_type = "Late Absent"
                    
                    # Get monthly statistics
                    cur.execute("""
                        SELECT 
                            COUNT(CASE WHEN strftime('%H', in_time) >= '09' THEN 1 END) as lates,
                            COUNT(CASE WHEN strftime('%H', in_time) >= '10' THEN 1 END) as late_absents,
                            COUNT(CASE WHEN strftime('%H', out_time) < '17' THEN 1 END) as early_leaves
                        FROM attendance_records 
                        WHERE username = ? 
                        AND strftime('%Y-%m', date) = ?
                    """, (username, today.strftime('%Y-%m')))
                    
                    stats = cur.fetchone()
                    
                    # Update labels
                    self.attendance_labels['username'].config(text=username.upper())
                    self.attendance_labels['date'].config(text=today.strftime('%d %b %Y'))
                    self.attendance_labels['login_type'].config(text=login_type)
                    self.attendance_labels['record_time'].config(
                        text=datetime.strptime(login_time, '%H:%M:%S').strftime('%I:%M %p'))
                    self.attendance_labels['attendance_type'].config(text=attendance_type)
                    self.attendance_labels['attendance_method'].config(
                        text="Camera" if with_camera else "Without Camera")
                    self.attendance_labels['total_lates'].config(text=str(stats[0]))
                    self.attendance_labels['total_late_absents'].config(text=str(stats[1]))
                    self.attendance_labels['total_early_leaves'].config(text=str(stats[2]))
                    
            #conn.close()
            
        except Exception as e:
            logging.error(f"Error updating attendance info: {e}")
            
            
    def update_attendance_display(self, username, login_type, timestamp, with_camera):
        """Update the attendance display in the system console"""
        try:
            if hasattr(self, 'attendance_labels'):
                # Get current date in proper format
                today = datetime.now()
                
                # Get type of attendance based on time
                login_hour = timestamp.hour
                attendance_type = "On Time"
                if login_hour >= 9:
                    attendance_type = "Late"
                if login_hour >= 10:
                    attendance_type = "Late Absent"
                    
                # Update all labels with new information
                self.attendance_labels['username'].config(text=username.upper())
                self.attendance_labels['date'].config(text=today.strftime('%d %b %Y'))
                self.attendance_labels['login_type'].config(text=login_type)
                self.attendance_labels['record_time'].config(text=timestamp.strftime('%I:%M %p'))
                self.attendance_labels['attendance_type'].config(text=attendance_type)
                self.attendance_labels['attendance_method'].config(text="Camera" if with_camera else "Without Camera")
                
                # Get monthly statistics from database
                conn = self.db_pool.get_local_connection()
                cur = conn.cursor()
                cur.execute("""
                    SELECT 
                        COUNT(CASE WHEN strftime('%H', in_time) >= '09' THEN 1 END) as lates,
                        COUNT(CASE WHEN strftime('%H', in_time) >= '10' THEN 1 END) as late_absents,
                        COUNT(CASE WHEN strftime('%H', out_time) < '17' THEN 1 END) as early_leaves
                    FROM attendance_records 
                    WHERE username = ? 
                    AND strftime('%Y-%m', date) = ?
                """, (username, today.strftime('%Y-%m')))
                
                stats = cur.fetchone()
                
                # Update statistics labels
                if stats:
                    self.attendance_labels['total_lates'].config(text=str(stats[0]))
                    self.attendance_labels['total_late_absents'].config(text=str(stats[1]))
                    self.attendance_labels['total_early_leaves'].config(text=str(stats[2]))
                
        except Exception as e:
            logging.error(f"Error updating attendance display: {e}")

    def _create_icon_button(self, parent, icon_text, command, color="#6c5ce7"):
        """Create a modern icon button with hover effect"""
        btn_frame = tk.Frame(parent, bg="#171721", width=32, height=32)  # Use theme color
        btn_frame.pack_propagate(False)
        
        # Create circular background - replace alpha with solid colors
        # Instead of f"{color}50", use a lighter version of the color
        base_color = self._hex_to_rgb(color)
        lighter_color = f"#{min(base_color[0]+30, 255):02x}{min(base_color[1]+30, 255):02x}{min(base_color[2]+30, 255):02x}"
        hover_color = f"#{min(base_color[0]+50, 255):02x}{min(base_color[1]+50, 255):02x}{min(base_color[2]+50, 255):02x}"
        active_color = f"#{min(base_color[0]+70, 255):02x}{min(base_color[1]+70, 255):02x}{min(base_color[2]+70, 255):02x}"
        
        btn_bg = tk.Frame(btn_frame, bg=lighter_color, width=30, height=30)
        btn_bg.place(relx=0.5, rely=0.5, anchor="center")
        btn_bg.bind("<Configure>", lambda e: self._make_round(btn_bg))
        
        # Icon
        icon = tk.Label(btn_bg, text=icon_text, font=("Segoe UI", 12), 
                    bg=lighter_color, fg=color)
        icon.place(relx=0.5, rely=0.5, anchor="center")
        
        # Hover effects
        def on_enter(e):
            btn_bg.config(bg=hover_color)
            icon.config(bg=hover_color, fg="#ffffff")
        
        def on_leave(e):
            btn_bg.config(bg=lighter_color)
            icon.config(bg=lighter_color, fg=color)
        
        def on_click(e):
            btn_bg.config(bg=active_color)
            icon.config(bg=active_color, fg="#ffffff")
            # Schedule return to normal state after click
            parent.after(100, lambda: [
                btn_bg.config(bg=lighter_color),
                icon.config(bg=lighter_color, fg=color)
            ])
            command()
        
        # Bind events
        for widget in [btn_frame, btn_bg, icon]:
            widget.bind("<Enter>", on_enter)
            widget.bind("<Leave>", on_leave)
            widget.bind("<Button-1>", on_click)
        
        return btn_frame

    def _create_modern_footer(self):
        """Create a modern footer with version and copyright info"""
        footer = tk.Frame(self.content_frame, bg="#171721")  # Use theme color
        footer.pack(fill='x', side='bottom', pady=(0, 15))
        
        # Add glass effect with less opacity
        self._add_glass_effect(footer, alpha=8)
        
        # Add version information
        version_label = tk.Label(footer, text="Version 2.0.5", 
                            font=("Consolas", 9), 
                            bg="#171721", fg="#a29bfe")
        version_label.pack(side='left', padx=40)
        
        # Add copyright
        copyright_label = tk.Label(footer, text="© 2025 Attendance Systems Inc.", 
                            font=("Consolas", 9), 
                            bg="#171721", fg="#dfe6e9")
        copyright_label.pack(side='right', padx=40)
        
        
    def _update_particle_bounds(self, width, height):
        """Update the boundaries for particle animation"""
        self._particle_bounds = (width, height)

    def _animate_particles(self):
        """Animate background particles"""
        width, height = self._particle_bounds
        
        for p in self._particles:
            # Get current position
            x1, y1, x2, y2 = self.bg_canvas.coords(p["id"])
            
            # Calculate new position
            dx = math.cos(p["direction"]) * p["speed"]
            dy = math.sin(p["direction"]) * p["speed"]
            
            # Move particle
            self.bg_canvas.move(p["id"], dx, dy)
            
            # Check boundaries and bounce
            new_x1, new_y1, new_x2, new_y2 = self.bg_canvas.coords(p["id"])
            
            if new_x1 < 0 or new_x2 > width:
                p["direction"] = math.pi - p["direction"]
            
            if new_y1 < 0 or new_y2 > height:
                p["direction"] = -p["direction"]
        
        # Continue animation
        self.main_window.after(50, self._animate_particles)

    

    def _draw_glowing_logo(self, canvas):
        """Draw a modern glowing logo with animation"""
        # Create glowing outer ring
        outer_glow = canvas.create_oval(2, 2, 58, 58, outline="#6c5ce7", width=2, tags="logo")
        
        # Create middle ring
        middle_ring = canvas.create_oval(12, 12, 48, 48, outline="#a29bfe", width=2, tags="logo")
        
        # Create inner circle with gradient
        inner_circle = canvas.create_oval(18, 18, 42, 42, fill="#6c5ce7", outline="", tags="logo")
        
        # Add highlight spot for 3D effect
        highlight = canvas.create_oval(22, 22, 32, 32, fill="#ffffff", outline="", tags="logo")
        
        # Store animation references
        self._logo_elements = {
            "outer": outer_glow,
            "middle": middle_ring,
            "inner": inner_circle,
            "highlight": highlight
        }
        
        # Set up pulsing animation
        self._animate_logo(canvas)

    def _animate_logo(self, canvas, pulse_size=0, direction=1):
        """Create subtle pulsing animation for logo"""
        # Pulse size oscillates between -2 and 2
        new_pulse = pulse_size + (0.2 * direction)
        
        if new_pulse > 2:
            new_pulse = 2
            direction = -1
        elif new_pulse < -2:
            new_pulse = -2
            direction = 1
        
        # Update outer ring
        canvas.coords(self._logo_elements["outer"], 
                    2-new_pulse, 2-new_pulse, 
                    58+new_pulse, 58+new_pulse)
        
        # Update middle ring with opposite phase
        canvas.coords(self._logo_elements["middle"], 
                    12+new_pulse/2, 12+new_pulse/2, 
                    48-new_pulse/2, 48-new_pulse/2)
        
        # Continue animation
        self.main_window.after(50, lambda: self._animate_logo(canvas, new_pulse, direction))

    def _update_animated_time(self):
        """Update the time with smooth digit animation"""
        now = datetime.now()
        time_str = now.strftime("%I:%M:%S %p")  # 12-hour format with AM/PM
        date_str = now.strftime("%A, %B %d")
        
        # If we don't have previous time or it changed
        if not hasattr(self, '_prev_time_str') or self._prev_time_str != time_str:
            self.time_label.config(text=time_str)
            self._prev_time_str = time_str
            
            # Flash separator effect
            self.time_label.config(fg="#a29bfe")
            self.main_window.after(100, lambda: self.time_label.config(fg="#ffffff"))
        
        self.date_label.config(text=date_str)
        
        # Schedule next update
        self.main_window.after(1000, self._update_animated_time)

    


    def _hex_to_rgb(self, hex_color):
        """Convert hex color to RGB values"""
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

    
    
    def _add_glass_effect(self, frame, alpha=20, highlight_top=False):
        """Add a glass morphism effect to a frame"""
        # Background with modified transparency - Tkinter doesn't support alpha channel
        # Convert alpha value to a darker shade instead
        alpha_factor = min(1.0, alpha / 100)
        base_color = (42, 42, 66)  # #2a2a42 in RGB
        
        # Make the color darker based on alpha
        adjusted_color = tuple(int(c * alpha_factor) for c in base_color)
        glass_bg_color = f"#{adjusted_color[0]:02x}{adjusted_color[1]:02x}{adjusted_color[2]:02x}"
        
        # Create frame with adjusted color
        glass_bg = tk.Frame(frame, bg=glass_bg_color)
        glass_bg.place(x=0, y=0, relwidth=1, relheight=1)
        
        # Border effect - replace "#ffffff10" with a light gray
        # Use different colors based on alpha to simulate transparency
        if alpha > 10:
            border_color = "#333344"  # Slightly lighter than background
        else:
            border_color = "#222233"  # Very subtle lighter than background
            
        border = tk.Frame(frame, bg=border_color, height=1)
        if highlight_top:
            # Highlight top edge for depth
            border.place(x=0, y=0, relwidth=1)
        else:
            # Highlight bottom edge for depth
            border.place(x=0, rely=1.0, relwidth=1, y=-1)

    
   

    def _draw_advanced_background(self, event):
        """Draw a professional gradient background with animated particles"""
        width = event.width
        height = event.height
        self.bg_canvas.delete("gradient")
        
        # Create a deep, rich gradient base
        for i in range(height + 50):
            # Dark blue to deep purple gradient
            r_val = int(23 + (i/height) * 15)
            g_val = int(23 + (i/height) * 5)
            b_val = int(33 + (i/height) * 18)
            color = f"#{r_val:02x}{g_val:02x}{b_val:02x}"
            self.bg_canvas.create_line(0, i, width, i, fill=color, tags="gradient")
        
        # Add subtle circular gradient overlay for depth
        center_x = width / 2
        center_y = height / 2
        radius = max(width, height) * 0.7
        
        # Create subtle glow at center - fix the alpha channel issue
        for r in range(int(radius), 0, -int(radius/8)):
            # Instead of using alpha, use darker colors for outer circles
            opacity_factor = r / radius
            glow_r = int(40 * opacity_factor)
            glow_g = int(32 * opacity_factor)
            glow_b = int(80 * opacity_factor)
            glow_color = f"#{glow_r:02x}{glow_g:02x}{glow_b:02x}"
            
            x0, y0 = center_x - r, center_y - r
            x1, y1 = center_x + r, center_y + r
            self.bg_canvas.create_oval(x0, y0, x1, y1, 
                                    fill=glow_color, 
                                    outline="", 
                                    tags="gradient")
        
        # Add subtle accent lines - fix the alpha channel issue
        for i in range(5):
            y_pos = height * (i/5)
            # Use a very dark white instead of transparent
            self.bg_canvas.create_line(0, y_pos, width, y_pos, 
                                    fill="#111111", width=1, 
                                    tags="gradient")
        
        # Update content frame size
        self.bg_canvas.itemconfig(self.content_frame_id, width=width, height=height)
        
        # Initialize particles if not already created
        if not hasattr(self, '_particles'):
            self._create_background_particles(width, height)
        else:
            # Update particle boundaries
            self._update_particle_bounds(width, height)

    def _create_background_particles(self, width, height):
        """Create subtle animated particles in background"""
        self._particles = []
        self._particle_bounds = (width, height)
        
        # Create 20 subtle particles
        for _ in range(20):
            x = random.randint(0, width)
            y = random.randint(0, height)
            size = random.randint(2, 5)
            speed = random.uniform(0.2, 1.0)
            
            # Fix alpha channel issue by using solid colors with varying brightness
            colors = ["#6c5ce7", "#a29bfe", "#74b9ff", "#55efc4"]
            color = random.choice(colors)
            
            particle = self.bg_canvas.create_oval(
                x, y, x+size, y+size,
                fill=color, outline="",
                tags="particle"
            )
            
            self._particles.append({
                "id": particle,
                "speed": speed,
                "direction": random.uniform(0, 2*math.pi),
                "size": size,
                "color": color
            })
        
        # Start animation
        self._animate_particles()

    
    
    def _style_modern_scrollbar(self, scrollbar):
        """Apply modern styling to scrollbar"""
        # This works on Windows and some Linux systems
        try:
            # Try to configure with custom style
            scrollbar.config(troughcolor="#10101c", bg="#6c5ce7", 
                            activebackground="#a29bfe", width=8)
        except:
            # Fallback for platforms that don't support all options
            scrollbar.config(bg="#6c5ce7")

    def _make_round(self, widget):
        """Make a widget round by configuring canvas or adding corner radius"""
        # Get widget dimensions
        width = widget.winfo_width()
        height = widget.winfo_height()
        
        # Make widget round by setting corner radius
        if hasattr(widget, 'radius'):
            return  # Already configured
            
        # For tkinter, we use a simple method to make it appear round
        if width > 0 and height > 0:
            # Find the minimum dimension for a perfect circle
            radius = min(width, height) // 2
            widget.config(bd=0, highlightthickness=0)
            widget.radius = radius
            
            # Try to add rounded corners if available on platform
            try:
                widget.config(relief="round", borderwidth=0, highlightthickness=0)
            except:
                pass  # Not all platforms support rounded corners

    def _type_welcome_message(self):
        """Type welcome message with animation effect"""
        welcome_text = [
            "\n  ╔══════════════════════════════════════════════════════════╗",
            "  ║                ATTENDANCE SYSTEM STARTED                 ║",
            "  ╚══════════════════════════════════════════════════════════╝\n",
            "  • System initialized successfully",
            "  • Loading AI Heavy Models",
            "  • Starting background services..."
        ]
        
        # Clear any existing text
        self.log_text.delete(1.0, tk.END)
        
        # Type text with delay
        def type_line(line_index=0, char_index=0):
            if line_index < len(welcome_text):
                current_line = welcome_text[line_index]
                
                if char_index < len(current_line):
                    # Add character
                    self.log_text.insert(tk.END, current_line[char_index])
                    self.log_text.see(tk.END)
                    
                    # Schedule next character
                    delay = 5 if current_line[char_index] in (" ,.:;•-_=+[]{}()|") else 10
                    self.main_window.after(delay, lambda: type_line(line_index, char_index + 1))
                else:
                    # Line completed, add newline
                    self.log_text.insert(tk.END, "\n")
                    self.log_text.see(tk.END)
                    
                    # Schedule next line with slightly longer delay
                    self.main_window.after(100, lambda: type_line(line_index + 1, 0))
        
        # Start typing animation
        self.main_window.after(200, type_line)

    
    def _setup_animations(self):
        """Setup global animations for the application"""
        # Animation for status updates
        self._animate_status_updates()

    def _animate_status_updates(self):
        """Animate the status updates with progress bars"""
        # Progress values for each status item (0-100)
        if not hasattr(self, '_progress_values'):
            self._progress_values = {
                "tensorflow": 0,
                "database": 0,
                "network": 0
            }
            self._status_states = {
                "tensorflow": "initializing",
                "database": "connecting",
                "network": "detecting"
            }
        
        # Update progress with randomized increments
        for key, value in self._progress_values.items():
            if value < 100:
                # Random increment between 1-5%
                increment = random.randint(1, 5)
                new_value = min(value + increment, 100)
                self._progress_values[key] = new_value
                
                # Update progress bar width
                parent_width = self.status_progress[key].master.winfo_width()
                progress_width = int((parent_width * new_value) / 100)
                self.status_progress[key].config(width=progress_width)
                
                # Update status text based on progress
                if new_value < 33:
                    status_text = f"{self._status_states[key].capitalize()}... ({new_value}%)"
                elif new_value < 66:
                    status_text = f"Almost ready... ({new_value}%)"
                elif new_value < 100:
                    status_text = f"Finalizing... ({new_value}%)"
                else:
                    status_text = "Online and ready"
                    # Change status indicator color
                    self.status_indicators[key].config(bg="#00b894")
                
                self.status_labels[key].config(text=status_text)
                
                # If completed, change color to green
                if new_value == 100:
                    self.status_progress[key].config(bg="#00b894")
        
        # Continue animation until all reach 100%
        if any(value < 100 for value in self._progress_values.values()):
            self.main_window.after(200, self._animate_status_updates)
        else:
            self._show_all_systems_ready()

    def _show_all_systems_ready(self):
        """Show a notification that all systems are ready"""
        # Create a floating notification
        notification = tk.Toplevel(self.main_window)
        notification.overrideredirect(True)  # Remove window decorations
        notification.attributes("-topmost", True)  # Keep on top
        notification.configure(bg="#6c5ce7")
        
        # Calculate position (bottom right of main window)
        main_x = self.main_window.winfo_x()
        main_y = self.main_window.winfo_y()
        main_width = self.main_window.winfo_width()
        main_height = self.main_window.winfo_height()
        
        # Position notification
        notification.geometry(f"420x80+{main_x + main_width - 450}+{main_y + main_height - 100}")
        
        # Add content
        content = tk.Frame(notification, bg="#6c5ce7", padx=15, pady=15)
        content.pack(fill='both', expand=True)
        
        title = tk.Label(content, text="  ✅ Ready to proceed—no wait needed", 
                    font=("Montserrat", 12, "bold"), 
                    bg="#6c5ce7", fg="white")
        title.pack(anchor='w')
        
        message = tk.Label(content, text="System is ready to process attendance", 
                        font=("Montserrat", 10), 
                        bg="#6c5ce7", fg="white")
        message.pack(anchor='w', pady=(5, 0))
        
        # Auto-close after 3 seconds
        self.main_window.after(3000, notification.destroy)

    
    def collect_all_logs(self):
        """Export all terminal events with timestamps ensuring complete log capture"""
        try:
            # Get current date and time for filename
            now = datetime.now()
            date_str = now.strftime("%Y%m%d_%H%M%S")
            filename = f"application_logs_{date_str}.txt"

            # Dictionary to store logs with timestamps as keys to prevent duplicates
            log_entries = {}

            # Special list for XLA logs that need to appear first
            xla_logs = []

            # 1. First collect startup logs (earliest logs)
            startup_log_file = os.path.join(os.path.expanduser('~'), 'attendance_complete.log')
            if os.path.exists(startup_log_file):
                with open(startup_log_file, 'r', encoding='utf-8') as f:
                    startup_logs = f.readlines()
                    for log in startup_logs:
                        try:
                            timestamp_str = log[:23]
                            timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S,%f')
                            log_entries[timestamp] = log.strip()
                        except:
                            log_entries[datetime.min] = log.strip()

            # 2. Collect early initialization and XLA logs from stderr
            if hasattr(sys.stderr, '_terminal_logs'):
                for log in sys.stderr._terminal_logs:
                    if ("WARNING: All log messages before absl::InitializeLog()" in log or
                        "XLA service" in log or 
                        "StreamExecutor device" in log or
                        "service.cc" in log):
                        xla_logs.append(log.strip())
                    else:
                        try:
                            timestamp_str = log[:23]
                            timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S,%f')
                            log_entries[timestamp] = log.strip()
                        except:
                            log_entries[datetime.min] = log.strip()

            # 3. Collect terminal logger output
            terminal_log_file = os.path.join(os.path.expanduser('~'), 'terminal_output.log')
            if os.path.exists(terminal_log_file):
                with open(terminal_log_file, 'r', encoding='utf-8') as f:
                    terminal_logs = f.readlines()
                    for log in terminal_logs:
                        try:
                            timestamp_str = log[:23]
                            timestamp = datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S,%f')
                            log_entries[timestamp] = log.strip()
                        except:
                            log_entries[datetime.min] = log.strip()

            # Write logs to file
            with open(filename, 'w', encoding='utf-8') as f:
                # Write header
                f.write("="*80 + "\n")
                f.write(f"Complete Application Log Export - Generated at {now}\n")
                f.write("="*80 + "\n\n")

                # First write XLA logs at the beginning
                for log in xla_logs:
                    f.write(f"{log}\n")
                if xla_logs:
                    f.write("\n")  # Add separator after XLA logs

                # Then write chronological logs
                for timestamp in sorted(log_entries.keys()):
                    f.write(f"{log_entries[timestamp]}\n")

            # Show success message in console
            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"\n  ✓ Complete application logs exported to {filename}\n")
                self.log_text.see(tk.END)

        except Exception as e:
            if hasattr(self, 'log_text'):
                self.log_text.insert(tk.END, f"\n  ✗ Failed to export logs: {str(e)}\n")
                self.log_text.see(tk.END)
            logging.error(f"Error exporting logs: {e}")

    def update_status(self, status_key, message, success=True):
        """Update status indicators in the modern UI
        
        Args:
            status_key (str): Key of the status item to update ('tensorflow', 'database', or 'network')
            message (str): Status message to display
            success (bool): Whether the operation was successful
        """
        # Check if we're running in the main thread
        if threading.current_thread() is not threading.main_thread():
            # If not in main thread, schedule this update to run in main thread
            self.main_window.after(0, lambda: self.update_status(status_key, message, success))
            return
        
        # Check if the UI elements exist
        if not hasattr(self, 'status_labels') or not hasattr(self, 'status_indicators'):
            logging.warning(f"Status update called before UI was initialized: {status_key} - {message}")
            return
            
        try:
            # Set progress value to 100% if successful, otherwise to 30%
            progress_value = 100 if success else 30
            
            # Update the status text
            if status_key in self.status_labels:
                self.status_labels[status_key].config(text=message)
                
            # Update the status indicator color
            if status_key in self.status_indicators:
                color = "#00b894" if success else "#ff7675"  # Green for success, red for failure
                self.status_indicators[status_key].config(bg=color)
                
            # Update the progress bar
            if hasattr(self, 'status_progress') and status_key in self.status_progress:
                # Get parent width for calculating progress width
                parent_width = self.status_progress[status_key].master.winfo_width()
                progress_width = int((parent_width * progress_value) / 100)
                
                # Update progress bar
                self.status_progress[status_key].config(width=progress_width)
                self.status_progress[status_key].config(bg="#00b894" if success else "#ff7675")
            
            # Log the status update
            log_method = logging.info if success else logging.error
            log_method(f"Status update - {status_key}: {message}")
            
            # Add to console log if it exists
            if hasattr(self, 'log_text'):
                status_icon = "✓" if success else "✗"
                self.log_text.insert(tk.END, f"\n  {status_icon} {status_key.capitalize()}: {message}")
                self.log_text.see(tk.END)
                
        except Exception as e:
            logging.error(f"Error updating status for {status_key}: {str(e)}")
            
        # Update UI values for animation consistency
        if hasattr(self, '_progress_values') and status_key in self._progress_values:
            self._progress_values[status_key] = progress_value
            
            

    def force_sync(self):
        """Force sync all pending data and photos to cloud"""
        try:
            if not self.db_pool or not self.db_pool.is_online:
                messagebox.showwarning("Offline", "Cannot sync - system is offline")
                return
                
            # Create progress window
            progress_window = tk.Toplevel(self.main_window)
            progress_window.title("Syncing")
            progress_window.geometry("300x150")
            progress_window.transient(self.main_window)
            progress_window.grab_set()
            
            label = tk.Label(progress_window, text="Checking for all pending photo uploads...\nPlease wait...")
            label.pack(pady=20)
            progress_window.update()
            
            # Get local connection
            local_conn = self.db_pool.get_local_connection()
            local_cur = local_conn.cursor()
            
            # Find ALL records with local files but no S3 URLs
            local_cur.execute("""
                SELECT ar.record_id, ar.username, ar.video_path_in, ar.video_path_out
                FROM attendance_records ar
                WHERE (ar.video_path_in IS NOT NULL AND ar.video_path_in NOT LIKE 's3://%')
                OR (ar.video_path_out IS NOT NULL AND ar.video_path_out NOT LIKE 's3://%')
            """)
            
            pending_records = local_cur.fetchall()
            
            if pending_records:
                label.config(text=f"Uploading {len(pending_records)} pending photos...")
                progress_window.update()
                
                aws_conn = self.db_pool.get_aws_connection()
                aws_cur = aws_conn.cursor()
                
                for record_id, username, in_path, out_path in pending_records:
                    try:
                        # Upload in_time photo if exists and not already in S3
                        if in_path and not in_path.startswith('s3://'):
                            if os.path.exists(in_path):
                                s3_url = self.db_pool.upload_video_to_s3(in_path, username, "IN_TIME")
                                if s3_url:
                                    local_cur.execute("""
                                        UPDATE attendance_records 
                                        SET video_path_in = ? 
                                        WHERE record_id = ?
                                    """, (s3_url, record_id))
                                    aws_cur.execute("""
                                        UPDATE attendance_records 
                                        SET video_path_in = %s 
                                        WHERE record_id = %s
                                    """, (s3_url, record_id))
                        
                        # Upload out_time photo if exists and not already in S3
                        if out_path and not out_path.startswith('s3://'):
                            if os.path.exists(out_path):
                                s3_url = self.db_pool.upload_video_to_s3(out_path, username, "OUT_TIME")
                                if s3_url:
                                    local_cur.execute("""
                                        UPDATE attendance_records 
                                        SET video_path_out = ? 
                                        WHERE record_id = ?
                                    """, (s3_url, record_id))
                                    aws_cur.execute("""
                                        UPDATE attendance_records 
                                        SET video_path_out = %s 
                                        WHERE record_id = %s
                                    """, (s3_url, record_id))
                        
                        local_conn.commit()
                        aws_conn.commit()
                        
                    except Exception as e:
                        logging.error(f"Error uploading photos for record {record_id}: {e}")
                        continue
                
                self.db_pool.return_aws_connection(aws_conn)
            
            # Update label for syncing status
            label.config(text="Syncing all pending data...\nPlease wait...")
            progress_window.update()
            
            # Do full sync of all data
            self.network_manager._sync_all_pending_data()
            
            # Close progress window
            progress_window.destroy()
            
            messagebox.showinfo("Success", "All data and photos synced successfully!")
            
        except Exception as e:
            logging.error(f"Error during manual sync: {e}")
            messagebox.showerror("Error", "Failed to sync data to cloud")
            
        

    def handle_no_camera_login(self):
        """Handle login without camera with modern UI design and improved error handling - FIXED VERSION"""
        try:
            # Create login window with modern styling
            login_window = tk.Toplevel(self.main_window)
            login_window.title("Login")
            login_window.geometry("420x600")
            login_window.resizable(False, False)
            
            # Make window modal and ensure it stays on top
            login_window.transient(self.main_window)
            login_window.grab_set()
            login_window.attributes('-topmost', True)
            login_window.lift()
            login_window.focus_force()
            
            # Center the window
            try:
                self.center_window_robust(login_window, 420, 600)
            except:
                # Fallback centering
                screen_width = login_window.winfo_screenwidth()
                screen_height = login_window.winfo_screenheight()
                x = (screen_width - 420) // 2
                y = (screen_height - 600) // 2
                login_window.geometry(f"420x600+{x}+{y}")
            
            # Create base containers with proper hierarchy
            main_frame = tk.Frame(login_window, bg="#1e293b")
            main_frame.pack(fill='both', expand=True)
            
            # Logo section
            logo_frame = tk.Frame(main_frame, bg="#1e293b", height=80)
            logo_frame.pack(fill='x', pady=(20, 30))
            logo_frame.pack_propagate(False)
            
            # Create canvas for logo
            logo_canvas = tk.Canvas(logo_frame, width=80, height=80, bg="#1e293b", highlightthickness=0)
            logo_canvas.place(relx=0.5, rely=0.5, anchor="center")
            self._draw_simple_logo(logo_canvas)
            
            # Title and subtitle
            title_frame = tk.Frame(main_frame, bg="#1e293b")
            title_frame.pack(pady=(0, 20))
            
            welcome_label = tk.Label(
                title_frame,
                text="In & Out without Camera",
                font=("Montserrat", 18, "bold"),
                fg="#f1f5f9",
                bg="#1e293b"
            )
            welcome_label.pack()
            
            subtitle_label = tk.Label(
                title_frame,
                text="Please enter your credentials to continue",
                font=("Montserrat", 10),
                fg="#94a3b8",
                bg="#1e293b"
            )
            subtitle_label.pack(pady=(5, 0))
            
            # Create modern inputs
            input_frame = tk.Frame(main_frame, bg="#1e293b")
            input_frame.pack(pady=(10, 20), fill="x")
            
            # Username field with icon effect
            username_container = tk.Frame(input_frame, bg="#1e293b")
            username_container.pack(pady=(0, 15), fill="x", padx=40)
            
            username_label = tk.Label(
                username_container,
                text="USERNAME",
                font=("Montserrat", 9),
                fg="#94a3b8",
                bg="#1e293b",
                anchor="w"
            )
            username_label.pack(anchor="w", padx=(3, 0))
            
            username_wrapper = tk.Frame(
                username_container,
                bg="#2d3748",
                highlightbackground="#4a5568",
                highlightthickness=1
            )
            username_wrapper.pack(fill="x", ipady=5)
            
            # Create a mini-canvas for the username icon
            username_icon_canvas = tk.Canvas(username_wrapper, width=30, height=20, bg="#2d3748", highlightthickness=0)
            username_icon_canvas.pack(side="left", padx=(8, 0))
            
            # Draw a simple user icon
            username_icon_canvas.create_oval(5, 5, 15, 15, outline="#a29bfe", width=2)
            username_icon_canvas.create_arc(5, 12, 15, 25, start=0, extent=180, outline="#a29bfe", width=2, style="arc")
            
            username_entry = tk.Entry(
                username_wrapper,
                font=("Montserrat", 11),
                bd=0,
                bg="#2d3748",
                fg="#e2e8f0",
                insertbackground="#e2e8f0"
            )
            username_entry.pack(side="left", expand=True, fill="x", padx=(5, 10))
            
            # Password field with icon effect
            password_container = tk.Frame(input_frame, bg="#1e293b")
            password_container.pack(pady=(0, 5), fill="x", padx=40)
            
            password_label = tk.Label(
                password_container,
                text="PASSWORD",
                font=("Montserrat", 9),
                fg="#94a3b8",
                bg="#1e293b",
                anchor="w"
            )
            password_label.pack(anchor="w", padx=(3, 0))
            
            password_wrapper = tk.Frame(
                password_container,
                bg="#2d3748",
                highlightbackground="#4a5568",
                highlightthickness=1
            )
            password_wrapper.pack(fill="x", ipady=5)
            
            # Create a mini-canvas for the password icon
            password_icon_canvas = tk.Canvas(password_wrapper, width=30, height=20, bg="#2d3748", highlightthickness=0)
            password_icon_canvas.pack(side="left", padx=(8, 0))
            
            # Draw a simple lock icon
            password_icon_canvas.create_rectangle(5, 7, 15, 15, outline="#a29bfe", width=2)
            password_icon_canvas.create_oval(8, 3, 12, 7, outline="#a29bfe", width=2)
            
            password_entry = tk.Entry(
                password_wrapper,
                font=("Montserrat", 11),
                bd=0,
                show="•",
                bg="#2d3748",
                fg="#e2e8f0",
                insertbackground="#e2e8f0"
            )
            password_entry.pack(side="left", expand=True, fill="x", padx=(5, 10))
            
            # Show/hide password toggle
            show_password_var = tk.BooleanVar()
            show_password_var.set(False)
            
            def toggle_password_visibility():
                if show_password_var.get():
                    password_entry.config(show="")
                else:
                    password_entry.config(show="•")
            
            show_password_frame = tk.Frame(password_container, bg="#1e293b")
            show_password_frame.pack(anchor="w", pady=(5, 0))
            
            show_password_checkbox = tk.Checkbutton(
                show_password_frame,
                text="Show password",
                variable=show_password_var,
                command=toggle_password_visibility,
                font=("Montserrat", 9),
                fg="#94a3b8",
                bg="#1e293b",
                activebackground="#1e293b",
                activeforeground="#a29bfe",
                selectcolor="#2d3748"
            )
            show_password_checkbox.pack(side="left")
            
            # Login button implementation
            button_frame = tk.Frame(main_frame, bg="#1e293b")
            button_frame.pack(pady=(15, 0), fill="x", padx=40)
            
            login_button = tk.Button(
                button_frame,
                text="LOGIN",
                font=("Montserrat", 12, "bold"),
                fg="white",
                width=30,
                height=2,
                relief=tk.FLAT,
                cursor="hand2",
                bg="#6c5ce7",
                activebackground="#8075e5"
            )
            login_button.pack(pady=(10, 0))
            
            def handle_login(event=None):
                # Disable button to prevent multiple clicks
                login_button.config(text="LOGGING IN...", state=tk.DISABLED)
                username = username_entry.get().strip().upper()
                password = password_entry.get().strip()
                
                # Schedule actual login process after visual feedback
                login_window.after(300, lambda: process_login(username, password))
            
            def process_login(username, password):
                """Process the actual login after visual effects - FIXED VERSION"""
                try:
                    if not username or not password:
                        # Restore button text and state
                        login_button.config(text="LOGIN", state=tk.NORMAL)
                        # Show error with modern styling
                        self._show_error_message(login_window, "Username and password are required!")
                        return
                    
                    # FIXED: Check if user exists first
                    conn = self.db_pool.get_local_connection()
                    cur = conn.cursor()
                    
                    # FIXED: Use correct column name (company_user_uuid)
                    cur.execute("""
                        SELECT company_user_uuid, username 
                        FROM users 
                        WHERE UPPER(username) = ? AND password = ?
                    """, (username, password))
                    result = cur.fetchone()
                    
                    if not result:
                        # Restore button text and state
                        login_button.config(text="LOGIN", state=tk.NORMAL)
                        # Show error with modern styling
                        self._show_error_message(login_window, "Invalid username or password!")
                        # Shake animation for wrong password
                        self._shake_widget(login_window)
                        return
                    
                    # FIXED: Use company_user_uuid from result
                    company_user_uuid, found_username = result
                    
                    # Get current time
                    current_time = datetime.now()
                    if current_time.tzinfo is None:
                        dhaka_tz = pytz.timezone('Asia/Dhaka')
                        current_time = dhaka_tz.localize(current_time)
                    
                    # FIXED: Use the database pool's determine_login_type method
                    login_type = self.db_pool.determine_login_type(username)
                    
                    # Generate blank image for no-camera login
                    filepath = self._generate_blank_image(username, login_type, current_time)
                    
                    if not filepath:
                        login_button.config(text="LOGIN", state=tk.NORMAL)
                        self._show_error_message(login_window, "Failed to create attendance record!")
                        return
                    
                    # FIXED: Save attendance with correct parameters
                    success, aws_success = self.db_pool.save_login_to_db(
                        username=username,
                        login_type=login_type,
                        timestamp=current_time,
                        filepath=filepath,
                        with_camera=False
                    )
                    
                    if success:
                        # Show success
                        login_button.config(text="SUCCESS!", bg="#28a745")
                        
                        # FIXED: Create instance manager to record attendance
                        if hasattr(self, 'instance_manager'):
                            self.instance_manager.record_attendance()
                        
                        # Show the modern success notification
                        self._show_success_notification(self.main_window, username, login_type, current_time)
                        
                        # Update the attendance display
                        if hasattr(self, 'update_attendance_display_after_login'):
                            self.update_attendance_display_after_login(username, login_type, current_time, False)
                        
                        # Close login window after success
                        def delayed_close():
                            try:
                                if login_window and login_window.winfo_exists():
                                    login_window.destroy()
                                
                                # FIXED: Trigger sync and close after successful login
                                self.main_window.after(500, lambda: self.on_closing())
                            except Exception as close_error:
                                logging.error(f"Error closing login window: {close_error}")
                        
                        # Close after showing success for 2 seconds
                        self.main_window.after(2000, delayed_close)
                        
                    else:
                        # Restore button text and state
                        login_button.config(text="LOGIN", state=tk.NORMAL)
                        # Show error with modern styling
                        self._show_error_message(login_window, "Failed to record attendance!")
                
                except Exception as e:
                    logging.error(f"No-camera login error: {e}")
                    # Restore button text and state
                    login_button.config(text="LOGIN", state=tk.NORMAL)
                    # Show error with modern styling
                    self._show_error_message(login_window, f"Login failed: {str(e)}")
            
            # Set the login command
            login_button.config(command=handle_login)
            
            # Bind Enter key to both entry widgets
            username_entry.bind('<Return>', handle_login)
            password_entry.bind('<Return>', handle_login)
            
            # Add "Forgot Password?" text link
            forgot_frame = tk.Frame(main_frame, bg="#1e293b")
            forgot_frame.pack(pady=(15, 0))
            
            def show_forgot_password():
                self._show_forgot_password_popup(login_window)
                
            forgot_label = tk.Label(
                forgot_frame,
                text="Forgot password?",
                font=("Montserrat", 9, "underline"),
                fg="#94a3b8",
                bg="#1e293b",
                cursor="hand2"
            )
            forgot_label.pack()
            forgot_label.bind("<Button-1>", lambda e: show_forgot_password())
            
            # Add footer with version info
            footer_frame = tk.Frame(login_window, bg="#1e293b")
            footer_frame.pack(side="bottom", fill="x", pady=10)
            
            version_label = tk.Label(
                footer_frame,
                text="Attendance System v2.0",
                font=("Montserrat", 8),
                fg="#64748b",
                bg="#1e293b"
            )
            version_label.pack(side="right", padx=15)
            
            # Set focus to username entry
            username_entry.focus_set()
            
            # Handle window close properly
            def on_window_close():
                try:
                    if login_window and login_window.winfo_exists():
                        login_window.destroy()
                except:
                    pass
            
            login_window.protocol("WM_DELETE_WINDOW", on_window_close)
            
        except Exception as e:
            logging.error(f"No-camera login error: {e}")
            messagebox.showerror("Error", "Login window failed to open!")

            
            
      
    def _generate_blank_image(self, username, login_type, timestamp):
        """
        Generate a blank image with text annotations for camera-less logins.
        Args:
            username (str): The username of the person logging in.
            login_type (str): "IN_TIME" or "OUT_TIME".
            timestamp (datetime): The timestamp of the login.
        Returns:
            str: Filepath where the blank image is saved.
        """
        try:
            # Define the directory structure
            year_folder = os.path.join(self.db_pool.db_dir, username, str(timestamp.year))
            month_folder = os.path.join(year_folder, timestamp.strftime("%B"))

            # Create directories if they don't exist
            os.makedirs(month_folder, exist_ok=True)
            logging.debug(f"Directory created/exists: {month_folder}")
            
            # Generate filename
            filename = f"{username.upper()}_{timestamp.strftime('%d_%b_%Y')}_{login_type}_{timestamp.strftime('%I_%M_%p')}.jpg"
            filepath = os.path.join(month_folder, filename)

                       # Validate the file path

            try:
                validate_filepath(filepath)
            except ValueError as e:
                logging.error(f"Invalid file path: {e}")
                messagebox.showerror("Error", "The file path is invalid. Please check the directory structure.")
                return None
           
           
            # Create a blank black image (640x480 pixels)
            blank_image = np.zeros((480, 640, 3), dtype=np.uint8)

            # Add text annotations to the blank image
            font = cv2.FONT_HERSHEY_SIMPLEX
            font_scale = 1
            font_thickness = 2
            color = (255, 255, 255)  # White color

            # Text content
            text_lines = [
                f"Username: {username}",
                f"Date: {timestamp.strftime('%d %b %Y')}",
                f"{login_type.replace('_', ' ')}: {timestamp.strftime('%I:%M %p')}"
            ]

            # Calculate starting position for text
            y_offset = 50  # Initial vertical position
            for line in text_lines:
                text_size = cv2.getTextSize(line, font, font_scale, font_thickness)[0]
                x_position = (blank_image.shape[1] - text_size[0]) // 2  # Center horizontally
                cv2.putText(blank_image, line, (x_position, y_offset), font, font_scale, color, font_thickness)
                y_offset += 50  # Move to the next line

            # Save the blank image
            cv2.imwrite(filepath, blank_image)

            return filepath

        except Exception as e:
            logging.error(f"Error generating blank image: {e}")
            return None 
                           
        
    def _show_error_message(self, parent_window, message):
        """Show a modern error message dialog"""
        try:
            error_window = tk.Toplevel(parent_window)
            error_window.title("Error")
            error_window.geometry("350x150")
            error_window.transient(parent_window)
            error_window.grab_set()
            
            # Center the error window
            self.center_window_robust(error_window, 350, 150)
            
            # Configure background
            error_window.configure(bg="#1e293b")
            
            # Create main frame
            main_frame = tk.Frame(error_window, bg="#1e293b", padx=20, pady=20)
            main_frame.pack(fill='both', expand=True)
            
            # Error icon (you can use emoji or text)
            icon_label = tk.Label(main_frame, text="⚠️", font=("Segoe UI", 24), 
                                bg="#1e293b", fg="#dc3545")
            icon_label.pack(pady=(0, 10))
            
            # Error message
            msg_label = tk.Label(main_frame, text=message, font=("Montserrat", 10),
                            bg="#1e293b", fg="#ffffff", wraplength=300)
            msg_label.pack(pady=(0, 15))
            
            # OK button
            ok_button = tk.Button(main_frame, text="OK", command=error_window.destroy,
                                bg="#dc3545", fg="white", font=("Montserrat", 9, "bold"),
                                width=10, cursor="hand2", relief="flat")
            ok_button.pack()
            
            # Make sure window is on top
            error_window.lift()
            error_window.focus_force()
            
        except Exception as e:
            logging.error(f"Error showing error message: {e}")
            # Fallback to simple messagebox
            messagebox.showerror("Error", message)

    def _shake_widget(self, widget, amplitude=10, duration=500):
        """Create a shake animation effect for wrong password"""
        try:
            original_x = widget.winfo_x()
            shake_count = 6
            shake_time = duration // (shake_count * 2)
            
            def shake(count, direction):
                if count > 0 and widget.winfo_exists():
                    new_x = original_x + (amplitude * direction)
                    widget.geometry(f"+{new_x}+{widget.winfo_y()}")
                    widget.after(shake_time, lambda: shake(count - 1, -direction))
                else:
                    # Return to original position
                    if widget.winfo_exists():
                        widget.geometry(f"+{original_x}+{widget.winfo_y()}")
            
            shake(shake_count, 1)
            
        except Exception as e:
            logging.error(f"Error in shake animation: {e}")

    def _show_forgot_password_popup(self, parent_window):
        """Show forgot password information"""
        try:
            info_window = tk.Toplevel(parent_window)
            info_window.title("Forgot Password")
            info_window.geometry("400x200")
            info_window.transient(parent_window)
            info_window.grab_set()
            
            # Center window
            self.center_window_robust(info_window, 400, 200)
            
            # Configure background
            info_window.configure(bg="#1e293b")
            
            # Create main frame
            main_frame = tk.Frame(info_window, bg="#1e293b", padx=25, pady=25)
            main_frame.pack(fill='both', expand=True)
            
            # Title
            title_label = tk.Label(main_frame, text="Password Recovery", 
                                font=("Montserrat", 14, "bold"),
                                bg="#1e293b", fg="#ffffff")
            title_label.pack(pady=(0, 15))
            
            # Information text
            info_text = "Please contact your HR department or system administrator to reset your password."
            info_label = tk.Label(main_frame, text=info_text, 
                                font=("Montserrat", 10),
                                bg="#1e293b", fg="#94a3b8",
                                wraplength=350, justify="center")
            info_label.pack(pady=(0, 20))
            
            # OK button
            ok_button = tk.Button(main_frame, text="OK", command=info_window.destroy,
                                bg="#6c5ce7", fg="white", 
                                font=("Montserrat", 10, "bold"),
                                width=15, height=2, cursor="hand2", relief="flat")
            ok_button.pack()
            
        except Exception as e:
            logging.error(f"Error showing forgot password popup: {e}")
            messagebox.showinfo("Forgot Password", 
                            "Please contact your HR department to reset your password.")

    def _draw_simple_logo(self, canvas):
        """Draw a modern logo without animation to prevent Tkinter errors"""
        # Create fixed outer ring
        canvas.create_oval(2, 2, 58, 58, outline="#6c5ce7", width=2)
        
        # Create middle ring
        canvas.create_oval(12, 12, 48, 48, outline="#a29bfe", width=2)
        
        # Create inner circle
        canvas.create_oval(18, 18, 42, 42, fill="#6c5ce7", outline="")
        
        # Add highlight spot for 3D effect
        canvas.create_oval(22, 22, 32, 32, fill="#ffffff", outline="")
        
        
    
        
    # REPLACE handle_qr_login method with this:
    def handle_qr_login(self):
        """Handle QR code based login"""
        try:
            # Create QR code window
            qr_window = tk.Toplevel(self.main_window)
            qr_window.title("QR Code Login")
            qr_window.geometry("500x400")
            qr_window.transient(self.main_window)
            qr_window.grab_set()
            
            # Center window safely
            try:
                self.center_window_robust(qr_window, 500, 400)
            except:
                # Fallback centering
                screen_width = qr_window.winfo_screenwidth()
                screen_height = qr_window.winfo_screenheight()
                x = (screen_width - 500) // 2
                y = (screen_height - 400) // 2
                qr_window.geometry(f"500x400+{x}+{y}")
            
            # Create QR code receiver
            qr_receiver = AndroidIDReceiver()
            
            def on_qr_window_close():
                qr_receiver.cleanup()
                qr_window.destroy()
                    
            qr_window.protocol("WM_DELETE_WINDOW", on_qr_window_close)
            
            # Start QR code receiver
            qr_receiver.root = qr_window
            qr_receiver.setup_ui()
            qr_receiver.start_server()
            threading.Thread(target=qr_receiver.setup_ngrok, daemon=True).start()
            
            def handle_android_id(android_id):
                try:
                    # Get current time
                    current_time = datetime.now()
                    
                    # Get username from android_id using proper connection
                    conn = self.db_pool.get_local_connection()
                    cur = conn.cursor()
                    cur.execute("SELECT username FROM users WHERE user_android_id = ?", (android_id,))
                    result = cur.fetchone()
                    
                    if not result:
                        messagebox.showerror("Error", "No user found with this device ID")
                        return
                        
                    username = result[0]
                    
                    # Determine login type
                    login_type = self.db_pool.determine_login_type(username)
                    
                    # Save login
                    success, aws_success = self.db_pool.save_login_to_db(
                        username=username,
                        login_type=login_type,
                        timestamp=current_time,
                        with_camera=False
                    )
                    
                    if success:
                        # Show success notification
                        self._show_success_notification(
                            self.main_window,
                            username,
                            login_type,
                            current_time
                        )
                        
                        # Update attendance display
                        self.update_attendance_display_after_login(
                            username,
                            login_type,
                            current_time,
                            with_camera=False
                        )
                        
                        # Close QR window
                        qr_window.destroy()
                    else:
                        messagebox.showerror("Error", "Failed to record attendance")
                        
                except Exception as e:
                    logging.error(f"Error in QR login: {e}")
                    messagebox.showerror("Error", f"Login failed: {str(e)}")
            
            # Set callback
            qr_receiver.android_id_callback = handle_android_id
            
        except Exception as e:
            logging.error(f"Error in QR code login: {e}")
            messagebox.showerror("Error", f"QR login failed: {str(e)}")


    def register_new_user_with_qr(self):
        """Register a new user with QR code following the same pattern as camera registration"""
        try:
            # Check if company is selected in local database
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            
            # Create company_info table if it doesn't exist
            cur.execute("""
                CREATE TABLE IF NOT EXISTS company_info (
                    company_id TEXT PRIMARY KEY,
                    company_name TEXT NOT NULL,
                    password TEXT NOT NULL,
                    address_road TEXT,
                    address_city TEXT,
                    contact_person_name TEXT,
                    contact_person_designation TEXT,
                    contact_person_number TEXT,
                    created_at TEXT,
                    last_updated TEXT
                )
            """)
            
            # Check if any company exists
            cur.execute("SELECT company_name, company_id FROM company_info LIMIT 1")
            company_data = cur.fetchone()
            
            if not company_data:
                messagebox.showinfo("Company Required", "You did not have selected any company. Please select a company first then continue...")
                self.manage_company_info()
                return
            
            company_name, company_uuid = company_data

            # Create registration window
            register_window = tk.Toplevel(self.main_window)
            register_window.title("QR Code Registration")
            register_window.geometry("500x600")
            register_window.transient(self.main_window)
            register_window.grab_set()
            
            # Center window safely
            try:
                self.center_window_robust(register_window, 500, 600)
            except:
                pass

            form_frame = tk.Frame(register_window, padx=20, pady=20)
            form_frame.pack(fill='both', expand=True)

            # Title
            title_label = tk.Label(form_frame, text="Register New User with QR Code", 
                                font=("Helvetica", 16, "bold"))
            title_label.pack(pady=(0, 20))

            # Username
            username_frame = tk.Frame(form_frame)
            username_frame.pack(fill='x', pady=10)
            tk.Label(username_frame, text="Username:", font=("Helvetica", 10)).pack(side='left')
            username_var = tk.StringVar()
            username_entry = tk.Entry(username_frame, textvariable=username_var, font=("Helvetica", 10))
            username_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            # Designation
            designation_frame = tk.Frame(form_frame)
            designation_frame.pack(fill='x', pady=10)
            tk.Label(designation_frame, text="Designation:", font=("Helvetica", 10)).pack(side='left')
            designation_var = tk.StringVar()
            designation_entry = tk.Entry(designation_frame, textvariable=designation_var, font=("Helvetica", 10))
            designation_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            # Password
            password_frame = tk.Frame(form_frame)
            password_frame.pack(fill='x', pady=10)
            tk.Label(password_frame, text="Password:", font=("Helvetica", 10)).pack(side='left')
            password_var = tk.StringVar()
            password_entry = tk.Entry(password_frame, textvariable=password_var, show="*", font=("Helvetica", 10))
            password_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            # Company (auto-selected, read-only)
            company_frame = tk.Frame(form_frame)
            company_frame.pack(fill='x', pady=10)
            tk.Label(company_frame, text="Company:", font=("Helvetica", 10)).pack(side='left')
            company_var = tk.StringVar(value=company_name)
            company_entry = tk.Entry(company_frame, textvariable=company_var, state='readonly', font=("Helvetica", 10))
            company_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            def handle_registration():
                username = username_var.get().strip().upper()  # ALWAYS UPPERCASE
                password = password_var.get().strip()
                designation = designation_var.get().strip()

                if not username or not password or not designation:
                    messagebox.showerror("Error", "Please fill in all fields")
                    return

                # Check if username exists
                if self._is_username_exists(username):
                    if messagebox.askyesno("User Exists", "This user already exists. Are you an existing employee?"):
                        # Existing user: update android_id after QR scan
                        open_qr_for_existing_user(username, password, designation)
                    else:
                        messagebox.showerror("Error", "Please choose a different username")
                    return

                # New user: open QR for registration
                open_qr_for_new_user(username, password, designation)

            def open_qr_for_existing_user(username, password, designation):
                """Handle existing user android_id update - FIXED VERSION"""
                qr_window = tk.Toplevel(register_window)
                qr_window.title("QR Code Registration - Existing User")
                qr_window.geometry("500x400")
                qr_window.transient(register_window)
                qr_window.grab_set()
                
                qr_receiver = AndroidIDReceiver()
                
                def on_qr_window_close():
                    qr_receiver.cleanup()
                    qr_window.destroy()
                    
                qr_window.protocol("WM_DELETE_WINDOW", on_qr_window_close)
                qr_receiver.root = qr_window
                qr_receiver.setup_ui()
                qr_receiver.start_server()
                threading.Thread(target=qr_receiver.setup_ngrok, daemon=True).start()
                
                def handle_android_id(android_id):
                    """COMPLETE ANDROID ID HANDLER FOR EXISTING USER"""
                    try:
                        # STEP 1: Check for existing Android ID globally (across all companies)
                        local_conn = self.db_pool.get_local_connection()
                        local_cur = local_conn.cursor()
                        
                        # Check if Android ID already exists anywhere
                        local_cur.execute("SELECT username, company_name, company_user_uuid FROM users WHERE user_android_id = ?", (android_id,))
                        existing_device = local_cur.fetchone()
                        
                        if existing_device:
                            existing_username, existing_company, existing_company_user_uuid = existing_device
                            
                            # Show popup asking if user wants to update info with existing company
                            confirm_window = tk.Toplevel(qr_window)
                            confirm_window.title("Device Already Registered")
                            confirm_window.geometry("450x200")
                            confirm_window.transient(qr_window)
                            confirm_window.grab_set()
                            
                            self.center_window_robust(confirm_window, 450, 200)
                            
                            main_frame = tk.Frame(confirm_window, padx=20, pady=20)
                            main_frame.pack(fill='both', expand=True)
                            
                            message_text = f"We found this smartphone is registered with username '{existing_username}' of company '{existing_company}'\n\nDo you want to update your info with this company?"
                            
                            tk.Label(main_frame, text=message_text, font=('Arial', 10), 
                                    wraplength=400, justify='center').pack(pady=20)
                            
                            def handle_update_existing():
                                confirm_window.destroy()
                                qr_window.destroy()
                                register_window.destroy()
                                
                                # Update existing user with same company_user_uuid
                                try:
                                    # Update local database
                                    local_cur.execute("""
                                        UPDATE users 
                                        SET username = ?,
                                            password = ?,
                                            user_designation = ?,
                                            last_updated = strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc')
                                        WHERE company_user_uuid = ?
                                    """, (username, password, designation, existing_company_user_uuid))
                                    local_conn.commit()
                                    
                                    # Update AWS if online
                                    if self._check_internet_connection():
                                        aws_conn = self.db_pool.get_aws_connection()
                                        if aws_conn:
                                            try:
                                                aws_cur = aws_conn.cursor()
                                                aws_cur.execute("""
                                                    UPDATE users 
                                                    SET username = %s,
                                                        password = %s,
                                                        user_designation = %s,
                                                        last_updated = TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS')
                                                    WHERE company_user_uuid = %s
                                                """, (username, password, designation, existing_company_user_uuid))
                                                aws_conn.commit()
                                            finally:
                                                self.db_pool.return_aws_connection(aws_conn)
                                    
                                    messagebox.showinfo("Success", f"Updated device registration for company '{existing_company}' successfully!")
                                    
                                except Exception as e:
                                    logging.error(f"Error updating existing device: {e}")
                                    messagebox.showerror("Error", f"Failed to update device registration: {str(e)}")
                            
                            def handle_cancel():
                                confirm_window.destroy()
                                messagebox.showinfo("Cancelled", "Device registration cancelled.")
                            
                            button_frame = tk.Frame(main_frame)
                            button_frame.pack(pady=20)
                            
                            tk.Button(button_frame, text="Yes, Update Info", command=handle_update_existing,
                                    bg="#28a745", fg="white", font=('Arial', 10, 'bold'),
                                    width=15, height=2).pack(side='left', padx=10)
                            
                            tk.Button(button_frame, text="Cancel", command=handle_cancel,
                                    bg="#dc3545", fg="white", font=('Arial', 10, 'bold'),
                                    width=10, height=2).pack(side='left', padx=10)
                            
                            return  # Exit here if device already exists
                        
                        # STEP 2: Get existing user's company_user_uuid (DON'T change it)
                        local_cur.execute("SELECT company_user_uuid, user_uuid FROM users WHERE username = ?", (username,))
                        result = local_cur.fetchone()
                        
                        if not result:
                            messagebox.showerror("Error", "User data not found!")
                            return
                        
                        company_user_uuid, existing_user_uuid = result
                        
                        # STEP 3: Update existing user - only update android_id and other changeable fields
                        local_cur.execute("""
                            UPDATE users 
                            SET user_android_id = ?,
                                password = ?,
                                company_name = ?,
                                company_uuid = ?,
                                user_designation = ?,
                                last_updated = strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc')
                            WHERE company_user_uuid = ?
                        """, (android_id, password, company_name, company_uuid, designation, company_user_uuid))
                        local_conn.commit()
                        
                        # Update AWS if we have internet connection
                        aws_success = False
                        if self._check_internet_connection():
                            aws_conn = self.db_pool.get_aws_connection()
                            if aws_conn:
                                try:
                                    aws_cur = aws_conn.cursor()
                                    aws_cur.execute("""
                                        UPDATE users 
                                        SET user_android_id = %s,
                                            password = %s,
                                            company_name = %s,
                                            company_uuid = %s,
                                            user_designation = %s,
                                            last_updated = TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS')
                                        WHERE company_user_uuid = %s
                                    """, (android_id, password, company_name, company_uuid, designation, company_user_uuid))
                                    aws_conn.commit()
                                    aws_success = True
                                    logging.info("Successfully updated user in AWS")
                                except Exception as e:
                                    logging.error(f"AWS update failed: {e}")
                                    aws_conn.rollback()
                                finally:
                                    self.db_pool.return_aws_connection(aws_conn)
                            else:
                                logging.warning("Could not get AWS connection")
                        else:
                            logging.warning("No internet connection - AWS update skipped")
                        
                        # Show appropriate success message
                        if aws_success:
                            messagebox.showinfo("Success", "Device registered and user info updated successfully in both local and cloud databases!")
                        else:
                            messagebox.showinfo("Success", "Device registered and user info updated successfully in local database! Cloud sync will happen when connection is available.")
                            
                        qr_window.destroy()
                        register_window.destroy()
                        
                    except Exception as e:
                        logging.error(f"Error updating existing user: {e}")
                        messagebox.showerror("Error", f"Failed to update user: {str(e)}")
                
                qr_receiver.android_id_callback = handle_android_id

            def open_qr_for_new_user(username, password, designation):
                """Handle new user registration with android_id"""
                # Generate UUIDs for new user
                user_uuid = str(uuid.uuid4())
                company_user_uuid = str(uuid.uuid4())
                
                qr_window = tk.Toplevel(register_window)
                qr_window.title("QR Code Registration - New User")
                qr_window.geometry("500x400")
                qr_window.transient(register_window)
                qr_window.grab_set()
                
                qr_receiver = AndroidIDReceiver()
                
                def on_qr_window_close():
                    qr_receiver.cleanup()
                    qr_window.destroy()
                    
                qr_window.protocol("WM_DELETE_WINDOW", on_qr_window_close)
                qr_receiver.root = qr_window
                qr_receiver.setup_ui()
                qr_receiver.start_server()
                threading.Thread(target=qr_receiver.setup_ngrok, daemon=True).start()
                
                def handle_android_id(android_id):
                    """COMPLETE ANDROID ID HANDLER FOR NEW USER"""
                    try:
                        # STEP 1: Check for existing Android ID globally (across all companies)
                        local_conn = self.db_pool.get_local_connection()
                        local_cur = local_conn.cursor()
                        
                        # Check if Android ID already exists anywhere
                        local_cur.execute("SELECT username, company_name, company_user_uuid FROM users WHERE user_android_id = ?", (android_id,))
                        existing_device = local_cur.fetchone()
                        
                        if existing_device:
                            existing_username, existing_company, existing_company_user_uuid = existing_device
                            
                            # Show popup asking if user wants to update info with existing company
                            confirm_window = tk.Toplevel(qr_window)
                            confirm_window.title("Device Already Registered")
                            confirm_window.geometry("450x200")
                            confirm_window.transient(qr_window)
                            confirm_window.grab_set()
                            
                            self.center_window_robust(confirm_window, 450, 200)
                            
                            main_frame = tk.Frame(confirm_window, padx=20, pady=20)
                            main_frame.pack(fill='both', expand=True)
                            
                            message_text = f"We found this smartphone is registered with username '{existing_username}' of company '{existing_company}'\n\nDo you want to update your info with this company?"
                            
                            tk.Label(main_frame, text=message_text, font=('Arial', 10), 
                                    wraplength=400, justify='center').pack(pady=20)
                            
                            def handle_update_existing():
                                confirm_window.destroy()
                                qr_window.destroy()
                                register_window.destroy()
                                
                                # Update existing user with same company_user_uuid
                                try:
                                    # Update local database
                                    local_cur.execute("""
                                        UPDATE users 
                                        SET username = ?,
                                            password = ?,
                                            user_designation = ?,
                                            last_updated = strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc')
                                        WHERE company_user_uuid = ?
                                    """, (username, password, designation, existing_company_user_uuid))
                                    local_conn.commit()
                                    
                                    # Update AWS if online
                                    if self._check_internet_connection():
                                        aws_conn = self.db_pool.get_aws_connection()
                                        if aws_conn:
                                            try:
                                                aws_cur = aws_conn.cursor()
                                                aws_cur.execute("""
                                                    UPDATE users 
                                                    SET username = %s,
                                                        password = %s,
                                                        user_designation = %s,
                                                        last_updated = TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS')
                                                    WHERE company_user_uuid = %s
                                                """, (username, password, designation, existing_company_user_uuid))
                                                aws_conn.commit()
                                            finally:
                                                self.db_pool.return_aws_connection(aws_conn)
                                    
                                    messagebox.showinfo("Success", f"Updated device registration for company '{existing_company}' successfully!")
                                    
                                except Exception as e:
                                    logging.error(f"Error updating existing device: {e}")
                                    messagebox.showerror("Error", f"Failed to update device registration: {str(e)}")
                            
                            def handle_cancel():
                                confirm_window.destroy()
                                messagebox.showinfo("Cancelled", "Device registration cancelled.")
                            
                            button_frame = tk.Frame(main_frame)
                            button_frame.pack(pady=20)
                            
                            tk.Button(button_frame, text="Yes, Update Info", command=handle_update_existing,
                                    bg="#28a745", fg="white", font=('Arial', 10, 'bold'),
                                    width=15, height=2).pack(side='left', padx=10)
                            
                            tk.Button(button_frame, text="Cancel", command=handle_cancel,
                                    bg="#dc3545", fg="white", font=('Arial', 10, 'bold'),
                                    width=10, height=2).pack(side='left', padx=10)
                            
                            return  # Exit here if device already exists
                        
                        # STEP 2: Check if username exists in current company only
                        local_cur.execute("SELECT company_user_uuid FROM users WHERE username = ? AND company_uuid = ?", 
                                        (username, company_uuid))
                        existing_user = local_cur.fetchone()
                        
                        if existing_user:
                            existing_company_user_uuid = existing_user[0]
                            
                            # Show popup asking if existing employee
                            confirm_window = tk.Toplevel(qr_window)
                            confirm_window.title("Existing Username")
                            confirm_window.geometry("350x150")
                            confirm_window.transient(qr_window)
                            confirm_window.grab_set()
                            
                            self.center_window_robust(confirm_window, 350, 150)
                            
                            main_frame = tk.Frame(confirm_window, padx=20, pady=20)
                            main_frame.pack(fill='both', expand=True)
                            
                            tk.Label(main_frame, text=f"Username '{username}' already exists in this company.\nAre you an existing employee?",
                                    font=('Arial', 10), wraplength=300, justify='center').pack(pady=10)
                            
                            def handle_existing_employee():
                                confirm_window.destroy()
                                qr_window.destroy()
                                register_window.destroy()
                                
                                # Update existing employee with Android ID
                                try:
                                    local_cur.execute("""
                                        UPDATE users 
                                        SET user_android_id = ?,
                                            password = ?,
                                            user_designation = ?,
                                            last_updated = strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc')
                                        WHERE company_user_uuid = ?
                                    """, (android_id, password, designation, existing_company_user_uuid))
                                    local_conn.commit()
                                    
                                    # Update AWS if online
                                    if self._check_internet_connection():
                                        aws_conn = self.db_pool.get_aws_connection()
                                        if aws_conn:
                                            try:
                                                aws_cur = aws_conn.cursor()
                                                aws_cur.execute("""
                                                    UPDATE users 
                                                    SET user_android_id = %s,
                                                        password = %s,
                                                        user_designation = %s,
                                                        last_updated = TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS')
                                                    WHERE company_user_uuid = %s
                                                """, (android_id, password, designation, existing_company_user_uuid))
                                                aws_conn.commit()
                                            finally:
                                                self.db_pool.return_aws_connection(aws_conn)
                                    
                                    messagebox.showinfo("Success", "Existing employee updated with device ID successfully!")
                                    
                                except Exception as e:
                                    logging.error(f"Error updating existing employee: {e}")
                                    messagebox.showerror("Error", f"Failed to update employee: {str(e)}")
                            
                            def handle_not_existing():
                                confirm_window.destroy()
                                messagebox.showinfo("Info", "Please choose a different username.")
                                return
                            
                            button_frame = tk.Frame(main_frame)
                            button_frame.pack(pady=10)
                            
                            tk.Button(button_frame, text="Yes", command=handle_existing_employee,
                                    bg="#28a745", fg="white", width=8).pack(side='left', padx=5)
                            
                            tk.Button(button_frame, text="No", command=handle_not_existing,
                                    bg="#dc3545", fg="white", width=8).pack(side='left', padx=5)
                            
                            return  # Exit here if username exists
                        
                        # STEP 3: Create new user - no conflicts found
                        # Create blank face encoding for QR registration
                        blank_encoding = np.zeros(512)
                        
                        # Insert new user in local database
                        local_cur.execute("""
                            INSERT INTO users (company_user_uuid, username, password, face_encoding,
                                            user_uuid, company_name, company_uuid, user_designation, 
                                            user_android_id, created_at, last_updated)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, 
                                strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc'), 
                                strftime('%Y-%m-%d %H:%M:%S', 'now', 'utc'))
                        """, (company_user_uuid, username, password, serialize_face_embedding(blank_encoding),
                            user_uuid, company_name, company_uuid, designation, android_id))
                        local_conn.commit()
                        
                        # Insert in AWS if we have internet connection
                        aws_success = False
                        if self._check_internet_connection():
                            aws_conn = self.db_pool.get_aws_connection()
                            if aws_conn:
                                try:
                                    aws_cur = aws_conn.cursor()
                                    aws_cur.execute("""
                                        INSERT INTO users (company_user_uuid, username, password, face_encoding,
                                                        user_uuid, company_name, company_uuid, user_designation,
                                                        user_android_id, created_at, last_updated)
                                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,
                                            TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'),
                                            TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS'))
                                        ON CONFLICT (company_user_uuid) DO UPDATE SET
                                            password = EXCLUDED.password,
                                            face_encoding = EXCLUDED.face_encoding,
                                            user_uuid = EXCLUDED.user_uuid,
                                            company_name = EXCLUDED.company_name,
                                            company_uuid = EXCLUDED.company_uuid,
                                            user_designation = EXCLUDED.user_designation,
                                            user_android_id = EXCLUDED.user_android_id,
                                            last_updated = TO_CHAR(CURRENT_TIMESTAMP AT TIME ZONE 'UTC', 'YYYY-MM-DD HH24:MI:SS')
                                    """, (company_user_uuid, username, password, serialize_face_embedding(blank_encoding),
                                        user_uuid, company_name, company_uuid, designation, android_id))
                                    aws_conn.commit()
                                    aws_success = True
                                    logging.info("Successfully created user in AWS")
                                except Exception as e:
                                    logging.error(f"AWS creation failed: {e}")
                                    aws_conn.rollback()
                                finally:
                                    self.db_pool.return_aws_connection(aws_conn)
                            else:
                                logging.warning("Could not get AWS connection")
                        else:
                            logging.warning("No internet connection - AWS creation skipped")
                        
                        # Show appropriate success message
                        if aws_success:
                            messagebox.showinfo("Success", "User registered successfully in both local and cloud databases!")
                        else:
                            messagebox.showinfo("Success", "User registered successfully in local database! Cloud sync will happen when connection is available.")
                            
                        qr_window.destroy()
                        register_window.destroy()
                        
                    except Exception as e:
                        logging.error(f"Error registering new user: {e}")
                        messagebox.showerror("Error", f"Failed to register user: {str(e)}")
                
                qr_receiver.android_id_callback = handle_android_id

            # Register button
            register_btn = tk.Button(
                form_frame,
                text="Register with QR Code",
                command=handle_registration,
                bg="#74b9ff",
                fg="white",
                padx=20,
                pady=5,
                font=("Helvetica", 10, "bold"),
                width=25,
                height=2
            )
            register_btn.pack(pady=20)

            # Set focus to username entry
            username_entry.focus()

        except Exception as e:
            logging.error(f"Error in QR code registration: {e}")
            messagebox.showerror("Error", f"QR registration failed: {str(e)}")



    def _gradient_color(self, color1, color2, ratio):
        """Calculate a color between two colors at given ratio (0-1)"""
        # Convert hex colors to RGB
        r1, g1, b1 = int(color1[1:3], 16), int(color1[3:5], 16), int(color1[5:7], 16)
        r2, g2, b2 = int(color2[1:3], 16), int(color2[3:5], 16), int(color2[5:7], 16)
        
        # Calculate intermediate color
        r = int(r1 * (1 - ratio) + r2 * ratio)
        g = int(g1 * (1 - ratio) + g2 * ratio)
        b = int(b1 * (1 - ratio) + b2 * ratio)
        
        # Convert back to hex
        return f"#{r:02x}{g:02x}{b:02x}"

    # Completely redesigned notification window
    def _show_success_notification(self, parent, username, login_type, current_time):
        """Show a more professional success notification with large clock and better checkmark"""
        # Create success notification window
        success_window = tk.Toplevel(parent)
        success_window.title("")
        success_window.configure(bg="#171721")
        success_window.overrideredirect(True)  # Remove window decorations
        
        # Calculate position (bottom right corner)
        screen_width = success_window.winfo_screenwidth()
        screen_height = success_window.winfo_screenheight()
        width = 530  # Increased width for better layout
        height = 220  # Increased height for better layout
        x = screen_width - width - 20
        y = screen_height - height - 60  # Leave space for taskbar
        
        # Start off-screen
        success_window.geometry(f"{width}x{height}+{screen_width}+{y}")
        
        # Create main background frame
        main_bg = tk.Frame(success_window, bg="#1e293b")
        main_bg.place(x=0, y=0, relwidth=1, relheight=1)
        
        # Status bar at left (green for success)
        status_bar = tk.Frame(main_bg, bg="#4caf50", width=8)
        status_bar.place(x=0, y=0, width=8, relheight=1)
        
        # Create a two-column layout
        left_frame = tk.Frame(main_bg, bg="#1e293b")
        left_frame.place(x=20, y=10, width=260, height=height-20)
        
        right_frame = tk.Frame(main_bg, bg="#1e293b")
        right_frame.place(x=280, y=10, width=200, height=height-20)
        
        # LEFT COLUMN: SUCCESS INFO AND USER DETAILS
        
        # Header with clear checkmark
        header_frame = tk.Frame(left_frame, bg="#1e293b", height=40)
        header_frame.pack(fill="x")
        
        # Clear checkmark in a circle (more visible)
        check_canvas = tk.Canvas(header_frame, width=30, height=30, 
                                bg="#1e293b", highlightthickness=0)
        check_canvas.pack(side="left", padx=(0, 10))
        
        # Draw a more visible checkmark
        check_canvas.create_oval(2, 2, 28, 28, fill="#4caf50", outline="white", width=1)
        check_canvas.create_line(8, 15, 12, 20, 22, 8, fill="white", width=3)
        
        # Success title
        title_label = tk.Label(
            header_frame,
            text="Attendance Recorded",
            font=("Montserrat", 12, "bold"),
            fg="#ffffff",
            bg="#1e293b"
        )
        title_label.pack(side="left", anchor="w")
        
        # Separator
        separator = tk.Frame(left_frame, height=1, bg="#3e4c6a")
        separator.pack(fill="x", pady=5)
        
        # User information section
        user_info = tk.Frame(left_frame, bg="#1e293b")
        user_info.pack(fill="x", anchor="w", pady=10)
        
        # Username - prominent display
        username_label = tk.Label(
            user_info,
            text=username,
            font=("Montserrat", 22, "bold"),
            fg="#f1f5f9",
            bg="#1e293b",
            anchor="w"
        )
        username_label.pack(anchor="w", pady=(0, 5))
        
        # Login type with more prominence
        login_type_frame = tk.Frame(user_info, bg="#1e293b")
        login_type_frame.pack(fill="x", anchor="w")
        
        login_type_label = tk.Label(
            login_type_frame,
            text=login_type,
            font=("Montserrat", 18, "bold"),
            fg="#a29bfe",
            bg="#1e293b"
        )
        login_type_label.pack(side="left")
        
        # Readable timestamp
        time_frame = tk.Frame(user_info, bg="#1e293b")
        time_frame.pack(fill="x", anchor="w", pady=(15, 0))
        
        timestamp_label = tk.Label(
            time_frame,
            text="Recorded at:",
            font=("Montserrat", 12),
            fg="#94a3b8",
            bg="#1e293b"
        )
        timestamp_label.pack(side="left")
        
        time_str = current_time.strftime('%I:%M %p - %b %d')
        time_label = tk.Label(
            time_frame,
            text=time_str,
            font=("Montserrat", 10, "bold"),
            fg="#f1f5f9",
            bg="#1e293b"
        )
        time_label.pack(side="left", padx=(6, 0))
        
        # RIGHT COLUMN: ANALOG CLOCK - MUCH LARGER
        
        # Create large analog clock (prominent)
        clock_frame = tk.Frame(right_frame, bg="#1e293b")
        clock_frame.pack(expand=True, fill="both")
        
        clock_canvas = tk.Canvas(clock_frame, bg="#1e293b", highlightthickness=0)
        clock_canvas.pack(expand=True, fill="both")
        
        # Function to draw clock that respects container size
        def draw_clock():
            # Get actual dimensions of canvas
            width = clock_canvas.winfo_width()
            height = clock_canvas.winfo_height()
            size = min(width, height) - 20  # Padding
            center_x = width / 2
            center_y = height / 2
            radius = size / 2
            
            # Clear previous drawings
            clock_canvas.delete("all")
            
            # Draw outer circle with glow effect
            clock_canvas.create_oval(
                center_x - radius, center_y - radius,
                center_x + radius, center_y + radius,
                fill="#2d3748", outline="#a29bfe", width=3, tags="clock"
            )
            
            # Draw hour markers
            for i in range(12):
                angle = math.radians(i * 30)
                # Outer points of hour markers
                outer_x = center_x + (radius - 5) * math.sin(angle)
                outer_y = center_y - (radius - 5) * math.cos(angle)
                # Inner points of hour markers
                inner_x = center_x + (radius - 15) * math.sin(angle)
                inner_y = center_y - (radius - 15) * math.cos(angle)
                
                # Make quarter hours (3, 6, 9, 12) more prominent
                if i % 3 == 0:
                    clock_canvas.create_line(
                        inner_x, inner_y, outer_x, outer_y,
                        fill="#f1f5f9", width=3, tags="clock"
                    )
                else:
                    clock_canvas.create_line(
                        inner_x, inner_y, outer_x, outer_y,
                        fill="#a29bfe", width=2, tags="clock"
                    )
            
            # Calculate hand angles based on current time
            hour = current_time.hour % 12
            minute = current_time.minute
            second = current_time.second
            
            # Draw hour hand
            hour_angle = math.radians((hour * 30) + (minute / 2))
            hour_length = radius * 0.5
            hour_x = center_x + hour_length * math.sin(hour_angle)
            hour_y = center_y - hour_length * math.cos(hour_angle)
            clock_canvas.create_line(
                center_x, center_y, hour_x, hour_y,
                fill="#f1f5f9", width=5, arrow=tk.LAST, arrowshape=(8, 10, 3), tags="clock"
            )
            
            # Draw minute hand
            minute_angle = math.radians(minute * 6)
            minute_length = radius * 0.7
            minute_x = center_x + minute_length * math.sin(minute_angle)
            minute_y = center_y - minute_length * math.cos(minute_angle)
            clock_canvas.create_line(
                center_x, center_y, minute_x, minute_y,
                fill="#f1f5f9", width=3, arrow=tk.LAST, arrowshape=(8, 10, 3), tags="clock"
            )
            
            # Draw second hand
            second_angle = math.radians(second * 6)
            second_length = radius * 0.8
            second_x = center_x + second_length * math.sin(second_angle)
            second_y = center_y - second_length * math.cos(second_angle)
            clock_canvas.create_line(
                center_x, center_y, second_x, second_y,
                fill="#4caf50", width=2, tags="clock"
            )
            
            # Draw center cap
            clock_canvas.create_oval(
                center_x - 6, center_y - 6,
                center_x + 6, center_y + 6,
                fill="#a29bfe", outline="#f1f5f9", width=1, tags="clock"
            )
        
        # Close button at the bottom
        button_frame = tk.Frame(main_bg, bg="#1e293b")
        # button_frame.place(x=width-100-20, y=height-40-10, width=50, height=20)
        button_frame.place(x=width-50-20, y=height-30-10, width=50, height=20)
        close_button = tk.Button(
            button_frame,
            text="Close",
            font=("Montserrat", 10, "bold"),
            fg="white",
            bg="#6c5ce7",
            activebackground="#8075e5",
            relief=tk.FLAT,
            command=lambda: close_notification()
        )
        close_button.pack(fill="both", expand=True)
        
        # Auto-close countdown at bottom left
        countdown_frame = tk.Frame(main_bg, bg="#1e293b")
        countdown_frame.place(x=20, y=height-25-10, width=150, height=25)
        
        countdown_label = tk.Label(
            countdown_frame,
            text="Closing in 5s",
            font=("Montserrat", 9),
            fg="#94a3b8",
            bg="#1e293b"
        )
        countdown_label.pack(anchor="w")
        
        # Animate entry from right
        def animate_entry(pos):
            if pos > x:
                success_window.geometry(f"{width}x{height}+{pos}+{y}")
                success_window.after(10, lambda: animate_entry(pos - 20))
            else:
                success_window.geometry(f"{width}x{height}+{x}+{y}")
                # Draw clock once window is positioned
                draw_clock()
                # Start countdown
                countdown = 5
                update_countdown(countdown)
        
        # Update countdown
        def update_countdown(count):
            if count > 0:
                countdown_label.config(text=f"Closing in {count}s")
                success_window.after(1000, lambda: update_countdown(count - 1))
            else:
                # Animate exit
                animate_exit(x)
        
        # Animate exit to right
        def animate_exit(pos):
            if pos < screen_width:
                success_window.geometry(f"{width}x{height}+{pos}+{y}")
                success_window.after(10, lambda: animate_exit(pos + 20))
            else:
                success_window.destroy()
        
        # Function to close the notification
        def close_notification():
            animate_exit(x)
        
        # Start entry animation
        animate_entry(screen_width)
        
        # Bind configure event to redraw clock when window size changes
        clock_canvas.bind("<Configure>", lambda e: draw_clock())

            

    

    def handle_hr_window(self):
        """Handle HR window opening with password protection FIRST, then company check"""
        try:
            # Check if HR authentication window already exists
            if hasattr(self, '_hr_auth_window') and self._hr_auth_window is not None:
                try:
                    if self._hr_auth_window.winfo_exists():
                        self._hr_auth_window.lift()
                        self._hr_auth_window.focus_force()
                        return
                except tk.TclError:
                    self._hr_auth_window = None

            # FIRST: Check if any company exists in local database
            local_conn = self.db_pool.get_local_connection()
            local_cur = local_conn.cursor()
            local_cur.execute("SELECT company_name, company_id FROM company_info LIMIT 1")
            company_data = local_cur.fetchone()
            
            # If no company exists, show company selection first
            if not company_data:
                messagebox.showinfo("Company Required", "No company selected. Please select a company first!")
                self.manage_company_info()
                return
                
            company_name, company_uuid = company_data

            # SECOND: Now ask for HR password
            try:
                self._hr_auth_window = tk.Toplevel(self.main_window)
                password_window = self._hr_auth_window
                
                password_window.title("HR Authentication")
                password_window.withdraw()
                password_window.transient(self.main_window)
                password_window.resizable(False, False)
                password_window.grab_set()
                password_window.geometry("400x300")
                
                self.center_window_robust(password_window, 400, 300)
                password_window.deiconify()
                
            except Exception as window_error:
                logging.error(f"Failed to create HR auth window: {window_error}")
                if hasattr(self, '_hr_auth_window'):
                    self._hr_auth_window = None
                return

            # Create main frame
            try:
                main_frame = tk.Frame(password_window, padx=20, pady=20, bg="#f0f0f0")
                main_frame.pack(fill='both', expand=True)
                
                # Company info display
                company_frame = tk.Frame(main_frame, bg="#f0f0f0")
                company_frame.pack(fill='x', pady=(0, 20))
                
                tk.Label(company_frame, text="Company:", font=('Arial', 10, 'bold'), 
                        bg="#f0f0f0").pack(anchor='w')
                tk.Label(company_frame, text=company_name, font=('Arial', 12), 
                        bg="#f0f0f0", fg="blue").pack(anchor='w')
                
                # Password entry section
                tk.Label(main_frame, text="Enter HR Password:", font=('Arial', 12, 'bold'), 
                        bg="#f0f0f0").pack(pady=(10, 5))
                
                password_entry = tk.Entry(main_frame, show="*", font=('Arial', 11), width=30)
                password_entry.pack(pady=5)
                
                status_label = tk.Label(main_frame, text="", fg="red", bg="#f0f0f0")
                status_label.pack(pady=5)
                
                # Button frame
                button_frame = tk.Frame(main_frame, bg="#f0f0f0")
                button_frame.pack(pady=20)
                
                # Initialize database flag
                db_ready = [False]
                
                def initialize_auth_system():
                    """Initialize authentication system for current company"""
                    try:
                        local_conn = self.db_pool.get_local_connection()
                        local_cur = local_conn.cursor()
                        
                        # Check if HR password exists for this company
                        local_cur.execute("""
                            SELECT password FROM hr_passwords 
                            WHERE company_uuid = ? 
                            ORDER BY last_updated DESC LIMIT 1
                        """, (company_uuid,))
                        
                        if not local_cur.fetchone():
                            # No password for this company, create default
                            default_password = "HR@123"
                            local_cur.execute("""
                                INSERT OR REPLACE INTO hr_passwords (company_uuid, company_name, password) 
                                VALUES (?, ?, ?)
                            """, (company_uuid, company_name, default_password))
                            local_conn.commit()
                            logging.info(f"Created default HR password for company: {company_name}")
                        
                        db_ready[0] = True
                        logging.info("HR authentication system initialized successfully")
                        
                    except Exception as e:
                        logging.error(f"Error initializing HR auth system: {e}")
                        db_ready[0] = False
                
                def verify_password():
                    """Verify password for current company"""
                    try:
                        if not self._hr_auth_window or not self._hr_auth_window.winfo_exists():
                            return
                            
                        entered_password = password_entry.get().strip()
                        if not entered_password:
                            status_label.config(text="Please enter password")
                            return
                        
                        if not db_ready[0]:
                            status_label.config(text="System initializing, please wait...")
                            return
                        
                        # Get stored password for this company
                        local_conn = self.db_pool.get_local_connection()
                        local_cur = local_conn.cursor()
                        
                        local_cur.execute("""
                            SELECT password FROM hr_passwords 
                            WHERE company_uuid = ? 
                            ORDER BY last_updated DESC LIMIT 1
                        """, (company_uuid,))
                        local_result = local_cur.fetchone()
                        
                        # Check local password first
                        if local_result and entered_password == local_result[0]:
                            self.close_auth_window_safely()
                            self.show_hr_window()
                            return
                        
                        # If online and not found locally, try AWS
                        if self.db_pool.is_online:
                            try:
                                aws_conn = self.db_pool.get_aws_connection()
                                if aws_conn:
                                    try:
                                        aws_cur = aws_conn.cursor()
                                        aws_cur.execute("""
                                            SELECT password FROM hr_passwords 
                                            WHERE company_uuid = %s 
                                            ORDER BY last_updated DESC LIMIT 1
                                        """, (company_uuid,))
                                        aws_result = aws_cur.fetchone()
                                        
                                        if aws_result and entered_password == aws_result[0]:
                                            # Update local database with AWS password
                                            local_cur.execute("""
                                                INSERT OR REPLACE INTO hr_passwords 
                                                (company_uuid, company_name, password) 
                                                VALUES (?, ?, ?)
                                            """, (company_uuid, company_name, aws_result[0]))
                                            local_conn.commit()
                                            self.close_auth_window_safely()
                                            self.show_hr_window()
                                            return
                                    finally:
                                        self.db_pool.return_aws_connection(aws_conn)
                            except Exception as e:
                                logging.error(f"Error checking AWS password: {e}")
                        
                        # Password didn't match
                        status_label.config(text="Invalid password! Please try again.")
                        password_entry.delete(0, tk.END)
                        password_entry.focus_set()
                        
                    except Exception as e:
                        logging.error(f"Error verifying HR password: {e}")
                        if self._hr_auth_window and self._hr_auth_window.winfo_exists():
                            status_label.config(text="Authentication error. Please try again.")
                
                def change_hr_password():
                    """Enhanced Change HR password with proper UI"""
                    try:
                        if not self._hr_auth_window or not self._hr_auth_window.winfo_exists():
                            return
                            
                        if not db_ready[0]:
                            status_label.config(text="System initializing, please wait...")
                            return
                        
                        # Create password change window
                        change_window = tk.Toplevel(password_window)
                        change_window.title("Change HR Password")
                        change_window.geometry("400x350")
                        change_window.transient(password_window)
                        change_window.grab_set()
                        change_window.configure(bg="#f8f9fa")
                        
                        self.center_window_robust(change_window, 400, 350)
                        
                        change_frame = tk.Frame(change_window, padx=25, pady=25, bg="#f8f9fa")
                        change_frame.pack(fill='both', expand=True)
                        
                        # Title
                        tk.Label(change_frame, text=f"Change HR Password", 
                                font=('Arial', 14, 'bold'), bg="#f8f9fa").pack(pady=(0, 10))
                        tk.Label(change_frame, text=f"Company: {company_name}", 
                                font=('Arial', 10), bg="#f8f9fa", fg="blue").pack(pady=(0, 20))
                        
                        # Current password
                        tk.Label(change_frame, text="Current Password:", 
                                font=('Arial', 10, 'bold'), bg="#f8f9fa").pack(anchor='w', pady=(0, 5))
                        current_entry = tk.Entry(change_frame, show="*", font=('Arial', 11), width=30)
                        current_entry.pack(pady=(0, 15))
                        
                        # New password
                        tk.Label(change_frame, text="New Password:", 
                                font=('Arial', 10, 'bold'), bg="#f8f9fa").pack(anchor='w', pady=(0, 5))
                        new_entry = tk.Entry(change_frame, show="*", font=('Arial', 11), width=30)
                        new_entry.pack(pady=(0, 15))
                        
                        # Confirm new password
                        tk.Label(change_frame, text="Confirm New Password:", 
                                font=('Arial', 10, 'bold'), bg="#f8f9fa").pack(anchor='w', pady=(0, 5))
                        confirm_entry = tk.Entry(change_frame, show="*", font=('Arial', 11), width=30)
                        confirm_entry.pack(pady=(0, 15))
                        
                        change_status = tk.Label(change_frame, text="", fg="red", bg="#f8f9fa")
                        change_status.pack(pady=5)
                        
                        def process_change():
                            current_password = current_entry.get().strip()
                            new_password = new_entry.get().strip()
                            confirm_password = confirm_entry.get().strip()
                            
                            if not all([current_password, new_password, confirm_password]):
                                change_status.config(text="All fields are required")
                                return
                            
                            if new_password != confirm_password:
                                change_status.config(text="New passwords do not match")
                                return
                            
                            if len(new_password) < 3:
                                change_status.config(text="Password must be at least 3 characters")
                                return
                            
                            try:
                                # Verify current password
                                local_conn = self.db_pool.get_local_connection()
                                local_cur = local_conn.cursor()
                                local_cur.execute("""
                                    SELECT password FROM hr_passwords 
                                    WHERE company_uuid = ? 
                                    ORDER BY last_updated DESC LIMIT 1
                                """, (company_uuid,))
                                result = local_cur.fetchone()
                                
                                if not result or current_password != result[0]:
                                    change_status.config(text="Invalid current password")
                                    return
                                
                                # Update password in local database
                                local_cur.execute("""
                                    INSERT OR REPLACE INTO hr_passwords 
                                    (company_uuid, company_name, password) 
                                    VALUES (?, ?, ?)
                                """, (company_uuid, company_name, new_password))
                                local_conn.commit()
                                
                                # Update AWS if online
                                if self.db_pool.is_online:
                                    aws_conn = self.db_pool.get_aws_connection()
                                    if aws_conn:
                                        try:
                                            aws_cur = aws_conn.cursor()
                                            aws_cur.execute("""
                                                INSERT INTO hr_passwords (company_uuid, company_name, password, last_updated) 
                                                VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
                                                ON CONFLICT (company_uuid) DO UPDATE SET
                                                    password = EXCLUDED.password,
                                                    last_updated = EXCLUDED.last_updated
                                            """, (company_uuid, company_name, new_password))
                                            aws_conn.commit()
                                        finally:
                                            self.db_pool.return_aws_connection(aws_conn)
                                
                                messagebox.showinfo("Success", "HR password updated successfully!")
                                change_window.destroy()
                                
                            except Exception as e:
                                logging.error(f"Error changing password: {e}")
                                change_status.config(text="Failed to change password")
                        
                        # Button frame
                        btn_frame = tk.Frame(change_frame, bg="#f8f9fa")
                        btn_frame.pack(pady=20)
                        
                        # OK Button
                        ok_btn = tk.Button(btn_frame, text="OK", command=process_change,
                                        bg="#28a745", fg="white", font=('Arial', 10, 'bold'),
                                        width=12, height=2, cursor="hand2")
                        ok_btn.pack(side='left', padx=5)
                        
                        # Cancel Button  
                        cancel_btn = tk.Button(btn_frame, text="Cancel", command=change_window.destroy,
                                            bg="#dc3545", fg="white", font=('Arial', 10, 'bold'),
                                            width=12, height=2, cursor="hand2")
                        cancel_btn.pack(side='left', padx=5)
                        
                        # Enter key binding
                        def on_enter(event):
                            process_change()
                        
                        current_entry.bind('<Return>', on_enter)
                        new_entry.bind('<Return>', on_enter)
                        confirm_entry.bind('<Return>', on_enter)
                        change_window.bind('<Return>', on_enter)
                        
                        current_entry.focus_set()
                        
                    except Exception as e:
                        logging.error(f"Error in change password: {e}")
                
                # Create buttons with proper styling
                login_btn = tk.Button(button_frame, text="Login", command=verify_password, 
                                    bg="#007bff", fg="white", font=('Arial', 10, 'bold'),
                                    width=12, height=2, cursor="hand2")
                login_btn.pack(side='left', padx=5)
                
                change_btn = tk.Button(button_frame, text="Change Password", command=change_hr_password,
                                    bg="#ffc107", fg="black", font=('Arial', 10, 'bold'),
                                    width=15, height=2, cursor="hand2")
                change_btn.pack(side='left', padx=5)
                
                cancel_btn = tk.Button(button_frame, text="Cancel", command=self.close_auth_window_safely,
                                    bg="#6c757d", fg="white", font=('Arial', 10, 'bold'),
                                    width=12, height=2, cursor="hand2")
                cancel_btn.pack(side='left', padx=5)
                
                # Bind Enter key
                def on_enter_key(event):
                    try:
                        if self._hr_auth_window and self._hr_auth_window.winfo_exists():
                            verify_password()
                    except tk.TclError:
                        pass
                
                password_entry.bind('<Return>', on_enter_key)
                password_window.bind('<Return>', on_enter_key)
                
                # Set focus
                password_window.after(100, lambda: password_entry.focus_set())
                
                # Start background initialization
                init_thread = threading.Thread(target=initialize_auth_system, daemon=True)
                init_thread.start()
                
                # Handle window close
                password_window.protocol("WM_DELETE_WINDOW", self.close_auth_window_safely)
                
            except Exception as ui_error:
                logging.error(f"Error creating HR auth UI: {ui_error}")
                self.close_auth_window_safely()
                
        except Exception as e:
            logging.error(f"Error in HR authentication: {e}")
            self.close_auth_window_safely()
            
            
            

    def close_auth_window_safely(self):
        """Safely close the HR authentication window"""
        try:
            if hasattr(self, '_hr_auth_window') and self._hr_auth_window is not None:
                try:
                    if self._hr_auth_window.winfo_exists():
                        self._hr_auth_window.grab_release()  # Release grab first
                        self._hr_auth_window.destroy()
                except tk.TclError:
                    pass  # Window already destroyed
                finally:
                    self._hr_auth_window = None
        except Exception as e:
            logging.error(f"Error closing auth window: {e}")
            self._hr_auth_window = None

    def center_window_robust(self, window, width, height):
        """Robust window centering with comprehensive error handling"""
        try:
            # Validate inputs
            if not window:
                logging.error("No window provided for centering")
                return
                
            # Check if window exists and is accessible
            try:
                if not window.winfo_exists():
                    logging.error("Window does not exist for centering")
                    return
            except tk.TclError as e:
                logging.error(f"Window is not accessible: {e}")
                return
            
            # Ensure window is updated
            try:
                window.update_idletasks()
            except tk.TclError as e:
                logging.warning(f"Could not update window tasks: {e}")
            
            # Get screen dimensions with fallbacks
            try:
                screen_width = window.winfo_screenwidth()
                screen_height = window.winfo_screenheight()
                
                # Validate screen dimensions
                if screen_width <= 0 or screen_height <= 0:
                    screen_width = 1920
                    screen_height = 1080
                    
            except tk.TclError:
                # Fallback dimensions
                screen_width = 1920
                screen_height = 1080
                logging.warning("Using fallback screen dimensions")
            
            # Calculate center position
            x = max(0, (screen_width - width) // 2)
            y = max(0, (screen_height - height) // 2)
            
            # Set geometry safely
            try:
                geometry_string = f"{width}x{height}+{x}+{y}"
                window.geometry(geometry_string)
                logging.debug(f"Window geometry set to: {geometry_string}")
            except tk.TclError as e:
                logging.error(f"Failed to set window geometry: {e}")
                return
            
            # Try to bring window to front (optional operations)
            try:
                window.lift()
            except tk.TclError as e:
                logging.warning(f"Could not lift window: {e}")
            
            try:
                window.focus_force()
            except tk.TclError as e:
                logging.warning(f"Could not focus window: {e}")
                
        except Exception as e:
            logging.error(f"Unexpected error in window centering: {e}")
            # Don't re-raise - just log and continue
            
 
    def center_window(self, window, width=None, height=None):
        """Center a window on the screen - delegates to safe version"""
        self.center_window_safe(window, width, height)


    def show_hr_window(self):
        """Show HR window after successful authentication"""
        try:
            hr_window = tk.Toplevel(self.main_window)
            hr_window.title("HR Management")
            hr_window.geometry("370x400")
            
            try:
                main_x = self.main_window.winfo_x()
                main_y = self.main_window.winfo_y()
                hr_window.geometry(f"370x400+{main_x + 50}+{main_y + 50}")
            except:
                pass

            # Create canvas and scrollbar
            canvas = tk.Canvas(hr_window)
            scrollbar = tk.Scrollbar(hr_window, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=scrollbar.set)

            # Create main frame inside canvas
            main_frame = tk.Frame(canvas, padx=20, pady=20)
            
            # Add title label
            title_label = tk.Label(main_frame, text="HR Management", font=('Arial', 16, 'bold'))
            title_label.pack(pady=10)

            # Create all buttons with consistent styling
            buttons = [
                ('Register with Camera', lambda: self.register_new_user_with_camera() if self.tensorflow_ready else messagebox.showinfo("Please Wait", "System is still initializing...")),
                ('Register without Camera', lambda: self.register_new_user_without_camera() if self.tensorflow_ready else messagebox.showinfo("Please Wait", "System is still initializing...")),
                ('Attendance Report', self.open_customized_report_window),
                ('Sync Users to Local Database', self.sync_users_to_local),
                ('Check Fake Attendance', self.open_verify_image_window), 
                ('Upload AI Model (if Failure)', self.upload_ai_model),
                ('Delete Face', lambda: self.face_recognition_system.reset_database() if self.tensorflow_ready else messagebox.showinfo("Please Wait", "System is still initializing...")),
                ('Generate Facial Features', lambda: self.generate_embeddings_report() if self.tensorflow_ready else messagebox.showinfo("Please Wait", "System is still initializing...")),
                ('Sync Users to Cloud Database', self.sync_users_to_cloud),
                ('Sync HR Password to Local', self.sync_hr_password_to_local),  # UPDATED: Now company-oriented
                ('Register New User With QR Code', self.register_new_user_with_qr),
                ('Company Info.', self.manage_company_info),
            ]
            
            # Create and pack all buttons
            for text, command in buttons:
                if text == 'Company Info.':
                    wrapped_command = command
                else:
                    wrapped_command = lambda cmd=command: (cmd(), hr_window.destroy())
                
                btn = tk.Button(
                    main_frame,
                    text=text,
                    command=wrapped_command,
                    height=2,
                    width=30,
                    font=('Microsoft Uighur', 12, 'bold'),
                    relief='raised',
                    bg='blue4',
                    fg="white"
                )
                btn.pack(pady=10)

            # Pack scrollbar and canvas
            scrollbar.pack(side="right", fill="y")
            canvas.pack(side="left", fill="both", expand=True)

            # Add main frame to canvas
            canvas.create_window((0, 0), window=main_frame, anchor="nw")

            # Configure canvas scrolling
            def configure_scroll_region(event):
                canvas.configure(scrollregion=canvas.bbox("all"))

            main_frame.bind("<Configure>", configure_scroll_region)

            # Handle window close with X button
            hr_window.protocol("WM_DELETE_WINDOW", hr_window.destroy)
            
        except Exception as e:
            logging.error(f"Error opening HR window: {e}")
            messagebox.showerror("Error", "Failed to open HR window")
            
  


    def manage_company_info(self):
        """Handle company information management"""
        try:
            # Check if online
            if not self.db_pool.is_online:
                messagebox.showinfo("Error", "No Internet connection. Please connect to the network.")
                return

            # Check if online - but don't rely solely on is_online flag
            try:
                test_conn = self.db_pool.get_aws_connection()
                if not test_conn:
                    messagebox.showinfo("Error", "No Internet connection. Please connect to the network.")
                    return
                try:
                    self.db_pool.return_aws_connection(test_conn)
                except Exception as ex:
                    logging.error(f"Error returning AWS connection: {ex}")
            except Exception as e:
                messagebox.showinfo("Error", "No Internet connection. Please connect to the network.")
                return

            # Create company management window instantly
            company_window = tk.Toplevel(self.main_window)
            company_window.title("Company Management")
            company_window.geometry("400x300")
            company_window.transient(self.main_window)
            company_window.lift()
            company_window.focus_force()
            self.center_window(company_window, 400, 300)
            
            # Add title
            title_label = tk.Label(company_window, text="Company Management", font=('Arial', 14, 'bold'))
            title_label.pack(pady=20)
            
            # Add placeholder/loading label
            loading_label = tk.Label(company_window, text="Loading options...", font=('Arial', 10, 'italic'))
            loading_label.pack(pady=10)

            def load_buttons():
                loading_label.destroy()
                select_btn = tk.Button(
                    company_window,
                    text="Select An Existing Company",
                    command=lambda: self.select_existing_company(company_window),
                    height=2,
                    width=30,
                    font=('Microsoft Uighur', 12, 'bold'),
                    relief='raised',
                    bg='blue4',
                    fg="white"
                )
                select_btn.pack(pady=10)

                create_btn = tk.Button(
                    company_window,
                    text="Create a New Company",
                    command=lambda: self.create_new_company(company_window),
                    height=2,
                    width=30,
                    font=('Microsoft Uighur', 12, 'bold'),
                    relief='raised',
                    bg='blue4',
                    fg="white"
                )
                create_btn.pack(pady=10)

                edit_btn = tk.Button(
                    company_window,
                    text="Edit Existing Company",
                    command=lambda: self.edit_existing_company(company_window),
                    height=2,
                    width=30,
                    font=('Microsoft Uighur', 12, 'bold'),
                    relief='raised',
                    bg='blue4',
                    fg="white"
                )
                edit_btn.pack(pady=10)

            # Schedule button loading after window is shown
            company_window.after(100, load_buttons)

        except Exception as e:
            logging.error(f"Error in manage_company_info: {e}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")

    def select_existing_company(self, company_window):
        """Select an existing company from AWS with password verification"""
        try:
            # Create company selection window
            select_window = tk.Toplevel(company_window)
            select_window.title("Select Company")
            select_window.geometry("500x500")
            select_window.transient(company_window)
            select_window.transient(self.main_window)
            select_window.lift()
            select_window.focus_force()
            select_window.grab_set()
            
            # Center window
            self.center_window(select_window, 500, 500)
            
            # Add search field
            search_frame = tk.Frame(select_window)
            search_frame.pack(pady=10, fill='x', padx=20)
            
            tk.Label(search_frame, text="Search:", font=('Arial', 10)).pack(side='left')
            search_var = tk.StringVar()
            search_entry = tk.Entry(search_frame, textvariable=search_var, width=25)
            search_entry.pack(side='left', padx=5)
            
            # Create listbox with scrollbar for companies
            list_frame = tk.Frame(select_window)
            list_frame.pack(pady=10, fill='both', expand=True, padx=20)
            
            scrollbar = tk.Scrollbar(list_frame)
            scrollbar.pack(side='right', fill='y')
            
            company_listbox = tk.Listbox(list_frame, width=50, height=12, font=('Arial', 9))
            company_listbox.pack(side='left', fill='both', expand=True)
            
            company_listbox.config(yscrollcommand=scrollbar.set)
            scrollbar.config(command=company_listbox.yview)
            
            # Dictionary to store company data
            company_data = {}
            
            # Loading label
            loading_label = tk.Label(select_window, text="Loading companies...", fg='blue')
            loading_label.pack(pady=5)
            
            # Function to load companies from AWS (optimized)
            def load_companies(search_term=""):
                loading_label.config(text="Loading companies...")
                select_window.update_idletasks()
                
                company_listbox.delete(0, tk.END)
                company_data.clear()
                
                # Get AWS connection
                aws_conn = self.db_pool.get_aws_connection()
                if not aws_conn:
                    loading_label.config(text="Error: Cannot connect to AWS database", fg='red')
                    messagebox.showerror("Connection Error", "Cannot connect to AWS database")
                    return False
                
                try:
                    aws_cur = aws_conn.cursor()
                    
                    # Optimized query with search filter
                    if search_term.strip():
                        aws_cur.execute("""
                            SELECT company_id, company_name, password 
                            FROM company_info 
                            WHERE company_name ILIKE %s 
                            ORDER BY company_name
                            LIMIT 100
                        """, (f'%{search_term.strip()}%',))
                    else:
                        aws_cur.execute("""
                            SELECT company_id, company_name, password 
                            FROM company_info 
                            ORDER BY company_name
                            LIMIT 100
                        """)
                    
                    companies = aws_cur.fetchall()
                    
                    if not companies:
                        loading_label.config(text="No companies found", fg='orange')
                        return False
                    
                    for company_id, company_name, password in companies:
                        company_listbox.insert(tk.END, company_name)
                        company_data[company_name] = {
                            'id': company_id,
                            'password': password
                        }
                    
                    loading_label.config(text=f"Loaded {len(companies)} companies", fg='green')
                    return True
                    
                except Exception as e:
                    logging.error(f"Error loading companies: {e}")
                    loading_label.config(text="Error loading companies", fg='red')
                    messagebox.showerror("Database Error", f"Failed to load companies: {str(e)}")
                    return False
                finally:
                    try:
                        self.db_pool.return_aws_connection(aws_conn)
                    except Exception as return_error:
                        logging.error(f"Error returning AWS connection: {return_error}")
            
            # Initial load of companies
            if not load_companies():
                select_window.destroy()
                return
            
            # Search function with debouncing
            search_timer = None
            def search_companies(event=None):
                nonlocal search_timer
                if search_timer:
                    select_window.after_cancel(search_timer)
                search_timer = select_window.after(500, lambda: load_companies(search_var.get()))
            
            # Bind search entry
            search_entry.bind('<KeyRelease>', search_companies)
            search_button = tk.Button(search_frame, text="Search", command=lambda: load_companies(search_var.get()))
            search_button.pack(side='left', padx=5)
            
            def verify_and_sync_company():
                """Verify password and sync selected company"""
                try:
                    # Get selected company
                    selection = company_listbox.curselection()
                    if not selection:
                        messagebox.showwarning("Selection Required", "Please select a company to sync")
                        return
                    
                    company_name = company_listbox.get(selection[0])
                    company_info = company_data.get(company_name)
                    
                    if not company_info:
                        messagebox.showerror("Error", "Company information not found")
                        return
                    
                    company_id = company_info['id']
                    stored_password = company_info['password']
                    
                    # Check if company already exists in local database
                    local_conn = self.db_pool.get_local_connection()
                    local_cur = local_conn.cursor()
                    
                    # Create table if it doesn't exist
                    local_cur.execute("""
                        CREATE TABLE IF NOT EXISTS company_info (
                            company_id TEXT PRIMARY KEY,
                            company_name TEXT NOT NULL,
                            password TEXT NOT NULL,
                            address_road TEXT,
                            address_city TEXT,
                            contact_person_name TEXT,
                            contact_person_designation TEXT,
                            contact_person_number TEXT,
                            created_at TEXT,
                            last_updated TEXT
                        )
                    """)
                    
                    # Check if ANY company exists in local database (not just this one)
                    local_cur.execute("SELECT company_name FROM company_info")
                    existing_companies = local_cur.fetchall()
                    
                    if existing_companies:
                        # Show confirmation dialog with existing company names
                        existing_names = [comp[0] for comp in existing_companies]
                        if len(existing_names) == 1:
                            message = f"We found this company exists in local database:\n'{existing_names[0]}'\n\nAre you sure you want to replace it with '{company_name}'?"
                        else:
                            names_list = '\n'.join([f"• {name}" for name in existing_names])
                            message = f"We found these companies exist in local database:\n{names_list}\n\nAre you sure you want to replace them with '{company_name}'?"
                        
                        result = messagebox.askyesno(
                            "Company Exists", 
                            message,
                            icon='question'
                        )
                        if not result:
                            return
                        action_type = "replace"
                    else:
                        action_type = "sync"
                    
                    # Create password verification window
                    password_window = tk.Toplevel(select_window)
                    password_window.title("Verify Company Password")
                    password_window.geometry("350x200")
                    password_window.transient(select_window)
                    password_window.grab_set()
                    
                    # Center password window
                    self.center_window(password_window, 350, 200)
                    
                    # Password verification form
                    main_frame = tk.Frame(password_window, padx=20, pady=20)
                    main_frame.pack(fill='both', expand=True)
                    
                    tk.Label(main_frame, text=f"Company: {company_name}", 
                            font=('Arial', 11, 'bold')).pack(pady=5)
                    
                    tk.Label(main_frame, text="Enter company password : \nAnd Hit Enter to continue", 
                            font=('Arial', 10)).pack(pady=(10, 5))
                    
                    password_var = tk.StringVar()
                    password_entry = tk.Entry(main_frame, textvariable=password_var, 
                                            show="*", width=25, font=('Arial', 10))
                    password_entry.pack(pady=5)
                    password_entry.focus_set()
                    
                    # Status label
                    status_label = tk.Label(main_frame, text="", fg='red')
                    status_label.pack(pady=5)
                    
                    def perform_sync():
                        """Perform the actual sync operation"""
                        entered_password = password_var.get().strip()
                        
                        if not entered_password:
                            status_label.config(text="Please enter password")
                            return
                        
                        if entered_password != stored_password:
                            status_label.config(text="Invalid password! Please try again.")
                            password_entry.delete(0, tk.END)
                            password_entry.focus_set()
                            return
                        
                        # Password is correct, proceed with sync
                        password_window.destroy()
                        
                        # Show progress window
                        progress_window = tk.Toplevel(select_window)
                        progress_window.title("Syncing Company")
                        progress_window.geometry("300x120")
                        progress_window.transient(select_window)
                        progress_window.grab_set()
                        
                        self.center_window(progress_window, 300, 120)
                        
                        progress_label = tk.Label(progress_window, 
                                                text=f"{'Updating' if action_type == 'update' else 'Syncing'} {company_name}...",
                                                font=('Arial', 10))
                        progress_label.pack(pady=20)
                        
                        # Progress bar simulation
                        progress_bar_frame = tk.Frame(progress_window)
                        progress_bar_frame.pack(pady=10)
                        
                        progress_steps = ["Connecting to AWS...", "Fetching company data...", 
                                        "Syncing to local database...", "Completing sync..."]
                        current_step = 0
                        
                        def update_progress():
                            nonlocal current_step
                            if current_step < len(progress_steps):
                                progress_label.config(text=progress_steps[current_step])
                                current_step += 1
                                progress_window.after(500, update_progress)
                            else:
                                complete_sync()
                        
                        def complete_sync():
                            aws_conn = None
                            try:
                                # Get complete company data from AWS for ONLY the selected company
                                aws_conn = self.db_pool.get_aws_connection()
                                if not aws_conn:
                                    raise Exception("Cannot connect to AWS database")
                                
                                aws_cur = aws_conn.cursor()
                                # This query gets ONLY the selected company by company_id
                                aws_cur.execute("SELECT * FROM company_info WHERE company_id = %s", (company_id,))
                                company_record = aws_cur.fetchone()
                                
                                if not company_record:
                                    raise Exception("Company data not found in AWS")
                                
                                # Get column names
                                column_names = [desc[0] for desc in aws_cur.description]
                                
                                # Create company dictionary
                                company_dict = dict(zip(column_names, company_record))
                                
                                # Sync to local database - ONLY this selected company
                                local_conn = self.db_pool.get_local_connection()
                                local_cur = local_conn.cursor()
                                
                                # Create table if it doesn't exist
                                local_cur.execute("""
                                    CREATE TABLE IF NOT EXISTS company_info (
                                        company_id TEXT PRIMARY KEY,
                                        company_name TEXT NOT NULL,
                                        password TEXT NOT NULL,
                                        address_road TEXT,
                                        address_city TEXT,
                                        contact_person_name TEXT,
                                        contact_person_designation TEXT,
                                        contact_person_number TEXT,
                                        created_at TEXT,
                                        last_updated TEXT
                                    )
                                """)
                                
                                # Clear ALL existing companies first (ensure only one company)
                                local_cur.execute("DELETE FROM company_info")
                                
                                # Insert the selected company as the only company
                                placeholders = ', '.join(['?'] * len(column_names))
                                local_cur.execute(f"""
                                    INSERT INTO company_info ({', '.join(column_names)})
                                    VALUES ({placeholders})
                                """, [company_dict[col] for col in column_names])
                                
                                local_conn.commit()
                                
                                # Determine success message based on action type
                                if action_type == "replace":
                                    success_message = f"Company '{company_name}' replaced successfully!"
                                else:
                                    success_message = f"Company '{company_name}' synced successfully!"
                                
                                # Show completion
                                progress_label.config(text="Sync completed!")
                                progress_window.after(1000, progress_window.destroy)
                                
                                # Show success message
                                messagebox.showinfo("Success", success_message)
                                
                                # Close selection window
                                select_window.destroy()

                                company_window.destroy()  # Close company management window
                                # Open registration window
                                
                                # FIXED: Open HR Management window instead of registration
                                self.show_hr_window()
                            
                       
                            except Exception as e:
                                logging.error(f"Error during company sync: {e}")
                                progress_window.destroy()
                                messagebox.showerror("Sync Error", f"Failed to sync company: {str(e)}")
                            finally:
                                if aws_conn:
                                    try:
                                        self.db_pool.return_aws_connection(aws_conn)
                                    except Exception as return_error:
                                        logging.error(f"Error returning AWS connection: {return_error}")
                        
                        # Start progress updates
                        update_progress()
                    
                    # Buttons frame
                    button_frame = tk.Frame(main_frame)
                    button_frame.pack(pady=15)
                    
                    verify_btn = tk.Button(button_frame, text="Verify & Sync", 
                                        command=perform_sync,
                                        bg='green4', fg='white', 
                                        font=('Arial', 10, 'bold'),
                                        width=12)
                    verify_btn.pack(side='left', padx=5)
                    
                    cancel_btn = tk.Button(button_frame, text="Cancel", 
                                        command=password_window.destroy,
                                        bg='red4', fg='white',
                                        width=8)
                    cancel_btn.pack(side='left', padx=5)
                    
                    # Enter key binding
                    def on_enter(event):
                        perform_sync()
                    
                    password_entry.bind('<Return>', on_enter)
                    password_window.bind('<Return>', on_enter)
                    
                except Exception as e:
                    logging.error(f"Error in verify_and_sync_company: {e}")
                    messagebox.showerror("Error", f"An error occurred: {str(e)}")
            
            # Add buttons at the bottom
            button_frame = tk.Frame(select_window)
            button_frame.pack(pady=15)
            
            sync_btn = tk.Button(button_frame, text="Sync Selected Company", 
                            command=verify_and_sync_company,
                            bg='green4', fg='white', 
                            font=('Arial', 10, 'bold'),
                            width=20, height=2)
            sync_btn.pack(side='left', padx=10)
            
            cancel_btn = tk.Button(button_frame, text="Cancel", 
                                command=select_window.destroy,
                                bg='red4', fg='white',
                                width=10, height=2)
            cancel_btn.pack(side='left', padx=10)
            
            # Double-click to select
            def on_double_click(event):
                verify_and_sync_company()
            
            company_listbox.bind('<Double-1>', on_double_click)
            
        except Exception as e:
            logging.error(f"Error in select_existing_company: {e}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")



    def create_new_company(self, company_window):
        """Create a new company"""
        try:
            # Create company form window
            form_window = tk.Toplevel(company_window)
            form_window.title("Create New Company")
            form_window.geometry("500x600")
            form_window.transient(company_window)
            form_window.grab_set()
            
            # Center the window
            try:
                self.center_window_robust(form_window, 500, 600)
            except:
                pass
            
            # Add form fields
            form_frame = tk.Frame(form_window, padx=20, pady=20)
            form_frame.pack(fill='both', expand=True)
            
            # Title
            title_label = tk.Label(form_frame, text="Create New Company", font=('Arial', 14, 'bold'))
            title_label.grid(row=0, column=0, columnspan=2, pady=10, sticky='w')
            
            # Form fields
            fields = [
                ("Company Name", None),
                ("Password", None),
                ("Address Road/House", None),
                ("Address City", None),
                ("Contact Person Name", None),
                ("Contact Person Designation", None),
                ("Contact Person Number", None)
            ]
            
            # Create entry widgets for each field
            entries = {}
            for i, (field_name, _) in enumerate(fields):
                tk.Label(form_frame, text=f"{field_name}:", anchor='w').grid(
                    row=i+1, column=0, sticky='w', pady=10)
                
                # Use password entry for password field
                if field_name == "Password":
                    entry = tk.Entry(form_frame, width=30, show="*")
                else:
                    entry = tk.Entry(form_frame, width=30)
                    
                entry.grid(row=i+1, column=1, sticky='w', padx=10)
                entries[field_name] = entry
            
            # Function to save company
            def save_company():
                # Validate required fields
                company_name = entries["Company Name"].get().strip()
                password = entries["Password"].get().strip()
                
                if not company_name or not password:
                    messagebox.showinfo("Required Fields", "Company Name and Password are required")
                    return
                
                # Generate UUID for company_id
                company_id = str(uuid.uuid4())
                
                # Get current timestamp
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Collect all field values
                company_data = {
                    "company_id": company_id,
                    "company_name": company_name,
                    "password": password,
                    "address_road": entries["Address Road/House"].get().strip(),
                    "address_city": entries["Address City"].get().strip(),
                    "contact_person_name": entries["Contact Person Name"].get().strip(),
                    "contact_person_designation": entries["Contact Person Designation"].get().strip(),
                    "contact_person_number": entries["Contact Person Number"].get().strip(),
                    "created_at": current_time,
                    "last_updated": current_time
                }
                
                # Show progress window
                progress_window = tk.Toplevel(form_window)
                progress_window.title("Saving Company")
                progress_window.geometry("300x100")
                progress_window.transient(form_window)
                progress_window.grab_set()
                
                try:
                    self.center_window_robust(progress_window, 300, 100)
                except:
                    pass
                
                progress_label = tk.Label(progress_window, text="Saving company information...")
                progress_label.pack(pady=20)
                progress_window.update()
                
                aws_conn = None
                try:
                    # Save to AWS only
                    aws_conn = self.db_pool.get_aws_connection()
                    if not aws_conn:
                        messagebox.showerror("Error", "Cannot connect to AWS database")
                        progress_window.destroy()
                        return
                    
                    aws_cur = aws_conn.cursor()
                    
                    # Check if company name already exists
                    aws_cur.execute("""
                        SELECT 1 FROM company_info WHERE company_name = %s
                    """, (company_name,))
                    
                    if aws_cur.fetchone():
                        messagebox.showinfo("Company Exists", "A company with this name already exists")
                        progress_window.destroy()
                        return
                    
                    # Insert into AWS only
                    columns = company_data.keys()
                    placeholders = ', '.join(['%s'] * len(columns))
                    
                    aws_cur.execute(f"""
                        INSERT INTO company_info ({', '.join(columns)})
                        VALUES ({placeholders})
                    """, list(company_data.values()))
                    
                    aws_conn.commit()
                    
                    # Update progress and show success
                    progress_label.config(text="Company saved successfully!")
                    progress_window.after(1500, progress_window.destroy)
                    
                    messagebox.showinfo("Success", "Company created successfully!")
                    form_window.destroy()
                    
                except Exception as e:
                    logging.error(f"Error saving company: {e}")
                    progress_window.destroy()
                    messagebox.showerror("Error", f"Failed to save company: {str(e)}")
                    if aws_conn:
                        try:
                            aws_conn.rollback()
                        except Exception as rollback_error:
                            logging.error(f"Error during rollback: {rollback_error}")
                finally:
                    if aws_conn:
                        try:
                            self.db_pool.return_aws_connection(aws_conn)
                        except Exception as return_error:
                            logging.error(f"Error returning AWS connection: {return_error}")
            
            # Function to cancel and close the form
            def cancel_form():
                form_window.destroy()
            
            # Add buttons at the bottom
            button_frame = tk.Frame(form_frame)
            button_frame.grid(row=len(fields)+2, column=0, columnspan=2, pady=20)
            
            # Save button
            save_btn = tk.Button(
                button_frame,
                text="Create Company",
                command=save_company,
                bg="#28a745",
                fg="white",
                font=('Arial', 10, 'bold'),
                width=15,
                height=2,
                cursor="hand2"
            )
            save_btn.pack(side='left', padx=10)
            
            # Cancel button
            cancel_btn = tk.Button(
                button_frame,
                text="Cancel",
                command=cancel_form,
                bg="#dc3545",
                fg="white",
                font=('Arial', 10, 'bold'),
                width=15,
                height=2,
                cursor="hand2"
            )
            cancel_btn.pack(side='left', padx=10)
            
            # Bind Enter key to save (when focus is on entry fields)
            def on_enter_key(event):
                save_company()
            
            for entry in entries.values():
                entry.bind('<Return>', on_enter_key)
            
            # Set focus to first entry field
            entries["Company Name"].focus_set()
            
            # Handle window close event
            def on_window_close():
                try:
                    form_window.destroy()
                except Exception as e:
                    logging.error(f"Error closing company creation window: {e}")
            
            form_window.protocol("WM_DELETE_WINDOW", on_window_close)
            
            # Add validation for real-time feedback
            def validate_company_name(*args):
                name = entries["Company Name"].get().strip()
                if len(name) > 0:
                    # Enable save button
                    save_btn.config(state='normal')
                else:
                    # Disable save button
                    save_btn.config(state='disabled')
            
            # Initially disable save button
            save_btn.config(state='disabled')
            
            # Bind validation to company name entry
            company_name_var = tk.StringVar()
            entries["Company Name"].config(textvariable=company_name_var)
            company_name_var.trace('w', validate_company_name)
            
            logging.info("Company creation form displayed successfully")
            
        except Exception as e:
            logging.error(f"Error in create_new_company: {e}")
            messagebox.showerror("Error", f"Failed to open company creation form: {str(e)}")
            try:
                if 'form_window' in locals() and form_window.winfo_exists():
                    form_window.destroy()
            except:
                pass
            


    def edit_existing_company(self, company_window):
        """Edit an existing company - AWS ONLY"""
        try:
            # First select a company
            select_window = tk.Toplevel(company_window)
            select_window.title("Select Company to Edit")
            select_window.geometry("500x400")
            
            # Add search field
            search_frame = tk.Frame(select_window)
            search_frame.pack(pady=10, fill='x', padx=20)
            
            tk.Label(search_frame, text="Search:").pack(side='left')
            search_var = tk.StringVar()
            search_entry = tk.Entry(search_frame, textvariable=search_var, width=30)
            search_entry.pack(side='left', padx=5)
            
            # Create listbox with scrollbar for companies
            list_frame = tk.Frame(select_window)
            list_frame.pack(pady=10, fill='both', expand=True, padx=20)
            
            scrollbar = tk.Scrollbar(list_frame)
            scrollbar.pack(side='right', fill='y')
            
            company_listbox = tk.Listbox(list_frame, width=50, height=15)
            company_listbox.pack(side='left', fill='both', expand=True)
            
            company_listbox.config(yscrollcommand=scrollbar.set)
            scrollbar.config(command=company_listbox.yview)
            
            # Dictionary to store company_id -> company_name mapping
            company_data = {}
            
            # Function to load companies from AWS
            def load_companies(search_term=""):
                company_listbox.delete(0, tk.END)
                company_data.clear()
                
                # Get AWS connection
                aws_conn = self.db_pool.get_aws_connection()
                if not aws_conn:
                    messagebox.showerror("Error", "Cannot connect to AWS database")
                    return
                
                try:
                    aws_cur = aws_conn.cursor()
                    
                    # Query with search filter if provided
                    if search_term:
                        aws_cur.execute("""
                            SELECT company_id, company_name FROM company_info 
                            WHERE company_name ILIKE %s 
                            ORDER BY company_name
                        """, (f'%{search_term}%',))
                    else:
                        aws_cur.execute("""
                            SELECT company_id, company_name FROM company_info 
                            ORDER BY company_name
                        """)
                    
                    companies = aws_cur.fetchall()
                    
                    for company_id, company_name in companies:
                        company_listbox.insert(tk.END, company_name)
                        company_data[company_name] = company_id
                        
                except Exception as e:
                    logging.error(f"Error loading companies: {e}")
                    messagebox.showerror("Error", f"Failed to load companies: {str(e)}")
                finally:
                    try:
                        self.db_pool.return_aws_connection(aws_conn)
                    except Exception as return_error:
                        logging.error(f"Error returning AWS connection: {return_error}")
            
            # Initial load of companies
            load_companies()
            
            # Search function
            def search_companies(event=None):
                search_term = search_var.get()
                load_companies(search_term)
            
            # Bind search entry to search function
            search_entry.bind('<Return>', search_companies)
            tk.Button(search_frame, text="Search", command=search_companies).pack(side='left', padx=5)
            
            # Complete the edit_selected_company function - AWS ONLY VERSION
            def edit_selected_company():
                selection = company_listbox.curselection()
                if not selection:
                    messagebox.showinfo("Selection Required", "Please select a company")
                    return
                
                company_name = company_listbox.get(selection[0])
                company_id = company_data[company_name]
                
                # Verify password
                password_window = tk.Toplevel(select_window)
                password_window.title("Verify Password")
                password_window.geometry("350x200")
                password_window.transient(select_window)
                password_window.grab_set()
                
                # Center password window
                self.center_window(password_window, 350, 200)
                
                # Password verification form
                main_frame = tk.Frame(password_window, padx=20, pady=20)
                main_frame.pack(fill='both', expand=True)
                
                tk.Label(main_frame, text=f"Company: {company_name}", 
                        font=('Arial', 11, 'bold')).pack(pady=5)
                
                tk.Label(main_frame, text="Enter company password to edit: \nAnd hit enter to continue", 
                        font=('Arial', 10)).pack(pady=(10, 5))
                
                password_var = tk.StringVar()
                password_entry = tk.Entry(main_frame, textvariable=password_var, 
                                        show="*", width=25, font=('Arial', 10))
                password_entry.pack(pady=5)
                password_entry.focus_set()
                
                # Status label for error messages
                status_label = tk.Label(main_frame, text="", fg='red')
                status_label.pack(pady=5)
                
                def verify_password_and_edit():
                    """Verify password and open edit form - AWS ONLY"""
                    entered_password = password_var.get().strip()
                    
                    if not entered_password:
                        status_label.config(text="Please enter password")
                        return
                    
                    # Get AWS connection for password verification
                    aws_conn = None
                    try:
                        aws_conn = self.db_pool.get_aws_connection()
                        if not aws_conn:
                            status_label.config(text="Cannot connect to AWS database")
                            return
                        
                        aws_cur = aws_conn.cursor()
                        
                        # Get stored password from AWS
                        aws_cur.execute("""
                            SELECT password FROM company_info WHERE company_id = %s
                        """, (company_id,))
                        
                        result = aws_cur.fetchone()
                        if not result:
                            status_label.config(text="Company not found in database")
                            return
                        
                        stored_password = result[0]
                        
                        if entered_password != stored_password:
                            status_label.config(text="Invalid password! Please try again.")
                            password_entry.delete(0, tk.END)
                            password_entry.focus_set()
                            return
                        
                        # Password correct, close password window and open edit form
                        password_window.destroy()
                        self.open_aws_edit_form(company_id, company_name)
                        
                    except Exception as e:
                        logging.error(f"Error verifying password: {e}")
                        status_label.config(text="Error verifying password")
                    finally:
                        if aws_conn:
                            try:
                                self.db_pool.return_aws_connection(aws_conn)
                            except Exception as return_error:
                                logging.error(f"Error returning AWS connection: {return_error}")
                
                # Add buttons
                button_frame = tk.Frame(main_frame)
                button_frame.pack(pady=15)
                
                verify_btn = tk.Button(button_frame, text="Verify & Edit", 
                                    command=verify_password_and_edit,
                                    bg='green4', fg='white', 
                                    font=('Arial', 10, 'bold'),
                                    width=12)
                verify_btn.pack(side='left', padx=5)
                
                cancel_btn = tk.Button(button_frame, text="Cancel", 
                                    command=password_window.destroy,
                                    bg='red4', fg='white',
                                    width=8)
                cancel_btn.pack(side='left', padx=5)
                
                # Enter key binding for password entry
                def on_enter(event):
                    verify_password_and_edit()
                
                password_entry.bind('<Return>', on_enter)
                password_window.bind('<Return>', on_enter)

            # Add buttons for the select window
            button_frame = tk.Frame(select_window)
            button_frame.pack(pady=15)

            edit_btn = tk.Button(button_frame, text="Edit Selected Company", 
                            command=edit_selected_company,
                            bg='orange', fg='white', 
                            font=('Arial', 10, 'bold'),
                            width=20, height=2)
            edit_btn.pack(side='left', padx=10)

            cancel_btn = tk.Button(button_frame, text="Cancel", 
                                command=select_window.destroy,
                                bg='red4', fg='white',
                                width=10, height=2)
            cancel_btn.pack(side='left', padx=10)

            # Double-click to edit
            def on_double_click(event):
                edit_selected_company()

            company_listbox.bind('<Double-1>', on_double_click)

        except Exception as e:
            logging.error(f"Error in edit_existing_company: {e}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


    def open_aws_edit_form(self, company_id, company_name):
        """Open form to edit company in AWS"""
        try:
            # Use context manager for better connection handling
            with self.db_pool.get_aws_connection_context() as (aws_conn, aws_cur):
                if not aws_conn or not aws_cur:
                    messagebox.showerror("Error", "Cannot connect to AWS database")
                    return
                
                # Get company data from AWS
                aws_cur.execute("""
                    SELECT * FROM company_info WHERE company_id = %s
                """, (company_id,))
                
                company_data = aws_cur.fetchone()
                if not company_data:
                    messagebox.showerror("Error", "Company data not found")
                    return
                
                # Get column names
                column_names = [desc[0] for desc in aws_cur.description]
                
                # Create a dictionary of column name -> value
                company_dict = dict(zip(column_names, company_data))
            
            # Create edit form window
            form_window = tk.Toplevel()
            form_window.title("Edit Company")
            form_window.geometry("500x600")
            
            # Add form fields
            form_frame = tk.Frame(form_window, padx=20, pady=20)
            form_frame.pack(fill='both', expand=True)
            
            # Title
            title_label = tk.Label(form_frame, text=f"Edit Company: {company_dict['company_name']}", font=('Arial', 14, 'bold'))
            title_label.grid(row=0, column=0, columnspan=2, pady=10, sticky='w')
            
            # Form fields mapping
            field_mapping = {
                "Company Name": "company_name",
                "Password": "password",
                "Address Road/House": "address_road",
                "Address City": "address_city",
                "Contact Person Name": "contact_person_name",
                "Contact Person Designation": "contact_person_designation",
                "Contact Person Number": "contact_person_number"
            }
            
            # Create entry widgets for each field
            entries = {}
            row = 1
            for display_name, db_field in field_mapping.items():
                tk.Label(form_frame, text=f"{display_name}:", anchor='w').grid(
                    row=row, column=0, sticky='w', pady=10)
                
                # Use password entry for password field
                if display_name == "Password":
                    entry = tk.Entry(form_frame, width=30, show="*")
                else:
                    entry = tk.Entry(form_frame, width=30)
                
                # Set current value
                if company_dict.get(db_field):
                    entry.insert(0, company_dict[db_field])
                    
                entry.grid(row=row, column=1, sticky='w', padx=10)
                entries[db_field] = entry
                row += 1
            
            # Function to update company
            def update_company():
                # Check if still online
                if not self.db_pool.is_online:
                    messagebox.showerror("Error", "No internet connection. Cannot update company.")
                    return
                
                # Validate required fields
                company_name = entries["company_name"].get().strip()
                password = entries["password"].get().strip()
                
                if not company_name or not password:
                    messagebox.showinfo("Required Fields", "Company Name and Password are required")
                    return
                
                # Get current timestamp
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                
                # Collect all field values
                updated_data = {
                    "company_name": company_name,
                    "password": password,
                    "address_road": entries["address_road"].get().strip(),
                    "address_city": entries["address_city"].get().strip(),
                    "contact_person_name": entries["contact_person_name"].get().strip(),
                    "contact_person_designation": entries["contact_person_designation"].get().strip(),
                    "contact_person_number": entries["contact_person_number"].get().strip(),
                    "last_updated": current_time
                }
                
                # Show progress window
                progress_window = tk.Toplevel(form_window)
                progress_window.title("Updating Company")
                progress_window.geometry("300x100")
                progress_label = tk.Label(progress_window, text="Updating company information...")
                progress_label.pack(pady=20)
                progress_window.update()
                
                try:
                    # Use context manager for update operation
                    with self.db_pool.get_aws_connection_context() as (update_aws_conn, update_aws_cur):
                        if not update_aws_conn or not update_aws_cur:
                            messagebox.showerror("Error", "Cannot connect to AWS database")
                            progress_window.destroy()
                            return
                        
                        # Check if another company with same name exists (excluding this one)
                        update_aws_cur.execute("""
                            SELECT 1 FROM company_info 
                            WHERE company_name = %s AND company_id != %s
                        """, (company_name, company_id))
                        
                        if update_aws_cur.fetchone():
                            messagebox.showinfo("Company Exists", "Another company with this name already exists")
                            progress_window.destroy()
                            return
                        
                        # Update AWS
                        update_fields = [f"{field} = %s" for field in updated_data.keys()]
                        update_values = list(updated_data.values())
                        update_values.append(company_id)  # For WHERE clause
                        
                        update_aws_cur.execute(f"""
                            UPDATE company_info 
                            SET {', '.join(update_fields)}
                            WHERE company_id = %s
                        """, update_values)
                        
                        update_aws_conn.commit()
                        
                        # Update progress
                        progress_label.config(text="Company updated successfully!")
                        progress_window.after(1000, progress_window.destroy)
                        
                        # Ask if user wants to sync to local
                        sync_choice = messagebox.askyesno(
                            "Sync to Local", 
                            "Company updated successfully in AWS!\n\nDo you want to sync this updated company to your local database?",
                            icon='question'
                        )
                        
                        if sync_choice:
                            self.sync_updated_company_to_local(company_id, company_name)
                        
                        messagebox.showinfo("Success", "Company updated successfully!")
                        form_window.destroy()
                        
                except Exception as e:
                    logging.error(f"Error updating company: {e}")
                    progress_window.destroy()
                    messagebox.showerror("Error", f"Failed to update company: {str(e)}")
            
            # Add buttons
            button_frame = tk.Frame(form_frame)
            button_frame.grid(row=row, column=0, columnspan=2, pady=20)
            
            tk.Button(
                button_frame, 
                text="Update Company", 
                command=update_company,
                bg='green4',
                fg='white',
                font=('Arial', 10, 'bold')
            ).pack(side='left', padx=5)
            
            tk.Button(
                button_frame, 
                text="Cancel", 
                command=form_window.destroy,
                bg='red4',
                fg='white'
            ).pack(side='left', padx=5)
            
            # Handle window close
            def on_window_close():
                form_window.destroy()
            
            form_window.protocol("WM_DELETE_WINDOW", on_window_close)
            
        except Exception as e:
            logging.error(f"Error in open_aws_edit_form: {e}")
            messagebox.showerror("Error", f"An error occurred: {str(e)}")


    def sync_updated_company_to_local(self, company_id, company_name):
        """Sync the updated company from AWS to local database"""
        try:
            # Show progress window
            progress_window = tk.Toplevel()
            progress_window.title("Syncing to Local")
            progress_window.geometry("300x120")
            progress_label = tk.Label(progress_window, text=f"Syncing {company_name} to local database...")
            progress_label.pack(pady=20)
            progress_window.update()
            
            # Get updated company data from AWS
            aws_conn = self.db_pool.get_aws_connection()
            if not aws_conn:
                progress_window.destroy()
                messagebox.showerror("Error", "Cannot connect to AWS database for sync")
                return
            
            try:
                aws_cur = aws_conn.cursor()
                aws_cur.execute("SELECT * FROM company_info WHERE company_id = %s", (company_id,))
                company_record = aws_cur.fetchone()
                
                if not company_record:
                    progress_window.destroy()
                    messagebox.showerror("Error", "Company data not found in AWS")
                    return
                
                # Get column names
                column_names = [desc[0] for desc in aws_cur.description]
                company_dict = dict(zip(column_names, company_record))
                
                # Sync to local database
                local_conn = self.db_pool.get_local_connection()
                local_cur = local_conn.cursor()
                
                # Create table if it doesn't exist
                local_cur.execute("""
                    CREATE TABLE IF NOT EXISTS company_info (
                        company_id TEXT PRIMARY KEY,
                        company_name TEXT NOT NULL,
                        password TEXT NOT NULL,
                        address_road TEXT,
                        address_city TEXT,
                        contact_person_name TEXT,
                        contact_person_designation TEXT,
                        contact_person_number TEXT,
                        created_at TEXT,
                        last_updated TEXT
                    )
                """)
                
                # Clear existing companies and insert updated one
                local_cur.execute("DELETE FROM company_info")
                
                # Insert the updated company
                placeholders = ', '.join(['?'] * len(column_names))
                local_cur.execute(f"""
                    INSERT INTO company_info ({', '.join(column_names)})
                    VALUES ({placeholders})
                """, [company_dict[col] for col in column_names])
                
                local_conn.commit()
                
                progress_label.config(text="Sync completed successfully!")
                progress_window.after(1500, progress_window.destroy)
                
                messagebox.showinfo("Sync Complete", f"Company '{company_name}' has been synced to local database!")
                
            finally:
                try:
                    self.db_pool.return_aws_connection(aws_conn)
                except Exception as ex:
                    logging.error(f"Error returning AWS connection during sync: {ex}")
                    try:
                        aws_conn.close()
                    except:
                        pass
                        
        except Exception as e:
            logging.error(f"Error syncing company to local: {e}")
            messagebox.showerror("Error", f"Failed to sync company to local database: {str(e)}")

            

    def sync_users_to_cloud(self):
        try:
            if not self.db_pool.is_online:
                messagebox.showinfo("Error", "No internet connection. Cannot sync with AWS.")
                return

            # Show progress window
            progress_window = tk.Toplevel(self.hr_window)
            progress_window.title("Sync Progress")
            progress_window.geometry("300x150")
            
            progress_label = tk.Label(progress_window, text="Checking users...", pady=10)
            progress_label.pack()
            
            count_label = tk.Label(progress_window, text="")
            count_label.pack()

            # Get local users
            local_conn = self.db_pool.get_local_connection()
            local_cur = local_conn.cursor()
            local_cur.execute("SELECT company_user_uuid, username, face_encoding, created_at, last_updated FROM users")
            local_users = local_cur.fetchall()

            # Get AWS connection
            aws_conn = self.db_pool.get_aws_connection()
            if not aws_conn:
                messagebox.showinfo("Error", "Cannot connect to AWS database")
                return

            try:
                aws_cur = aws_conn.cursor()
                synced_count = 0
                skipped_count = 0

                for user in local_users:
                    company_user_uuid, username, face_encoding, created_at, last_updated = user
                    
                    # Check if user exists in AWS
                    aws_cur.execute("SELECT 1 FROM users WHERE company_user_uuid = %s", (company_user_uuid,))
                    user_exists = aws_cur.fetchone()

                    if not user_exists:
                        # User doesn't exist in AWS, sync them
                        try:
                            aws_cur.execute("""
                                INSERT INTO users 
                                (company_user_uuid, username, face_encoding, created_at, last_updated)
                                VALUES (%s, %s, %s, %s, %s)
                            """, (company_user_uuid, username, face_encoding, created_at, last_updated))
                            aws_conn.commit()
                            synced_count += 1
                        except Exception as e:
                            logging.error(f"Error syncing user {username}: {e}")
                            continue
                    else:
                        skipped_count += 1

                    # Update progress
                    count_label.config(
                        text=f"Synced: {synced_count}\nSkipped: {skipped_count}\nProcessing: {username}"
                    )
                    progress_window.update()

                progress_window.destroy()
                messagebox.showinfo(
                    "Sync Complete", 
                    f"Sync completed!\nNew users synced: {synced_count}\nExisting users skipped: {skipped_count}"
                )

            finally:
                if aws_conn:
                    self.db_pool.return_aws_connection(aws_conn)

        except Exception as e:
            logging.error(f"Error in cloud sync: {e}")
            messagebox.showerror("Error", "Failed to sync with cloud database")

    def sync_hr_password_to_local(self):
        """Sync HR password for current company from AWS to local"""
        try:
            # Get current company info
            local_conn = self.db_pool.get_local_connection()
            local_cur = local_conn.cursor()
            local_cur.execute("SELECT company_name, company_id FROM company_info LIMIT 1")
            company_data = local_cur.fetchone()
            
            if not company_data:
                messagebox.showinfo("Error", "No company selected. Please select a company first!")
                return
                
            company_name, company_uuid = company_data

            # Show progress window
            progress_window = tk.Toplevel()
            progress_window.title("Sync HR Password")
            progress_window.geometry("350x120")
            progress_label = tk.Label(progress_window, 
                                    text=f"Syncing HR password for {company_name}...")
            progress_label.pack(pady=20)
            progress_window.update()

            if not self.db_pool.is_online:
                progress_label.config(text="Cannot sync - No internet connection")
                progress_window.after(2000, progress_window.destroy)
                return

            # Get password for current company from AWS
            aws_conn = self.db_pool.get_aws_connection()
            if not aws_conn:
                progress_label.config(text="Cannot connect to AWS database")
                progress_window.after(2000, progress_window.destroy)
                return

            try:
                aws_cur = aws_conn.cursor()
                aws_cur.execute("""
                    SELECT password, created_at, last_updated 
                    FROM hr_passwords 
                    WHERE company_uuid = %s
                    ORDER BY last_updated DESC 
                    LIMIT 1
                """, (company_uuid,))
                result = aws_cur.fetchone()

                if not result:
                    progress_label.config(text=f"No HR password found for {company_name} in AWS")
                    progress_window.after(3000, progress_window.destroy)
                    return

                # Save to local database for current company
                local_cur.execute("""
                    INSERT OR REPLACE INTO hr_passwords 
                    (company_uuid, company_name, password, created_at, last_updated)
                    VALUES (?, ?, ?, datetime('now'), datetime('now'))
                """, (company_uuid, company_name, result[0]))
                
                local_conn.commit()

                progress_label.config(text=f"HR password synced successfully for {company_name}!")
                progress_window.after(2000, progress_window.destroy)
                messagebox.showinfo("Success", f"HR password synced for {company_name}")

            except Exception as e:
                logging.error(f"Error syncing HR password: {e}")
                progress_label.config(text="Error syncing password")
                progress_window.after(2000, progress_window.destroy)
                messagebox.showerror("Error", "Failed to sync HR password")
            finally:
                if aws_conn:
                    self.db_pool.return_aws_connection(aws_conn)

        except Exception as e:
            logging.error(f"Error in HR password sync: {e}")
            messagebox.showerror("Error", "Failed to sync HR password")




            
            
  
    def handle_camera_login(self):
        """Handle camera login with immediate camera access"""
        try:
            cap = cv2.VideoCapture(0)
            if not cap.isOpened():
                messagebox.showerror("Error", "Cannot access camera!")
                return

            # Create preview window
            preview_window = tk.Toplevel(self.main_window)
            preview_window.title("Camera Preview")
            preview_label = tk.Label(preview_window)
            preview_label.pack()
            
            status_label = tk.Label(preview_window, 
                                text="Camera Ready - Click Capture when ready",
                                fg="green")
            status_label.pack()

            def update_preview():
                ret, frame = cap.read()
                if ret:
                    frame = cv2.flip(frame, 1)
                    frame = cv2.resize(frame, (640, 480))
                    img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    img = Image.fromarray(img)
                    imgtk = ImageTk.PhotoImage(image=img)
                    preview_label.imgtk = imgtk
                    preview_label.configure(image=imgtk)
                    if not preview_window.winfo_exists():
                        return
                    preview_window.after(10, update_preview)

            def capture_image():
                ret, frame = cap.read()
                if ret:
                    self.captured_frame = frame
                    status_label.config(text="Image Captured! Processing...", fg="orange")
                    
                    # Process the captured frame
                    self.process_captured_frame()
                    
                    # Update attendance display with "---" while processing
                    self.update_attendance_display(
                        username="---",
                        login_type="---",
                        timestamp=datetime.now(),
                        with_camera=True
                    )
                    
                    cap.release()
                    preview_window.destroy()

            # Add larger capture button with custom styling
            capture_button = tk.Button(
                preview_window, 
                text="CAPTURE", 
                command=capture_image,
                width=20,
                height=3,
                font=('Arial', 16, 'bold'),
                bg='#4CAF50',
                fg='white',
                relief='raised',
                cursor='hand2'
            )
            capture_button.pack(pady=20)

            update_preview()

        except Exception as e:
            logging.error(f"Camera error: {e}")
            messagebox.showerror("Error", "Camera access failed!")
    def process_captured_frame(self):
        """Process captured frame when TensorFlow is ready"""
        if self.captured_frame is None:
            return

        def check_tensorflow_status():
            if self.tensorflow_ready:
                try:
                    # Store the captured frame in face_recognition_system
                    self.face_recognition_system.current_frame = self.captured_frame
                    
                    # Verify identity and get result
                    result = self.face_recognition_system.verify_identity()
                    
                    # If verification successful, update attendance display immediately
                    if result and hasattr(result, 'username'):
                        current_time = self.time_manager.get_accurate_time()
                        login_type = self.db_pool.determine_login_type(result.username)
                        
                        # Update attendance info display
                        self.attendance_labels['username'].config(text=result.username.upper())
                        self.attendance_labels['date'].config(text=current_time.strftime('%d %b %Y'))
                        self.attendance_labels['login_type'].config(text=login_type)
                        self.attendance_labels['record_time'].config(text=current_time.strftime('%I:%M %p'))
                        
                        # Determine attendance type
                        login_hour = current_time.hour
                        attendance_type = "On Time"
                        if login_hour >= 9:
                            attendance_type = "Late"
                        if login_hour >= 10:
                            attendance_type = "Late Absent"
                        
                        self.attendance_labels['attendance_type'].config(text=attendance_type)
                        self.attendance_labels['attendance_method'].config(text="Camera")
                        
                        # Update monthly statistics
                        conn = self.db_pool.get_local_connection()
                        cur = conn.cursor()
                        cur.execute("""
                            SELECT 
                                COUNT(CASE WHEN strftime('%H', in_time) >= '09' THEN 1 END) as lates,
                                COUNT(CASE WHEN strftime('%H', in_time) >= '10' THEN 1 END) as late_absents,
                                COUNT(CASE WHEN strftime('%H', out_time) < '17' THEN 1 END) as early_leaves
                            FROM attendance_records 
                            WHERE username = ? 
                            AND strftime('%Y-%m', date) = ?
                        """, (result.username, current_time.strftime('%Y-%m')))
                        
                        stats = cur.fetchone()
                        if stats:
                            self.attendance_labels['total_lates'].config(text=str(stats[0]))
                            self.attendance_labels['total_late_absents'].config(text=str(stats[1]))
                            self.attendance_labels['total_early_leaves'].config(text=str(stats[2]))
                        
                except Exception as e:
                    logging.error(f"Processing error: {e}")
                    messagebox.showerror("Error", "Face processing failed!")
            else:
                # Check again in 1 second
                self.main_window.after(1000, check_tensorflow_status)
                self.status_labels['tensorflow'].config(
                    text="Still Loading AI Heavy Models...... Please wait.", fg="orange")

        # Start checking TensorFlow status
        check_tensorflow_status()
           

        def check_tensorflow_status():
            if self.tensorflow_ready:
                try:
                    # Process the frame with face recognition
                    self.face_recognition_system.verify_identity(self.captured_frame)
                except Exception as e:
                    logging.error(f"Processing error: {e}")
                    messagebox.showerror("Error", "Face processing failed!")
            else:
                # Check again in 1 second
                self.main_window.after(1000, check_tensorflow_status)
                self.status_labels['tensorflow'].config(
                    text="Still Loading AI Heavy Models...... Please wait.", fg="orange")

        # Start checking TensorFlow status
        check_tensorflow_status()



    def initialize_database(self):
        try:
            self.db_pool = DatabasePool.get_instance()
            self.db_pool.initialize_database()
            self.update_status("database", "Connected", success=True)
        except Exception as e:
            self.update_status("database", "Connection failed", success=False)


            

    def safe_verify_identity(self):
        """Safe wrapper for verify_identity"""
        if not self.is_system_ready():
            messagebox.showinfo("Please Wait", "System is still initializing...")
            return
        self.face_recognition_system.verify_identity()


    def safe_handle_forced_login(self):
        """Safe wrapper for handle_forced_login"""
        if not self.is_system_ready():
            messagebox.showinfo("Please Wait", "System is still initializing...")
            return
        self.handle_forced_login()

    def safe_open_hr_window(self):
        """Safe wrapper for open_hr_window"""
        if not self.is_system_ready():
            messagebox.showinfo("Please Wait", "System is still initializing...")
            return
        self.open_hr_window()



    def is_system_ready(self):
        """Check if all components are initialized"""
        return all([
            self.face_recognition_system is not None,
            self.db_pool is not None,
            self.network_manager is not None
        ])

    
    
    

    def on_closing(self):
        """Handle application cleanup on closing - syncs only last 7 days with a futuristic UI"""
        try:
            self.closing = True
            
            # Check if cleanup has already been called
            if hasattr(self, '_cleanup_called'):
                logging.info("Cleanup already in progress, destroying main window")
                try:
                    if hasattr(self, 'main_window') and self.main_window:
                        self.main_window.destroy()
                except:
                    pass
                return
            
            if self.db_pool and self.db_pool.is_online:            
            

                import math
                import time
                
                # Create a modern progress window with glassmorphism design
                progress_window = tk.Toplevel(self.main_window)
                progress_window.title("Sync Dashboard")
                progress_window.geometry("750x550")
                progress_window.transient(self.main_window)
                progress_window.grab_set()
                
                # Set window transparency (if available on platform)
                try:
                    progress_window.attributes("-alpha", 0.97)  # Slight transparency
                except:
                    pass
                    
                # Dark mode theme
                bg_color = "#111214"
                text_color = "#ffffff"
                accent_color = "#3584e4"
                card_bg = "#1a1b1e"
                success_color = "#57e389"
                warning_color = "#f8e45c"
                error_color = "#ed5e68"
                neutral_color = "#77767b"
                
                progress_window.configure(bg=bg_color)
                
                # Prevent window from being closed by X button
                progress_window.protocol("WM_DELETE_WINDOW", lambda: None)

                # Create canvas for rounded corners and shadows
                style = ttk.Style()
                style.configure("TProgressbar", 
                                thickness=8, 
                                troughcolor=card_bg,
                                background=accent_color,
                                borderwidth=0)
                
                # Create a futuristic header with animated title
                header_frame = tk.Frame(progress_window, bg=bg_color, height=80)
                header_frame.pack(fill=tk.X, pady=(15, 5), padx=30)
                
                # Create animated sync icon
                self.sync_angle = 0
                
                def draw_sync_icon(canvas, x, y, size=24, color=accent_color):
                    # Clear previous drawing
                    canvas.delete("sync_icon")
                    
                    # Calculate positions for the circular sync arrows
                    angle = self.sync_angle
                    radius = size/2
                    
                    # Draw rotating arrows
                    arrow1_x1 = x + radius * math.cos(math.radians(angle))
                    arrow1_y1 = y + radius * math.sin(math.radians(angle))
                    arrow1_x2 = x + radius * math.cos(math.radians(angle + 120))
                    arrow1_y2 = y + radius * math.sin(math.radians(angle + 120))
                    
                    arrow2_x1 = x + radius * math.cos(math.radians(angle + 180))
                    arrow2_y1 = y + radius * math.sin(math.radians(angle + 180))
                    arrow2_x2 = x + radius * math.cos(math.radians(angle + 300))
                    arrow2_y2 = y + radius * math.sin(math.radians(angle + 300))
                    
                    # Draw arrows with rounded ends
                    canvas.create_line(arrow1_x1, arrow1_y1, arrow1_x2, arrow1_y2, 
                                    fill=color, width=2.5, arrow=tk.LAST, 
                                    arrowshape=(8, 10, 5), smooth=True, tags="sync_icon")
                    canvas.create_line(arrow2_x1, arrow2_y1, arrow2_x2, arrow2_y2, 
                                    fill=color, width=2.5, arrow=tk.LAST, 
                                    arrowshape=(8, 10, 5), smooth=True, tags="sync_icon")
                    
                    # Update the angle for the next frame
                    self.sync_angle = (self.sync_angle + 5) % 360
                    
                    # Schedule the next update if window still exists
                    if progress_window.winfo_exists():
                        progress_window.after(50, lambda: draw_sync_icon(canvas, x, y, size, color))
                
                icon_canvas = tk.Canvas(header_frame, width=40, height=40, 
                                    bg=bg_color, highlightthickness=0)
                icon_canvas.pack(side=tk.LEFT, padx=(0, 15))
                
                # Start the animation
                progress_window.after(50, lambda: draw_sync_icon(icon_canvas, 20, 20, 30, accent_color))
                
                # Animated title using typing effect
                title_text = "CLOUD SYNC DASHBOARD"
                title_index = [0]  # Using list to make it mutable in nested function
                title_label = tk.Label(header_frame, text="", font=("Segoe UI", 18, "bold"), 
                                bg=bg_color, fg=text_color)
                title_label.pack(side=tk.LEFT)
                
                def type_title():
                    if title_index[0] < len(title_text):
                        current_text = title_label.cget("text")
                        title_label.config(text=current_text + title_text[title_index[0]])
                        title_index[0] += 1
                        progress_window.after(40, type_title)
                
                # Start title animation
                progress_window.after(100, type_title)
                
                # Create main content frame with glassmorphism effect
                main_frame = tk.Frame(progress_window, bg=bg_color)
                main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=10)
                
                # Left panel for status and progress
                left_panel = tk.Frame(main_frame, bg=bg_color, width=300)
                left_panel.pack(side=tk.LEFT, fill=tk.BOTH, padx=(0, 15))
                
                # Right panel for log
                right_panel = tk.Frame(main_frame, bg=bg_color)
                right_panel.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
                
                # Status card with glow effect
                status_card = tk.Frame(left_panel, bg=card_bg, padx=15, pady=15)
                status_card.pack(fill=tk.X, pady=(0, 15))
                
                # Create pulsing glow effect for active cards
                def pulse_glow(frame, intensity=0, increasing=True):
                    if not progress_window.winfo_exists():
                        return
                        
                    if increasing:
                        intensity += 1
                        if intensity >= 10:
                            increasing = False
                    else:
                        intensity -= 1
                        if intensity <= 0:
                            increasing = True
                    
                    # Subtle border glow using multiple frames
                    border_color = f"#{int(17 + intensity * 3):02x}{int(53 + intensity * 3):02x}{int(228 - intensity * 5):02x}"
                    frame.config(highlightbackground=border_color, highlightcolor=border_color)
                    
                    # Schedule next pulse
                    progress_window.after(50, lambda: pulse_glow(frame, intensity, increasing))
                
                # Apply glowing border to the status card
                status_card.config(highlightthickness=1, highlightbackground=accent_color)
                progress_window.after(100, lambda: pulse_glow(status_card))
                
                # Status content with icon
                status_header = tk.Frame(status_card, bg=card_bg)
                status_header.pack(fill=tk.X, pady=(0, 10))
                
                status_icon_label = tk.Label(status_header, text="⚡", font=("Segoe UI", 14), 
                                        bg=card_bg, fg=accent_color)
                status_icon_label.pack(side=tk.LEFT)
                
                status_title = tk.Label(status_header, text="SYNC STATUS", font=("Segoe UI", 12, "bold"), 
                                    bg=card_bg, fg=text_color)
                status_title.pack(side=tk.LEFT, padx=(5, 0))
                
                # Status info with modern label
                status_content = tk.Frame(status_card, bg=card_bg)
                status_content.pack(fill=tk.X)
                
                current_action = tk.Label(status_content, text="Initializing...", 
                                    font=("Segoe UI", 11), bg=card_bg, 
                                    fg=accent_color, anchor="w", justify=tk.LEFT)
                current_action.pack(fill=tk.X, pady=(0, 10))
                
                # Create circular progress indicator
                progress_canvas = tk.Canvas(status_card, width=180, height=180, 
                                        bg=card_bg, highlightthickness=0)
                progress_canvas.pack(pady=(5, 10))
                
                # Variables for progress animation
                self.progress_value = 0
                self.target_progress = 0
                
                def draw_progress(canvas, progress=0):
                    if not progress_window.winfo_exists():
                        return
                        
                    canvas.delete("progress")
                    
                    # Draw background circle
                    canvas.create_oval(10, 10, 170, 170, fill=bg_color, 
                                    outline=neutral_color, width=2, tags="progress")
                    
                    # Draw progress arc
                    if progress > 0:
                        angle = 360 * (progress / 100)
                        canvas.create_arc(20, 20, 160, 160, start=90, extent=-angle, 
                                        outline=accent_color, width=8, 
                                        style=tk.ARC, tags="progress")
                    
                    # Draw center details
                    canvas.create_oval(45, 45, 135, 135, fill=card_bg, 
                                    outline="", tags="progress")
                    
                    # Draw percentage text
                    canvas.create_text(90, 90, text=f"{int(progress)}%", 
                                    font=("Segoe UI", 22, "bold"), 
                                    fill=text_color, tags="progress")

                    # Schedule next animation frame if not at target
                    if abs(self.progress_value - self.target_progress) > 0.5:
                        # Animate smoothly towards target
                        self.progress_value += (self.target_progress - self.progress_value) * 0.1
                        progress_window.after(30, lambda: draw_progress(canvas, self.progress_value))
                
                # Initialize progress
                draw_progress(progress_canvas, 0)
                
                # Stats indicators with modern design
                stats_frame = tk.Frame(status_card, bg=card_bg)
                stats_frame.pack(fill=tk.X, pady=(10, 0))
                
                # Records counter with pill shape
                records_container = tk.Frame(stats_frame, bg=card_bg, padx=5, pady=5)
                records_container.pack(side=tk.LEFT, fill=tk.X, expand=True)
                
                records_label = tk.Label(records_container, text="PENDING", 
                                    font=("Segoe UI", 8), bg=card_bg, fg=neutral_color)
                records_label.pack(anchor="w")
                
                records_count = tk.Label(records_container, text="0", 
                                    font=("Segoe UI", 16, "bold"), bg=card_bg, fg=text_color)
                records_count.pack(anchor="w")
                
                # Completed counter with pill shape
                completed_container = tk.Frame(stats_frame, bg=card_bg, padx=5, pady=5)
                completed_container.pack(side=tk.RIGHT, fill=tk.X, expand=True)
                
                completed_label = tk.Label(completed_container, text="SYNCED", 
                                        font=("Segoe UI", 8), bg=card_bg, fg=neutral_color)
                completed_label.pack(anchor="e")
                
                completed_count = tk.Label(completed_container, text="0", 
                                        font=("Segoe UI", 16, "bold"), bg=card_bg, fg=success_color)
                completed_count.pack(anchor="e")
                
                # Log card with modern design
                log_card = tk.Frame(right_panel, bg=card_bg, padx=15, pady=15)
                log_card.pack(fill=tk.BOTH, expand=True)
                
                log_header = tk.Frame(log_card, bg=card_bg)
                log_header.pack(fill=tk.X, pady=(0, 10))
                
                log_icon_label = tk.Label(log_header, text="📊", font=("Segoe UI", 14), 
                                        bg=card_bg, fg=text_color)
                log_icon_label.pack(side=tk.LEFT)
                
                log_title = tk.Label(log_header, text="ACTIVITY LOG", font=("Segoe UI", 12, "bold"), 
                                bg=card_bg, fg=text_color)
                log_title.pack(side=tk.LEFT, padx=(5, 0))
                
                # Create modern text widget with custom styling for logs
                log_container = tk.Frame(log_card, bg=card_bg)
                log_container.pack(fill=tk.BOTH, expand=True)
                
                # Custom styled Text widget
                log_text = tk.Text(log_container, font=("Consolas", 9),
                                bg="#0d0e10", fg="#e0e0e0", 
                                relief=tk.FLAT, padx=10, pady=10,
                                insertbackground=accent_color,
                                wrap=tk.WORD, height=15)
                log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
                
                # Modern scrollbar
                scrollbar_frame = tk.Frame(log_container, width=8, bg=card_bg)
                scrollbar_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=(5, 0))
                
                # Custom scrollbar coloring
                style.configure("Modern.Vertical.TScrollbar", 
                            background=accent_color, 
                            troughcolor="#1e1f23",
                            borderwidth=0,
                            arrowsize=0)
                
                scrollbar = ttk.Scrollbar(scrollbar_frame, command=log_text.yview, 
                                        style="Modern.Vertical.TScrollbar")
                scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
                log_text.config(yscrollcommand=scrollbar.set)
                
                # Add futuristic line indicators on the left side of the log
                log_text.tag_configure("left_margin", lmargin1=20, lmargin2=20)
                
                # Footer with animated dots for "processing" indicator
                footer_frame = tk.Frame(progress_window, bg=bg_color, height=40)
                footer_frame.pack(fill=tk.X, padx=30, pady=(5, 20))
                
                # Sync time estimator
                time_label = tk.Label(footer_frame, text="00:00", font=("Segoe UI", 10), 
                                bg=bg_color, fg=neutral_color)
                time_label.pack(side=tk.RIGHT)
                
                sync_start_time = time.time()
                
                def update_timer():
                    if not progress_window.winfo_exists():
                        return
                        
                    elapsed = int(time.time() - sync_start_time)
                    mins = elapsed // 60
                    secs = elapsed % 60
                    time_label.config(text=f"⏱️ {mins:02d}:{secs:02d}")
                    progress_window.after(1000, update_timer)
                
                # Start timer
                progress_window.after(1000, update_timer)
                
                # Initialize counters
                completed_uploads = 0
                total_pending = 0
                
                # Processing indicator with animated dots
                processing_label = tk.Label(footer_frame, text="PROCESSING", font=("Segoe UI", 10), 
                                        bg=bg_color, fg=accent_color)
                processing_label.pack(side=tk.LEFT)
                
                # Animated dots function
                def animate_dots():
                    if not progress_window.winfo_exists():
                        return
                    
                    current_text = processing_label.cget("text")
                    base_text = "PROCESSING"
                    
                    # Get current dot count
                    dot_count = current_text.count(".")
                    
                    # Update dots (cycle through 0-3 dots)
                    new_dot_count = (dot_count + 1) % 4
                    new_text = base_text + "." * new_dot_count
                    
                    # Update label
                    processing_label.config(text=new_text)
                    
                    # Schedule next animation frame
                    progress_window.after(500, animate_dots)
                
                # Start dot animation
                progress_window.after(500, animate_dots)
                
                # Configure text tags for styling the log with modern colors
                log_text.tag_configure("error", foreground=error_color)
                log_text.tag_configure("info", foreground=accent_color)
                log_text.tag_configure("success", foreground=success_color, font=("Consolas", 9, "bold"))
                log_text.tag_configure("warning", foreground=warning_color)
                log_text.tag_configure("timestamp", foreground="#77767b")
                
                # Create custom logging handler with animated entries
                class WindowLogHandler(logging.Handler):
                    def emit(self, record):
                        msg = self.format(record)
                        # Use after() to ensure thread safety
                        progress_window.after(0, update_log, msg, record.levelname)
                
                # Add the custom handler to the logger
                window_handler = WindowLogHandler()
                window_handler.setFormatter(logging.Formatter('%(message)s'))
                logging.getLogger().addHandler(window_handler)

                def update_log(message, message_type="INFO"):
                    nonlocal completed_uploads
                    timestamp = datetime.now().strftime("%H:%M:%S")
                    
                    # Check if window still exists
                    if not progress_window.winfo_exists():
                        return
                    
                    # Format based on message type
                    if "Found" in message and "pending" in message:
                        try:
                            count = message.split()[1]
                            total_pending_records = int(count)
                            records_count.config(text=f"{total_pending_records}")
                            formatted_msg = f"{timestamp} │ 📊 Found {count} records to sync"
                            tag_to_use = "info"
                        except:
                            formatted_msg = f"{timestamp} │ 📊 {message}"
                            tag_to_use = "info"
                    
                    elif "Successfully synced record" in message or "Synced:" in message:
                        # Extract username and date
                        parts = message.split('_')
                        if len(parts) >= 4:
                            date = f"{parts[-3]} {parts[-2]} {parts[-1].split()[0]}"
                            username = parts[0].split()[-1]
                            formatted_msg = f"{timestamp} │ ✓ Synced record for {username} - {date}"
                        else:
                            formatted_msg = f"{timestamp} │ ✓ {message}"
                        tag_to_use = "success"
                    
                    elif "Uploaded" in message and "photo" in message:
                        completed_uploads += 1
                        completed_count.config(text=f"{completed_uploads}")
                        
                        # Format S3 upload messages
                        if "s3://" in message:
                            parts = message.split('/')
                            username = parts[-4] if len(parts) >= 4 else "Unknown"
                            date_parts = parts[-1].split('_')
                            date = f"{date_parts[0]} {date_parts[1]} {date_parts[2]}" if len(date_parts) >= 3 else "Unknown"
                            formatted_msg = f"{timestamp} │ 📤 Uploaded photo for {username} - {date}"
                        else:
                            formatted_msg = f"{timestamp} │ 📤 {message}"
                        tag_to_use = "info"
                    
                    elif "Error" in message:
                        formatted_msg = f"{timestamp} │ ❌ {message}"
                        tag_to_use = "error"
                    
                    elif "Starting sync" in message:
                        current_action.config(text="Initializing sync process...")
                        formatted_msg = f"{timestamp} │ 🔄 Starting sync process"
                        tag_to_use = "info"
                    
                    elif "Syncing last 7 days" in message:
                        current_action.config(text="Syncing attendance records...")
                        formatted_msg = f"{timestamp} │ 📅 Syncing last 7 days of data"
                        tag_to_use = "info"
                    
                    elif "Sync process completed" in message:
                        current_action.config(text="Sync completed successfully!")
                        formatted_msg = f"{timestamp} │ ✅ Sync process completed successfully!"
                        tag_to_use = "success"
                        
                        # Set progress to 100%
                        self.target_progress = 100
                        processing_label.config(text="COMPLETED", fg=success_color)
                        
                    else:
                        formatted_msg = f"{timestamp} │ ℹ️ {message}"
                        tag_to_use = "neutral"

                    # Insert with visual animation effect
                    log_text.insert(tk.END, formatted_msg + "\n", (tag_to_use, "left_margin"))
                    log_text.see(tk.END)
                    
                    # Animate new log entry with a subtle highlight effect
                    def highlight_new_entry(iteration=5):
                        if iteration > 0 and progress_window.winfo_exists():
                            line_start = log_text.index(f"end-2c linestart")
                            line_end = log_text.index(f"end-1c")
                            
                            # Toggle background highlight
                            if iteration % 2 == 1:
                                log_text.tag_add("highlight", line_start, line_end)
                            else:
                                log_text.tag_remove("highlight", line_start, line_end)
                                
                            progress_window.after(100, lambda: highlight_new_entry(iteration - 1))
                    
                    # Configure highlight tag
                    highlight_color = "#1e3a66" if tag_to_use != "error" else "#661e1e"
                    log_text.tag_configure("highlight", background=highlight_color)
                    
                    # Start highlight animation for new entry
                    highlight_new_entry()
                    
                    # Update progress based on log messages
                    if total_pending > 0:
                        # Calculate approximate progress
                        if "Starting sync" in message:
                            self.target_progress = 5
                        elif "Found" in message and "pending" in message:
                            self.target_progress = 10
                        elif completed_uploads > 0:
                            # Progress based on completed uploads
                            progress_pct = min(95, 10 + (completed_uploads / max(1, total_pending)) * 85)
                            self.target_progress = progress_pct
                        elif "Sync process completed" in message:
                            self.target_progress = 100
                    
                    # Update progress display if window exists
                    if progress_window.winfo_exists():
                        draw_progress(progress_canvas, self.progress_value)

                def perform_sync():
                    try:
                        update_log("Starting sync process...")
                        
                        # Calculate date 7 days ago
                        seven_days_ago = (datetime.now() - timedelta(days=7)).date()
                        
                        # Get local connection in the same thread
                        local_conn = sqlite3.connect(self.db_pool.local_db_path)
                        local_cur = local_conn.cursor()
                        
                        # Find records from last 7 days with local files but no S3 URLs
                        local_cur.execute("""
                            SELECT ar.record_id, ar.username, ar.video_path_in, ar.video_path_out
                            FROM attendance_records ar
                            WHERE date(ar.date) >= date(?)
                            AND ((ar.video_path_in IS NOT NULL AND ar.video_path_in NOT LIKE 's3://%')
                            OR (ar.video_path_out IS NOT NULL AND ar.video_path_out NOT LIKE 's3://%'))
                        """, (seven_days_ago.isoformat(),))
                        
                        pending_records = local_cur.fetchall()
                        nonlocal total_pending
                        total_pending = len(pending_records)
                        
                        if pending_records:
                            update_log(f"Found {len(pending_records)} pending photo uploads...")
                            
                            aws_conn = self.db_pool.get_aws_connection()
                            aws_cur = aws_conn.cursor()
                            
                            for i, (record_id, username, in_path, out_path) in enumerate(pending_records):
                                try:
                                    # Update status with progress
                                    current_action.config(text=f"Uploading data for {username}...")
                                    
                                    # Upload in_time photo if exists
                                    if in_path and not in_path.startswith('s3://'):
                                        if os.path.exists(in_path):
                                            s3_url = self.db_pool.upload_video_to_s3(in_path, username, "IN_TIME")
                                            if s3_url:
                                                update_log(f"Uploaded IN_TIME photo: {s3_url}")
                                    
                                    # Upload out_time photo if exists
                                    if out_path and not out_path.startswith('s3://'):
                                        if os.path.exists(out_path):
                                            s3_url = self.db_pool.upload_video_to_s3(out_path, username, "OUT_TIME")
                                            if s3_url:
                                                update_log(f"Uploaded OUT_TIME photo: {s3_url}")
                                    
                                except Exception as e:
                                    update_log(f"Error uploading photos for record {record_id}: {e}")
                                    continue
                            
                            self.db_pool.return_aws_connection(aws_conn)
                        
                        local_conn.close()
                        
                        update_log("\nSyncing recent pending data...")
                        current_action.config(text="Running final data synchronization...")
                        
                        # Do regular sync only once - for last 7 days
                        if not hasattr(self, '_sync_done'):
                            update_log("Syncing last 7 days data before exit...")
                            self.network_manager._sync_all_pending_data(days_filter=7)
                            self._sync_done = True

                        update_log("\nSync process completed!")
                        
                        # Create animated success checkmark
                        def draw_checkmark():
                            # Clear canvas
                            progress_canvas.delete("checkmark")
                            
                            # Draw checkmark
                            progress_canvas.create_line(50, 90, 80, 120, width=10, 
                                                    fill=success_color, tags="checkmark",
                                                    capstyle=tk.ROUND, joinstyle=tk.ROUND)
                            progress_canvas.create_line(80, 120, 130, 60, width=10, 
                                                    fill=success_color, tags="checkmark",
                                                    capstyle=tk.ROUND, joinstyle=tk.ROUND)
                        
                        # Show success animation after progress hits 100%
                        def check_for_completion():
                            if self.progress_value >= 99:
                                draw_checkmark()
                            elif progress_window.winfo_exists():
                                progress_window.after(100, check_for_completion)
                        
                        # Start checking for completion
                        check_for_completion()
                        
                        # Show completed status for a moment before closing
                        # Create animated fade out
                        def fade_out(alpha=1.0, self=self):
                            if alpha <= 0:
                                # Clean up and close
                                try:
                                    logging.getLogger().removeHandler(window_handler)
                                except:
                                    pass
                                
                                try:
                                    progress_window.destroy()
                                except:
                                    pass
                                
                                try:
                                    self.cleanup()
                                except Exception as e:
                                    logging.error(f"Error during cleanup: {e}")
                                
                                # Add a check before destroying main_window
                                try:
                                    if hasattr(self, 'main_window') and self.main_window is not None:
                                        # Check if the main_window still exists using a safer approach
                                        self.main_window.destroy()
                                except Exception as e:
                                    logging.error(f"Error destroying main window: {e}")
                            else:
                                try:
                                    # Check if the progress window still exists
                                    if progress_window.winfo_exists():
                                        progress_window.attributes("-alpha", alpha)
                                        progress_window.after(50, lambda: fade_out(alpha - 0.05, self))
                                    else:
                                        # Window was already destroyed, clean up
                                        try:
                                            self.cleanup()
                                        except:
                                            pass
                                except Exception as e:
                                    # If any error occurs, just clean up and exit
                                    logging.error(f"Error during fade out: {e}")
                                    try:
                                        logging.getLogger().removeHandler(window_handler)
                                    except:
                                        pass
                                    try:
                                        progress_window.destroy()
                                    except:
                                        pass
                                    try:
                                        self.cleanup()
                                    except:
                                        pass
                        
                        # Start fadeout after delay
                        progress_window.after(3000, lambda: fade_out(1.0, self))

                    except Exception as e:
                        update_log(f"Error during sync: {e}")
                        current_action.config(text="Sync failed! Closing...")
                        
                        # Show error indication
                        self.target_progress = 100
                        progress_canvas.itemconfig("progress", fill=error_color if "progress" in progress_canvas.find_all() else "")
                        processing_label.config(text="FAILED", fg=error_color)
                        
                        # Still destroy window after error with each statement on a separate line
                        progress_window.after(3000, lambda: [
                            logging.getLogger().removeHandler(window_handler),
                            progress_window.destroy(),
                            self.cleanup(),
                            self.main_window.destroy() if hasattr(self, 'main_window') and 
                            self.main_window.winfo_exists() else None
                        ])

                # Apply some visual effects when window opens
                progress_window.attributes("-alpha", 0.0)
                
                def fade_in(alpha=0.0, self=self):
                    alpha += 0.1
                    try:
                        progress_window.attributes("-alpha", min(alpha, 0.97))
                        if alpha < 0.97:
                            progress_window.after(20, lambda: fade_in(alpha, self))
                        else:
                            # Start sync in a separate thread once window is visible
                            sync_thread = threading.Thread(target=perform_sync)
                            sync_thread.daemon = True
                            sync_thread.start()
                    except:
                        # If transparency not supported, just show window and start sync
                        sync_thread = threading.Thread(target=perform_sync)
                        sync_thread.daemon = True
                        sync_thread.start()
                
                # Start fade-in animation
                progress_window.after(10, lambda: fade_in(0.0, self))

         
                
            else:
                self.cleanup()
                if hasattr(self, 'main_window') and self.main_window:
                    try:
                        self.main_window.destroy()
                    except:
                        pass
                    
         
        except Exception as e:
            logging.error(f"Error during application shutdown: {e}")
        finally:
            # Ensure cleanup happens regardless of errors
            try:
                if not hasattr(self, '_cleanup_called'):
                    self.cleanup()
            except Exception as e:
                logging.error(f"Error during final cleanup: {e}")


            
    def cleanup(self):
        """FIXED: Enhanced cleanup that doesn't cause destroy errors"""
        try:
            # Set flag to prevent multiple cleanup calls
            if hasattr(self, '_cleanup_called'):
                logging.info("Cleanup already called, skipping")
                return
            self._cleanup_called = True
            
            logging.info("Starting application cleanup...")
            
            # Stop network monitoring FIRST (most important)
            if hasattr(self, 'network_manager') and self.network_manager:
                try:
                    self.network_manager.cleanup()
                    logging.info("Network manager cleaned up successfully")
                except Exception as e:
                    logging.error(f"Error cleaning up network manager: {e}")
            
            # Cleanup instance manager
            if hasattr(self, 'instance_manager'):
                try:
                    self.instance_manager.cleanup()
                    logging.info("Instance manager cleaned up")
                except Exception as e:
                    logging.error(f"Error cleaning up instance manager: {e}")
            
            # Clean up database connections LAST
            if hasattr(self, 'db_pool') and self.db_pool:
                try:
                    self.db_pool.cleanup()
                    logging.info("Database pool cleaned up")
                except Exception as e:
                    logging.error(f"Error cleaning up database pool: {e}")
            
            logging.info("Application cleanup completed successfully")
            
        except Exception as e:
            logging.error(f"Error during application cleanup: {e}")
            
            
            
            
    def upload_ai_model(self):
        try:
            # Create the deepface weights directory in user's home
            home_dir = os.path.expanduser('~')
            deepface_dir = os.path.join(home_dir, '.deepface')
            weights_dir = os.path.join(deepface_dir, 'weights')
            facenet_dir = os.path.join(weights_dir, 'facenet512')
            
            os.makedirs(weights_dir, exist_ok=True)
            os.makedirs(facenet_dir, exist_ok=True)
            
            # Open file dialog for selecting .h5 file
            file_path = tk.filedialog.askopenfilename(
                title='Select facenet512_weights.h5 file',
                filetypes=[('H5 files', '*.h5')]
            )
            
            if file_path:
                # Define target paths
                target_paths = [
                    os.path.join(weights_dir, 'facenet512_weights.h5'),
                    os.path.join(facenet_dir, 'facenet512_weights.h5')
                ]
                
                # Copy file to both locations
                for target_path in target_paths:
                    shutil.copy2(file_path, target_path)
                
                messagebox.showinfo("Success", 
                    "AI Model file copied successfully to required locations!\n"
                    "Please restart the application.")
                
                # Exit application
                self.main_window.destroy()
                os._exit(0)
                
        except Exception as e:
            logging.error(f"Error uploading AI model: {e}")
            messagebox.showerror("Error", f"Failed to upload AI model: {str(e)}")
            
            
            
    def open_customized_report_window(self):
        report_window = tk.Toplevel(self.main_window)
        report_window.title("Customized Attendance Report")
        report_window.geometry("400x700")
        
        # Create main frame
        main_frame = tk.Frame(report_window, padx=20, pady=20)
        main_frame.pack(expand=True, fill='both')
        
        # Get all usernames from database
        users = self.get_all_usernames()
        users.insert(0, "All Users")
        
        # Username dropdown
        tk.Label(main_frame, text="Select User:", font=('Arial', 10, 'bold')).pack(pady=5)
        username_var = tk.StringVar(value=users[0])
        username_dropdown = ttk.Combobox(main_frame, textvariable=username_var, values=users)
        username_dropdown.pack(pady=5)
        
        # Year dropdown (2025-2035)
        tk.Label(main_frame, text="Select Year:", font=('Arial', 10, 'bold')).pack(pady=5)
        years = list(range(2025, 2036))
        year_var = tk.StringVar(value="2025")
        year_dropdown = ttk.Combobox(main_frame, textvariable=year_var, values=years)
        year_dropdown.pack(pady=5)
        
        # Month dropdown
        tk.Label(main_frame, text="Select Month:", font=('Arial', 10, 'bold')).pack(pady=5)
        months = list(calendar.month_name)[1:]
        month_var = tk.StringVar(value=months[0])
        month_dropdown = ttk.Combobox(main_frame, textvariable=month_var, values=months)
        month_dropdown.pack(pady=5)
        
        # From date dropdown
        tk.Label(main_frame, text="From Date:", font=('Arial', 10, 'bold')).pack(pady=5)
        dates = list(range(1, 32))
        from_date_var = tk.StringVar(value="1")
        from_date_dropdown = ttk.Combobox(main_frame, textvariable=from_date_var, values=dates)
        from_date_dropdown.pack(pady=5)
        
        # To date dropdown
        tk.Label(main_frame, text="To Date:", font=('Arial', 10, 'bold')).pack(pady=5)
        to_date_var = tk.StringVar(value="1")
        to_date_dropdown = ttk.Combobox(main_frame, textvariable=to_date_var, values=dates)
        to_date_dropdown.pack(pady=5)
        
        # Add Office Time Settings Section
        time_frame = tk.LabelFrame(main_frame, text="Office Time Settings", font=('Arial', 10, 'bold'), pady=10)
        time_frame.pack(fill='x', pady=10)
        
        # In Time Settings
        in_time_frame = tk.Frame(time_frame)
        in_time_frame.pack(fill='x', pady=5)
        tk.Label(in_time_frame, text="Set In Time:", font=('Arial', 9, 'bold')).pack(side='left', padx=5)
        
        # Hour dropdown for in time
        hours = [str(i).zfill(2) for i in range(1, 13)]
        in_hour_var = tk.StringVar()
        in_hour_dropdown = ttk.Combobox(in_time_frame, textvariable=in_hour_var, values=hours, width=5)
        in_hour_dropdown.set("Hour")
        in_hour_dropdown.pack(side='left', padx=2)
        
        # Minute dropdown for in time
        minutes = [str(i).zfill(2) for i in range(60)]
        in_minute_var = tk.StringVar()
        in_minute_dropdown = ttk.Combobox(in_time_frame, textvariable=in_minute_var, values=minutes, width=5)
        in_minute_dropdown.set("Min")
        in_minute_dropdown.pack(side='left', padx=2)
        
        # AM/PM dropdown for in time
        ampm = ["AM", "PM"]
        in_ampm_var = tk.StringVar()
        in_ampm_dropdown = ttk.Combobox(in_time_frame, textvariable=in_ampm_var, values=ampm, width=5)
        in_ampm_dropdown.set("AM/PM")
        in_ampm_dropdown.pack(side='left', padx=2)
        
        # Out Time Settings
        out_time_frame = tk.Frame(time_frame)
        out_time_frame.pack(fill='x', pady=5)
        tk.Label(out_time_frame, text="Set Out Time:", font=('Arial', 9, 'bold')).pack(side='left', padx=5)
        
        # Hour dropdown for out time
        out_hour_var = tk.StringVar()
        out_hour_dropdown = ttk.Combobox(out_time_frame, textvariable=out_hour_var, values=hours, width=5)
        out_hour_dropdown.set("Hour")
        out_hour_dropdown.pack(side='left', padx=2)
        
        # Minute dropdown for out time
        out_minute_var = tk.StringVar()
        out_minute_dropdown = ttk.Combobox(out_time_frame, textvariable=out_minute_var, values=minutes, width=5)
        out_minute_dropdown.set("Min")
        out_minute_dropdown.pack(side='left', padx=2)
        
        # AM/PM dropdown for out time
        out_ampm_var = tk.StringVar()
        out_ampm_dropdown = ttk.Combobox(out_time_frame, textvariable=out_ampm_var, values=ampm, width=5)
        out_ampm_dropdown.set("AM/PM")
        out_ampm_dropdown.pack(side='left', padx=2)
        
        def generate_custom_report():
            try:
                # Get selected values
                username = username_var.get()
                year = int(year_var.get())
                month = month_var.get()
                from_date = int(from_date_var.get())
                to_date = int(to_date_var.get())
                
                # Validate time inputs
                if (in_hour_var.get() in ["", "Hour"] or 
                    in_minute_var.get() in ["", "Min"] or 
                    in_ampm_var.get() in ["", "AM/PM"] or
                    out_hour_var.get() in ["", "Hour"] or 
                    out_minute_var.get() in ["", "Min"] or 
                    out_ampm_var.get() in ["", "AM/PM"]):
                    messagebox.showerror("Error", "Please set both In Time and Out Time")
                    return
                    
                # Get time settings
                try:
                    in_time = datetime.strptime(
                        f"{in_hour_var.get()}:{in_minute_var.get()} {in_ampm_var.get()}", 
                        "%I:%M %p"
                    ).time()
                    
                    out_time = datetime.strptime(
                        f"{out_hour_var.get()}:{out_minute_var.get()} {out_ampm_var.get()}", 
                        "%I:%M %p"
                    ).time()
                except ValueError:
                    messagebox.showerror("Error", "Invalid time format")
                    return
                
                # Validate dates
                if from_date > to_date:
                    messagebox.showerror("Error", "From date cannot be greater than To date")
                    return
                
                # Show progress window
                progress_window = tk.Toplevel(report_window)
                progress_window.title("Generating Report")
                progress_window.geometry("300x150")
                
                progress_label = tk.Label(progress_window, text="Generating report...")
                progress_label.pack(pady=20)
                
                # Start report generation in separate thread with custom time settings
                thread = threading.Thread(
                    target=self.generate_customized_report,
                    args=(username, year, month, from_date, to_date, progress_window, in_time, out_time)
                )
                thread.start()
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to generate report: {str(e)}")
        
        # Generate Report Button
        tk.Button(
            main_frame,
            text="Generate Report",
            command=generate_custom_report,
            width=20,
            height=2
        ).pack(pady=20)
        
        
    def generate_customized_report(self, username, year, month, from_date, to_date, progress_window, in_time, out_time):
        try:
            dhaka_tz = pytz.timezone('Asia/Dhaka')
            
            # Convert month name to number
            month_num = list(calendar.month_name).index(month)
            if month_num == 0:
                raise ValueError("Invalid month name")
            
            # Calculate working days in the selected date range
            start_date = datetime(year, month_num, from_date)
            end_date = datetime(year, month_num, to_date)
            
            aws_conn = self.db_pool.get_aws_connection()
            if not aws_conn:
                self.update_progress_label(progress_window, "Could not connect to AWS database")
                return

            try:
                cur = aws_conn.cursor()
                
                # Get working days count
                cur.execute("""
                    SELECT COUNT(DISTINCT date::date) 
                    FROM attendance_records 
                    WHERE date::date BETWEEN %s::date AND %s::date
                    AND (in_time IS NOT NULL OR out_time IS NOT NULL)
                """, (start_date.date(), end_date.date()))
                
                working_days = cur.fetchone()[0]
                total_office_hours = working_days * 8  # 8 hours per working day
                
                # Get users based on selection
                if username == "All Users":
                    cur.execute("SELECT company_user_uuid, username FROM users ORDER BY username")
                else:
                    cur.execute("SELECT company_user_uuid, username FROM users WHERE username = %s", (username,))
                users = cur.fetchall()
                
                if not users:
                    self.update_progress_label(progress_window, "No users found")
                    return

                wb = Workbook()
                ws = wb.active
                ws.title = f"Attendance {month} {year} ({from_date}-{to_date})"

                # Write headers
                headers = ['User ID', 'Username']
                for day in range(from_date, to_date + 1):
                    headers.extend([
                        f"{day} - In Time",
                        f"{day} - Out Time",
                        f"{day} - Camera In",
                        f"{day} - Camera Out",
                        f"{day} - Late",
                        f"{day} - Early Leave"
                    ])
                # Add summary columns
                headers.extend([
                    'Total Lates',
                    'Total Early Leaves',
                    'Total In Time With Camera',
                    'Total Out Time With Camera',
                    'Working Days in Period',
                    'Total Office Hours Required',
                    'Number of Absent Days',
                    'Actual Office Stay (Hours)'
                ])

                # Write headers
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                    ws.cell(row=1, column=col).font = Font(bold=True)

                for row, (company_user_uuid, username) in enumerate(users, 2):
                    ws.cell(row=row, column=1, value=company_user_uuid)
                    ws.cell(row=row, column=2, value=username)
                    
                    # Get attendance records for date range
                    cur.execute("""
                        SELECT 
                            date::date as attendance_date,
                            CAST(in_time AS timestamp) as in_time,
                            CAST(out_time AS timestamp) as out_time,
                            in_time_with_camera,
                            out_time_with_camera,
                            video_path_in,
                            video_path_out
                        FROM attendance_records 
                        WHERE company_user_uuid = %s 
                        AND date::date BETWEEN %s::date AND %s::date
                        ORDER BY date
                    """, (company_user_uuid, start_date.date(), end_date.date()))
                    
                    attendance_records = cur.fetchall()
                    attendance_dict = {}
                    
                    # Initialize counters
                    total_lates = 0
                    total_early_leaves = 0
                    total_in_camera = 0
                    total_out_camera = 0
                    total_actual_hours = 0
                    present_days = 0
                    
                    for record in attendance_records:
                        date_str = str(record[0])
                        
                        try:
                            in_time_record = parser.parse(str(record[1])) if record[1] else None
                            if in_time_record:
                                in_time_record = pytz.utc.localize(in_time_record).astimezone(dhaka_tz)
                        except (TypeError, ValueError):
                            in_time_record = None

                        try:
                            out_time_record = parser.parse(str(record[2])) if record[2] else None
                            if out_time_record:
                                out_time_record = pytz.utc.localize(out_time_record).astimezone(dhaka_tz)
                        except (TypeError, ValueError):
                            out_time_record = None

                        # Count camera usage
                        if record[3] == 'Y':
                            total_in_camera += 1
                        if record[4] == 'Y':
                            total_out_camera += 1
                            
                        # Calculate actual hours if both in and out times exist
                        if in_time_record and out_time_record:
                            present_days += 1
                            duration = out_time_record - in_time_record
                            total_actual_hours += duration.total_seconds() / 3600

                        attendance_dict[date_str] = (
                            in_time_record, 
                            out_time_record, 
                            record[3], 
                            record[4],
                            record[5],  # video_path_in
                            record[6]   # video_path_out
                        )

                    col = 3  # Start after User ID and Username
                    for day in range(from_date, to_date + 1):
                        date = datetime(year, month_num, day).date()
                        record = attendance_dict.get(str(date))

                        if record:
                            in_time_record, out_time_record, in_camera, out_camera, _, _ = record
                            
                            in_time_str = in_time_record.strftime('%I:%M %p') if in_time_record else '-'
                            out_time_str = out_time_record.strftime('%I:%M %p') if out_time_record else '-'
                            
                            is_late = 'Yes' if (in_time_record and in_time_record.time() > in_time) else 'No'
                            is_early = 'Yes' if (out_time_record and out_time_record.time() < out_time) else 'No'

                            if is_late == 'Yes':
                                total_lates += 1
                            if is_early == 'Yes':
                                total_early_leaves += 1

                            ws.cell(row=row, column=col, value=in_time_str)
                            ws.cell(row=row, column=col + 1, value=out_time_str)
                            ws.cell(row=row, column=col + 2, value=in_camera or 'N')
                            ws.cell(row=row, column=col + 3, value=out_camera or 'N')
                            ws.cell(row=row, column=col + 4, value=is_late)
                            ws.cell(row=row, column=col + 5, value=is_early)
                        else:
                            for i in range(6):
                                ws.cell(row=row, column=col + i, value='-')
                        
                        col += 6

                    # Calculate absent days
                    absent_days = working_days - present_days

                    # Write summary columns
                    summary_col = len(headers) - 7  # Position for summary columns
                    ws.cell(row=row, column=summary_col, value=total_lates)
                    ws.cell(row=row, column=summary_col + 1, value=total_early_leaves)
                    ws.cell(row=row, column=summary_col + 2, value=total_in_camera)
                    ws.cell(row=row, column=summary_col + 3, value=total_out_camera)
                    ws.cell(row=row, column=summary_col + 4, value=working_days)
                    ws.cell(row=row, column=summary_col + 5, value=f"{total_office_hours} hours")
                    ws.cell(row=row, column=summary_col + 6, value=absent_days)
                    ws.cell(row=row, column=summary_col + 7, value=f"{total_actual_hours:.2f} hours")

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column = [cell for cell in column]
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width

                # Save the report
                report_path = os.path.join(
                    os.path.expanduser("~"),
                    "Desktop",
                    f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                )
                
                wb.save(report_path)
                progress_window.destroy()
                messagebox.showinfo("Success", f"Report generated successfully!\nSaved as: {report_path}")

            except Exception as e:
                logging.error(f"Database operation error: {str(e)}")
                self.update_progress_label(progress_window, f"Error: {str(e)}")
                raise

            finally:
                if aws_conn:
                    self.db_pool.return_aws_connection(aws_conn)

        except Exception as e:
            logging.error(f"Error generating customized report: {str(e)}")
            self.update_progress_label(progress_window, f"Error: {str(e)}")

    def sync_record_to_local(self, local_cursor, record):
        """Sync a single record to local database"""
        try:
            local_cursor.execute("""
                INSERT OR REPLACE INTO attendance_records (
                    record_id, username, date, in_time, out_time,
                    video_path_in, video_path_out, latitude_in, longitude_in,
                    latitude_out, longitude_out, device_id_in, device_id_out,
                    created_at, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, record[13:])  # record[13:] contains all fields from record_id onwards
        except Exception as e:
            logging.error(f"Error syncing record to local: {e}")
                
            
    def open_verify_image_window(self):
        verify_window = tk.Toplevel(self.main_window)
        verify_window.title("Verify Image")
        verify_window.geometry("400x500")
        
        # Create main frame
        main_frame = tk.Frame(verify_window, padx=20, pady=20)
        main_frame.pack(expand=True, fill='both')
        
        # Get all usernames from database
        users = self.get_all_usernames()
        users.insert(0, "All Users")  # Add "All Users" option
        
        # Username dropdown
        tk.Label(main_frame, text="Select User:").pack(pady=5)
        username_var = tk.StringVar(value=users[0])
        username_dropdown = ttk.Combobox(main_frame, textvariable=username_var, values=users)
        username_dropdown.pack(pady=5)
        
        # Year dropdown (2025-2035)
        tk.Label(main_frame, text="Select Year:").pack(pady=5)
        years = list(range(2025, 2036))
        year_var = tk.StringVar(value="2025")
        year_dropdown = ttk.Combobox(main_frame, textvariable=year_var, values=years)
        year_dropdown.pack(pady=5)
        
        # Month dropdown
        tk.Label(main_frame, text="Select Month:").pack(pady=5)
        months = list(calendar.month_name)[1:]  # Get month names
        month_var = tk.StringVar(value=months[0])
        month_dropdown = ttk.Combobox(main_frame, textvariable=month_var, values=months)
        month_dropdown.pack(pady=5)
        
        # From date dropdown
        tk.Label(main_frame, text="From Date:").pack(pady=5)
        dates = list(range(1, 32))
        from_date_var = tk.StringVar(value="1")
        from_date_dropdown = ttk.Combobox(main_frame, textvariable=from_date_var, values=dates)
        from_date_dropdown.pack(pady=5)
        
        # To date dropdown
        tk.Label(main_frame, text="To Date:").pack(pady=5)
        to_date_var = tk.StringVar(value="1")
        to_date_dropdown = ttk.Combobox(main_frame, textvariable=to_date_var, values=dates)
        to_date_dropdown.pack(pady=5)
        
        def validate_and_fetch():
            try:
                # Get selected values
                username = username_var.get()
                year = year_var.get()
                month = month_var.get()
                from_date = int(from_date_var.get())
                to_date = int(to_date_var.get())
                
                # Validate dates
                if from_date > to_date:
                    messagebox.showerror("Error", "From date cannot be greater than To date")
                    return
                    
                # Show progress window
                progress_window = tk.Toplevel(verify_window)
                progress_window.title("Download Progress")
                progress_window.geometry("300x150")
                
                progress_label = tk.Label(progress_window, text="Downloading images...")
                progress_label.pack(pady=20)
                
                # Start download in separate thread
                thread = threading.Thread(
                    target=self.download_s3_images,
                    args=(username, int(year), month, from_date, to_date, progress_window)
                )
                thread.start()
                
            except Exception as e:
                messagebox.showerror("Error", f"Failed to process request: {str(e)}")
        
        # OK Button
        tk.Button(
            main_frame,
            text="OK",
            command=validate_and_fetch,
            width=20,
            height=2
        ).pack(pady=20)

    def get_all_usernames(self):
        """Get all usernames from database"""
        try:
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            cur.execute("SELECT DISTINCT username FROM users ORDER BY username")
            usernames = [row[0] for row in cur.fetchall()]
            return usernames
        except Exception as e:
            logging.error(f"Error fetching usernames: {e}")
            return []

    def download_s3_images(self, username, year, month, from_date, to_date, progress_window):
        """Download images from S3 based on criteria"""
        try:
            # Create local directory for downloads
            download_dir = os.path.join(
                "downloaded_images",
                str(year),
                month,
                username if username != "All Users" else "all_users"
            )
            os.makedirs(download_dir, exist_ok=True)
            
            # Update progress
            progress_window.after(0, lambda: self.update_progress_label(
                progress_window, "Connecting to S3..."))
            
            # Get S3 client
            s3_client = self.db_pool.s3_client
            bucket_name = self.db_pool.aws_config.s3_bucket
            
            # List objects in bucket
            prefix = f"{username}/" if username != "All Users" else ""
            response = s3_client.list_objects_v2(
                Bucket=bucket_name,
                Prefix=prefix
            )
            
            if 'Contents' not in response:
                progress_window.after(0, lambda: self.update_progress_label(
                    progress_window, "No images found!"))
                return
                
            # Filter and download matching files
            downloaded = 0
            for obj in response['Contents']:
                key = obj['Key']
                # Parse filename to get date information
                filename = os.path.basename(key)
                try:
                    # Extract date from filename (adjust parsing based on your filename format)
                    file_date = self.parse_filename_date(filename)
                    if file_date:
                        file_year = file_date.year
                        file_month = file_date.strftime("%B")
                        file_day = file_date.day
                        
                        # Check if file matches criteria
                        if (file_year == year and
                            file_month == month and
                            from_date <= file_day <= to_date):
                            
                            # Download file
                            download_path = os.path.join(download_dir, filename)
                            s3_client.download_file(bucket_name, key, download_path)
                            downloaded += 1
                            
                            # Update progress
                            progress_window.after(0, lambda: self.update_progress_label(
                                progress_window, f"Downloaded {downloaded} images..."))
                            
                except Exception as e:
                    logging.error(f"Error processing file {filename}: {e}")
                    continue
                    
            # Show completion message
            final_message = f"Downloaded {downloaded} images to {download_dir}"
            progress_window.after(0, lambda: self.update_progress_label(
                progress_window, final_message))
            
            # Add close button to progress window
            progress_window.after(0, lambda: self.add_close_button(progress_window))
            
        except Exception as e:
            logging.error(f"Error downloading images: {e}")
            progress_window.after(0, lambda: self.update_progress_label(
                progress_window, f"Error: {str(e)}"))

    def update_progress_label(self, window, text):
        """Update progress window label"""
        for widget in window.winfo_children():
            if isinstance(widget, tk.Label):
                widget.config(text=text)
                break

    def add_close_button(self, window):
        """Add close button to window"""
        tk.Button(
            window,
            text="Close",
            command=window.destroy
        ).pack(pady=10)

    def parse_filename_date(self, filename):
        """Parse date from filename"""
        try:
            # Adjust this based on your actual filename format
            # Example format: "USERNAME_DD_MMM_YYYY_TYPE_HH_MM_AM/PM.jpg"
            parts = filename.split('_')
            date_str = f"{parts[1]} {parts[2]} {parts[3]}"  # DD MMM YYYY
            return datetime.strptime(date_str, "%d %b %Y")
        except Exception as e:
            logging.error(f"Error parsing date from filename {filename}: {e}")
            return None
        
        
    def sync_users_to_local(self):
        """Sync users from AWS to local database, skipping existing users"""
        try:
            if not self.db_pool.is_online:
                messagebox.showinfo("Error", "No internet connection. Cannot sync with AWS.")
                return

            # Get AWS connection
            aws_conn = self.db_pool.get_aws_connection()
            if not aws_conn:
                messagebox.showinfo("Error", "Cannot connect to AWS database")
                return

            try:
                # Get local connection
                local_conn = self.db_pool.get_local_connection()
                local_cur = local_conn.cursor()
                aws_cur = aws_conn.cursor()

                # Get existing local users
                local_cur.execute("SELECT username FROM users")
                existing_users = {row[0] for row in local_cur.fetchall()}

                # Get all AWS users
                aws_cur.execute("""
                    SELECT company_user_uuid, username, face_encoding, created_at, last_updated 
                    FROM users 
                    ORDER BY username
                """)
                aws_users = aws_cur.fetchall()

                # Initialize counters
                synced_count = 0
                skipped_count = 0
                
                # Create progress window
                progress_window = tk.Toplevel(self.main_window)
                progress_window.title("Sync Progress")
                progress_window.geometry("300x150")
                
                progress_label = tk.Label(progress_window, text="Syncing users...", pady=10)
                progress_label.pack()
                
                count_label = tk.Label(progress_window, text="")
                count_label.pack()

                # Sync each user
                for user in aws_users:
                    company_user_uuid, username, face_encoding, created_at, last_updated = user
                    
                    if username in existing_users:
                        skipped_count += 1
                        count_label.config(
                            text=f"Synced: {synced_count}\nSkipped: {skipped_count}\nProcessing: {username}"
                        )
                        progress_window.update()
                        continue

                    try:
                        local_cur.execute("""
                            INSERT INTO users 
                            (company_user_uuid, username, face_encoding, created_at, last_updated)
                            VALUES (?, ?, ?, ?, ?)
                        """, (company_user_uuid, username, face_encoding, created_at, last_updated))
                        synced_count += 1
                        count_label.config(
                            text=f"Synced: {synced_count}\nSkipped: {skipped_count}\nProcessing: {username}"
                        )
                        progress_window.update()
                    except Exception as e:
                        logging.error(f"Error syncing user {username}: {e}")
                        continue

                local_conn.commit()
                progress_window.destroy()

                # Show final results
                messagebox.showinfo(
                    "Sync Complete", 
                    f"Sync completed!\nNew users synced: {synced_count}\nExisting users skipped: {skipped_count}"
                )

            except Exception as e:
                logging.error(f"Database operation error: {e}")
                messagebox.showerror("Error", f"Failed to sync users: {str(e)}")
                if local_conn:
                    local_conn.rollback()

            finally:
                if aws_conn:
                    self.db_pool.return_aws_connection(aws_conn)

        except Exception as e:
            logging.error(f"Sync error: {e}")
            messagebox.showerror("Error", "Failed to start sync process")
            
            
    
    def create_unique_window(self, parent, title, width=400, height=300):
        """Create a window ensuring only one instance with the given title exists"""
        # Check if window with this title already exists
        if title in self.open_windows and self.open_windows[title].winfo_exists():
            # Window exists, bring it to front
            window = self.open_windows[title]
            window.lift()
            window.focus_force()
            return window
        
        # Create new window
        window = tk.Toplevel(parent)
        window.title(title)
        window.geometry(f"{width}x{height}")
        
        # Store reference
        self.open_windows[title] = window
        
        # Remove from tracking when closed
        def on_close():
            if title in self.open_windows:
                del self.open_windows[title]
            window.destroy()
        
        window.protocol("WM_DELETE_WINDOW", on_close)
        
        # Center and make visible
        self.center_window(window, width, height)
        
        return window


    def center_window(self, window, width=None, height=None):
        """Center a window on the screen and ensure it's visible"""
        if not window or not window.winfo_exists():
            logging.error("Cannot center window: window does not exist.")
            return
    
        try:
                
            if width is None:
                width = window.winfo_width()
            if height is None:
                height = window.winfo_height()
            
            # Get screen dimensions
            screen_width = window.winfo_screenwidth()
            screen_height = window.winfo_screenheight()
            
            # Calculate position
            x = (screen_width - width) // 2
            y = (screen_height - height) // 2
            
            # Set window position
            window.geometry(f"{width}x{height}+{x}+{y}")
            
            # Ensure window is visible
            window.attributes('-topmost', True)
            window.update()
            window.attributes('-topmost', False)
            
            # Bring window to front
            window.lift()
            window.focus_force()

        except Exception as e:
            logging.error(f"Error centering window: {e}")
    
    
    def register_new_user_with_camera(self):
        """Initial registration window for collecting user information"""
        try:
            # First check if company is selected
            local_conn = self.db_pool.get_local_connection()
            local_cur = local_conn.cursor()
            
            # Check if company exists in local database
            local_cur.execute("SELECT company_name, company_id FROM company_info")
            company_data = local_cur.fetchone()
            
            if not company_data:
                messagebox.showinfo("Company Required", "Please select a company first!")
                self.manage_company_info()
                return
                
            company_name, company_id = company_data
            
            # Create registration window
            register_window = tk.Toplevel()
            register_window.title("Register New User")
            register_window.geometry("500x600")
            register_window.grab_set()
            
            # CRITICAL FIX: Keep registration window on top
            register_window.attributes('-topmost', True)
            register_window.lift()
            register_window.focus_force()
            
            # Create main frame with padding
            main_frame = tk.Frame(register_window, padx=30, pady=30)
            main_frame.pack(fill='both', expand=True)
            
            # Title
            title_label = tk.Label(main_frame, text="Register New User", font=("Helvetica", 18, "bold"))
            title_label.pack(pady=(0, 30))
            
            # Username field
            username_frame = tk.Frame(main_frame)
            username_frame.pack(fill='x', pady=10)
            tk.Label(username_frame, text="Username:", font=("Helvetica", 10)).pack(side='left')
            username_var = tk.StringVar()
            username_entry = tk.Entry(username_frame, textvariable=username_var, font=("Helvetica", 10))
            username_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))
            
            # Company name (read-only)
            company_frame = tk.Frame(main_frame)
            company_frame.pack(fill='x', pady=10)
            tk.Label(company_frame, text="Company:", font=("Helvetica", 10)).pack(side='left')
            company_var = tk.StringVar(value=company_name)
            company_entry = tk.Entry(company_frame, textvariable=company_var, state='readonly', font=("Helvetica", 10))
            company_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))
            
            # Designation field
            designation_frame = tk.Frame(main_frame)
            designation_frame.pack(fill='x', pady=10)
            tk.Label(designation_frame, text="Designation:", font=("Helvetica", 10)).pack(side='left')
            designation_var = tk.StringVar()
            designation_entry = tk.Entry(designation_frame, textvariable=designation_var, font=("Helvetica", 10))
            designation_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            # Password field
            password_frame = tk.Frame(main_frame)
            password_frame.pack(fill='x', pady=10)
            tk.Label(password_frame, text="Password:", font=("Helvetica", 10)).pack(side='left')
            password_var = tk.StringVar()
            password_entry = tk.Entry(password_frame, textvariable=password_var, show="*", font=("Helvetica", 10))
            password_entry.pack(side='right', expand=True, fill='x', padx=(20, 0))

            def proceed_with_registration():
                username = username_var.get().strip().upper()
                password = password_var.get().strip()
                designation = designation_var.get().strip()
                
                if not all([username, password, designation]):
                    messagebox.showerror("Error", "All fields are required!")
                    return
                
                # Store credentials for the registration process
                self.face_recognition_system.username = username
                self.face_recognition_system.temp_password = password
                self.face_recognition_system.temp_designation = designation
                self.face_recognition_system.temp_company_name = company_name
                self.face_recognition_system.temp_company_uuid = company_id
                
                # Close registration window
                register_window.destroy()
                
                # CRITICAL FIX: Ensure camera window stays on top
                # Force focus to main window first, then start camera
                self.main_window.focus_force()
                self.main_window.lift()
                
                # Start camera capture with a slight delay to ensure proper window management
                self.main_window.after(100, lambda: self.face_recognition_system.register_new_user(capture_with_camera=True))

            # Register button
            register_button = tk.Button(
                main_frame,
                text="Proceed with Camera Registration",
                command=proceed_with_registration,
                width=30,
                height=2,
                font=("Helvetica", 10, "bold")
            )
            register_button.pack(pady=30)
            
            # Set focus to username entry
            username_entry.focus()
            
        except Exception as e:
            logging.error(f"Error in register_new_user_with_camera: {e}")
            messagebox.showerror("Error", "An error occurred during registration")


        
        
        

    def register_new_user_without_camera(self):
        """Register new user without camera"""
        try:
            # Create popup window for username and password
            register_window = tk.Toplevel()
            register_window.title("Register New User")
            register_window.geometry("300x200")
            
            # Username entry
            tk.Label(register_window, text="Username:").pack(pady=5)
            username_entry = tk.Entry(register_window)
            username_entry.pack(pady=5)
            
            # Password entry
            tk.Label(register_window, text="Password:").pack(pady=5)
            password_entry = tk.Entry(register_window, show="*")  # Hide password
            password_entry.pack(pady=5)

            def handle_registration():
                username = username_entry.get().upper()
                password = password_entry.get()
                
                if not username or not password:
                    messagebox.showerror("Error", "Username and password are required!")
                    return
                    
                # Check if username exists
                if self._is_username_exists(username):
                    # Create popup window asking if existing employee
                    confirm_window = tk.Toplevel()
                    confirm_window.title("Existing Username")
                    confirm_window.geometry("300x150")
                    
                    tk.Label(
                        confirm_window, 
                        text="This username already exists.\nAre you an existing/old employee?",
                        font=('Arial', 10),
                        pady=10
                    ).pack()
                    
                    def handle_yes():
                        confirm_window.destroy()
                        register_window.destroy()
                        # Get existing company_user_uuid and update password
                        conn = self.db_pool.get_local_connection()
                        cur = conn.cursor()
                        cur.execute("SELECT company_user_uuid FROM users WHERE username = ?", (username,))
                        existing_company_user_uuid = cur.fetchone()[0]
                        
                        # Update password in local database
                        cur.execute("""
                            UPDATE users 
                            SET password = ?,
                                last_updated = datetime('now')
                            WHERE company_user_uuid = ?
                        """, (password, existing_company_user_uuid))
                        conn.commit()
                        
                        # Update AWS if online
                        if self.db_pool.is_online:
                            aws_conn = self.db_pool.get_aws_connection()
                            if aws_conn:
                                try:
                                    aws_cur = aws_conn.cursor()
                                    aws_cur.execute("""
                                        UPDATE users 
                                        SET password = %s,
                                            last_updated = CURRENT_TIMESTAMP
                                        WHERE company_user_uuid = %s
                                    """, (password, existing_company_user_uuid))
                                    aws_conn.commit()
                                finally:
                                    self.db_pool.return_aws_connection(aws_conn)
                        
                        messagebox.showinfo("Success", "Password updated successfully!")
                    
                    def handle_no():
                        confirm_window.destroy()
                        # Register as new user
                        register_new_user()
                    
                    tk.Button(confirm_window, text="Yes", command=handle_yes).pack(pady=5)
                    tk.Button(confirm_window, text="No", command=handle_no).pack(pady=5)
                    return
                
                register_new_user()

            def register_new_user():
                username = username_entry.get().upper()
                password = password_entry.get()
                company_user_uuid = str(uuid.uuid4())
                blank_encoding = np.zeros(512)
                serialized_encoding = serialize_face_embedding(blank_encoding)
                
                # Save to local database
                conn = self.db_pool.get_local_connection()
                cur = conn.cursor()
                
                try:
                    cur.execute("""
                        INSERT INTO users 
                        (company_user_uuid, username, password, face_encoding, created_at, last_updated)
                        VALUES (?, ?, ?, ?, datetime('now'), datetime('now'))
                    """, (company_user_uuid, username, password, serialized_encoding))
                    conn.commit()
                    
                    # Save to AWS if online
                    if self.db_pool.is_online:
                        aws_conn = self.db_pool.get_aws_connection()
                        if aws_conn:
                            try:
                                aws_cur = aws_conn.cursor()
                                aws_cur.execute("""
                                    INSERT INTO users 
                                    (company_user_uuid, username, password, face_encoding, created_at, last_updated)
                                    VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP, CURRENT_TIMESTAMP)
                                """, (company_user_uuid, username, password, serialized_encoding))
                                aws_conn.commit()
                            finally:
                                self.db_pool.return_aws_connection(aws_conn)
                    
                    messagebox.showinfo("Success", f"User {username} registered successfully!")
                    register_window.destroy()
                    
                except Exception as e:
                    logging.error(f"Error registering user: {e}")
                    messagebox.showerror("Error", "Failed to register user")
                    
            # Register button
            tk.Button(
                register_window,
                text="Register",
                command=handle_registration,
                width=20,
                height=2
            ).pack(pady=20)
            
        except Exception as e:
            logging.error(f"Error in registration: {e}")
            messagebox.showerror("Error", "Registration failed")
            
            
        
    # ADD this method to check actual internet connectivity
    def _check_internet_connection(self):
        """Check actual internet connectivity, not just AWS/S3 status"""
        try:
            # Test with multiple reliable servers
            test_hosts = [
                ("8.8.8.8", 53),      # Google DNS
                ("1.1.1.1", 53),      # Cloudflare DNS
                ("208.67.222.222", 53) # OpenDNS
            ]
            
            for host, port in test_hosts:
                try:
                    socket.create_connection((host, port), timeout=3)
                    logging.info("Internet connection confirmed")
                    return True
                except (socket.timeout, socket.error):
                    continue
                    
            logging.warning("No internet connection detected")
            return False
            
        except Exception as e:
            logging.error(f"Error checking internet connection: {e}")
            return False


    # REPLACE the existing _is_username_exists method with this FIXED version:
    def _is_username_exists(self, username):
        """Check if username exists in either local or AWS database"""
        try:
            # Convert input username to uppercase for consistency
            username = username.upper()
            
            # Check local database first
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            cur.execute("SELECT 1 FROM users WHERE UPPER(username) = ?", (username,))
            if cur.fetchone():
                logging.info(f"Username {username} found in local database")
                return True

            # Check AWS if we have internet connection (not just if db_pool says online)
            if self._check_internet_connection():
                aws_conn = self.db_pool.get_aws_connection()
                if aws_conn:
                    try:
                        aws_cur = aws_conn.cursor()
                        aws_cur.execute("SELECT 1 FROM users WHERE UPPER(username) = %s", (username,))
                        if aws_cur.fetchone():
                            logging.info(f"Username {username} found in AWS database")
                            return True
                    except Exception as e:
                        logging.error(f"Error checking AWS for username: {e}")
                    finally:
                        self.db_pool.return_aws_connection(aws_conn)
                else:
                    logging.warning("Could not get AWS connection for username check")
            else:
                logging.info("No internet connection - skipping AWS username check")
            
            logging.info(f"Username {username} not found in any database")
            return False
            
        except Exception as e:
            logging.error(f"Error checking username existence: {e}")
            return False




    def serialize_face_embedding(embedding):
        """Convert numpy array to JSON string"""
        return json.dumps(embedding.tolist())


                
    def generate_embeddings_report(self):
        """Wrapper method to call generate_embeddings_report from FaceRecognitionSystem."""
        self.face_recognition_system.generate_embeddings_report()
                
    
  
    def handle_forced_login(self):
        """Handle forced login without face verification"""
        username = simpledialog.askstring("Forced Login", "Enter your username:")
        if not username:
            messagebox.showerror("Error", "Username is required")
            return

        try:
            # Check if user exists
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            cur.execute("SELECT company_user_uuid FROM users WHERE username = ?", (username,))
            result = cur.fetchone()

            if not result:
                messagebox.showerror("Error", "Username not found in database")
                return

            company_user_uuid = result[0]
            current_time = self.time_manager.get_accurate_time()
            login_type = self.db_pool.determine_login_type(username)

            # Save attendance record with blank image since it's forced login
            filepath = self.save_login_video(username, login_type, frame=None)  # Explicitly pass None for frame
            
            if not filepath:
                logging.error("Failed to save blank image for forced login.")

            success, aws_success = self.db_pool.save_login_to_db(
                username,
                login_type,
                current_time, 
                filepath,  # Pass the generated blank image path here
                with_camera=False  # Forced login is always without camera
            )

            if success:
                messagebox.showinfo(
                    "Success",
                    f"{username.upper()}, YOUR {login_type} IS RECORDED AT {current_time.strftime('%I:%M %p')}"
                )
            else:
                messagebox.showerror("Error", "Failed to record attendance")

        except Exception as e:
            logging.error(f"Forced login error: {e}")
            messagebox.showerror("Error", "An error occurred during forced login")
            

    
    
 
 
    
          
        
    def create_gradient(self, canvas, width, height):
        """Draws a vertical gradient on the given canvas."""
        colors = ["#0000FF", "#FF0000"]  # Blue to Red gradient
        limit = height
        (r1, g1, b1) = self.hex_to_rgb(colors[0])
        (r2, g2, b2) = self.hex_to_rgb(colors[1])
        for i in range(limit):
            r = int(r1 + (float(i) / limit) * (r2 - r1))
            g = int(g1 + (float(i) / limit) * (g2 - g1))
            b = int(b1 + (float(i) / limit) * (b2 - b1))
            color = '#%02x%02x%02x' % (r, g, b)
            canvas.create_line(0, i, width, i, fill=color)

    def hex_to_rgb(self, hex_color):
        """Converts a hex color string to RGB tuple."""
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

    def create_gradient_button(self, parent, text, command):
        """Creates a button with a gradient background."""
        # Create a gradient image
        width, height = 200, 50
        gradient_image = Image.new("RGB", (width, height))
        for y in range(height):
            r = int(255 * (1 - y / height))  # Red decreases from top to bottom
            b = int(255 * (y / height))     # Blue increases from top to bottom
            gradient_image.putpixel((0, y), (r, 0, b))  # Apply gradient
            for x in range(1, width):
                gradient_image.putpixel((x, y), (r, 0, b))

        # Convert the gradient image to a Tkinter-compatible format
        gradient_photo = ImageTk.PhotoImage(gradient_image)

        # Create the button with the gradient image
        button = tk.Button(
            parent,
            text=text,
            image=gradient_photo,
            compound="center",
            command=command,
            font=("Arial", 12, "bold"),
            fg="white",  # Text color
            bd=0  # Remove border
        )
        button.image = gradient_photo  # Keep a reference to avoid garbage collection
        button.pack(pady=10)


    
        
        
    def handle_forced_login(self):
        """Handle forced login without face verification"""
        username = simpledialog.askstring("Forced Login", "Enter your username:")
        if not username:
            messagebox.showerror("Error", "Username is required")
            return

        try:
            # Check if user exists
            conn = self.db_pool.get_local_connection()
            cur = conn.cursor()
            cur.execute("SELECT company_user_uuid FROM users WHERE username = ?", (username,))
            result = cur.fetchone()

            if not result:
                messagebox.showerror("Error", "Username not found in database")
                return

            company_user_uuid = result[0]
            current_time = self.time_manager.get_accurate_time()
            login_type = self.db_pool.determine_login_type(username)

            # Save attendance record with blank image since it's forced login
            filepath = self.save_login_video(username, login_type, frame=None)  # Explicitly pass None for frame
            
            if not filepath:
                logging.error("Failed to save blank image for forced login.")

            success, aws_success = self.db_pool.save_login_to_db(
                username,
                login_type,
                current_time, 
                filepath,  # Pass the generated blank image path here
                with_camera=False  # Forced login is always without camera
            )

            if success:
                messagebox.showinfo(
                    "Success",
                    f"{username.upper()}, YOUR {login_type} IS RECORDED AT {current_time.strftime('%I:%M %p')}"
                )
            else:
                messagebox.showerror("Error", "Failed to record attendance")

        except Exception as e:
            logging.error(f"Forced login error: {e}")
            messagebox.showerror("Error", "An error occurred during forced login")
            

    
    
    
    
    
                

    def handle_login(self):
        """Handle login button click"""
        username = simpledialog.askstring("Login", "Enter your username:")
        if username:
            self.username = username
            self.face_recognition_system.username = username  # Set username in face recognition system
            self.face_recognition_system.verify_identity()
        else:
            messagebox.showerror("Error", "Username is required")

    def verify_identity(self):
        """Verify user identity"""
        # First make sure to load the latest FAISS index
        self.face_recognition_system.load_faiss_index()
        self.face_recognition_system.verify_identity()
            

    def create_unique_window(self, parent, title, width=400, height=300):
        """Create a window ensuring only one instance with the given title exists"""
        # Check if window with this title already exists
        if title in self.open_windows and self.open_windows[title].winfo_exists():
            # Window exists, bring it to front
            window = self.open_windows[title]
            window.lift()
            window.focus_force()
            return window
        
        # Create new window
        window = tk.Toplevel(parent)
        window.title(title)
        window.geometry(f"{width}x{height}")
        
        # Store reference
        self.open_windows[title] = window
        
        # Remove from tracking when closed
        def on_close():
            if title in self.open_windows:
                del self.open_windows[title]
            window.destroy()
        
        window.protocol("WM_DELETE_WINDOW", on_close)
        
        # Center and make visible
        self.center_window(window, width, height)
        
        return window


    def force_sync_attendance_records(self):
        """Forcefully sync attendance records at application startup."""
        try:
            if self.db_pool.is_online:
                logging.info("Forcing attendance records sync at startup...")
                self.network_manager._sync_all_pending_data()
                logging.info("Attendance records sync completed.")
            else:
                logging.warning("Cannot sync attendance records: No internet connection.")
        except Exception as e:
            logging.error(f"Error during forced attendance records sync: {e}")
            messagebox.showerror("Sync Error", "Failed to sync attendance records at startup.")
            
 

        
    def check_sync_status(self, record_id):
        success, message = self.network_manager.verify_sync_status(record_id)
        if success:
            logging.info(f"Record {record_id} is properly synced")
        else:
            logging.error(f"Sync issue for record {record_id}: {message}")
            
            
    def _verify_timezone_config(self):
        """Ensure proper timezone configuration"""
        try:
            if self.db_pool.is_online:
                aws_conn = self.db_pool.get_aws_connection()
                if aws_conn:
                    cur = aws_conn.cursor()
                    cur.execute("SHOW timezone")
                    timezone = cur.fetchone()[0]
                    if timezone != 'UTC':
                        logging.warning(f"AWS database timezone is {timezone}, should be UTC")
                    self.db_pool.return_aws_connection(aws_conn)
        except Exception as e:
            logging.error(f"Failed to verify timezone configuration: {e}")
            self.countdown_label.config(text="Capture complete!")

    def stop_capture(self):
        self.is_capturing = False

    def verify_time(self):
        try:
            ntp_client = ntplib.NTPClient()
            response = ntp_client.request('pool.ntp.org', version=3)
            current_time = datetime.fromtimestamp(response.tx_time)
            system_time = datetime.now()
            time_difference = abs((current_time - system_time).total_seconds())
            if time_difference > 60:  # Time difference exceeds 1 minute
                messagebox.showinfo('TIME ERROR', 'TIME IS NOT UPDATED. PLEASE UPDATE TIME.')
                return False
            return True
        except:
            return True  # Assume correct time for offline cases

    def login(self):
        if not self.verify_time():
            return

        self.start_capture('login')



    def process_login(self):
        if not self.video_frames:
            messagebox.showinfo('ERROR', 'No video captured. Please try again.')
            return

        try:
            from deepface import DeepFace
            current_time = self.time_manager.get_accurate_time()
            
            
            frame = self.video_frames[-1]  # Use last frame
            obj = DeepFace.represent(
                img_path=frame,
                model_name='Facenet512',
                detector_backend='mtcnn',
                enforce_detection=True,
                align=True
            )
            
            if obj:
                unknown_embedding = obj[0]['embedding']
                
                # Get all users from database
                conn = self.db_pool.get_local_connection()
                cur = conn.cursor()
                cur.execute("SELECT username, face_encoding FROM users")
                
                for username, embedding_json in cur.fetchall():
                    if not embedding_json:
                        continue
                    
                    known_embedding = deserialize_face_embedding(embedding_json)
                    
                    # Compare embeddings using cosine similarity
                    similarity = np.dot(unknown_embedding, known_embedding) / (
                        np.linalg.norm(unknown_embedding) * np.linalg.norm(known_embedding)
                    )
                    
                    if similarity > 0.6:  # Threshold for matching
                        # Determine login type and save attendance
                        login_type = self.db_pool.determine_login_type(username)
                        success, _ = self.db_pool.save_login_to_db(username, login_type, current_time)
                        
                        if success:
                            msg = f"{username.upper()}, YOUR {login_type} IS {current_time.strftime('%I:%M %p')}"
                            messagebox.showinfo('SUCCESS', msg)
                        else:
                            messagebox.showinfo('ERROR', 'Failed to save attendance')
                        return
                        
                messagebox.showinfo('ACCESS DENIED', 'Unknown user')
                
        except Exception as e:
            logging.error(f"Login error: {e}")
            messagebox.showinfo('ERROR', 'Login failed. Please try again.')
            

  



    
    def save_login_video(self, name, login_type, frame=None):
        with self.file_lock:
            try:
                now = datetime.now()
                
                # Use os.path.join for proper path construction
                year_folder = os.path.join(self.db_pool.db_dir, name, str(now.year))
                month_folder = os.path.join(year_folder, now.strftime("%B"))
                
                # Create directories
                os.makedirs(month_folder, exist_ok=True)
                
                # Generate filename using safe characters
                filename = f"{name.upper()}_{now.strftime('%d_%b_%Y')}_{login_type}_{now.strftime('%I_%M_%p')}.jpg"
                filename = filename.replace(':', '_')  # Replace colons with underscores
                
                filepath = os.path.join(month_folder, filename)
                
                # Validate the file path
                validate_filepath(filepath)
                
                
                # If no frame is provided, create a blank image with text annotations
                if frame is None:
                    # Define reduced image size
                    height, width =200, 400
                    
                    # Create a multi-color gradient background
                    gradient = np.zeros((height, width, 3), dtype=np.uint8)
                    for y in range(height):
                        # Calculate color intensities based on the row (y-coordinate)
                        red = int(255 * (1 - y / height))  # Red decreases from top to middle
                        blue = int(255 * (y / height))    # Blue increases from top to bottom
                        black = 0                         # Black remains constant
                        
                        # Apply the gradient colors
                        gradient[y, :, :] = [blue, black, red]  # BGR format (OpenCV uses BGR)
                    
                    # Add text annotations on the gradient background
                    font = cv2.FONT_HERSHEY_SIMPLEX
                    text_lines = [
                        f"Username: {name}",
                        f"Date: {now.strftime('%d %b %Y')}",
                        f"{login_type}: {now.strftime('%I:%M %p')}"
                    ]
                    
                    y_offset = 50
                    for line in text_lines:
                        text_size = cv2.getTextSize(line, font, 1, 2)[0]
                        x_position = (width - text_size[0]) // 2
                        cv2.putText(gradient, line, (x_position, y_offset), 
                                font, 1, (255, 255, 255), 2)  # White text
                        y_offset += 50
                    
                    # Save the gradient image
                    cv2.imwrite(filepath, gradient)
                else:
                    # Save the provided frame
                    cv2.imwrite(filepath, frame)
                
                return filepath
                
            except Exception as e:
                logging.error(f"Error saving login video: {e}")
                return None
            
        
    
    def show_popup(self, message, title="Information"):
        popup_window = tk.Toplevel(self.main_window)
        popup_window.title(title)
        tk.Label(popup_window, text=message, padx=20, pady=20).pack()
        tk.Button(popup_window, text="OK", command=popup_window.destroy).pack(pady=10)

    def try_again_register_new_user(self):
        self.stop_capture()
        self.register_new_user_window.destroy()

        

    def complete_registration(self):
        if not self.video_frames or not self.username:
            messagebox.showinfo('ERROR', 'Registration failed. Please try again.')
            return

        try:
            from deepface import DeepFace
            
            
            # Use the last captured frame for embedding generation
            frame = self.video_frames[-1]
            obj = DeepFace.represent(
                img_path=frame,
                model_name='Facenet512',
                detector_backend='mtcnn',
                enforce_detection=True,
                align=True
            )

            if obj:
                # Extract embedding and serialize it
                embedding = obj[0]['embedding']
                serialized_embedding = serialize_face_embedding(embedding)

                # Save or update user in the database
                success = self.db_pool.save_or_update_user(self.username, serialized_embedding)
                if success:
                    messagebox.showinfo('SUCCESS', 'User registered successfully!')
                else:
                    messagebox.showinfo('ERROR', 'Failed to save user data.')
            else:
                messagebox.showinfo('ERROR', 'No face detected. Please try again.')
        except Exception as e:
            logging.error(f"Registration error: {e}")
            messagebox.showinfo('ERROR', 'Registration failed. Please try again.')
    
  
    def start_network(self):
        try:
            if self.db_pool:
                self.network_manager = NetworkSyncManager(self.db_pool)
                self.network_manager.start_monitoring()
                self.update_status("network", "Online", success=True)
        except Exception as e:
            self.update_status("network", "Offline", success=False)




    def initialize_components(self):
        """Initialize all components asynchronously in parallel"""
        try:
            # Create separate threads for each initialization

            tensorflow_thread = threading.Thread(target=self.load_tensorflow, daemon=True)
            db_thread = threading.Thread(target=self.initialize_database, daemon=True)
            network_thread = threading.Thread(target=self.start_network, daemon=True)
            
            
            # Start all threads simultaneously
            tensorflow_thread.start()
            db_thread.start()
            network_thread.start()

            
            logging.info("All initialization processes started in parallel")
            
        except Exception as e:
            logging.error(f"Error starting initialization threads: {e}")

                
    def login(self):
        # Retrieve the username and pass it to the FaceRecognitionSystem
        if hasattr(self, 'username') and self.username:  # Add username check
            self.face_recognition_system.verify_identity()
        else:
            logging.error("No username available for login")
            messagebox.showerror("Error", "Username not found")

 
 
 
    def setup_global_hotkeys(self):
        """Setup global hotkeys for special functions - FIXED CTRL+H"""
        try:
            # Register Ctrl+Alt+Shift combination for emergency exit
            def emergency_exit(event=None):
                try:
                    # Release camera if exists
                    if hasattr(self, 'cap') and self.cap is not None:
                        self.cap.release()
                    
                    # Destroy all opencv windows
                    cv2.destroyAllWindows()
                    
                    # Force close all tkinter windows without calling on_closing
                    for widget in self.main_window.winfo_children():
                        if isinstance(widget, tk.Toplevel):
                            try:
                                widget.destroy()
                            except:
                                pass
                    
                    # Force exit
                    os._exit(0)
                except:
                    os._exit(0)

            # Register Ctrl+Alt+Shift combination for hiding windows
            def hide_all_windows(event=None):
                try:
                    self.main_window.withdraw()  # Hide main window
                    # Hide all child windows
                    for widget in self.main_window.winfo_children():
                        if isinstance(widget, tk.Toplevel):
                            try:
                                widget.withdraw()
                            except:
                                pass
                except Exception as e:
                    logging.error(f"Error hiding windows: {e}")

            # FIXED: Register Ctrl+H for HR authentication (not direct HR window)
            def show_hr_auth(event=None):
                try:
                    # Call the authentication window instead of direct HR window
                    self.handle_hr_window()  # This opens the password authentication first
                except Exception as e:
                    logging.error(f"Error showing HR authentication: {e}")

            # Bind the hotkeys
            self.main_window.bind('<Control-Alt-Shift-KeyPress>', emergency_exit)
            self.main_window.bind('<Control-Alt-Shift-h>', hide_all_windows)  # 'h' for hide
            
            # FIXED: Ctrl+H now opens authentication window
            self.main_window.bind('<Control-h>', show_hr_auth)  # Ctrl+H for HR authentication
            
            logging.info("Global hotkeys configured: Ctrl+H = HR Authentication")

        except Exception as e:
            logging.error(f"Error setting up global hotkeys: {e}")
 


    def load_tensorflow(self):
        try:
            # Pre-configuration for faster loading with advanced optimizations
            import os
            os.environ['TF_CPP_MIN_LOG_LEVEL'] = '3'  # Suppress all warnings
            os.environ['CUDA_VISIBLE_DEVICES'] = '0'
            os.environ['TF_GPU_THREAD_MODE'] = 'gpu_private'
            os.environ['TF_USE_CUDNN'] = '1'
            os.environ['TF_ENABLE_ONEDNN_OPTS'] = '1'
            os.environ['TF_FORCE_GPU_ALLOW_GROWTH'] = 'true'
            os.environ['TF_FUNCTION_JIT_COMPILE_DEFAULT'] = '1'
            os.environ['TF_XLA_FLAGS'] = '--tf_xla_enable_xla_devices --tf_xla_cpu_global_jit'
            # Additional optimizations
            os.environ['TF_GPU_ALLOCATOR'] = 'cuda_malloc_async'
            os.environ['TF_DISABLE_NVTX_RANGES'] = '1'  # Disable profiling overhead
            
            # Advanced optimizations for faster startup and inference
            os.environ['TF_CUDNN_DETERMINISTIC'] = '0'  # Faster but non-deterministic ops
            os.environ['TF_AUTOTUNE_THRESHOLD'] = '3'   # More aggressive op fusion
            os.environ['TF_MLIR_ENABLE_MERGE_CONTROL_FLOW_PASS'] = '1'  # Better control flow optimization
            os.environ['TF_GPU_CUPTI_FORCE_DISABLED'] = '1'  # Disable profiling tools
            os.environ['CUDA_CACHE_DISABLE'] = '0'  # Enable JIT caching
            os.environ['TF_KERAS_ENABLE_EAGER_CLIENT'] = '1'  # Use C++ eager client 
            os.environ['TF_DISABLE_MKL'] = '0'  # Use MKL optimizations if available
            os.environ['TF_MKL_OPTIMIZE_PRIMITIVIES'] = '1'
            
            # Set fixed seed to improve caching behavior
            os.environ['TF_DETERMINISTIC_OPS'] = '0'  # Allow non-deterministic ops for performance
            os.environ['PYTHONHASHSEED'] = '0'  # Fixed hash seed
            
            # Import necessary modules
            import logging
            logging.info("Fast TensorFlow configuration starting...")
            
            # Import TensorFlow only when needed
            import tensorflow as tf
            
            # IMPORTANT: Do NOT disable eager execution as MTCNN requires it
            # The MTCNN library calls .numpy() on tensors which requires eager execution
            logging.info("Keeping eager execution enabled for MTCNN compatibility")
            
            # Minimal GPU setup
            gpus = tf.config.experimental.list_physical_devices('GPU')
            if gpus:
                try:
                    # More efficient memory growth setting
                    for gpu in gpus:
                        tf.config.experimental.set_memory_growth(gpu, True)
                    
                    # Only set virtual device config if needed
                    if len(gpus) > 0:
                        tf.config.experimental.set_virtual_device_configuration(
                            gpus[0],
                            [tf.config.experimental.VirtualDeviceConfiguration(memory_limit=1024)]
                        )
                    logging.info(f"Configured {len(gpus)} GPUs for TensorFlow")
                except RuntimeError as e:
                    logging.warning(f"GPU configuration warning: {e}")
            else:
                logging.info("No GPUs detected, using CPU mode")
            
            # Advanced thread optimization based on system detection
            import multiprocessing
            cpu_count = multiprocessing.cpu_count()
            
            # Optimize thread allocation based on available CPUs
            if cpu_count >= 8:
                # For multi-core systems: use more intra-op threads for better parallelism
                tf.config.threading.set_inter_op_parallelism_threads(max(2, cpu_count // 4))
                tf.config.threading.set_intra_op_parallelism_threads(max(2, cpu_count // 2))
            else:
                # For lower-end systems: use minimal threading to avoid overhead
                tf.config.threading.set_inter_op_parallelism_threads(1)
                tf.config.threading.set_intra_op_parallelism_threads(2)
                
            logging.info(f"Thread configuration: inter_op={tf.config.threading.get_inter_op_parallelism_threads()}, " 
                        f"intra_op={tf.config.threading.get_intra_op_parallelism_threads()}")
            
            # Enhanced optimizer settings for maximum performance
            tf.config.optimizer.set_jit(True)  # Enable XLA
            tf.config.optimizer.set_experimental_options({
                'layout_optimizer': True,
                'constant_folding': True,
                'debug_stripper': True,
                'auto_mixed_precision': False,  # Disable for faster initialization
                'shape_optimization': True,
                'dependency_optimization': True,
                'loop_optimization': True,
                'function_optimization': True,
                'implementation_selector': True,
                'remapping': True,
                'scoped_allocator_optimization': True,
            })
            
            # Enhanced initialization with memory pre-allocation strategy
            # Create a small computational graph first to warm up TF
            with tf.device('/CPU:0'):
                # Warm up TF runtime with small ops
                _ = tf.zeros([1])
                _ = tf.random.normal([1, 1])
                _ = tf.matmul(tf.ones([1, 1]), tf.ones([1, 1]))
                
            # Pre-allocate some GPU memory with a dummy tensor to avoid fragmentation
            if gpus:
                try:
                    with tf.device('/GPU:0'):
                        # Allocate and immediately release to reserve space and avoid later fragmentation
                        dummy_tensor = tf.random.normal([64, 64])
                        _ = tf.reduce_sum(dummy_tensor).numpy()
                        del dummy_tensor
                        
                        # Run a small graph compile for XLA warmup
                        @tf.function(jit_compile=True)
                        def warmup_func(x):
                            return tf.nn.relu(tf.matmul(x, x))
                        
                        warmup_func(tf.random.normal([10, 10]))
                        logging.info("GPU memory pre-allocation completed")
                except Exception as e:
                    logging.warning(f"GPU memory pre-allocation failed: {e}")
            
            self.is_initialized = True
            logging.info("Fast TensorFlow configuration completed")
            
            logging.info("Starting AI model initialization...")
            
            import deepface
            import sys
            import shutil
            
            # Define all possible model locations in priority order
            model_file = 'facenet512_weights.h5'
            possible_locations = [
                # 1. Original project directory
                os.path.join(os.path.abspath('.'), 'weights', model_file),
                os.path.join(os.path.dirname(os.path.abspath(__file__)), 'weights', model_file),
                # 2. User's home directory
                os.path.join(os.path.expanduser('~'), '.deepface', 'weights', model_file),
                # 3. PyInstaller temp directory (last resort)
                os.path.join(sys._MEIPASS, 'weights', model_file) if hasattr(sys, '_MEIPASS') else None
            ]
            
            # Filter out None values
            possible_locations = [loc for loc in possible_locations if loc]
            
            # Log all possible locations
            for loc in possible_locations:
                logging.info(f"Checking for model at: {loc}")
                if os.path.exists(loc):
                    logging.info(f"Found model at: {loc}")
                    source_path = loc
                    break
            else:
                raise FileNotFoundError(f"Model file not found in any of the expected locations")

            # Setup DeepFace directories
            home_dir = os.path.expanduser('~')
            deepface_dir = os.path.join(home_dir, '.deepface')
            weights_dir = os.path.join(deepface_dir, 'weights')
            facenet_dir = os.path.join(weights_dir, 'facenet512')
            
            # Create directories
            os.makedirs(weights_dir, exist_ok=True)
            os.makedirs(facenet_dir, exist_ok=True)
            
            # Define target paths
            target_paths = [
                os.path.join(weights_dir, model_file),
                os.path.join(facenet_dir, model_file)
            ]
            
            # Copy model to required locations
            for target_path in target_paths:
                if not os.path.exists(target_path):
                    shutil.copy2(source_path, target_path)
                    logging.info(f"Copied model to: {target_path}")
                else:
                    logging.info(f"Model already exists at: {target_path}")
            
            # Initialize DeepFace with optimized loading approach
            from deepface import DeepFace
            import numpy as np
            import time
            
            # Cache model initialization with a timeout strategy
            logging.info("Attempting to load DeepFace models with optimized approach...")
            
            # Create a better dummy image with gradients for more efficient model initialization
            # Using a gradient image instead of zeros helps initialize convolution layers better
            dummy_img = np.zeros((112, 112, 3), dtype=np.uint8)  # Standard input size
            for i in range(112):
                for j in range(112):
                    # Create a gradient pattern
                    dummy_img[i, j, 0] = i * 2  # R channel
                    dummy_img[i, j, 1] = j * 2  # G channel
                    dummy_img[i, j, 2] = (i + j) // 2  # B channel
            
            DeepFace.represent(
                img_path=dummy_img,
                model_name='Facenet512',
                detector_backend='mtcnn',
                enforce_detection=False
            )
            logging.info("DeepFace models loaded successfully")
            
            # Initialize face recognition system
            self.face_recognition_system = FaceRecognitionSystem(
                main_window=self.main_window,
                db_pool=self.db_pool,
                app=self
            )
            
            self.tensorflow_ready = True
            # At the end of load_tensorflow method
            self.on_tensorflow_loaded()
            
            self.update_status("status", "Heavy AI Models: All Models loaded successfully", success=True)
        except Exception as e:
            self.update_status("status", f"Error loading AI models: {str(e)}", success=False)
            import traceback
            logging.error(traceback.format_exc())
            raise
        
        
        
        
if __name__ == "__main__":
    app = None
    instance_manager = None
    
    try:
        # Create instance manager
        instance_manager = SingleInstanceManager()
        
        # Check if another instance is running or recent attendance
        if not instance_manager.check_instance():
            logging.debug("Instance check failed - exiting")
            sys.exit(0)
            
        # Check if high priority argument is passed
        if "--priority" in sys.argv:
            process = psutil.Process()
            process.nice(psutil.HIGH_PRIORITY_CLASS)
            logging.debug("Set high priority mode")
            
        app = App()
        app.instance_manager = instance_manager  # Store reference for cleanup
        
        # Start the application
        app.main_window.mainloop()
        
    except Exception as e:
        logging.error(f"Application startup error: {e}")
        # Show error to user
        import tkinter as tk
        from tkinter import messagebox
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("Error", f"Application failed to start: {str(e)}")
            root.destroy()
        except:
            pass
    finally:
        if 'app' in locals():
            app.cleanup()